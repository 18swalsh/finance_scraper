#02035
from bs4 import BeautifulSoup, Comment
import requests
import pandas as pd
import numpy as np
import itertools
from tabulate import tabulate
import types
import datetime
import time
import xlsxwriter
import csv
from openpyxl import load_workbook
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()


#dictionary to store all data
data_dict = {}

d_valid = True
#---------------------------- Get earnings announcements for a given date ----------------------------------------------
# date = ""
# while d_valid == True:
#     date = str(raw_input("Enter a date or type 'exit' (Format: YYYY-MM-DD): "))
#     try:
#         datetime.datetime.strptime(date, '%Y-%m-%d')
#         d_valid = True
#     except:
#         if date =='exit':
#             print "bye"
#             exit()
#         else:
#             print "Nah" + "\n"

start_time = time.time()
time_one = time.time() #----------------------------------------------------------------------------------------------------------------------

def pretty_print(table_title, table_to_print):
    df = pd.read_html(str(table_to_print))
    print table_title, "\n", (tabulate(df[0], headers='keys', tablefmt='psql'))

def replace_with_newlines(element):
    text = ''
    for elem in element.recursiveChildGenerator():
        if isinstance(elem, types.StringTypes):
            text += elem.strip()
        elif elem.name == 'br':
            text += '\n'
    return text

def removekey(d, key):
    r = dict(d)
    del r[key]
    return r

date = "2017-07-20"

url = "finance.yahoo.com/calendar/earnings?day=" + date
r  = requests.get("https://" + url)
data = r.text
soup = BeautifulSoup(data, "html5lib" )
data_table = soup.find('table', {'class':'data-table'})

headers = []
cells = []
tickers = []

#pretty_print('Earnings announcements on ' + date, data_table)
data_dict['Tickers'] = pd.read_html(str(data_table))

cols = data_table.find_all('td')
#ths = data_table.find_all('th')
#for th in ths[1::]:
#    headers.append(th.get_text())
for col in cols[1::]:
    if col != "-":
        cells.append(col.get_text())
    else:
        cells.append("NULL")

w = ''
spl = [list(y) for x, y in itertools.groupby(cells, lambda z: z == w) if not x]

for line in spl:
    tickers.append(line[0])
print "tickers found"
#print len(tickers)
time_two = time.time() #----------------------------------------------------------------------------------------------------------------------

#-----------------------add more info for known tickers-----------------------------------------------------------------
#https://finance.yahoo.com/quote/TICKER?p=TICKER - stock page

#print "-----------------------------SUMMARY-----------------------------"

ticker_urls = []

#Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "?p=" + ticker)
    #initialize dictionary structure
    data_dict[ticker] = []

x = -1
#Run for each URL
for ticker_url in ticker_urls:
    x += 1
    r  = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    summary_table = soup.find_all('table',['class','W(100%)'])



    #Summary------------------------------------------------------------------------------------------------------------

    #Error if the url is bad
    try:

        cur_price = soup.find('span',['class','Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)'])
        # Filter out stocks under $20
        if float(cur_price.get_text()) < 20:
            #print tickers[x], "was removed from search (price below $20)"
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        # remove companies with a market cap over 10B
        market_cap = 0
        market_caps = summary_table[1].find_all('tr')
        market_cap_string = market_caps[0].get_text()[10:]

        if market_cap_string[-1:] == 'M':
            market_cap = float(market_cap_string[:-1]) * 1000000
        elif market_cap_string[-1:] == 'B':
            market_cap = float(market_cap_string[:-1]) * 1000000000

        if market_cap > 10000000000:
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        #print "Company: " + c_name.get_text()
        #print "Current Price: " + cur_price.get_text()
        #pretty_print("Summary", summary_table)
        data_dict[tickers[x]].extend((c_name.get_text(),cur_price.get_text(),("Summary",pd.read_html(str(summary_table)))))

    except:
        #print "Stock Quote Unavailable"
        #print tickers[x], "removed from search"
        data_dict = removekey(data_dict, tickers[x])
        tickers.remove(tickers[x])
        x -= 1


print "summary complete"
#print len(tickers)

time_three = time.time() #----------------------------------------------------------------------------------------------------------------------

#print "-----------------------------ANALYSTS-----------------------------"
#Analysts------------------------------------------------------------------------------------------------------------

ticker_urls = []

# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/analysts?p=" + ticker)
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        #print c_name.get_text()

        data_dict[tickers[x]].append(c_name.get_text())
        analyst_est_sect = soup.find('section', id='quote-leaf-comp')
        analyst_tables = analyst_est_sect.find_all('table', ['class', 'W(100%)'])
        if len(analyst_tables) == 0:
            #print tickers[x], "was removed (no analyst estimates available)"
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        for a_table in analyst_tables:
            #pretty_print("",a_table)
            data_dict[tickers[x]].extend(pd.read_html(str(a_table)))

            # here is an alternative if the  4 lines above break
            # all_tables = soup.find_all('table', class_='W(100%)')
            # for table in all_tables:
            #     pretty_print("This Table", table

    except:
        data_dict = removekey(data_dict, tickers[x])
        #print tickers[x], "removed (analyst estimates unavailable)"
        tickers.remove(tickers[x])
        x -= 1

print "analysts complete"
#print len(tickers)

#print "-----------------------------STATISTICS-----------------------------"
#Statistics---------------------------------------------------------------------------------------------------------
ticker_urls = []
stat_heads = []
# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/key-statistics?p=" + ticker)

z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    z += 1 #serves as x
    x = 0
    try:
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        stat_tables = soup.find_all('table', ['class', 'table-qsp-stats Mt(10px)'])

        #print c_name.get_text()
        data_dict[tickers[z]].append(c_name.get_text())
        stat_headers = soup.find_all(True, {'class':['Pt(20px)','Pt(6px) Pstart(20px)','Fz(s) Mt(20px)']})
        for stat_header in stat_headers:
            stat_heads.append(stat_header.get_text())


        for stat_table in stat_tables:
            if stat_heads[x] == "Financial Highlights" or stat_heads[x] == "Trading Information":
                #pretty_print(stat_heads[x] + "\n" + "\n" + stat_heads[x+1],stat_table)
                data_dict[tickers[z]].extend((stat_heads[x] + "\n" + "\n" + stat_heads[x+1], pd.read_html(str(stat_table))))
                x += 1
            else:
                #pretty_print(stat_heads[x], stat_table)
                data_dict[tickers[z]].extend((stat_heads[x], pd.read_html(str(stat_table))))
            x += 1
    except:
        #print "Stock Stats Unavailable"
        pass

print "statistics complete"
time_four = time.time() #----------------------------------------------------------------------------------------------------------------------
#print "-----------------------------PROFILE-----------------------------"
#Profile------------------------------------------------------------------------------------------------------------
ticker_urls = []

# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/profile?p=" + ticker)

profile_info = []
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # strip comments
        comments = soup.find_all(text=lambda text: isinstance(text, Comment))
        [comment.extract() for comment in comments]

        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        geo_con_web = replace_with_newlines(soup.find('p', ['class', 'D(ib) W(47.727%) Pend(40px)']))
        sec_ind_emp = replace_with_newlines(soup.find('p', ['class', 'D(ib) Va(t)']))
        execs_table = soup.find_all('table', ['class', 'W(100%)'])
        #execs = execs_table[0].findChildren()
        descript = replace_with_newlines(soup.find('p',['class','Mt(15px) Lh(1.6)']))

        #print c_name.get_text()
        #pretty_print("Executives", execs_table)
        #print geo_con_web.splitlines()
        #print sec_ind_emp.splitlines()
        #print "Description: " + descript
        data_dict[tickers[x]].extend((c_name.get_text(), pd.read_html(str(execs_table)), geo_con_web.splitlines(),sec_ind_emp.splitlines(),("Description", descript)))
    except:
        #print "Stock Profile Unavailable"
        pass

print "profile complete"

time_five = time.time() #----------------------------------------------------------------------------------------------------------------------
#print "-----------------------------FINANCIALS-----------------------------"
# Financials (marketwatch.com)----------------------------------------------------------------------------------------
ticker_urls = []

# INCOME STATEMENT-----------------------------------------------------
for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials")
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Income Statement"
        data_dict[tickers[x]].extend((c_name.get_text(),"Income Statement"))
        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            #pretty_print("", fin_table)
            data_dict[tickers[x]].append(pd.read_html(str(fin_table)))
    except:
        #print "Stock Financials Unavailable"
        pass


print "income statements complete"

time_six = time.time()  # ----------------------------------------------------------------------------------------------------------------------
# BALANCE SHEET ----------------------------------------------------
ticker_urls = []

for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials/balance-sheet")
z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    fin_heads = []
    x = 0
    y = 0
    z += 1 #serving as x
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Balance Sheet"
        data_dict[tickers[z]].extend((c_name.get_text(), "Balance Sheet"))
        fin_headers = soup.find_all('h2')
        for fin_header in fin_headers:
            fin_heads.append(fin_header.get_text())
        # only want the 2nd and 3rd
        fin_heads.pop(0)

        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            if len(fin_tables) == 3:
                if x == 0 or x == 2:
                    #pretty_print(fin_heads[y], fin_table)
                    data_dict[tickers[z]].extend((fin_heads[y], pd.read_html(str(fin_table))))
                    y += 1
                else:
                    #pretty_print("", fin_table)
                    data_dict[tickers[z]].append(pd.read_html(str(fin_table)))
            elif len(fin_tables) == 2:
                #pretty_print(fin_heads[x], fin_table)
                data_dict[tickers[z]].extend((fin_heads[x], pd.read_html(str(fin_table))))
            else:
                #pretty_print("", fin_table)
                data_dict[tickers[z]].append(pd.read_html(str(fin_table)))
            x += 1
    except:
        pass
        #print "Stock Financials Unavailable"

print "balance sheets complete"

time_seven = time.time()  # ----------------------------------------------------------------------------------------------------------------------
# CASH FLOW STATEMENT -----------------------------------------------

ticker_urls = []

for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials/cash-flow")
z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    fin_heads = []
    x = 0
    y = 0
    z += 1 #serving as x
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Cash Flow Statement"
        data_dict[tickers[z]].extend((c_name.get_text(), "Cash Flow Statement"))
        fin_headers = soup.find_all('h2')
        for fin_header in fin_headers:
            fin_heads.append(fin_header.get_text())
        # only want the 2nd and 3rd
        fin_heads.pop(0)

        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            #pretty_print(fin_heads[x], fin_table)
            data_dict[tickers[z]].extend((fin_heads[x], pd.read_html(str(fin_table))))
            x += 1

    except:
        #print "Stock Financials Unavailable"
        pass

print "cash flow statements complete"

time_eight = time.time()  # ----------------------------------------------------------------------------------------------------------------------

#Historical Data ----- (potentially download) ----------------------------------------------------------------------
#print  "-----------------------------HISTORICAL DATA (download file)-----------------------------"
#print "To be completed"

#print tickers
print "Program took", (time.time() - start_time)/60 , "minutes to run"

time_nine = time.time() #----------------------------------------------------------------------------------------------------------------------

print time_two - time_one, time_three - time_two, time_four - time_three, time_five - time_four, time_six - time_five, time_seven - time_six, time_eight - time_seven, time_nine - time_eight
#print data_dict

x = 0
#Test querying the dicitonary

#Test 3
# try:
#     print "test3 - Summary with data table", data_dict[tickers[x]][2]
# except:
#     pass
#
# #Test 4
# try:
#     print "test3 - 'Summary'", pd.DataFrame(data_dict[tickers[x]][2][1])
# except:
#     pass


#Look into "Stock Financials Unavailable"


def is_df(var):
    if isinstance(var, pd.DataFrame):
        return True
    else:
        return False

#Write to Excel ------------------------------------------------------------



#create excel workbook
workbook = xlsxwriter.Workbook('Output.xlsx')

#add formats
format_dict = {}

x = -1
# create one sheet for each ticker
for ticker in tickers:
    x += 1
    workbook.add_worksheet(ticker)
    worksheet = workbook.get_worksheet_by_name(ticker)
    # write to each sheet
    worksheet.write('B1', "Company Name")
    worksheet.write('C1', data_dict[tickers[x]][0])
    worksheet.write('B2', "Current Price")
    worksheet.write('C2', data_dict[tickers[x]][1])
    worksheet.write('B4',data_dict[tickers[x]][2][0])
    #worksheet.write('E1', pd.DataFrame(data_dict[tickers[x]][2][1]))
    #adjust column width
    worksheet.set_column('A:A', 0)
    worksheet.set_column('B:G', 20)
    worksheet.write('D12','=LEFT(C12,FIND("-",C12)-2)')
    worksheet.write('E12','=TRIM(RIGHT(C12,FIND("-",C12)-1))')
    worksheet.write('K2', '=LEFT(C1,FIND("(",C1) - 2)')
    worksheet.write('D1','=TRIM(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("PLC",UPPER(LEFT(C1,FIND("(",C1) - 2))),3,"" ), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND(", INC.",UPPER(LEFT(C1,FIND("(",C1) - 2))),6,"" )), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORPORATION",UPPER(LEFT(C1,FIND("(",C1) - 2))),11,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORP.",UPPER(LEFT(C1,FIND("(",C1) - 2))),5,"" )), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("COMPANY",UPPER(LEFT(C1,FIND("(",C1) - 2))),7,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("INC.",UPPER(LEFT(C1,FIND("(",C1) - 2))),4,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORP",UPPER(LEFT(C1,FIND("(",C1) - 2))),4,"" )),LEFT(C1,FIND("(",C1) - 2)))')
    worksheet.write('K3', '=" is scheduled to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('K4', '="The stock is currently trading at " & TEXT(C2,"$####.00") & ", " & IF(C2-C7=0, "at the same price" & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down from" & TEXT((C7-C8)/C7*1,"##.##%")) & " yesterday\'s close")), IF(C2-C7>0, "up " & TEXT((C7-C2)/C7*-1,"##.##%") & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down from" & TEXT((C7-C8)/C7*1,"##.##%")) & " yesterday\'s close")), "down " & TEXT((C7-C2)/C7*1,"##.##%") & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down from" & TEXT((C7-C8)/C7*1,"##.##%")) & " yesterday\'s close")) ))')
    worksheet.write('K5', '="The one year target estimate for " & D1 & " is " & TEXT(C23,"$####.00")')
    worksheet.write('K6', '=" which would be " & IF(OR(LEFT(ABS((C23-C2)/C2*100),1)="8",LEFT(ABS((C23-C2)/C2*100),2)="18"), "an ", "a ")  &TEXT(ABS((C23-C2)/C2),"####.00%")&IF((C23-C2)>0," increase over"," decrease from")&" the current price"')
    worksheet.write('K7', '="Earnings are expected to " & IF(C28=D28, "remain constant over the next quarter", IF( D28>C28,  "increase by " & TEXT((D28-C28)/C28*100,"##.##") & "% over last quarter", "decrease by " & TEXT((D28-C28)/C28*-100,"##.##") & "% from last quarter")) & " based on the average of " & $C$27 & " analyst estimates (Yahoo Finance)"')
    worksheet.write('K8', '=IF(VALUE(C2)=D12, "The stock is trading at an all-time low",IF(VALUE(C2) =E12,"The stock is trading at an all-time high",IF(VALUE(C2)<D12+(E12-D12)/3, "The stock is trading in the low end of its 52-week range",IF(VALUE(C2)<D12+2*(E12-D12)/3, "The stock is trading near the middle of its 52 week range", "The stock is trading in the high end of its 52-week range"))) )')
    worksheet.write('K9', '="Over the last 4 quarters, we\'ve seen a positive earnings surprise " & 4 -COUNTIF(C45:F45,"-*") & IF(4 - COUNTIF(C45:F45,"-*")=1, " time,"," times,") & " and a negative earnings surprise " & COUNTIF(C45:F45,"-*") & IF(COUNTIF(C45:F45,"-*")=1, " time", " times")')
    #Paragraph 1
    worksheet.write('K17','=K2 & K3 & ". " & K4 & ". " & K5 & K6 & ". " & K7 & ". " & K8 & ". " & K9 & "."')

    #converts financial statement text into numbers
    for y in range(140,400):
        worksheet.write('J' + str(y), '=IF(TRIM(C' + str(y) + ')="-", "N/A", IF(RIGHT(C' + str(y) + ',1)="M",1000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="B",1000000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="%",0.01*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),C' + str(y) + '))))')
        worksheet.write('K' + str(y), '=IF(TRIM(D' + str(y) + ')="-", "N/A", IF(RIGHT(D' + str(y) + ',1)="M",1000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="B",1000000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="%",0.01*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),D' + str(y) + '))))')
        worksheet.write('L' + str(y), '=IF(TRIM(E' + str(y) + ')="-", "N/A", IF(RIGHT(E' + str(y) + ',1)="M",1000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="B",1000000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="%",0.01*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),E' + str(y) + '))))')
        worksheet.write('M' + str(y), '=IF(TRIM(F' + str(y) + ')="-", "N/A", IF(RIGHT(F' + str(y) + ',1)="M",1000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="B",1000000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="%",0.01*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),F' + str(y) + '))))')
        worksheet.write('N' + str(y), '=IF(TRIM(G' + str(y) + ')="-", "N/A", IF(RIGHT(G' + str(y) + ',1)="M",1000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="B",1000000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="%",0.01*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),G' + str(y) + '))))')


        #worksheet.write('J100:N400','=IF(TRIM(C147)="-", "N/A", IF(RIGHT(C147,1)="M",1000000*VALUE(LEFT(C147,LEN(C147)-1)),IF(RIGHT(C147,1)="B",1000000000*VALUE(LEFT(C147,LEN(C147)-1)),IF(RIGHT(C147,1)="%",0.01*VALUE(LEFT(C147,LEN(C147)-1)),C147))))')



    #additons
    #=IF(STDEV.P(J146:N146)<0.1,IF(COUNTIF(J146:N146,">0")=5,"pos_trend"),"") where J146:N146 is a range of annual data
        # checks if the data isn't deviating too much, and if it isn't, whether there is a positive trend

workbook.close()

writer = pd.ExcelWriter('Output.xlsx', engine='openpyxl')
wb = load_workbook('Output.xlsx')

# create pandas excel writer for dfs
writer.book = wb

# have to tell pandas that we alerady have sheets, and what they are
writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

# insert dataframes

x = -1
for ticker in tickers:
    offset = 6
    x += 1

    try:
        df = pd.DataFrame(data_dict[tickers[x]][2][1][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1

        df = pd.DataFrame(data_dict[tickers[x]][2][1][1])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 2
    except:
        pass

    try:
        df = pd.DataFrame(data_dict[tickers[x]][4])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][4])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][5])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][5])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][6])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][6])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][7])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][7])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][8])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][8])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][9])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][9])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][12][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][12][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][14][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][14][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][16][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][16][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][18][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][18][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][20][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][20][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][22][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][22][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][24][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][24][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][26][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][26][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][28][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][28][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][30][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][30][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][32][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][32][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][38][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][38][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][39][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][39][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][43][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][43][0])"

    #for the 2 asset tables issue
    try:
        if is_df(data_dict[tickers[x]][44][0]):
            try:
                df = pd.DataFrame(data_dict[tickers[x]][44][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][44][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][46][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][46][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][50][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][50][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][52][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][52][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][54][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][54][0])"

        else:
            try:
                df = pd.DataFrame(data_dict[tickers[x]][45][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][45][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][47][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][47][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][51][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][51][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][53][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][53][0])"
    except:
        print "index 44 out of range"

    # #Company Bio
    # worksheet = writer.sheets[ticker]
    # try:
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][33][0])
    #     offset += 1
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][33][1])
    #     offset += 1
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][33][2])
    #     offset += 1
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][33][3])
    #     offset += 1
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][33][4])
    #     offset += 1
    # except:
    #     print 'not printed'
    #
    # try:
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][34][0])
    #     offset += 1
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][34][1])
    #     offset += 1
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][34][2])
    #     offset += 1
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][34][3])
    #     offset += 1
    # except:
    #     print 'not printed'
    #
    # try:
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][35][0])
    #     offset += 1
    #     worksheet.write_string(offset, 1, data_dict[tickers[x]][35][1])
    #     offset += 1
    # except:
    #     print 'not printed'

    # try:
    #     df = data_dict[tickers[x]][33][0]
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][33][0])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][33][1])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][33][1])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][33][2])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][33][2])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][33][3])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][33][3])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][33][4])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][33][4])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][34][0])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][34][0])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][34][1])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][34][1])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][34][2])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][34][2])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][34][3])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][34][3])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][35][0])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][35][0])"
    #
    # try:
    #     df = pd.DataFrame(data_dict[tickers[x]][35][1])
    #     df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
    #     offset += len(df.index) + 1
    # except:
    #     print "pd.DataFrame(data_dict[tickers[x]][33][0])"










writer.save()

















