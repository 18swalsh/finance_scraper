 #02035
from bs4 import BeautifulSoup, Comment
import requests
import pandas as pd
import numpy as np
import itertools
from tabulate import tabulate
import types
import datetime
import time
import xlsxwriter
import csv
from openpyxl import load_workbook
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()


#dictionary to store all data
data_dict = {}

d_valid = True
#---------------------------- Get earnings announcements for a given date ----------------------------------------------
# date = ""
# while d_valid == True:
#     date = str(raw_input("Enter a date or type 'exit' (Format: YYYY-MM-DD): "))
#     try:
#         datetime.datetime.strptime(date, '%Y-%m-%d')
#         d_valid = True
#     except:
#         if date =='exit':
#             print "bye"
#             exit()
#         else:
#             print "Nah" + "\n"

start_time = time.time()
time_one = time.time() #----------------------------------------------------------------------------------------------------------------------

def pretty_print(table_title, table_to_print):
    df = pd.read_html(str(table_to_print))
    print table_title, "\n", (tabulate(df[0], headers='keys', tablefmt='psql'))

def replace_with_newlines(element):
    text = ''
    for elem in element.recursiveChildGenerator():
        if isinstance(elem, types.StringTypes):
            text += elem.strip()
        elif elem.name == 'br':
            text += '\n'
    return text

def removekey(d, key):
    r = dict(d)
    del r[key]
    return r

date = "2017-07-28"

url = "finance.yahoo.com/calendar/earnings?day=" + date
r  = requests.get("https://" + url)
data = r.text
soup = BeautifulSoup(data, "html5lib" )
data_table = soup.find('table', {'class':'data-table'})

headers = []
cells = []
tickers = []

#pretty_print('Earnings announcements on ' + date, data_table)
data_dict['Tickers'] = pd.read_html(str(data_table))

cols = data_table.find_all('td')
#ths = data_table.find_all('th')
#for th in ths[1::]:
#    headers.append(th.get_text())
for col in cols[1::]:
    if col != "-":
        cells.append(col.get_text())
    else:
        cells.append("NULL")

w = ''
spl = [list(y) for x, y in itertools.groupby(cells, lambda z: z == w) if not x]

cap = 0
for line in spl:
    tickers.append(line[0])
    cap += 1
    if cap > 100:
        break

print "tickers found"
#print len(tickers)
time_two = time.time() #----------------------------------------------------------------------------------------------------------------------

#-----------------------add more info for known tickers-----------------------------------------------------------------
#https://finance.yahoo.com/quote/TICKER?p=TICKER - stock page

#print "-----------------------------SUMMARY-----------------------------"

ticker_urls = []

#Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "?p=" + ticker)
    #initialize dictionary structure
    data_dict[ticker] = []

x = -1
#Run for each URL
for ticker_url in ticker_urls:
    x += 1
    r  = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    summary_table = soup.find_all('table',['class','W(100%)'])



    #Summary------------------------------------------------------------------------------------------------------------

    #Error if the url is bad
    try:

        cur_price = soup.find('span',['class','Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)'])
        # Filter out stocks under $20
        if float(cur_price.get_text()) < 20:
            #print tickers[x], "was removed from search (price below $20)"
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        # remove companies with a market cap over 10B
        market_cap = 0
        market_caps = summary_table[1].find_all('tr')
        market_cap_string = market_caps[0].get_text()[10:]

        if market_cap_string[-1:] == 'M':
            market_cap = float(market_cap_string[:-1]) * 1000000
        elif market_cap_string[-1:] == 'B':
            market_cap = float(market_cap_string[:-1]) * 1000000000

        if market_cap > 10000000000:
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        #print "Company: " + c_name.get_text()
        #print "Current Price: " + cur_price.get_text()
        #pretty_print("Summary", summary_table)
        data_dict[tickers[x]].extend((c_name.get_text(),cur_price.get_text(),("Summary",pd.read_html(str(summary_table)))))

    except:
        #print "Stock Quote Unavailable"
        #print tickers[x], "removed from search"
        data_dict = removekey(data_dict, tickers[x])
        tickers.remove(tickers[x])
        x -= 1


print "summary complete"
#print len(tickers)

time_three = time.time() #----------------------------------------------------------------------------------------------------------------------

#print "-----------------------------ANALYSTS-----------------------------"
#Analysts------------------------------------------------------------------------------------------------------------

ticker_urls = []

# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/analysts?p=" + ticker)
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        #print c_name.get_text()

        data_dict[tickers[x]].append(c_name.get_text())
        analyst_est_sect = soup.find('section', id='quote-leaf-comp')
        analyst_tables = analyst_est_sect.find_all('table', ['class', 'W(100%)'])
        if len(analyst_tables) == 0:
            #print tickers[x], "was removed (no analyst estimates available)"
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        for a_table in analyst_tables:
            #pretty_print("",a_table)
            data_dict[tickers[x]].extend(pd.read_html(str(a_table)))

            # here is an alternative if the  4 lines above break
            # all_tables = soup.find_all('table', class_='W(100%)')
            # for table in all_tables:
            #     pretty_print("This Table", table

    except:
        data_dict = removekey(data_dict, tickers[x])
        #print tickers[x], "removed (analyst estimates unavailable)"
        tickers.remove(tickers[x])
        x -= 1

print "analysts complete"
#print len(tickers)

#print "-----------------------------STATISTICS-----------------------------"
#Statistics---------------------------------------------------------------------------------------------------------
ticker_urls = []
stat_heads = []
# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/key-statistics?p=" + ticker)

z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    z += 1 #serves as x
    x = 0
    try:
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        stat_tables = soup.find_all('table', ['class', 'table-qsp-stats Mt(10px)'])

        #print c_name.get_text()
        data_dict[tickers[z]].append(c_name.get_text())
        stat_headers = soup.find_all(True, {'class':['Pt(20px)','Pt(6px) Pstart(20px)','Fz(s) Mt(20px)']})
        for stat_header in stat_headers:
            stat_heads.append(stat_header.get_text())


        for stat_table in stat_tables:
            if stat_heads[x] == "Financial Highlights" or stat_heads[x] == "Trading Information":
                #pretty_print(stat_heads[x] + "\n" + "\n" + stat_heads[x+1],stat_table)
                data_dict[tickers[z]].extend((stat_heads[x] + "\n" + "\n" + stat_heads[x+1], pd.read_html(str(stat_table))))
                x += 1
            else:
                #pretty_print(stat_heads[x], stat_table)
                data_dict[tickers[z]].extend((stat_heads[x], pd.read_html(str(stat_table))))
            x += 1
    except:
        #print "Stock Stats Unavailable"
        pass

print "statistics complete"
time_four = time.time() #----------------------------------------------------------------------------------------------------------------------
#print "-----------------------------PROFILE-----------------------------"
#Profile------------------------------------------------------------------------------------------------------------
ticker_urls = []

# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/profile?p=" + ticker)

profile_info = []
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # strip comments
        comments = soup.find_all(text=lambda text: isinstance(text, Comment))
        [comment.extract() for comment in comments]

        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        geo_con_web = replace_with_newlines(soup.find('p', ['class', 'D(ib) W(47.727%) Pend(40px)']))
        sec_ind_emp = replace_with_newlines(soup.find('p', ['class', 'D(ib) Va(t)']))
        execs_table = soup.find_all('table', ['class', 'W(100%)'])
        #execs = execs_table[0].findChildren()
        descript = replace_with_newlines(soup.find('p',['class','Mt(15px) Lh(1.6)']))

        #print c_name.get_text()
        #pretty_print("Executives", execs_table)
        #print geo_con_web.splitlines()
        #print sec_ind_emp.splitlines()
        #print "Description: " + descript
        data_dict[tickers[x]].extend((c_name.get_text(), pd.read_html(str(execs_table)), geo_con_web.splitlines(),sec_ind_emp.splitlines(),("Description", descript)))
    except:
        #print "Stock Profile Unavailable"
        pass

print "profile complete"

time_five = time.time() #----------------------------------------------------------------------------------------------------------------------
#print "-----------------------------FINANCIALS-----------------------------"
# Financials (marketwatch.com)----------------------------------------------------------------------------------------
ticker_urls = []

# INCOME STATEMENT-----------------------------------------------------
for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials")
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Income Statement"
        data_dict[tickers[x]].extend((c_name.get_text(),"Income Statement"))
        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            #pretty_print("", fin_table)
            data_dict[tickers[x]].append(pd.read_html(str(fin_table)))
    except:
        #print "Stock Financials Unavailable"
        pass


print "income statements complete"

time_six = time.time()  # ----------------------------------------------------------------------------------------------------------------------
# BALANCE SHEET ----------------------------------------------------
ticker_urls = []

for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials/balance-sheet")
z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    fin_heads = []
    x = 0
    y = 0
    z += 1 #serving as x
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Balance Sheet"
        data_dict[tickers[z]].extend((c_name.get_text(), "Balance Sheet"))
        fin_headers = soup.find_all('h2')
        for fin_header in fin_headers:
            fin_heads.append(fin_header.get_text())
        # only want the 2nd and 3rd
        fin_heads.pop(0)

        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            if len(fin_tables) == 3:
                if x == 0 or x == 2:
                    #pretty_print(fin_heads[y], fin_table)
                    data_dict[tickers[z]].extend((fin_heads[y], pd.read_html(str(fin_table))))
                    y += 1
                else:
                    #pretty_print("", fin_table)
                    data_dict[tickers[z]].append(pd.read_html(str(fin_table)))
            elif len(fin_tables) == 2:
                #pretty_print(fin_heads[x], fin_table)
                data_dict[tickers[z]].extend((fin_heads[x], pd.read_html(str(fin_table))))
            else:
                #pretty_print("", fin_table)
                data_dict[tickers[z]].append(pd.read_html(str(fin_table)))
            x += 1
    except:
        pass
        #print "Stock Financials Unavailable"

print "balance sheets complete"

time_seven = time.time()  # ----------------------------------------------------------------------------------------------------------------------
# CASH FLOW STATEMENT -----------------------------------------------

ticker_urls = []

for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials/cash-flow")
z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    fin_heads = []
    x = 0
    y = 0
    z += 1 #serving as x
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Cash Flow Statement"
        data_dict[tickers[z]].extend((c_name.get_text(), "Cash Flow Statement"))
        fin_headers = soup.find_all('h2')
        for fin_header in fin_headers:
            fin_heads.append(fin_header.get_text())
        # only want the 2nd and 3rd
        fin_heads.pop(0)

        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            #pretty_print(fin_heads[x], fin_table)
            data_dict[tickers[z]].extend((fin_heads[x], pd.read_html(str(fin_table))))
            x += 1

    except:
        #print "Stock Financials Unavailable"
        pass

print "cash flow statements complete"

time_eight = time.time()  # ----------------------------------------------------------------------------------------------------------------------

#Historical Data ----- (potentially download) ----------------------------------------------------------------------
#print  "-----------------------------HISTORICAL DATA (download file)-----------------------------"
#print "To be completed"

time_nine = time.time() #----------------------------------------------------------------------------------------------------------------------

print time_two - time_one, time_three - time_two, time_four - time_three, time_five - time_four, time_six - time_five, time_seven - time_six, time_eight - time_seven, time_nine - time_eight
#print data_dict



# Find comps (ycharts.com)------------------------------------------------------------------------------------------------------------------------------------
ticker_urls = []
for ticker in tickers:
    ticker_urls.append("http://www.ycharts.com/companies/" + ticker)

comp_dict = {}
comps = []
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    x +=1
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    comp_urls = []
    comps = []
    try:
        comp_table = soup.find('table', ['class', 'relCompSect'])
        table = pd.read_html(str(comp_table))
        data_dict[tickers[x]].append(table)
        #print table[0][0] prints the first column (comp tickers)
        #print table[0][1] prints the second column (comp names)

        #this next part doesn't work
        for tick_num in range(0,4):
            comp_urls.append("https://finance.yahoo.com/quote/" + table[0][0][tick_num] + "/key-statistics?p=" + table[0][0][tick_num])
        w = -1
        for comp_url in comp_urls:
            w += 1
            r = requests.get(comp_url)
            data = r.text
            soup = BeautifulSoup(data, "html5lib")
            try:
                # Company Name
                c_name = soup.find('h1', ['class', 'D(ib)'])
                stat_tables = soup.find_all('table', ['class', 'table-qsp-stats Mt(10px)'])
                data_dict[tickers[x]].extend((c_name.get_text(), pd.read_html(str(stat_tables[0]))))
                comps.append(c_name.get_text())
                #left off
            except:
                print "comp info not found"


        comp_dict[tickers[x]] = comps
    except:
        print "no comps found"
    #data_dict[tickers[x]].append(comps)
    #print comps

print "comps complete"

print data_dict
def is_df(var):
    if isinstance(var, pd.DataFrame):
        return True
    else:
        return False

#Write to Excel ------------------------------------------------------------



#create excel workbook
workbook = xlsxwriter.Workbook('Output.xlsx')

#Create output sheet
workbook.add_worksheet("Output Text")

#add formats
format_dict = {}

x = -1
# create one sheet for each ticker
for ticker in tickers:
    x += 1
    workbook.add_worksheet(ticker)
    worksheet = workbook.get_worksheet_by_name(ticker)
    # write to each sheet
    worksheet.write('B1', "Company Name")
    worksheet.write('C1', data_dict[tickers[x]][0])
    worksheet.write('B2', "Current Price")
    worksheet.write('C2', data_dict[tickers[x]][1])
    worksheet.write('B4',data_dict[tickers[x]][2][0])
    #worksheet.write('E1', pd.DataFrame(data_dict[tickers[x]][2][1]))

    #worksheet.write('F1', '="Comps"')
    #try:
    #    for comp in range(0,len(comp_dict[ticker[x]])):
    #        worksheet.write('F' + str(comp+2),comp_dict[ticker[x]][comp])
    #except:
    #    pass


    #adjust column width
    worksheet.set_column('A:A', 0)
    worksheet.set_column('B:G', 20)
    worksheet.write('D12','=LEFT(C12,FIND("-",C12)-2)')
    worksheet.write('E12','=TRIM(RIGHT(C12,FIND("-",C12)-1))')
    worksheet.write('K2', '=LEFT(C1,FIND("(",C1) - 2)')
    worksheet.write('D1','=TRIM(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("PLC",UPPER(LEFT(C1,FIND("(",C1) - 2))),3,"" ), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND(", INC.",UPPER(LEFT(C1,FIND("(",C1) - 2))),6,"" )), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORPORATION",UPPER(LEFT(C1,FIND("(",C1) - 2))),11,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORP.",UPPER(LEFT(C1,FIND("(",C1) - 2))),5,"" )), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("COMPANY",UPPER(LEFT(C1,FIND("(",C1) - 2))),7,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("INC.",UPPER(LEFT(C1,FIND("(",C1) - 2))),4,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORP",UPPER(LEFT(C1,FIND("(",C1) - 2))),4,"" )),LEFT(C1,FIND("(",C1) - 2)))')

    worksheet.write('K3', '=" is scheduled to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('L3', '=" is slated to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('M3', '=" will report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('N3', '=" reports earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('O3', '=" plans to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('P3', '=" is going to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')

    worksheet.write('K4', '="The stock is currently trading at " & TEXT(C2,"$####.00") & ", " & IF(C2-C7=0, "at the same price" & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s close")), IF(C2-C7>0, "up " & TEXT((C7-C2)/C7*-1,"##.##%") & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s close")), "down " & TEXT((C7-C2)/C7*1,"##.##%") & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s close")) ))')
    worksheet.write('L4', '="The current stock price is " & TEXT(C2,"$####.00") & ", " & IF(C2-C7=0, "at the same price" & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s closing price", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "a fraction", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s closing price", IF((C7-C8)/C7 <0.01, "a bit below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s closing price")), IF(C2-C7>0, "up " & TEXT((C7-C2)/C7*-1,"##.##%") & " after opening " & IF(C8-C7=0, "at yesterday\'s closing price", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "a fraction", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s closing price", IF((C7-C8)/C7 <0.01, "a bit below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s closing price")), "down " & TEXT((C7-C2)/C7*1,"##.##%") & " after opening " & IF(C8-C7=0, "at yesterday\'s closing price", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "a fraction", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s closing price", IF((C7-C8)/C7 <0.01, "a bit below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s closing price")) ))')

    worksheet.write('K5', '="The one year target estimate for " & D1 & " is " & TEXT(C23,"$####.00")')
    worksheet.write('L5', '=D1 &" has a one year target estimate of " & TEXT(C23,"$####.00")')
    worksheet.write('L5', '=D1 & " is expected to be trading at " & TEXT(C23, "$####.00") & ", based on target estimates"')

    worksheet.write('K6', '=" which would be " & IF(OR(LEFT(ABS((C23-C2)/C2*100),1)="8",LEFT(ABS((C23-C2)/C2*100),2)="18"), "an ", "a ")  &TEXT(ABS((C23-C2)/C2),"####.00%")&IF((C23-C2)>0," increase over"," decrease from")&" the current price"')
    worksheet.write('K7', '="Earnings are expected to " & IF(C28=D28, "remain constant over the next quarter", IF( D28>C28,  "increase by " & TEXT((D28-C28)/C28*100,"##.##") & "% over last quarter", "decrease by " & TEXT((D28-C28)/C28*-100,"##.##") & "% from last quarter")) & " based on the average of " & $C$27 & " analyst estimates (Yahoo Finance)"')
    worksheet.write('K8', '=IF(VALUE(C2)=D12, "The stock is trading at an all-time low",IF(VALUE(C2) =E12,"The stock is trading at an all-time high",IF(VALUE(C2)<D12+(E12-D12)/3, "The stock is trading in the low end of its 52-week range",IF(VALUE(C2)<D12+2*(E12-D12)/3, "The stock is trading near the middle of its 52 week range", "The stock is trading in the high end of its 52-week range"))) )')

    worksheet.write('K9', '="Over the last 4 quarters, we\'ve seen a positive earnings surprise " & 4 -COUNTIF(C45:F45,"-*") & IF(4 - COUNTIF(C45:F45,"-*")=1, " time,"," times,") & " and a negative earnings surprise " & COUNTIF(C45:F45,"-*") & IF(COUNTIF(C45:F45,"-*")=1, " time", " times")')
    worksheet.write('K9', '="Over the last 4 quarters, there" & IF(4 - COUNTIF(C45:F45,"-*")=1, " has"," have") & " been" & IF(4 - COUNTIF(C45:F45,"-*")=1, " a,","") & " positive earnings surprise" & IF(4 - COUNTIF(C45:F45,"-*")=1, " ","s ") & 4 -COUNTIF(C45:F45,"-*") & IF(4 - COUNTIF(C45:F45,"-*")=1, " time,"," times,") & " and a negative earnings surprise " & COUNTIF(C45:F45,"-*") & IF(COUNTIF(C45:F45,"-*")=1, " time", " times")')

    worksheet.write('K10','=IF(F48=F52,"",IF(F48>F52, "EPS estimates have increased by " & TEXT(F48-F52,"$0.00") & " in the 2 months leading up to the earnings report", "EPS estimates have decreased by " & TEXT(ABS(F48-F52),"$0.00") & " in the 2 months leading up to the earnings report"))')
    worksheet.write('K11','=IF(B145="Interest Income",U42, K42)')

    worksheet.write('J3','=RANDBETWEEN(1,6)')
    worksheet.write('J4','=RANDBETWEEN(1,2)')
    worksheet.write('J5','=RANDBETWEEN(1,2)')

    #Paragraph 1
    worksheet.write('K17','=K2 & IF(J3=1, K3,IF(J3=2,L3,IF(J3=3,M3,IF(J3=4,N3,IF(J3=5,O3,IF(J3=6,P3)))))) & ". " & IF(J4=1,K4,IF(J4=2,L4)) & ". " & IF(J5=1,K5,IF(J5=2,L5)) & K6 & ". " & K7 & ". " & K8 & ". " & K9 & "."')

    #converts financial statement text into numbers
    for y in range(60,400):
        #worksheet.write('J' + str(y), '=IF(TRIM(C' + str(y) + ')="-", "N/A", IF(RIGHT(C' + str(y) + ',1)="M",1000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="B",1000000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="%",0.01*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),C' + str(y) + '))))')
        #worksheet.write('K' + str(y), '=IF(TRIM(D' + str(y) + ')="-", "N/A", IF(RIGHT(D' + str(y) + ',1)="M",1000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="B",1000000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="%",0.01*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),D' + str(y) + '))))')
        #worksheet.write('L' + str(y), '=IF(TRIM(E' + str(y) + ')="-", "N/A", IF(RIGHT(E' + str(y) + ',1)="M",1000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="B",1000000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="%",0.01*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),E' + str(y) + '))))')
        #worksheet.write('M' + str(y), '=IF(TRIM(F' + str(y) + ')="-", "N/A", IF(RIGHT(F' + str(y) + ',1)="M",1000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="B",1000000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="%",0.01*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),F' + str(y) + '))))')
        #worksheet.write('N' + str(y), '=IF(TRIM(G' + str(y) + ')="-", "N/A", IF(RIGHT(G' + str(y) + ',1)="M",1000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="B",1000000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="%",0.01*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),G' + str(y) + '))))')

        #bug fix
        worksheet.write('J' + str(y), '=IFERROR(IF(TRIM(C' + str(y) + ')="-", "N/A", IF(RIGHT(C' + str(y) + ',1)=")",IF(RIGHT(C' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(C' + str(y) + ',2,LEN(C' + str(y) + ')-3)),IF(RIGHT(C' + str(y) + ',2)="M)",-1000000*VALUE(MID(C' + str(y) + ',2,LEN(C' + str(y) + ')-3)),IF(RIGHT(C' + str(y) + ',2)="B)",-1000000000*VALUE(MID(C' + str(y) + ',2,LEN(C' + str(y) + ')-3)),IF(RIGHT(C' + str(y) + ',2)="k)",-1000*VALUE(MID(C' + str(y) + ',2,LEN(C' + str(y) + ')-3)),VALUE(SUBSTITUTE(C' + str(y) + ',",","")))))),IF(RIGHT(C' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="M",1000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="B",1000000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="%",0.01*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="k",1000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),VALUE(SUBSTITUTE(C' + str(y) + ',",",""))))))))),"N/A")')
        worksheet.write('K' + str(y), '=IFERROR(IF(TRIM(D' + str(y) + ')="-", "N/A", IF(RIGHT(D' + str(y) + ',1)=")",IF(RIGHT(D' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(D' + str(y) + ',2,LEN(D' + str(y) + ')-3)),IF(RIGHT(D' + str(y) + ',2)="M)",-1000000*VALUE(MID(D' + str(y) + ',2,LEN(D' + str(y) + ')-3)),IF(RIGHT(D' + str(y) + ',2)="B)",-1000000000*VALUE(MID(D' + str(y) + ',2,LEN(D' + str(y) + ')-3)),IF(RIGHT(D' + str(y) + ',2)="k)",-1000*VALUE(MID(D' + str(y) + ',2,LEN(D' + str(y) + ')-3)),VALUE(SUBSTITUTE(D' + str(y) + ',",","")))))),IF(RIGHT(D' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="M",1000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="B",1000000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="%",0.01*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="k",1000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),VALUE(SUBSTITUTE(D' + str(y) + ',",",""))))))))),"N/A")')
        worksheet.write('L' + str(y), '=IFERROR(IF(TRIM(E' + str(y) + ')="-", "N/A", IF(RIGHT(E' + str(y) + ',1)=")",IF(RIGHT(E' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(E' + str(y) + ',2,LEN(E' + str(y) + ')-3)),IF(RIGHT(E' + str(y) + ',2)="M)",-1000000*VALUE(MID(E' + str(y) + ',2,LEN(E' + str(y) + ')-3)),IF(RIGHT(E' + str(y) + ',2)="B)",-1000000000*VALUE(MID(E' + str(y) + ',2,LEN(E' + str(y) + ')-3)),IF(RIGHT(E' + str(y) + ',2)="k)",-1000*VALUE(MID(E' + str(y) + ',2,LEN(E' + str(y) + ')-3)),VALUE(SUBSTITUTE(E' + str(y) + ',",","")))))),IF(RIGHT(E' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="M",1000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="B",1000000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="%",0.01*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="k",1000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),VALUE(SUBSTITUTE(E' + str(y) + ',",",""))))))))),"N/A")')
        worksheet.write('M' + str(y), '=IFERROR(IF(TRIM(F' + str(y) + ')="-", "N/A", IF(RIGHT(F' + str(y) + ',1)=")",IF(RIGHT(F' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(F' + str(y) + ',2,LEN(F' + str(y) + ')-3)),IF(RIGHT(F' + str(y) + ',2)="M)",-1000000*VALUE(MID(F' + str(y) + ',2,LEN(F' + str(y) + ')-3)),IF(RIGHT(F' + str(y) + ',2)="B)",-1000000000*VALUE(MID(F' + str(y) + ',2,LEN(F' + str(y) + ')-3)),IF(RIGHT(F' + str(y) + ',2)="k)",-1000*VALUE(MID(F' + str(y) + ',2,LEN(F' + str(y) + ')-3)),VALUE(SUBSTITUTE(F' + str(y) + ',",","")))))),IF(RIGHT(F' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="M",1000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="B",1000000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="%",0.01*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="k",1000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),VALUE(SUBSTITUTE(F' + str(y) + ',",",""))))))))),"N/A")')
        worksheet.write('N' + str(y), '=IFERROR(IF(TRIM(G' + str(y) + ')="-", "N/A", IF(RIGHT(G' + str(y) + ',1)=")",IF(RIGHT(G' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(G' + str(y) + ',2,LEN(G' + str(y) + ')-3)),IF(RIGHT(G' + str(y) + ',2)="M)",-1000000*VALUE(MID(G' + str(y) + ',2,LEN(G' + str(y) + ')-3)),IF(RIGHT(G' + str(y) + ',2)="B)",-1000000000*VALUE(MID(G' + str(y) + ',2,LEN(G' + str(y) + ')-3)),IF(RIGHT(G' + str(y) + ',2)="k)",-1000*VALUE(MID(G' + str(y) + ',2,LEN(G' + str(y) + ')-3)),VALUE(SUBSTITUTE(G' + str(y) + ',",","")))))),IF(RIGHT(G' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="M",1000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="B",1000000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="%",0.01*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="k",1000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),VALUE(SUBSTITUTE(G' + str(y) + ',",",""))))))))),"N/A")')

        worksheet.write('I' + str(y), '=IF(AND(K' + str(y) + '> J' + str(y)+ ', L' + str(y) + '> K' +str(y)+ ', M' + str(y) + '> L' +str(y)+ ', N' + str(y) +'> M' + str(y) + '), "pos_trend", IF(AND(K' + str(y) +'< J' +str(y)+ ', L' + str(y) +'< K' +str(y)+ ', M' + str(y) +'< L' +str(y)+ ', N' + str(y) + '< M' + str(y) + '), "neg_trend", "N/A"))')

    #trend analysis
    for row in range(22, 39):
        worksheet.write('J' + str(row), '=IF(K' + str(row) + ' <> "",' + str(row - 21) + ', 0)')
        worksheet.write('L' + str(row), '=IF(EXACT(K' + str(row) + ',UPPER(K' + str(row) + ')),K' + str(row) + ',LOWER(K' + str(row) + '))')
        worksheet.write('M22', '=L22')
        worksheet.write('M' + str(row + 1), '=IF(L' + str(row + 1) + '<>"", M' + str(row) + ' & ", " & L' + str(row + 1) + ',M' + str(row) + ')')
        #bank version
        worksheet.write('T' + str(row), '=IF(U' + str(row) + ' <> "",' + str(row - 21) + ', 0)')
        worksheet.write('V' + str(row), '=IF(EXACT(U' + str(row) + ',UPPER(U' + str(row) + ')),U' + str(row) + ',LOWER(U' + str(row) + '))')
        worksheet.write('W22', '=V22')
        worksheet.write('W' + str(row + 1), '=IF(V' + str(row + 1) + '<>"", W' + str(row) + ' & ", " & V' + str(row + 1) + ',W' + str(row) + ')')

    worksheet.write('K22', '=IF(I145="pos_trend","Revenue","")')
    worksheet.write('K23', '=IF(I146="pos_trend",B146,"")')
    worksheet.write('K24', '=IF(I153="pos_trend",B153,"")')
    worksheet.write('K25', '=IF(I154="pos_trend",B154,"")')
    worksheet.write('K26', '=IF(I155="pos_trend",B155,"")')
    worksheet.write('K27', '=IF(I172="pos_trend",B172,"")')
    worksheet.write('K28', '=IF(I173="pos_trend",B173,"")')
    worksheet.write('K29', '=IF(I174="pos_trend",B174,"")')
    worksheet.write('K30', '=IF(I185="pos_trend",B185,"")')
    worksheet.write('K31', '=IF(I186="pos_trend",B186,"")')
    worksheet.write('K32', '=IF(I187="pos_trend",B187,"")')
    worksheet.write('K33', '=IF(I195="pos_trend",B195,"")')
    worksheet.write('K34', '=IF(I196="pos_trend",B196,"")')
    worksheet.write('K35', '=IF(I201="pos_trend",B201,"")')
    worksheet.write('K36', '=IF(I202="pos_trend",B202,"")')
    worksheet.write('K37', '=IF(I203="pos_trend",B203,"")')
    worksheet.write('K38', '=IF(I351="pos_trend",B351,"")')
    worksheet.write('K39', '=IF(I352="pos_trend",B352,"")')
    worksheet.write('K40', '=VLOOKUP(J40,J22:K39,2,FALSE)')
    worksheet.write('K42', '=SUBSTITUTE(IF(M40<>"", D1 & " has managed to increase " & M40 & " each year since " & C144, "No positive trends")," , "," ")')

    worksheet.write('J40', '=MAX(J22:J39)')
    worksheet.write('M40', '=IF(IFERROR(FIND(",",M39),TRUE)=TRUE,M39,IF(NOT(EXACT(K40,UPPER(K40))),SUBSTITUTE(M39,LOWER(K40),"and "&LOWER(K40)),SUBSTITUTE(M39,K40,"and "&K40)))')

    #bank version
    worksheet.write('U22', '=IF(AND(B145 = "Interest Income",I145="pos_trend"), "Interest Income","")')
    worksheet.write('U23', '=IF(I151="pos_trend",B151,"")')
    worksheet.write('U24', '=IF(I161="pos_trend",B161,"")')
    worksheet.write('U25', '=IF(I162="pos_trend",B162,"")')
    worksheet.write('U26', '=IF(I167="pos_trend",B167,"")')
    worksheet.write('U27', '=IF(I170="pos_trend",B170,"")')
    worksheet.write('U28', '=IF(I171="pos_trend",B171,"")')
    worksheet.write('U29', '=IF(I172="pos_trend",B172,"")')
    worksheet.write('U30', '=IF(I178="pos_trend",B178,"")')
    worksheet.write('U31', '=IF(I199="pos_trend",B199,"")')
    worksheet.write('U32', '=IF(I209="pos_trend",B209,"")')
    worksheet.write('U33', '=IF(I231="pos_trend",B231,"")')
    worksheet.write('U34', '=IF(I251="pos_trend",B251,"")')
    worksheet.write('U35', '=IF(I279="pos_trend",B279,"")')
    worksheet.write('U36', '=IF(I336="pos_trend",B336,"")')
    worksheet.write('U37', '=IF(I235="pos_trend",B235,"")')
    worksheet.write('U38', '=IF(I236="pos_trend",B236,"")')
    worksheet.write('U40', '=VLOOKUP(T40,T22:U39,2,FALSE)')
    worksheet.write('U42', '=SUBSTITUTE(IF(W40<>"", D1 & " has managed to increase " & W40 & " each year since " & C144, "No positive trends")," , "," ")')

    worksheet.write('T40', '=MAX(T22:T39)')
    worksheet.write('W40', '=IF(IFERROR(FIND(",",W39),TRUE)=TRUE,W39,IF(NOT(EXACT(U40,UPPER(U40))),SUBSTITUTE(W39,LOWER(U40),"and "&LOWER(U40)),SUBSTITUTE(W39,U40,"and "&U40)))')

    #for row in range(348,356):
    #    worksheet.write('E' + str(row),'=TRIM(IF(ISNUMBER(VALUE(RIGHT(B'+str(row)+',1))),REPLACE(B'+str(row)+',LEN(B'+str(row)+'),1,""),B'+str(row)+'))')
    #    worksheet.write('E' + str(row),'=AVERAGE(VALUE(J'+str(row)+'),VALUE(J'+str(row+11)+'),VALUE(J'+str(row+22)+'),VALUE(J'+str(row+33)+'))')


    # #comps
    # for z in range(500,510):
    #     try:
    #         worksheet.write('C' + str(z), data_dict[tickers[x]][56][z-500])
    #     except:
    #         pass
    #
    #     try:
    #         worksheet.write('C' + str(z), data_dict[tickers[x]][55][z-500])
    #     except:
    #         pass

    worksheet.write('D67','Comp Average')
    worksheet.write('D68', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B68,B69:B500,0)+68))),VALUE(INDIRECT("J"&(MATCH(B68,B69:B500,0)+79))),VALUE(INDIRECT("J"&(MATCH(B68,B69:B500,0)+90))),VALUE(INDIRECT("J"&(MATCH(B68,B69:B500,0)+101)))),"")')
    worksheet.write('D69', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B69,B70:B501,0)+69))),VALUE(INDIRECT("J"&(MATCH(B69,B70:B501,0)+80))),VALUE(INDIRECT("J"&(MATCH(B69,B70:B501,0)+91))),VALUE(INDIRECT("J"&(MATCH(B69,B70:B501,0)+102)))),"")')
    worksheet.write('D70', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B70,B71:B502,0)+70))),VALUE(INDIRECT("J"&(MATCH(B70,B71:B502,0)+81))),VALUE(INDIRECT("J"&(MATCH(B70,B71:B502,0)+92))),VALUE(INDIRECT("J"&(MATCH(B70,B71:B502,0)+103)))),"")')
    worksheet.write('D71', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B71,B72:B503,0)+71))),VALUE(INDIRECT("J"&(MATCH(B71,B72:B503,0)+82))),VALUE(INDIRECT("J"&(MATCH(B71,B72:B503,0)+93))),VALUE(INDIRECT("J"&(MATCH(B71,B72:B503,0)+104)))),"")')
    worksheet.write('D72', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B72,B73:B504,0)+72))),VALUE(INDIRECT("J"&(MATCH(B72,B73:B504,0)+83))),VALUE(INDIRECT("J"&(MATCH(B72,B73:B504,0)+94))),VALUE(INDIRECT("J"&(MATCH(B72,B73:B504,0)+105)))),"")')
    worksheet.write('D73', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B73,B74:B505,0)+73))),VALUE(INDIRECT("J"&(MATCH(B73,B74:B505,0)+84))),VALUE(INDIRECT("J"&(MATCH(B73,B74:B505,0)+95))),VALUE(INDIRECT("J"&(MATCH(B73,B74:B505,0)+106)))),"")')
    worksheet.write('D74', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B74,B75:B506,0)+74))),VALUE(INDIRECT("J"&(MATCH(B74,B75:B506,0)+85))),VALUE(INDIRECT("J"&(MATCH(B74,B75:B506,0)+96))),VALUE(INDIRECT("J"&(MATCH(B74,B75:B506,0)+107)))),"")')
    worksheet.write('D75', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B75,B76:B507,0)+75))),VALUE(INDIRECT("J"&(MATCH(B75,B76:B507,0)+86))),VALUE(INDIRECT("J"&(MATCH(B75,B76:B507,0)+97))),VALUE(INDIRECT("J"&(MATCH(B75,B76:B507,0)+108)))),"")')
    worksheet.write('D76', '=IFERROR(AVERAGE(VALUE(INDIRECT("J"&(MATCH(B76,B77:B508,0)+76))),VALUE(INDIRECT("J"&(MATCH(B76,B77:B508,0)+87))),VALUE(INDIRECT("J"&(MATCH(B76,B77:B508,0)+98))),VALUE(INDIRECT("J"&(MATCH(B76,B77:B508,0)+109)))),"")')

    worksheet.write('E67', '=C1')
    worksheet.write('E68', '=IFERROR(IF(AND(C68<>"",D68<>0),IF(VALUE(J68)>VALUE(K68),"above average","below average"),"no data"),"no data")')
    worksheet.write('E69', '=IFERROR(IF(AND(C69<>"",D69<>0),IF(VALUE(J69)>VALUE(K69),"above average","below average"),"no data"),"no data")')
    worksheet.write('E70', '=IFERROR(IF(AND(C70<>"",D70<>0),IF(VALUE(J70)>VALUE(K70),"above average","below average"),"no data"),"no data")')
    worksheet.write('E71', '=IFERROR(IF(AND(C71<>"",D71<>0),IF(VALUE(J71)>VALUE(K71),"above average","below average"),"no data"),"no data")')
    worksheet.write('E72', '=IFERROR(IF(AND(C72<>"",D72<>0),IF(VALUE(J72)>VALUE(K72),"above average","below average"),"no data"),"no data")')
    worksheet.write('E73', '=IFERROR(IF(AND(C73<>"",D73<>0),IF(VALUE(J73)>VALUE(K73),"above average","below average"),"no data"),"no data")')
    worksheet.write('E74', '=IFERROR(IF(AND(C74<>"",D74<>0),IF(VALUE(J74)>VALUE(K74),"above average","below average"),"no data"),"no data")')
    worksheet.write('E75', '=IFERROR(IF(AND(C75<>"",D75<>0),IF(VALUE(J75)>VALUE(K75),"above average","below average"),"no data"),"no data")')
    worksheet.write('E76', '=IFERROR(IF(AND(C76<>"",D76<>0),IF(VALUE(J76)>VALUE(K76),"above average","below average"),"no data"),"no data")')

    worksheet.write('F70', '=IF(E70="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B70,1))),REPLACE(B70,LEN(B70),1,""),B70))),"")')
    worksheet.write('F71', '=IF(E71="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B71,1))),REPLACE(B71,LEN(B71),1,""),B71))),"")')
    worksheet.write('F72', '=IF(E72="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B72,1))),REPLACE(B72,LEN(B72),1,""),B72))),"")')
    worksheet.write('F73', '=IF(E73="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B73,1))),REPLACE(B73,LEN(B73),1,""),B73))),"")')
    worksheet.write('F74', '=IF(E74="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B74,1))),REPLACE(B74,LEN(B74),1,""),B74))),"")')
    worksheet.write('F75', '=IF(E75="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B75,1))),REPLACE(B75,LEN(B75),1,""),B75))),"")')
    worksheet.write('F76', '=IF(E76="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B76,1))),REPLACE(B76,LEN(B76),1,""),B76))),"")')
    worksheet.write('F77', '=IF(F76="",IF(F75="",IF(F74="",IF(F73="",IF(F72="",IF(F71="",IFERROR(LEFT(F70,FIND("(",F70) - 2),F70),IFERROR(LEFT(F71,FIND("(",F71) - 2),F71)),IFERROR(LEFT(F72,FIND("(",F72) - 2),F72)),IFERROR(LEFT(F73,FIND("(",F73) - 2),F73)),IFERROR(LEFT(F74,FIND("(",F74) - 2),F74)),IFERROR(LEFT(F75,FIND("(",F75) - 2),F75)),IFERROR(LEFT(F76,FIND("(",F76) - 2),F76))')

    worksheet.write('G70', '=IFERROR(LEFT(F70,FIND("(",F70) - 2),F70)')
    worksheet.write('G71', '=IF(F71<>"", G70 & ", " & IFERROR(LEFT(F71,FIND("(",F71) - 2),F71),G70)')
    worksheet.write('G72', '=IF(F72<>"", G71 & ", " & IFERROR(LEFT(F72,FIND("(",F72) - 2),F72),G71)')
    worksheet.write('G73', '=IF(F73<>"", G72 & ", " & IFERROR(LEFT(F73,FIND("(",F73) - 2),F73),G72)')
    worksheet.write('G74', '=IF(F74<>"", G73 & ", " & IFERROR(LEFT(F74,FIND("(",F74) - 2),F74),G73)')
    worksheet.write('G75', '=IF(F75<>"", G74 & ", " & IFERROR(LEFT(F75,FIND("(",F75) - 2),F75),G74)')
    worksheet.write('G76', '=IF(F76<>"", G75 & ", " & IFERROR(LEFT(F76,FIND("(",F76) - 2),F76),G75)')
    worksheet.write('G77', '=TRIM(IF(LEFT(G76,1)=",",REPLACE(G76,1,1,""),SUBSTITUTE(G76,F77, "and " & F77)))')

    worksheet.write('D78', '=IF(COUNTIF(E70:E76,"=above average")>0,"There are some indications that "&D1&" may be overvalued. The company has a higher " & G77 & " than the comparable average", "Inconclusive")')

    worksheet.write('F81','=IF(E70="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B70,1))),REPLACE(B70,LEN(B70),1,""),B70))),"")')
    worksheet.write('F82','=IF(E71="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B71,1))),REPLACE(B71,LEN(B71),1,""),B71))),"")')
    worksheet.write('F83','=IF(E72="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B72,1))),REPLACE(B72,LEN(B72),1,""),B72))),"")')
    worksheet.write('F84','=IF(E73="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B73,1))),REPLACE(B73,LEN(B73),1,""),B73))),"")')
    worksheet.write('F85','=IF(E74="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B74,1))),REPLACE(B74,LEN(B74),1,""),B74))),"")')
    worksheet.write('F86','=IF(E75="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B75,1))),REPLACE(B75,LEN(B75),1,""),B75))),"")')
    worksheet.write('F87','=IF(E76="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B76,1))),REPLACE(B76,LEN(B76),1,""),B76))),"")')
    worksheet.write('F88','=IF(F87="",IF(F86="",IF(F85="",IF(F84="",IF(F83="",IF(F82="",IFERROR(LEFT(F81,FIND("(",F81) - 2),F81),IFERROR(LEFT(F82,FIND("(",F82) - 2),F82)),IFERROR(LEFT(F83,FIND("(",F83) - 2),F83)),IFERROR(LEFT(F84,FIND("(",F84) - 2),F84)),IFERROR(LEFT(F85,FIND("(",F85) - 2),F85)),IFERROR(LEFT(F86,FIND("(",F86) - 2),F86)),IFERROR(LEFT(F87,FIND("(",F87) - 2),F87))')

    worksheet.write('G81','=IFERROR(LEFT(F81,FIND("(",F81) - 2),F81)')
    worksheet.write('G82','=IF(F82<>"", G81 & ", " & IFERROR(LEFT(F82,FIND("(",F82) - 2),F82),G81)')
    worksheet.write('G83','=IF(F83<>"", G82 & ", " & IFERROR(LEFT(F83,FIND("(",F83) - 2),F83),G82)')
    worksheet.write('G84','=IF(F84<>"", G83 & ", " & IFERROR(LEFT(F84,FIND("(",F84) - 2),F84),G83)')
    worksheet.write('G85','=IF(F85<>"", G84 & ", " & IFERROR(LEFT(F85,FIND("(",F85) - 2),F85),G84)')
    worksheet.write('G86','=IF(F86<>"", G85 & ", " & IFERROR(LEFT(F86,FIND("(",F86) - 2),F86),G85)')
    worksheet.write('G87','=IF(F87<>"", G86 & ", " & IFERROR(LEFT(F87,FIND("(",F87) - 2),F87),G86)')
    worksheet.write('G88','=TRIM(IF(LEFT(G87,1)=",",REPLACE(G87,1,1,""),SUBSTITUTE(G87,F88, "and " & F88)))')

    worksheet.write('D89','=IF(COUNTIF(E70:E76,"=below average")>0,"There are some indications that "&D1&" may be undervalued. The company has a lower " & G88 & " than the comparable average", "Inconclusive")')


    #Financial Statement Statistical Analysis -----------------------------------------------------------------------------------------------------------------------------------------------

    worksheet.write('V143', '="z-score"')
    worksheet.write('P144', '="Max"')
    worksheet.write('P145', '=MAX(J145:N145)')
    worksheet.write('P146', '=MAX(J146:N146)')
    worksheet.write('P147', '=MAX(J147:N147)')
    worksheet.write('P148', '=MAX(J148:N148)')
    worksheet.write('P149', '=MAX(J149:N149)')
    worksheet.write('P150', '=MAX(J150:N150)')
    worksheet.write('P151', '=MAX(J151:N151)')
    worksheet.write('P152', '=MAX(J152:N152)')
    worksheet.write('P153', '=MAX(J153:N153)')
    worksheet.write('P154', '=MAX(J154:N154)')
    worksheet.write('P155', '=MAX(J155:N155)')
    worksheet.write('P156', '=MAX(J156:N156)')
    worksheet.write('P157', '=MAX(J157:N157)')
    worksheet.write('P158', '=MAX(J158:N158)')
    worksheet.write('P159', '=MAX(J159:N159)')
    worksheet.write('P160', '=MAX(J160:N160)')
    worksheet.write('P161', '=MAX(J161:N161)')
    worksheet.write('P162', '=MAX(J162:N162)')
    worksheet.write('P163', '=MAX(J163:N163)')
    worksheet.write('P164', '=MAX(J164:N164)')
    worksheet.write('P165', '=MAX(J165:N165)')
    worksheet.write('P166', '=MAX(J166:N166)')
    worksheet.write('P167', '=MAX(J167:N167)')
    worksheet.write('P168', '=MAX(J168:N168)')
    worksheet.write('P169', '=MAX(J169:N169)')
    worksheet.write('P170', '=MAX(J170:N170)')
    worksheet.write('P171', '=MAX(J171:N171)')
    worksheet.write('P172', '=MAX(J172:N172)')
    worksheet.write('P173', '=MAX(J173:N173)')
    worksheet.write('P174', '=MAX(J174:N174)')
    worksheet.write('P175', '=MAX(J175:N175)')
    worksheet.write('P176', '=MAX(J176:N176)')
    worksheet.write('P177', '=MAX(J177:N177)')
    worksheet.write('P178', '=MAX(J178:N178)')
    worksheet.write('P179', '=MAX(J179:N179)')
    worksheet.write('P180', '=MAX(J180:N180)')
    worksheet.write('P181', '=MAX(J181:N181)')
    worksheet.write('P182', '=MAX(J182:N182)')
    worksheet.write('P183', '=MAX(J183:N183)')
    worksheet.write('P184', '=MAX(J184:N184)')
    worksheet.write('P185', '=MAX(J185:N185)')
    worksheet.write('P186', '=MAX(J186:N186)')
    worksheet.write('P187', '=MAX(J187:N187)')
    worksheet.write('P188', '=MAX(J188:N188)')
    worksheet.write('P189', '=MAX(J189:N189)')
    worksheet.write('P190', '=MAX(J190:N190)')
    worksheet.write('P191', '=MAX(J191:N191)')
    worksheet.write('P192', '=MAX(J192:N192)')
    worksheet.write('P193', '=MAX(J193:N193)')
    worksheet.write('P194', '=MAX(J194:N194)')
    worksheet.write('P195', '=MAX(J195:N195)')
    worksheet.write('P196', '=MAX(J196:N196)')
    worksheet.write('P197', '=MAX(J197:N197)')
    worksheet.write('P198', '=MAX(J198:N198)')
    worksheet.write('P199', '=MAX(J199:N199)')
    worksheet.write('P200', '=MAX(J200:N200)')
    worksheet.write('P201', '=MAX(J201:N201)')
    worksheet.write('P202', '=MAX(J202:N202)')
    worksheet.write('P203', '=MAX(J203:N203)')
    worksheet.write('P204', '=MAX(J204:N204)')
    worksheet.write('P205', '=MAX(J205:N205)')
    worksheet.write('P206', '=MAX(J206:N206)')
    worksheet.write('P207', '=MAX(J207:N207)')
    worksheet.write('P208', '=MAX(J208:N208)')
    worksheet.write('P209', '=MAX(J209:N209)')
    worksheet.write('P210', '=MAX(J210:N210)')
    worksheet.write('P211', '=MAX(J211:N211)')
    worksheet.write('P212', '=MAX(J212:N212)')
    worksheet.write('P213', '=MAX(J213:N213)')
    worksheet.write('P214', '=MAX(J214:N214)')
    worksheet.write('P215', '=MAX(J215:N215)')
    worksheet.write('P216', '=MAX(J216:N216)')
    worksheet.write('P217', '=MAX(J217:N217)')
    worksheet.write('P218', '=MAX(J218:N218)')
    worksheet.write('P219', '=MAX(J219:N219)')
    worksheet.write('P220', '=MAX(J220:N220)')
    worksheet.write('P221', '=MAX(J221:N221)')
    worksheet.write('P222', '=MAX(J222:N222)')
    worksheet.write('P223', '=MAX(J223:N223)')
    worksheet.write('P224', '=MAX(J224:N224)')
    worksheet.write('P225', '=MAX(J225:N225)')
    worksheet.write('P226', '=MAX(J226:N226)')
    worksheet.write('P227', '=MAX(J227:N227)')
    worksheet.write('P228', '=MAX(J228:N228)')
    worksheet.write('P229', '=MAX(J229:N229)')
    worksheet.write('P230', '=MAX(J230:N230)')
    worksheet.write('P231', '=MAX(J231:N231)')
    worksheet.write('P232', '=MAX(J232:N232)')
    worksheet.write('P233', '=MAX(J233:N233)')
    worksheet.write('P234', '=MAX(J234:N234)')
    worksheet.write('P235', '=MAX(J235:N235)')
    worksheet.write('P236', '=MAX(J236:N236)')
    worksheet.write('P237', '=MAX(J237:N237)')
    worksheet.write('P238', '=MAX(J238:N238)')
    worksheet.write('P239', '=MAX(J239:N239)')
    worksheet.write('P240', '=MAX(J240:N240)')
    worksheet.write('P241', '=MAX(J241:N241)')
    worksheet.write('P242', '=MAX(J242:N242)')
    worksheet.write('P243', '=MAX(J243:N243)')
    worksheet.write('P244', '=MAX(J244:N244)')
    worksheet.write('P245', '=MAX(J245:N245)')
    worksheet.write('P246', '=MAX(J246:N246)')
    worksheet.write('P247', '=MAX(J247:N247)')
    worksheet.write('P248', '=MAX(J248:N248)')
    worksheet.write('P249', '=MAX(J249:N249)')
    worksheet.write('P250', '=MAX(J250:N250)')
    worksheet.write('P251', '=MAX(J251:N251)')
    worksheet.write('P252', '=MAX(J252:N252)')
    worksheet.write('P253', '=MAX(J253:N253)')
    worksheet.write('P254', '=MAX(J254:N254)')
    worksheet.write('P255', '=MAX(J255:N255)')
    worksheet.write('P256', '=MAX(J256:N256)')
    worksheet.write('P257', '=MAX(J257:N257)')
    worksheet.write('P258', '=MAX(J258:N258)')
    worksheet.write('P259', '=MAX(J259:N259)')
    worksheet.write('P260', '=MAX(J260:N260)')
    worksheet.write('P261', '=MAX(J261:N261)')
    worksheet.write('P262', '=MAX(J262:N262)')
    worksheet.write('P263', '=MAX(J263:N263)')
    worksheet.write('P264', '=MAX(J264:N264)')
    worksheet.write('P265', '=MAX(J265:N265)')
    worksheet.write('P266', '=MAX(J266:N266)')
    worksheet.write('P267', '=MAX(J267:N267)')
    worksheet.write('P268', '=MAX(J268:N268)')
    worksheet.write('P269', '=MAX(J269:N269)')
    worksheet.write('P270', '=MAX(J270:N270)')
    worksheet.write('P271', '=MAX(J271:N271)')
    worksheet.write('P272', '=MAX(J272:N272)')
    worksheet.write('P273', '=MAX(J273:N273)')
    worksheet.write('P274', '=MAX(J274:N274)')
    worksheet.write('P275', '=MAX(J275:N275)')
    worksheet.write('P276', '=MAX(J276:N276)')
    worksheet.write('P277', '=MAX(J277:N277)')
    worksheet.write('P278', '=MAX(J278:N278)')
    worksheet.write('P279', '=MAX(J279:N279)')
    worksheet.write('P280', '=MAX(J280:N280)')
    worksheet.write('P281', '=MAX(J281:N281)')
    worksheet.write('P282', '=MAX(J282:N282)')
    worksheet.write('P283', '=MAX(J283:N283)')
    worksheet.write('P284', '=MAX(J284:N284)')
    worksheet.write('P285', '=MAX(J285:N285)')
    worksheet.write('P286', '=MAX(J286:N286)')
    worksheet.write('P287', '=MAX(J287:N287)')
    worksheet.write('P288', '=MAX(J288:N288)')
    worksheet.write('P289', '=MAX(J289:N289)')
    worksheet.write('P290', '=MAX(J290:N290)')
    worksheet.write('P291', '=MAX(J291:N291)')
    worksheet.write('P292', '=MAX(J292:N292)')
    worksheet.write('P293', '=MAX(J293:N293)')
    worksheet.write('P294', '=MAX(J294:N294)')
    worksheet.write('P295', '=MAX(J295:N295)')
    worksheet.write('P296', '=MAX(J296:N296)')
    worksheet.write('P297', '=MAX(J297:N297)')
    worksheet.write('P298', '=MAX(J298:N298)')
    worksheet.write('P299', '=MAX(J299:N299)')
    worksheet.write('P300', '=MAX(J300:N300)')
    worksheet.write('P301', '=MAX(J301:N301)')
    worksheet.write('P302', '=MAX(J302:N302)')
    worksheet.write('P303', '=MAX(J303:N303)')
    worksheet.write('P304', '=MAX(J304:N304)')
    worksheet.write('P305', '=MAX(J305:N305)')
    worksheet.write('P306', '=MAX(J306:N306)')
    worksheet.write('P307', '=MAX(J307:N307)')
    worksheet.write('P308', '=MAX(J308:N308)')
    worksheet.write('P309', '=MAX(J309:N309)')
    worksheet.write('P310', '=MAX(J310:N310)')
    worksheet.write('P311', '=MAX(J311:N311)')
    worksheet.write('P312', '=MAX(J312:N312)')
    worksheet.write('P313', '=MAX(J313:N313)')
    worksheet.write('P314', '=MAX(J314:N314)')
    worksheet.write('P315', '=MAX(J315:N315)')
    worksheet.write('P316', '=MAX(J316:N316)')
    worksheet.write('P317', '=MAX(J317:N317)')
    worksheet.write('P318', '=MAX(J318:N318)')
    worksheet.write('P319', '=MAX(J319:N319)')
    worksheet.write('P320', '=MAX(J320:N320)')
    worksheet.write('P321', '=MAX(J321:N321)')
    worksheet.write('P322', '=MAX(J322:N322)')
    worksheet.write('P323', '=MAX(J323:N323)')
    worksheet.write('P324', '=MAX(J324:N324)')
    worksheet.write('P325', '=MAX(J325:N325)')
    worksheet.write('P326', '=MAX(J326:N326)')
    worksheet.write('P327', '=MAX(J327:N327)')
    worksheet.write('P328', '=MAX(J328:N328)')
    worksheet.write('P329', '=MAX(J329:N329)')
    worksheet.write('P330', '=MAX(J330:N330)')
    worksheet.write('P331', '=MAX(J331:N331)')
    worksheet.write('P332', '=MAX(J332:N332)')
    worksheet.write('P333', '=MAX(J333:N333)')
    worksheet.write('P334', '=MAX(J334:N334)')
    worksheet.write('P335', '=MAX(J335:N335)')
    worksheet.write('P336', '=MAX(J336:N336)')
    worksheet.write('P337', '=MAX(J337:N337)')
    worksheet.write('P338', '=MAX(J338:N338)')
    worksheet.write('P339', '=MAX(J339:N339)')
    worksheet.write('P340', '=MAX(J340:N340)')
    worksheet.write('P341', '=MAX(J341:N341)')
    worksheet.write('P342', '=MAX(J342:N342)')
    worksheet.write('P343', '=MAX(J343:N343)')
    worksheet.write('P344', '=MAX(J344:N344)')
    worksheet.write('P345', '=MAX(J345:N345)')
    worksheet.write('P346', '=MAX(J346:N346)')
    worksheet.write('P347', '=MAX(J347:N347)')
    worksheet.write('P348', '=MAX(J348:N348)')
    worksheet.write('P349', '=MAX(J349:N349)')
    worksheet.write('P350', '=MAX(J350:N350)')
    worksheet.write('P351', '=MAX(J351:N351)')
    worksheet.write('P352', '=MAX(J352:N352)')
    worksheet.write('P353', '=MAX(J353:N353)')
    worksheet.write('Q144', '="Max Year"')
    worksheet.write('Q145', '=IFERROR(J144+MATCH(P145,J145:N145,0)-1,"")')
    worksheet.write('Q146', '=IFERROR(J144+MATCH(P146,J146:N146,0)-1,"")')
    worksheet.write('Q147', '=IFERROR(J144+MATCH(P147,J147:N147,0)-1,"")')
    worksheet.write('Q148', '=IFERROR(J144+MATCH(P148,J148:N148,0)-1,"")')
    worksheet.write('Q149', '=IFERROR(J144+MATCH(P149,J149:N149,0)-1,"")')
    worksheet.write('Q150', '=IFERROR(J144+MATCH(P150,J150:N150,0)-1,"")')
    worksheet.write('Q151', '=IFERROR(J144+MATCH(P151,J151:N151,0)-1,"")')
    worksheet.write('Q152', '=IFERROR(J144+MATCH(P152,J152:N152,0)-1,"")')
    worksheet.write('Q153', '=IFERROR(J144+MATCH(P153,J153:N153,0)-1,"")')
    worksheet.write('Q154', '=IFERROR(J144+MATCH(P154,J154:N154,0)-1,"")')
    worksheet.write('Q155', '=IFERROR(J144+MATCH(P155,J155:N155,0)-1,"")')
    worksheet.write('Q156', '=IFERROR(J144+MATCH(P156,J156:N156,0)-1,"")')
    worksheet.write('Q157', '=IFERROR(J144+MATCH(P157,J157:N157,0)-1,"")')
    worksheet.write('Q158', '=IFERROR(J144+MATCH(P158,J158:N158,0)-1,"")')
    worksheet.write('Q159', '=IFERROR(J144+MATCH(P159,J159:N159,0)-1,"")')
    worksheet.write('Q160', '=IFERROR(J144+MATCH(P160,J160:N160,0)-1,"")')
    worksheet.write('Q161', '=IFERROR(J144+MATCH(P161,J161:N161,0)-1,"")')
    worksheet.write('Q162', '=IFERROR(J144+MATCH(P162,J162:N162,0)-1,"")')
    worksheet.write('Q163', '=IFERROR(J144+MATCH(P163,J163:N163,0)-1,"")')
    worksheet.write('Q164', '=IFERROR(J144+MATCH(P164,J164:N164,0)-1,"")')
    worksheet.write('Q165', '=IFERROR(J144+MATCH(P165,J165:N165,0)-1,"")')
    worksheet.write('Q166', '=IFERROR(J144+MATCH(P166,J166:N166,0)-1,"")')
    worksheet.write('Q167', '=IFERROR(J144+MATCH(P167,J167:N167,0)-1,"")')
    worksheet.write('Q168', '=IFERROR(J144+MATCH(P168,J168:N168,0)-1,"")')
    worksheet.write('Q169', '=IFERROR(J144+MATCH(P169,J169:N169,0)-1,"")')
    worksheet.write('Q170', '=IFERROR(J144+MATCH(P170,J170:N170,0)-1,"")')
    worksheet.write('Q171', '=IFERROR(J144+MATCH(P171,J171:N171,0)-1,"")')
    worksheet.write('Q172', '=IFERROR(J144+MATCH(P172,J172:N172,0)-1,"")')
    worksheet.write('Q173', '=IFERROR(J144+MATCH(P173,J173:N173,0)-1,"")')
    worksheet.write('Q174', '=IFERROR(J144+MATCH(P174,J174:N174,0)-1,"")')
    worksheet.write('Q175', '=IFERROR(J144+MATCH(P175,J175:N175,0)-1,"")')
    worksheet.write('Q176', '=IFERROR(J144+MATCH(P176,J176:N176,0)-1,"")')
    worksheet.write('Q177', '=IFERROR(J144+MATCH(P177,J177:N177,0)-1,"")')
    worksheet.write('Q178', '=IFERROR(J144+MATCH(P178,J178:N178,0)-1,"")')
    worksheet.write('Q179', '=IFERROR(J144+MATCH(P179,J179:N179,0)-1,"")')
    worksheet.write('Q180', '=IFERROR(J144+MATCH(P180,J180:N180,0)-1,"")')
    worksheet.write('Q181', '=IFERROR(J144+MATCH(P181,J181:N181,0)-1,"")')
    worksheet.write('Q182', '=IFERROR(J144+MATCH(P182,J182:N182,0)-1,"")')
    worksheet.write('Q183', '=IFERROR(J144+MATCH(P183,J183:N183,0)-1,"")')
    worksheet.write('Q184', '=IFERROR(J144+MATCH(P184,J184:N184,0)-1,"")')
    worksheet.write('Q185', '=IFERROR(J144+MATCH(P185,J185:N185,0)-1,"")')
    worksheet.write('Q186', '=IFERROR(J144+MATCH(P186,J186:N186,0)-1,"")')
    worksheet.write('Q187', '=IFERROR(J144+MATCH(P187,J187:N187,0)-1,"")')
    worksheet.write('Q188', '=IFERROR(J144+MATCH(P188,J188:N188,0)-1,"")')
    worksheet.write('Q189', '=IFERROR(J144+MATCH(P189,J189:N189,0)-1,"")')
    worksheet.write('Q190', '=IFERROR(J144+MATCH(P190,J190:N190,0)-1,"")')
    worksheet.write('Q191', '=IFERROR(J144+MATCH(P191,J191:N191,0)-1,"")')
    worksheet.write('Q192', '=IFERROR(J144+MATCH(P192,J192:N192,0)-1,"")')
    worksheet.write('Q193', '=IFERROR(J144+MATCH(P193,J193:N193,0)-1,"")')
    worksheet.write('Q194', '=IFERROR(J144+MATCH(P194,J194:N194,0)-1,"")')
    worksheet.write('Q195', '=IFERROR(J144+MATCH(P195,J195:N195,0)-1,"")')
    worksheet.write('Q196', '=IFERROR(J144+MATCH(P196,J196:N196,0)-1,"")')
    worksheet.write('Q197', '=IFERROR(J144+MATCH(P197,J197:N197,0)-1,"")')
    worksheet.write('Q198', '=IFERROR(J144+MATCH(P198,J198:N198,0)-1,"")')
    worksheet.write('Q199', '=IFERROR(J144+MATCH(P199,J199:N199,0)-1,"")')
    worksheet.write('Q200', '=IFERROR(J144+MATCH(P200,J200:N200,0)-1,"")')
    worksheet.write('Q201', '=IFERROR(J144+MATCH(P201,J201:N201,0)-1,"")')
    worksheet.write('Q202', '=IFERROR(J144+MATCH(P202,J202:N202,0)-1,"")')
    worksheet.write('Q203', '=IFERROR(J144+MATCH(P203,J203:N203,0)-1,"")')
    worksheet.write('Q204', '=IFERROR(J144+MATCH(P204,J204:N204,0)-1,"")')
    worksheet.write('Q205', '=IFERROR(J144+MATCH(P205,J205:N205,0)-1,"")')
    worksheet.write('Q206', '=IFERROR(J144+MATCH(P206,J206:N206,0)-1,"")')
    worksheet.write('Q207', '=IFERROR(J144+MATCH(P207,J207:N207,0)-1,"")')
    worksheet.write('Q208', '=IFERROR(J144+MATCH(P208,J208:N208,0)-1,"")')
    worksheet.write('Q209', '=IFERROR(J144+MATCH(P209,J209:N209,0)-1,"")')
    worksheet.write('Q210', '=IFERROR(J144+MATCH(P210,J210:N210,0)-1,"")')
    worksheet.write('Q211', '=IFERROR(J144+MATCH(P211,J211:N211,0)-1,"")')
    worksheet.write('Q212', '=IFERROR(J144+MATCH(P212,J212:N212,0)-1,"")')
    worksheet.write('Q213', '=IFERROR(J144+MATCH(P213,J213:N213,0)-1,"")')
    worksheet.write('Q214', '=IFERROR(J144+MATCH(P214,J214:N214,0)-1,"")')
    worksheet.write('Q215', '=IFERROR(J144+MATCH(P215,J215:N215,0)-1,"")')
    worksheet.write('Q216', '=IFERROR(J144+MATCH(P216,J216:N216,0)-1,"")')
    worksheet.write('Q217', '=IFERROR(J144+MATCH(P217,J217:N217,0)-1,"")')
    worksheet.write('Q218', '=IFERROR(J144+MATCH(P218,J218:N218,0)-1,"")')
    worksheet.write('Q219', '=IFERROR(J144+MATCH(P219,J219:N219,0)-1,"")')
    worksheet.write('Q220', '=IFERROR(J144+MATCH(P220,J220:N220,0)-1,"")')
    worksheet.write('Q221', '=IFERROR(J144+MATCH(P221,J221:N221,0)-1,"")')
    worksheet.write('Q222', '=IFERROR(J144+MATCH(P222,J222:N222,0)-1,"")')
    worksheet.write('Q223', '=IFERROR(J144+MATCH(P223,J223:N223,0)-1,"")')
    worksheet.write('Q224', '=IFERROR(J144+MATCH(P224,J224:N224,0)-1,"")')
    worksheet.write('Q225', '=IFERROR(J144+MATCH(P225,J225:N225,0)-1,"")')
    worksheet.write('Q226', '=IFERROR(J144+MATCH(P226,J226:N226,0)-1,"")')
    worksheet.write('Q227', '=IFERROR(J144+MATCH(P227,J227:N227,0)-1,"")')
    worksheet.write('Q228', '=IFERROR(J144+MATCH(P228,J228:N228,0)-1,"")')
    worksheet.write('Q229', '=IFERROR(J144+MATCH(P229,J229:N229,0)-1,"")')
    worksheet.write('Q230', '=IFERROR(J144+MATCH(P230,J230:N230,0)-1,"")')
    worksheet.write('Q231', '=IFERROR(J144+MATCH(P231,J231:N231,0)-1,"")')
    worksheet.write('Q232', '=IFERROR(J144+MATCH(P232,J232:N232,0)-1,"")')
    worksheet.write('Q233', '=IFERROR(J144+MATCH(P233,J233:N233,0)-1,"")')
    worksheet.write('Q234', '=IFERROR(J144+MATCH(P234,J234:N234,0)-1,"")')
    worksheet.write('Q235', '=IFERROR(J144+MATCH(P235,J235:N235,0)-1,"")')
    worksheet.write('Q236', '=IFERROR(J144+MATCH(P236,J236:N236,0)-1,"")')
    worksheet.write('Q237', '=IFERROR(J144+MATCH(P237,J237:N237,0)-1,"")')
    worksheet.write('Q238', '=IFERROR(J144+MATCH(P238,J238:N238,0)-1,"")')
    worksheet.write('Q239', '=IFERROR(J144+MATCH(P239,J239:N239,0)-1,"")')
    worksheet.write('Q240', '=IFERROR(J144+MATCH(P240,J240:N240,0)-1,"")')
    worksheet.write('Q241', '=IFERROR(J144+MATCH(P241,J241:N241,0)-1,"")')
    worksheet.write('Q242', '=IFERROR(J144+MATCH(P242,J242:N242,0)-1,"")')
    worksheet.write('Q243', '=IFERROR(J144+MATCH(P243,J243:N243,0)-1,"")')
    worksheet.write('Q244', '=IFERROR(J144+MATCH(P244,J244:N244,0)-1,"")')
    worksheet.write('Q245', '=IFERROR(J144+MATCH(P245,J245:N245,0)-1,"")')
    worksheet.write('Q246', '=IFERROR(J144+MATCH(P246,J246:N246,0)-1,"")')
    worksheet.write('Q247', '=IFERROR(J144+MATCH(P247,J247:N247,0)-1,"")')
    worksheet.write('Q248', '=IFERROR(J144+MATCH(P248,J248:N248,0)-1,"")')
    worksheet.write('Q249', '=IFERROR(J144+MATCH(P249,J249:N249,0)-1,"")')
    worksheet.write('Q250', '=IFERROR(J144+MATCH(P250,J250:N250,0)-1,"")')
    worksheet.write('Q251', '=IFERROR(J144+MATCH(P251,J251:N251,0)-1,"")')
    worksheet.write('Q252', '=IFERROR(J144+MATCH(P252,J252:N252,0)-1,"")')
    worksheet.write('Q253', '=IFERROR(J144+MATCH(P253,J253:N253,0)-1,"")')
    worksheet.write('Q254', '=IFERROR(J144+MATCH(P254,J254:N254,0)-1,"")')
    worksheet.write('Q255', '=IFERROR(J144+MATCH(P255,J255:N255,0)-1,"")')
    worksheet.write('Q256', '=IFERROR(J144+MATCH(P256,J256:N256,0)-1,"")')
    worksheet.write('Q257', '=IFERROR(J144+MATCH(P257,J257:N257,0)-1,"")')
    worksheet.write('Q258', '=IFERROR(J144+MATCH(P258,J258:N258,0)-1,"")')
    worksheet.write('Q259', '=IFERROR(J144+MATCH(P259,J259:N259,0)-1,"")')
    worksheet.write('Q260', '=IFERROR(J144+MATCH(P260,J260:N260,0)-1,"")')
    worksheet.write('Q261', '=IFERROR(J144+MATCH(P261,J261:N261,0)-1,"")')
    worksheet.write('Q262', '=IFERROR(J144+MATCH(P262,J262:N262,0)-1,"")')
    worksheet.write('Q263', '=IFERROR(J144+MATCH(P263,J263:N263,0)-1,"")')
    worksheet.write('Q264', '=IFERROR(J144+MATCH(P264,J264:N264,0)-1,"")')
    worksheet.write('Q265', '=IFERROR(J144+MATCH(P265,J265:N265,0)-1,"")')
    worksheet.write('Q266', '=IFERROR(J144+MATCH(P266,J266:N266,0)-1,"")')
    worksheet.write('Q267', '=IFERROR(J144+MATCH(P267,J267:N267,0)-1,"")')
    worksheet.write('Q268', '=IFERROR(J144+MATCH(P268,J268:N268,0)-1,"")')
    worksheet.write('Q269', '=IFERROR(J144+MATCH(P269,J269:N269,0)-1,"")')
    worksheet.write('Q270', '=IFERROR(J144+MATCH(P270,J270:N270,0)-1,"")')
    worksheet.write('Q271', '=IFERROR(J144+MATCH(P271,J271:N271,0)-1,"")')
    worksheet.write('Q272', '=IFERROR(J144+MATCH(P272,J272:N272,0)-1,"")')
    worksheet.write('Q273', '=IFERROR(J144+MATCH(P273,J273:N273,0)-1,"")')
    worksheet.write('Q274', '=IFERROR(J144+MATCH(P274,J274:N274,0)-1,"")')
    worksheet.write('Q275', '=IFERROR(J144+MATCH(P275,J275:N275,0)-1,"")')
    worksheet.write('Q276', '=IFERROR(J144+MATCH(P276,J276:N276,0)-1,"")')
    worksheet.write('Q277', '=IFERROR(J144+MATCH(P277,J277:N277,0)-1,"")')
    worksheet.write('Q278', '=IFERROR(J144+MATCH(P278,J278:N278,0)-1,"")')
    worksheet.write('Q279', '=IFERROR(J144+MATCH(P279,J279:N279,0)-1,"")')
    worksheet.write('Q280', '=IFERROR(J144+MATCH(P280,J280:N280,0)-1,"")')
    worksheet.write('Q281', '=IFERROR(J144+MATCH(P281,J281:N281,0)-1,"")')
    worksheet.write('Q282', '=IFERROR(J144+MATCH(P282,J282:N282,0)-1,"")')
    worksheet.write('Q283', '=IFERROR(J144+MATCH(P283,J283:N283,0)-1,"")')
    worksheet.write('Q284', '=IFERROR(J144+MATCH(P284,J284:N284,0)-1,"")')
    worksheet.write('Q285', '=IFERROR(J144+MATCH(P285,J285:N285,0)-1,"")')
    worksheet.write('Q286', '=IFERROR(J144+MATCH(P286,J286:N286,0)-1,"")')
    worksheet.write('Q287', '=IFERROR(J144+MATCH(P287,J287:N287,0)-1,"")')
    worksheet.write('Q288', '=IFERROR(J144+MATCH(P288,J288:N288,0)-1,"")')
    worksheet.write('Q289', '=IFERROR(J144+MATCH(P289,J289:N289,0)-1,"")')
    worksheet.write('Q290', '=IFERROR(J144+MATCH(P290,J290:N290,0)-1,"")')
    worksheet.write('Q291', '=IFERROR(J144+MATCH(P291,J291:N291,0)-1,"")')
    worksheet.write('Q292', '=IFERROR(J144+MATCH(P292,J292:N292,0)-1,"")')
    worksheet.write('Q293', '=IFERROR(J144+MATCH(P293,J293:N293,0)-1,"")')
    worksheet.write('Q294', '=IFERROR(J144+MATCH(P294,J294:N294,0)-1,"")')
    worksheet.write('Q295', '=IFERROR(J144+MATCH(P295,J295:N295,0)-1,"")')
    worksheet.write('Q296', '=IFERROR(J144+MATCH(P296,J296:N296,0)-1,"")')
    worksheet.write('Q297', '=IFERROR(J144+MATCH(P297,J297:N297,0)-1,"")')
    worksheet.write('Q298', '=IFERROR(J144+MATCH(P298,J298:N298,0)-1,"")')
    worksheet.write('Q299', '=IFERROR(J144+MATCH(P299,J299:N299,0)-1,"")')
    worksheet.write('Q300', '=IFERROR(J144+MATCH(P300,J300:N300,0)-1,"")')
    worksheet.write('Q301', '=IFERROR(J144+MATCH(P301,J301:N301,0)-1,"")')
    worksheet.write('Q302', '=IFERROR(J144+MATCH(P302,J302:N302,0)-1,"")')
    worksheet.write('Q303', '=IFERROR(J144+MATCH(P303,J303:N303,0)-1,"")')
    worksheet.write('Q304', '=IFERROR(J144+MATCH(P304,J304:N304,0)-1,"")')
    worksheet.write('Q305', '=IFERROR(J144+MATCH(P305,J305:N305,0)-1,"")')
    worksheet.write('Q306', '=IFERROR(J144+MATCH(P306,J306:N306,0)-1,"")')
    worksheet.write('Q307', '=IFERROR(J144+MATCH(P307,J307:N307,0)-1,"")')
    worksheet.write('Q308', '=IFERROR(J144+MATCH(P308,J308:N308,0)-1,"")')
    worksheet.write('Q309', '=IFERROR(J144+MATCH(P309,J309:N309,0)-1,"")')
    worksheet.write('Q310', '=IFERROR(J144+MATCH(P310,J310:N310,0)-1,"")')
    worksheet.write('Q311', '=IFERROR(J144+MATCH(P311,J311:N311,0)-1,"")')
    worksheet.write('Q312', '=IFERROR(J144+MATCH(P312,J312:N312,0)-1,"")')
    worksheet.write('Q313', '=IFERROR(J144+MATCH(P313,J313:N313,0)-1,"")')
    worksheet.write('Q314', '=IFERROR(J144+MATCH(P314,J314:N314,0)-1,"")')
    worksheet.write('Q315', '=IFERROR(J144+MATCH(P315,J315:N315,0)-1,"")')
    worksheet.write('Q316', '=IFERROR(J144+MATCH(P316,J316:N316,0)-1,"")')
    worksheet.write('Q317', '=IFERROR(J144+MATCH(P317,J317:N317,0)-1,"")')
    worksheet.write('Q318', '=IFERROR(J144+MATCH(P318,J318:N318,0)-1,"")')
    worksheet.write('Q319', '=IFERROR(J144+MATCH(P319,J319:N319,0)-1,"")')
    worksheet.write('Q320', '=IFERROR(J144+MATCH(P320,J320:N320,0)-1,"")')
    worksheet.write('Q321', '=IFERROR(J144+MATCH(P321,J321:N321,0)-1,"")')
    worksheet.write('Q322', '=IFERROR(J144+MATCH(P322,J322:N322,0)-1,"")')
    worksheet.write('Q323', '=IFERROR(J144+MATCH(P323,J323:N323,0)-1,"")')
    worksheet.write('Q324', '=IFERROR(J144+MATCH(P324,J324:N324,0)-1,"")')
    worksheet.write('Q325', '=IFERROR(J144+MATCH(P325,J325:N325,0)-1,"")')
    worksheet.write('Q326', '=IFERROR(J144+MATCH(P326,J326:N326,0)-1,"")')
    worksheet.write('Q327', '=IFERROR(J144+MATCH(P327,J327:N327,0)-1,"")')
    worksheet.write('Q328', '=IFERROR(J144+MATCH(P328,J328:N328,0)-1,"")')
    worksheet.write('Q329', '=IFERROR(J144+MATCH(P329,J329:N329,0)-1,"")')
    worksheet.write('Q330', '=IFERROR(J144+MATCH(P330,J330:N330,0)-1,"")')
    worksheet.write('Q331', '=IFERROR(J144+MATCH(P331,J331:N331,0)-1,"")')
    worksheet.write('Q332', '=IFERROR(J144+MATCH(P332,J332:N332,0)-1,"")')
    worksheet.write('Q333', '=IFERROR(J144+MATCH(P333,J333:N333,0)-1,"")')
    worksheet.write('Q334', '=IFERROR(J144+MATCH(P334,J334:N334,0)-1,"")')
    worksheet.write('Q335', '=IFERROR(J144+MATCH(P335,J335:N335,0)-1,"")')
    worksheet.write('Q336', '=IFERROR(J144+MATCH(P336,J336:N336,0)-1,"")')
    worksheet.write('Q337', '=IFERROR(J144+MATCH(P337,J337:N337,0)-1,"")')
    worksheet.write('Q338', '=IFERROR(J144+MATCH(P338,J338:N338,0)-1,"")')
    worksheet.write('Q339', '=IFERROR(J144+MATCH(P339,J339:N339,0)-1,"")')
    worksheet.write('Q340', '=IFERROR(J144+MATCH(P340,J340:N340,0)-1,"")')
    worksheet.write('Q341', '=IFERROR(J144+MATCH(P341,J341:N341,0)-1,"")')
    worksheet.write('Q342', '=IFERROR(J144+MATCH(P342,J342:N342,0)-1,"")')
    worksheet.write('Q343', '=IFERROR(J144+MATCH(P343,J343:N343,0)-1,"")')
    worksheet.write('Q344', '=IFERROR(J144+MATCH(P344,J344:N344,0)-1,"")')
    worksheet.write('Q345', '=IFERROR(J144+MATCH(P345,J345:N345,0)-1,"")')
    worksheet.write('Q346', '=IFERROR(J144+MATCH(P346,J346:N346,0)-1,"")')
    worksheet.write('Q347', '=IFERROR(J144+MATCH(P347,J347:N347,0)-1,"")')
    worksheet.write('Q348', '=IFERROR(J144+MATCH(P348,J348:N348,0)-1,"")')
    worksheet.write('Q349', '=IFERROR(J144+MATCH(P349,J349:N349,0)-1,"")')
    worksheet.write('Q350', '=IFERROR(J144+MATCH(P350,J350:N350,0)-1,"")')
    worksheet.write('Q351', '=IFERROR(J144+MATCH(P351,J351:N351,0)-1,"")')
    worksheet.write('Q352', '=IFERROR(J144+MATCH(P352,J352:N352,0)-1,"")')
    worksheet.write('Q353', '=IFERROR(J144+MATCH(P353,J353:N353,0)-1,"")')
    worksheet.write('R144', '="Min"')
    worksheet.write('R145', '=IF(Q145="","",MIN(J145:N145))')
    worksheet.write('R146', '=IF(Q146="","",MIN(J146:N146))')
    worksheet.write('R147', '=IF(Q147="","",MIN(J147:N147))')
    worksheet.write('R148', '=IF(Q148="","",MIN(J148:N148))')
    worksheet.write('R149', '=IF(Q149="","",MIN(J149:N149))')
    worksheet.write('R150', '=IF(Q150="","",MIN(J150:N150))')
    worksheet.write('R151', '=IF(Q151="","",MIN(J151:N151))')
    worksheet.write('R152', '=IF(Q152="","",MIN(J152:N152))')
    worksheet.write('R153', '=IF(Q153="","",MIN(J153:N153))')
    worksheet.write('R154', '=IF(Q154="","",MIN(J154:N154))')
    worksheet.write('R155', '=IF(Q155="","",MIN(J155:N155))')
    worksheet.write('R156', '=IF(Q156="","",MIN(J156:N156))')
    worksheet.write('R157', '=IF(Q157="","",MIN(J157:N157))')
    worksheet.write('R158', '=IF(Q158="","",MIN(J158:N158))')
    worksheet.write('R159', '=IF(Q159="","",MIN(J159:N159))')
    worksheet.write('R160', '=IF(Q160="","",MIN(J160:N160))')
    worksheet.write('R161', '=IF(Q161="","",MIN(J161:N161))')
    worksheet.write('R162', '=IF(Q162="","",MIN(J162:N162))')
    worksheet.write('R163', '=IF(Q163="","",MIN(J163:N163))')
    worksheet.write('R164', '=IF(Q164="","",MIN(J164:N164))')
    worksheet.write('R165', '=IF(Q165="","",MIN(J165:N165))')
    worksheet.write('R166', '=IF(Q166="","",MIN(J166:N166))')
    worksheet.write('R167', '=IF(Q167="","",MIN(J167:N167))')
    worksheet.write('R168', '=IF(Q168="","",MIN(J168:N168))')
    worksheet.write('R169', '=IF(Q169="","",MIN(J169:N169))')
    worksheet.write('R170', '=IF(Q170="","",MIN(J170:N170))')
    worksheet.write('R171', '=IF(Q171="","",MIN(J171:N171))')
    worksheet.write('R172', '=IF(Q172="","",MIN(J172:N172))')
    worksheet.write('R173', '=IF(Q173="","",MIN(J173:N173))')
    worksheet.write('R174', '=IF(Q174="","",MIN(J174:N174))')
    worksheet.write('R175', '=IF(Q175="","",MIN(J175:N175))')
    worksheet.write('R176', '=IF(Q176="","",MIN(J176:N176))')
    worksheet.write('R177', '=IF(Q177="","",MIN(J177:N177))')
    worksheet.write('R178', '=IF(Q178="","",MIN(J178:N178))')
    worksheet.write('R179', '=IF(Q179="","",MIN(J179:N179))')
    worksheet.write('R180', '=IF(Q180="","",MIN(J180:N180))')
    worksheet.write('R181', '=IF(Q181="","",MIN(J181:N181))')
    worksheet.write('R182', '=IF(Q182="","",MIN(J182:N182))')
    worksheet.write('R183', '=IF(Q183="","",MIN(J183:N183))')
    worksheet.write('R184', '=IF(Q184="","",MIN(J184:N184))')
    worksheet.write('R185', '=IF(Q185="","",MIN(J185:N185))')
    worksheet.write('R186', '=IF(Q186="","",MIN(J186:N186))')
    worksheet.write('R187', '=IF(Q187="","",MIN(J187:N187))')
    worksheet.write('R188', '=IF(Q188="","",MIN(J188:N188))')
    worksheet.write('R189', '=IF(Q189="","",MIN(J189:N189))')
    worksheet.write('R190', '=IF(Q190="","",MIN(J190:N190))')
    worksheet.write('R191', '=IF(Q191="","",MIN(J191:N191))')
    worksheet.write('R192', '=IF(Q192="","",MIN(J192:N192))')
    worksheet.write('R193', '=IF(Q193="","",MIN(J193:N193))')
    worksheet.write('R194', '=IF(Q194="","",MIN(J194:N194))')
    worksheet.write('R195', '=IF(Q195="","",MIN(J195:N195))')
    worksheet.write('R196', '=IF(Q196="","",MIN(J196:N196))')
    worksheet.write('R197', '=IF(Q197="","",MIN(J197:N197))')
    worksheet.write('R198', '=IF(Q198="","",MIN(J198:N198))')
    worksheet.write('R199', '=IF(Q199="","",MIN(J199:N199))')
    worksheet.write('R200', '=IF(Q200="","",MIN(J200:N200))')
    worksheet.write('R201', '=IF(Q201="","",MIN(J201:N201))')
    worksheet.write('R202', '=IF(Q202="","",MIN(J202:N202))')
    worksheet.write('R203', '=IF(Q203="","",MIN(J203:N203))')
    worksheet.write('R204', '=IF(Q204="","",MIN(J204:N204))')
    worksheet.write('R205', '=IF(Q205="","",MIN(J205:N205))')
    worksheet.write('R206', '=IF(Q206="","",MIN(J206:N206))')
    worksheet.write('R207', '=IF(Q207="","",MIN(J207:N207))')
    worksheet.write('R208', '=IF(Q208="","",MIN(J208:N208))')
    worksheet.write('R209', '=IF(Q209="","",MIN(J209:N209))')
    worksheet.write('R210', '=IF(Q210="","",MIN(J210:N210))')
    worksheet.write('R211', '=IF(Q211="","",MIN(J211:N211))')
    worksheet.write('R212', '=IF(Q212="","",MIN(J212:N212))')
    worksheet.write('R213', '=IF(Q213="","",MIN(J213:N213))')
    worksheet.write('R214', '=IF(Q214="","",MIN(J214:N214))')
    worksheet.write('R215', '=IF(Q215="","",MIN(J215:N215))')
    worksheet.write('R216', '=IF(Q216="","",MIN(J216:N216))')
    worksheet.write('R217', '=IF(Q217="","",MIN(J217:N217))')
    worksheet.write('R218', '=IF(Q218="","",MIN(J218:N218))')
    worksheet.write('R219', '=IF(Q219="","",MIN(J219:N219))')
    worksheet.write('R220', '=IF(Q220="","",MIN(J220:N220))')
    worksheet.write('R221', '=IF(Q221="","",MIN(J221:N221))')
    worksheet.write('R222', '=IF(Q222="","",MIN(J222:N222))')
    worksheet.write('R223', '=IF(Q223="","",MIN(J223:N223))')
    worksheet.write('R224', '=IF(Q224="","",MIN(J224:N224))')
    worksheet.write('R225', '=IF(Q225="","",MIN(J225:N225))')
    worksheet.write('R226', '=IF(Q226="","",MIN(J226:N226))')
    worksheet.write('R227', '=IF(Q227="","",MIN(J227:N227))')
    worksheet.write('R228', '=IF(Q228="","",MIN(J228:N228))')
    worksheet.write('R229', '=IF(Q229="","",MIN(J229:N229))')
    worksheet.write('R230', '=IF(Q230="","",MIN(J230:N230))')
    worksheet.write('R231', '=IF(Q231="","",MIN(J231:N231))')
    worksheet.write('R232', '=IF(Q232="","",MIN(J232:N232))')
    worksheet.write('R233', '=IF(Q233="","",MIN(J233:N233))')
    worksheet.write('R234', '=IF(Q234="","",MIN(J234:N234))')
    worksheet.write('R235', '=IF(Q235="","",MIN(J235:N235))')
    worksheet.write('R236', '=IF(Q236="","",MIN(J236:N236))')
    worksheet.write('R237', '=IF(Q237="","",MIN(J237:N237))')
    worksheet.write('R238', '=IF(Q238="","",MIN(J238:N238))')
    worksheet.write('R239', '=IF(Q239="","",MIN(J239:N239))')
    worksheet.write('R240', '=IF(Q240="","",MIN(J240:N240))')
    worksheet.write('R241', '=IF(Q241="","",MIN(J241:N241))')
    worksheet.write('R242', '=IF(Q242="","",MIN(J242:N242))')
    worksheet.write('R243', '=IF(Q243="","",MIN(J243:N243))')
    worksheet.write('R244', '=IF(Q244="","",MIN(J244:N244))')
    worksheet.write('R245', '=IF(Q245="","",MIN(J245:N245))')
    worksheet.write('R246', '=IF(Q246="","",MIN(J246:N246))')
    worksheet.write('R247', '=IF(Q247="","",MIN(J247:N247))')
    worksheet.write('R248', '=IF(Q248="","",MIN(J248:N248))')
    worksheet.write('R249', '=IF(Q249="","",MIN(J249:N249))')
    worksheet.write('R250', '=IF(Q250="","",MIN(J250:N250))')
    worksheet.write('R251', '=IF(Q251="","",MIN(J251:N251))')
    worksheet.write('R252', '=IF(Q252="","",MIN(J252:N252))')
    worksheet.write('R253', '=IF(Q253="","",MIN(J253:N253))')
    worksheet.write('R254', '=IF(Q254="","",MIN(J254:N254))')
    worksheet.write('R255', '=IF(Q255="","",MIN(J255:N255))')
    worksheet.write('R256', '=IF(Q256="","",MIN(J256:N256))')
    worksheet.write('R257', '=IF(Q257="","",MIN(J257:N257))')
    worksheet.write('R258', '=IF(Q258="","",MIN(J258:N258))')
    worksheet.write('R259', '=IF(Q259="","",MIN(J259:N259))')
    worksheet.write('R260', '=IF(Q260="","",MIN(J260:N260))')
    worksheet.write('R261', '=IF(Q261="","",MIN(J261:N261))')
    worksheet.write('R262', '=IF(Q262="","",MIN(J262:N262))')
    worksheet.write('R263', '=IF(Q263="","",MIN(J263:N263))')
    worksheet.write('R264', '=IF(Q264="","",MIN(J264:N264))')
    worksheet.write('R265', '=IF(Q265="","",MIN(J265:N265))')
    worksheet.write('R266', '=IF(Q266="","",MIN(J266:N266))')
    worksheet.write('R267', '=IF(Q267="","",MIN(J267:N267))')
    worksheet.write('R268', '=IF(Q268="","",MIN(J268:N268))')
    worksheet.write('R269', '=IF(Q269="","",MIN(J269:N269))')
    worksheet.write('R270', '=IF(Q270="","",MIN(J270:N270))')
    worksheet.write('R271', '=IF(Q271="","",MIN(J271:N271))')
    worksheet.write('R272', '=IF(Q272="","",MIN(J272:N272))')
    worksheet.write('R273', '=IF(Q273="","",MIN(J273:N273))')
    worksheet.write('R274', '=IF(Q274="","",MIN(J274:N274))')
    worksheet.write('R275', '=IF(Q275="","",MIN(J275:N275))')
    worksheet.write('R276', '=IF(Q276="","",MIN(J276:N276))')
    worksheet.write('R277', '=IF(Q277="","",MIN(J277:N277))')
    worksheet.write('R278', '=IF(Q278="","",MIN(J278:N278))')
    worksheet.write('R279', '=IF(Q279="","",MIN(J279:N279))')
    worksheet.write('R280', '=IF(Q280="","",MIN(J280:N280))')
    worksheet.write('R281', '=IF(Q281="","",MIN(J281:N281))')
    worksheet.write('R282', '=IF(Q282="","",MIN(J282:N282))')
    worksheet.write('R283', '=IF(Q283="","",MIN(J283:N283))')
    worksheet.write('R284', '=IF(Q284="","",MIN(J284:N284))')
    worksheet.write('R285', '=IF(Q285="","",MIN(J285:N285))')
    worksheet.write('R286', '=IF(Q286="","",MIN(J286:N286))')
    worksheet.write('R287', '=IF(Q287="","",MIN(J287:N287))')
    worksheet.write('R288', '=IF(Q288="","",MIN(J288:N288))')
    worksheet.write('R289', '=IF(Q289="","",MIN(J289:N289))')
    worksheet.write('R290', '=IF(Q290="","",MIN(J290:N290))')
    worksheet.write('R291', '=IF(Q291="","",MIN(J291:N291))')
    worksheet.write('R292', '=IF(Q292="","",MIN(J292:N292))')
    worksheet.write('R293', '=IF(Q293="","",MIN(J293:N293))')
    worksheet.write('R294', '=IF(Q294="","",MIN(J294:N294))')
    worksheet.write('R295', '=IF(Q295="","",MIN(J295:N295))')
    worksheet.write('R296', '=IF(Q296="","",MIN(J296:N296))')
    worksheet.write('R297', '=IF(Q297="","",MIN(J297:N297))')
    worksheet.write('R298', '=IF(Q298="","",MIN(J298:N298))')
    worksheet.write('R299', '=IF(Q299="","",MIN(J299:N299))')
    worksheet.write('R300', '=IF(Q300="","",MIN(J300:N300))')
    worksheet.write('R301', '=IF(Q301="","",MIN(J301:N301))')
    worksheet.write('R302', '=IF(Q302="","",MIN(J302:N302))')
    worksheet.write('R303', '=IF(Q303="","",MIN(J303:N303))')
    worksheet.write('R304', '=IF(Q304="","",MIN(J304:N304))')
    worksheet.write('R305', '=IF(Q305="","",MIN(J305:N305))')
    worksheet.write('R306', '=IF(Q306="","",MIN(J306:N306))')
    worksheet.write('R307', '=IF(Q307="","",MIN(J307:N307))')
    worksheet.write('R308', '=IF(Q308="","",MIN(J308:N308))')
    worksheet.write('R309', '=IF(Q309="","",MIN(J309:N309))')
    worksheet.write('R310', '=IF(Q310="","",MIN(J310:N310))')
    worksheet.write('R311', '=IF(Q311="","",MIN(J311:N311))')
    worksheet.write('R312', '=IF(Q312="","",MIN(J312:N312))')
    worksheet.write('R313', '=IF(Q313="","",MIN(J313:N313))')
    worksheet.write('R314', '=IF(Q314="","",MIN(J314:N314))')
    worksheet.write('R315', '=IF(Q315="","",MIN(J315:N315))')
    worksheet.write('R316', '=IF(Q316="","",MIN(J316:N316))')
    worksheet.write('R317', '=IF(Q317="","",MIN(J317:N317))')
    worksheet.write('R318', '=IF(Q318="","",MIN(J318:N318))')
    worksheet.write('R319', '=IF(Q319="","",MIN(J319:N319))')
    worksheet.write('R320', '=IF(Q320="","",MIN(J320:N320))')
    worksheet.write('R321', '=IF(Q321="","",MIN(J321:N321))')
    worksheet.write('R322', '=IF(Q322="","",MIN(J322:N322))')
    worksheet.write('R323', '=IF(Q323="","",MIN(J323:N323))')
    worksheet.write('R324', '=IF(Q324="","",MIN(J324:N324))')
    worksheet.write('R325', '=IF(Q325="","",MIN(J325:N325))')
    worksheet.write('R326', '=IF(Q326="","",MIN(J326:N326))')
    worksheet.write('R327', '=IF(Q327="","",MIN(J327:N327))')
    worksheet.write('R328', '=IF(Q328="","",MIN(J328:N328))')
    worksheet.write('R329', '=IF(Q329="","",MIN(J329:N329))')
    worksheet.write('R330', '=IF(Q330="","",MIN(J330:N330))')
    worksheet.write('R331', '=IF(Q331="","",MIN(J331:N331))')
    worksheet.write('R332', '=IF(Q332="","",MIN(J332:N332))')
    worksheet.write('R333', '=IF(Q333="","",MIN(J333:N333))')
    worksheet.write('R334', '=IF(Q334="","",MIN(J334:N334))')
    worksheet.write('R335', '=IF(Q335="","",MIN(J335:N335))')
    worksheet.write('R336', '=IF(Q336="","",MIN(J336:N336))')
    worksheet.write('R337', '=IF(Q337="","",MIN(J337:N337))')
    worksheet.write('R338', '=IF(Q338="","",MIN(J338:N338))')
    worksheet.write('R339', '=IF(Q339="","",MIN(J339:N339))')
    worksheet.write('R340', '=IF(Q340="","",MIN(J340:N340))')
    worksheet.write('R341', '=IF(Q341="","",MIN(J341:N341))')
    worksheet.write('R342', '=IF(Q342="","",MIN(J342:N342))')
    worksheet.write('R343', '=IF(Q343="","",MIN(J343:N343))')
    worksheet.write('R344', '=IF(Q344="","",MIN(J344:N344))')
    worksheet.write('R345', '=IF(Q345="","",MIN(J345:N345))')
    worksheet.write('R346', '=IF(Q346="","",MIN(J346:N346))')
    worksheet.write('R347', '=IF(Q347="","",MIN(J347:N347))')
    worksheet.write('R348', '=IF(Q348="","",MIN(J348:N348))')
    worksheet.write('R349', '=IF(Q349="","",MIN(J349:N349))')
    worksheet.write('R350', '=IF(Q350="","",MIN(J350:N350))')
    worksheet.write('R351', '=IF(Q351="","",MIN(J351:N351))')
    worksheet.write('R352', '=IF(Q352="","",MIN(J352:N352))')
    worksheet.write('R353', '=IF(Q353="","",MIN(J353:N353))')
    worksheet.write('S144', '="Min Year"')
    worksheet.write('S145', '=IFERROR(J144+MATCH(R145,J145:N145)-1,"")')
    worksheet.write('S146', '=IFERROR(J144+MATCH(R146,J146:N146)-1,"")')
    worksheet.write('S147', '=IFERROR(J144+MATCH(R147,J147:N147)-1,"")')
    worksheet.write('S148', '=IFERROR(J144+MATCH(R148,J148:N148)-1,"")')
    worksheet.write('S149', '=IFERROR(J144+MATCH(R149,J149:N149)-1,"")')
    worksheet.write('S150', '=IFERROR(J144+MATCH(R150,J150:N150)-1,"")')
    worksheet.write('S151', '=IFERROR(J144+MATCH(R151,J151:N151)-1,"")')
    worksheet.write('S152', '=IFERROR(J144+MATCH(R152,J152:N152)-1,"")')
    worksheet.write('S153', '=IFERROR(J144+MATCH(R153,J153:N153)-1,"")')
    worksheet.write('S154', '=IFERROR(J144+MATCH(R154,J154:N154)-1,"")')
    worksheet.write('S155', '=IFERROR(J144+MATCH(R155,J155:N155)-1,"")')
    worksheet.write('S156', '=IFERROR(J144+MATCH(R156,J156:N156)-1,"")')
    worksheet.write('S157', '=IFERROR(J144+MATCH(R157,J157:N157)-1,"")')
    worksheet.write('S158', '=IFERROR(J144+MATCH(R158,J158:N158)-1,"")')
    worksheet.write('S159', '=IFERROR(J144+MATCH(R159,J159:N159)-1,"")')
    worksheet.write('S160', '=IFERROR(J144+MATCH(R160,J160:N160)-1,"")')
    worksheet.write('S161', '=IFERROR(J144+MATCH(R161,J161:N161)-1,"")')
    worksheet.write('S162', '=IFERROR(J144+MATCH(R162,J162:N162)-1,"")')
    worksheet.write('S163', '=IFERROR(J144+MATCH(R163,J163:N163)-1,"")')
    worksheet.write('S164', '=IFERROR(J144+MATCH(R164,J164:N164)-1,"")')
    worksheet.write('S165', '=IFERROR(J144+MATCH(R165,J165:N165)-1,"")')
    worksheet.write('S166', '=IFERROR(J144+MATCH(R166,J166:N166)-1,"")')
    worksheet.write('S167', '=IFERROR(J144+MATCH(R167,J167:N167)-1,"")')
    worksheet.write('S168', '=IFERROR(J144+MATCH(R168,J168:N168)-1,"")')
    worksheet.write('S169', '=IFERROR(J144+MATCH(R169,J169:N169)-1,"")')
    worksheet.write('S170', '=IFERROR(J144+MATCH(R170,J170:N170)-1,"")')
    worksheet.write('S171', '=IFERROR(J144+MATCH(R171,J171:N171)-1,"")')
    worksheet.write('S172', '=IFERROR(J144+MATCH(R172,J172:N172)-1,"")')
    worksheet.write('S173', '=IFERROR(J144+MATCH(R173,J173:N173)-1,"")')
    worksheet.write('S174', '=IFERROR(J144+MATCH(R174,J174:N174)-1,"")')
    worksheet.write('S175', '=IFERROR(J144+MATCH(R175,J175:N175)-1,"")')
    worksheet.write('S176', '=IFERROR(J144+MATCH(R176,J176:N176)-1,"")')
    worksheet.write('S177', '=IFERROR(J144+MATCH(R177,J177:N177)-1,"")')
    worksheet.write('S178', '=IFERROR(J144+MATCH(R178,J178:N178)-1,"")')
    worksheet.write('S179', '=IFERROR(J144+MATCH(R179,J179:N179)-1,"")')
    worksheet.write('S180', '=IFERROR(J144+MATCH(R180,J180:N180)-1,"")')
    worksheet.write('S181', '=IFERROR(J144+MATCH(R181,J181:N181)-1,"")')
    worksheet.write('S182', '=IFERROR(J144+MATCH(R182,J182:N182)-1,"")')
    worksheet.write('S183', '=IFERROR(J144+MATCH(R183,J183:N183)-1,"")')
    worksheet.write('S184', '=IFERROR(J144+MATCH(R184,J184:N184)-1,"")')
    worksheet.write('S185', '=IFERROR(J144+MATCH(R185,J185:N185)-1,"")')
    worksheet.write('S186', '=IFERROR(J144+MATCH(R186,J186:N186)-1,"")')
    worksheet.write('S187', '=IFERROR(J144+MATCH(R187,J187:N187)-1,"")')
    worksheet.write('S188', '=IFERROR(J144+MATCH(R188,J188:N188)-1,"")')
    worksheet.write('S189', '=IFERROR(J144+MATCH(R189,J189:N189)-1,"")')
    worksheet.write('S190', '=IFERROR(J144+MATCH(R190,J190:N190)-1,"")')
    worksheet.write('S191', '=IFERROR(J144+MATCH(R191,J191:N191)-1,"")')
    worksheet.write('S192', '=IFERROR(J144+MATCH(R192,J192:N192)-1,"")')
    worksheet.write('S193', '=IFERROR(J144+MATCH(R193,J193:N193)-1,"")')
    worksheet.write('S194', '=IFERROR(J144+MATCH(R194,J194:N194)-1,"")')
    worksheet.write('S195', '=IFERROR(J144+MATCH(R195,J195:N195)-1,"")')
    worksheet.write('S196', '=IFERROR(J144+MATCH(R196,J196:N196)-1,"")')
    worksheet.write('S197', '=IFERROR(J144+MATCH(R197,J197:N197)-1,"")')
    worksheet.write('S198', '=IFERROR(J144+MATCH(R198,J198:N198)-1,"")')
    worksheet.write('S199', '=IFERROR(J144+MATCH(R199,J199:N199)-1,"")')
    worksheet.write('S200', '=IFERROR(J144+MATCH(R200,J200:N200)-1,"")')
    worksheet.write('S201', '=IFERROR(J144+MATCH(R201,J201:N201)-1,"")')
    worksheet.write('S202', '=IFERROR(J144+MATCH(R202,J202:N202)-1,"")')
    worksheet.write('S203', '=IFERROR(J144+MATCH(R203,J203:N203)-1,"")')
    worksheet.write('S204', '=IFERROR(J144+MATCH(R204,J204:N204)-1,"")')
    worksheet.write('S205', '=IFERROR(J144+MATCH(R205,J205:N205)-1,"")')
    worksheet.write('S206', '=IFERROR(J144+MATCH(R206,J206:N206)-1,"")')
    worksheet.write('S207', '=IFERROR(J144+MATCH(R207,J207:N207)-1,"")')
    worksheet.write('S208', '=IFERROR(J144+MATCH(R208,J208:N208)-1,"")')
    worksheet.write('S209', '=IFERROR(J144+MATCH(R209,J209:N209)-1,"")')
    worksheet.write('S210', '=IFERROR(J144+MATCH(R210,J210:N210)-1,"")')
    worksheet.write('S211', '=IFERROR(J144+MATCH(R211,J211:N211)-1,"")')
    worksheet.write('S212', '=IFERROR(J144+MATCH(R212,J212:N212)-1,"")')
    worksheet.write('S213', '=IFERROR(J144+MATCH(R213,J213:N213)-1,"")')
    worksheet.write('S214', '=IFERROR(J144+MATCH(R214,J214:N214)-1,"")')
    worksheet.write('S215', '=IFERROR(J144+MATCH(R215,J215:N215)-1,"")')
    worksheet.write('S216', '=IFERROR(J144+MATCH(R216,J216:N216)-1,"")')
    worksheet.write('S217', '=IFERROR(J144+MATCH(R217,J217:N217)-1,"")')
    worksheet.write('S218', '=IFERROR(J144+MATCH(R218,J218:N218)-1,"")')
    worksheet.write('S219', '=IFERROR(J144+MATCH(R219,J219:N219)-1,"")')
    worksheet.write('S220', '=IFERROR(J144+MATCH(R220,J220:N220)-1,"")')
    worksheet.write('S221', '=IFERROR(J144+MATCH(R221,J221:N221)-1,"")')
    worksheet.write('S222', '=IFERROR(J144+MATCH(R222,J222:N222)-1,"")')
    worksheet.write('S223', '=IFERROR(J144+MATCH(R223,J223:N223)-1,"")')
    worksheet.write('S224', '=IFERROR(J144+MATCH(R224,J224:N224)-1,"")')
    worksheet.write('S225', '=IFERROR(J144+MATCH(R225,J225:N225)-1,"")')
    worksheet.write('S226', '=IFERROR(J144+MATCH(R226,J226:N226)-1,"")')
    worksheet.write('S227', '=IFERROR(J144+MATCH(R227,J227:N227)-1,"")')
    worksheet.write('S228', '=IFERROR(J144+MATCH(R228,J228:N228)-1,"")')
    worksheet.write('S229', '=IFERROR(J144+MATCH(R229,J229:N229)-1,"")')
    worksheet.write('S230', '=IFERROR(J144+MATCH(R230,J230:N230)-1,"")')
    worksheet.write('S231', '=IFERROR(J144+MATCH(R231,J231:N231)-1,"")')
    worksheet.write('S232', '=IFERROR(J144+MATCH(R232,J232:N232)-1,"")')
    worksheet.write('S233', '=IFERROR(J144+MATCH(R233,J233:N233)-1,"")')
    worksheet.write('S234', '=IFERROR(J144+MATCH(R234,J234:N234)-1,"")')
    worksheet.write('S235', '=IFERROR(J144+MATCH(R235,J235:N235)-1,"")')
    worksheet.write('S236', '=IFERROR(J144+MATCH(R236,J236:N236)-1,"")')
    worksheet.write('S237', '=IFERROR(J144+MATCH(R237,J237:N237)-1,"")')
    worksheet.write('S238', '=IFERROR(J144+MATCH(R238,J238:N238)-1,"")')
    worksheet.write('S239', '=IFERROR(J144+MATCH(R239,J239:N239)-1,"")')
    worksheet.write('S240', '=IFERROR(J144+MATCH(R240,J240:N240)-1,"")')
    worksheet.write('S241', '=IFERROR(J144+MATCH(R241,J241:N241)-1,"")')
    worksheet.write('S242', '=IFERROR(J144+MATCH(R242,J242:N242)-1,"")')
    worksheet.write('S243', '=IFERROR(J144+MATCH(R243,J243:N243)-1,"")')
    worksheet.write('S244', '=IFERROR(J144+MATCH(R244,J244:N244)-1,"")')
    worksheet.write('S245', '=IFERROR(J144+MATCH(R245,J245:N245)-1,"")')
    worksheet.write('S246', '=IFERROR(J144+MATCH(R246,J246:N246)-1,"")')
    worksheet.write('S247', '=IFERROR(J144+MATCH(R247,J247:N247)-1,"")')
    worksheet.write('S248', '=IFERROR(J144+MATCH(R248,J248:N248)-1,"")')
    worksheet.write('S249', '=IFERROR(J144+MATCH(R249,J249:N249)-1,"")')
    worksheet.write('S250', '=IFERROR(J144+MATCH(R250,J250:N250)-1,"")')
    worksheet.write('S251', '=IFERROR(J144+MATCH(R251,J251:N251)-1,"")')
    worksheet.write('S252', '=IFERROR(J144+MATCH(R252,J252:N252)-1,"")')
    worksheet.write('S253', '=IFERROR(J144+MATCH(R253,J253:N253)-1,"")')
    worksheet.write('S254', '=IFERROR(J144+MATCH(R254,J254:N254)-1,"")')
    worksheet.write('S255', '=IFERROR(J144+MATCH(R255,J255:N255)-1,"")')
    worksheet.write('S256', '=IFERROR(J144+MATCH(R256,J256:N256)-1,"")')
    worksheet.write('S257', '=IFERROR(J144+MATCH(R257,J257:N257)-1,"")')
    worksheet.write('S258', '=IFERROR(J144+MATCH(R258,J258:N258)-1,"")')
    worksheet.write('S259', '=IFERROR(J144+MATCH(R259,J259:N259)-1,"")')
    worksheet.write('S260', '=IFERROR(J144+MATCH(R260,J260:N260)-1,"")')
    worksheet.write('S261', '=IFERROR(J144+MATCH(R261,J261:N261)-1,"")')
    worksheet.write('S262', '=IFERROR(J144+MATCH(R262,J262:N262)-1,"")')
    worksheet.write('S263', '=IFERROR(J144+MATCH(R263,J263:N263)-1,"")')
    worksheet.write('S264', '=IFERROR(J144+MATCH(R264,J264:N264)-1,"")')
    worksheet.write('S265', '=IFERROR(J144+MATCH(R265,J265:N265)-1,"")')
    worksheet.write('S266', '=IFERROR(J144+MATCH(R266,J266:N266)-1,"")')
    worksheet.write('S267', '=IFERROR(J144+MATCH(R267,J267:N267)-1,"")')
    worksheet.write('S268', '=IFERROR(J144+MATCH(R268,J268:N268)-1,"")')
    worksheet.write('S269', '=IFERROR(J144+MATCH(R269,J269:N269)-1,"")')
    worksheet.write('S270', '=IFERROR(J144+MATCH(R270,J270:N270)-1,"")')
    worksheet.write('S271', '=IFERROR(J144+MATCH(R271,J271:N271)-1,"")')
    worksheet.write('S272', '=IFERROR(J144+MATCH(R272,J272:N272)-1,"")')
    worksheet.write('S273', '=IFERROR(J144+MATCH(R273,J273:N273)-1,"")')
    worksheet.write('S274', '=IFERROR(J144+MATCH(R274,J274:N274)-1,"")')
    worksheet.write('S275', '=IFERROR(J144+MATCH(R275,J275:N275)-1,"")')
    worksheet.write('S276', '=IFERROR(J144+MATCH(R276,J276:N276)-1,"")')
    worksheet.write('S277', '=IFERROR(J144+MATCH(R277,J277:N277)-1,"")')
    worksheet.write('S278', '=IFERROR(J144+MATCH(R278,J278:N278)-1,"")')
    worksheet.write('S279', '=IFERROR(J144+MATCH(R279,J279:N279)-1,"")')
    worksheet.write('S280', '=IFERROR(J144+MATCH(R280,J280:N280)-1,"")')
    worksheet.write('S281', '=IFERROR(J144+MATCH(R281,J281:N281)-1,"")')
    worksheet.write('S282', '=IFERROR(J144+MATCH(R282,J282:N282)-1,"")')
    worksheet.write('S283', '=IFERROR(J144+MATCH(R283,J283:N283)-1,"")')
    worksheet.write('S284', '=IFERROR(J144+MATCH(R284,J284:N284)-1,"")')
    worksheet.write('S285', '=IFERROR(J144+MATCH(R285,J285:N285)-1,"")')
    worksheet.write('S286', '=IFERROR(J144+MATCH(R286,J286:N286)-1,"")')
    worksheet.write('S287', '=IFERROR(J144+MATCH(R287,J287:N287)-1,"")')
    worksheet.write('S288', '=IFERROR(J144+MATCH(R288,J288:N288)-1,"")')
    worksheet.write('S289', '=IFERROR(J144+MATCH(R289,J289:N289)-1,"")')
    worksheet.write('S290', '=IFERROR(J144+MATCH(R290,J290:N290)-1,"")')
    worksheet.write('S291', '=IFERROR(J144+MATCH(R291,J291:N291)-1,"")')
    worksheet.write('S292', '=IFERROR(J144+MATCH(R292,J292:N292)-1,"")')
    worksheet.write('S293', '=IFERROR(J144+MATCH(R293,J293:N293)-1,"")')
    worksheet.write('S294', '=IFERROR(J144+MATCH(R294,J294:N294)-1,"")')
    worksheet.write('S295', '=IFERROR(J144+MATCH(R295,J295:N295)-1,"")')
    worksheet.write('S296', '=IFERROR(J144+MATCH(R296,J296:N296)-1,"")')
    worksheet.write('S297', '=IFERROR(J144+MATCH(R297,J297:N297)-1,"")')
    worksheet.write('S298', '=IFERROR(J144+MATCH(R298,J298:N298)-1,"")')
    worksheet.write('S299', '=IFERROR(J144+MATCH(R299,J299:N299)-1,"")')
    worksheet.write('S300', '=IFERROR(J144+MATCH(R300,J300:N300)-1,"")')
    worksheet.write('S301', '=IFERROR(J144+MATCH(R301,J301:N301)-1,"")')
    worksheet.write('S302', '=IFERROR(J144+MATCH(R302,J302:N302)-1,"")')
    worksheet.write('S303', '=IFERROR(J144+MATCH(R303,J303:N303)-1,"")')
    worksheet.write('S304', '=IFERROR(J144+MATCH(R304,J304:N304)-1,"")')
    worksheet.write('S305', '=IFERROR(J144+MATCH(R305,J305:N305)-1,"")')
    worksheet.write('S306', '=IFERROR(J144+MATCH(R306,J306:N306)-1,"")')
    worksheet.write('S307', '=IFERROR(J144+MATCH(R307,J307:N307)-1,"")')
    worksheet.write('S308', '=IFERROR(J144+MATCH(R308,J308:N308)-1,"")')
    worksheet.write('S309', '=IFERROR(J144+MATCH(R309,J309:N309)-1,"")')
    worksheet.write('S310', '=IFERROR(J144+MATCH(R310,J310:N310)-1,"")')
    worksheet.write('S311', '=IFERROR(J144+MATCH(R311,J311:N311)-1,"")')
    worksheet.write('S312', '=IFERROR(J144+MATCH(R312,J312:N312)-1,"")')
    worksheet.write('S313', '=IFERROR(J144+MATCH(R313,J313:N313)-1,"")')
    worksheet.write('S314', '=IFERROR(J144+MATCH(R314,J314:N314)-1,"")')
    worksheet.write('S315', '=IFERROR(J144+MATCH(R315,J315:N315)-1,"")')
    worksheet.write('S316', '=IFERROR(J144+MATCH(R316,J316:N316)-1,"")')
    worksheet.write('S317', '=IFERROR(J144+MATCH(R317,J317:N317)-1,"")')
    worksheet.write('S318', '=IFERROR(J144+MATCH(R318,J318:N318)-1,"")')
    worksheet.write('S319', '=IFERROR(J144+MATCH(R319,J319:N319)-1,"")')
    worksheet.write('S320', '=IFERROR(J144+MATCH(R320,J320:N320)-1,"")')
    worksheet.write('S321', '=IFERROR(J144+MATCH(R321,J321:N321)-1,"")')
    worksheet.write('S322', '=IFERROR(J144+MATCH(R322,J322:N322)-1,"")')
    worksheet.write('S323', '=IFERROR(J144+MATCH(R323,J323:N323)-1,"")')
    worksheet.write('S324', '=IFERROR(J144+MATCH(R324,J324:N324)-1,"")')
    worksheet.write('S325', '=IFERROR(J144+MATCH(R325,J325:N325)-1,"")')
    worksheet.write('S326', '=IFERROR(J144+MATCH(R326,J326:N326)-1,"")')
    worksheet.write('S327', '=IFERROR(J144+MATCH(R327,J327:N327)-1,"")')
    worksheet.write('S328', '=IFERROR(J144+MATCH(R328,J328:N328)-1,"")')
    worksheet.write('S329', '=IFERROR(J144+MATCH(R329,J329:N329)-1,"")')
    worksheet.write('S330', '=IFERROR(J144+MATCH(R330,J330:N330)-1,"")')
    worksheet.write('S331', '=IFERROR(J144+MATCH(R331,J331:N331)-1,"")')
    worksheet.write('S332', '=IFERROR(J144+MATCH(R332,J332:N332)-1,"")')
    worksheet.write('S333', '=IFERROR(J144+MATCH(R333,J333:N333)-1,"")')
    worksheet.write('S334', '=IFERROR(J144+MATCH(R334,J334:N334)-1,"")')
    worksheet.write('S335', '=IFERROR(J144+MATCH(R335,J335:N335)-1,"")')
    worksheet.write('S336', '=IFERROR(J144+MATCH(R336,J336:N336)-1,"")')
    worksheet.write('S337', '=IFERROR(J144+MATCH(R337,J337:N337)-1,"")')
    worksheet.write('S338', '=IFERROR(J144+MATCH(R338,J338:N338)-1,"")')
    worksheet.write('S339', '=IFERROR(J144+MATCH(R339,J339:N339)-1,"")')
    worksheet.write('S340', '=IFERROR(J144+MATCH(R340,J340:N340)-1,"")')
    worksheet.write('S341', '=IFERROR(J144+MATCH(R341,J341:N341)-1,"")')
    worksheet.write('S342', '=IFERROR(J144+MATCH(R342,J342:N342)-1,"")')
    worksheet.write('S343', '=IFERROR(J144+MATCH(R343,J343:N343)-1,"")')
    worksheet.write('S344', '=IFERROR(J144+MATCH(R344,J344:N344)-1,"")')
    worksheet.write('S345', '=IFERROR(J144+MATCH(R345,J345:N345)-1,"")')
    worksheet.write('S346', '=IFERROR(J144+MATCH(R346,J346:N346)-1,"")')
    worksheet.write('S347', '=IFERROR(J144+MATCH(R347,J347:N347)-1,"")')
    worksheet.write('S348', '=IFERROR(J144+MATCH(R348,J348:N348)-1,"")')
    worksheet.write('S349', '=IFERROR(J144+MATCH(R349,J349:N349)-1,"")')
    worksheet.write('S350', '=IFERROR(J144+MATCH(R350,J350:N350)-1,"")')
    worksheet.write('S351', '=IFERROR(J144+MATCH(R351,J351:N351)-1,"")')
    worksheet.write('S352', '=IFERROR(J144+MATCH(R352,J352:N352)-1,"")')
    worksheet.write('S353', '=IFERROR(J144+MATCH(R353,J353:N353)-1,"")')
    worksheet.write('T144', '="Average"')
    worksheet.write('T145', '=IFERROR(AVERAGE(J145:N145),"")')
    worksheet.write('T146', '=IFERROR(AVERAGE(J146:N146),"")')
    worksheet.write('T147', '=IFERROR(AVERAGE(J147:N147),"")')
    worksheet.write('T148', '=IFERROR(AVERAGE(J148:N148),"")')
    worksheet.write('T149', '=IFERROR(AVERAGE(J149:N149),"")')
    worksheet.write('T150', '=IFERROR(AVERAGE(J150:N150),"")')
    worksheet.write('T151', '=IFERROR(AVERAGE(J151:N151),"")')
    worksheet.write('T152', '=IFERROR(AVERAGE(J152:N152),"")')
    worksheet.write('T153', '=IFERROR(AVERAGE(J153:N153),"")')
    worksheet.write('T154', '=IFERROR(AVERAGE(J154:N154),"")')
    worksheet.write('T155', '=IFERROR(AVERAGE(J155:N155),"")')
    worksheet.write('T156', '=IFERROR(AVERAGE(J156:N156),"")')
    worksheet.write('T157', '=IFERROR(AVERAGE(J157:N157),"")')
    worksheet.write('T158', '=IFERROR(AVERAGE(J158:N158),"")')
    worksheet.write('T159', '=IFERROR(AVERAGE(J159:N159),"")')
    worksheet.write('T160', '=IFERROR(AVERAGE(J160:N160),"")')
    worksheet.write('T161', '=IFERROR(AVERAGE(J161:N161),"")')
    worksheet.write('T162', '=IFERROR(AVERAGE(J162:N162),"")')
    worksheet.write('T163', '=IFERROR(AVERAGE(J163:N163),"")')
    worksheet.write('T164', '=IFERROR(AVERAGE(J164:N164),"")')
    worksheet.write('T165', '=IFERROR(AVERAGE(J165:N165),"")')
    worksheet.write('T166', '=IFERROR(AVERAGE(J166:N166),"")')
    worksheet.write('T167', '=IFERROR(AVERAGE(J167:N167),"")')
    worksheet.write('T168', '=IFERROR(AVERAGE(J168:N168),"")')
    worksheet.write('T169', '=IFERROR(AVERAGE(J169:N169),"")')
    worksheet.write('T170', '=IFERROR(AVERAGE(J170:N170),"")')
    worksheet.write('T171', '=IFERROR(AVERAGE(J171:N171),"")')
    worksheet.write('T172', '=IFERROR(AVERAGE(J172:N172),"")')
    worksheet.write('T173', '=IFERROR(AVERAGE(J173:N173),"")')
    worksheet.write('T174', '=IFERROR(AVERAGE(J174:N174),"")')
    worksheet.write('T175', '=IFERROR(AVERAGE(J175:N175),"")')
    worksheet.write('T176', '=IFERROR(AVERAGE(J176:N176),"")')
    worksheet.write('T177', '=IFERROR(AVERAGE(J177:N177),"")')
    worksheet.write('T178', '=IFERROR(AVERAGE(J178:N178),"")')
    worksheet.write('T179', '=IFERROR(AVERAGE(J179:N179),"")')
    worksheet.write('T180', '=IFERROR(AVERAGE(J180:N180),"")')
    worksheet.write('T181', '=IFERROR(AVERAGE(J181:N181),"")')
    worksheet.write('T182', '=IFERROR(AVERAGE(J182:N182),"")')
    worksheet.write('T183', '=IFERROR(AVERAGE(J183:N183),"")')
    worksheet.write('T184', '=IFERROR(AVERAGE(J184:N184),"")')
    worksheet.write('T185', '=IFERROR(AVERAGE(J185:N185),"")')
    worksheet.write('T186', '=IFERROR(AVERAGE(J186:N186),"")')
    worksheet.write('T187', '=IFERROR(AVERAGE(J187:N187),"")')
    worksheet.write('T188', '=IFERROR(AVERAGE(J188:N188),"")')
    worksheet.write('T189', '=IFERROR(AVERAGE(J189:N189),"")')
    worksheet.write('T190', '=IFERROR(AVERAGE(J190:N190),"")')
    worksheet.write('T191', '=IFERROR(AVERAGE(J191:N191),"")')
    worksheet.write('T192', '=IFERROR(AVERAGE(J192:N192),"")')
    worksheet.write('T193', '=IFERROR(AVERAGE(J193:N193),"")')
    worksheet.write('T194', '=IFERROR(AVERAGE(J194:N194),"")')
    worksheet.write('T195', '=IFERROR(AVERAGE(J195:N195),"")')
    worksheet.write('T196', '=IFERROR(AVERAGE(J196:N196),"")')
    worksheet.write('T197', '=IFERROR(AVERAGE(J197:N197),"")')
    worksheet.write('T198', '=IFERROR(AVERAGE(J198:N198),"")')
    worksheet.write('T199', '=IFERROR(AVERAGE(J199:N199),"")')
    worksheet.write('T200', '=IFERROR(AVERAGE(J200:N200),"")')
    worksheet.write('T201', '=IFERROR(AVERAGE(J201:N201),"")')
    worksheet.write('T202', '=IFERROR(AVERAGE(J202:N202),"")')
    worksheet.write('T203', '=IFERROR(AVERAGE(J203:N203),"")')
    worksheet.write('T204', '=IFERROR(AVERAGE(J204:N204),"")')
    worksheet.write('T205', '=IFERROR(AVERAGE(J205:N205),"")')
    worksheet.write('T206', '=IFERROR(AVERAGE(J206:N206),"")')
    worksheet.write('T207', '=IFERROR(AVERAGE(J207:N207),"")')
    worksheet.write('T208', '=IFERROR(AVERAGE(J208:N208),"")')
    worksheet.write('T209', '=IFERROR(AVERAGE(J209:N209),"")')
    worksheet.write('T210', '=IFERROR(AVERAGE(J210:N210),"")')
    worksheet.write('T211', '=IFERROR(AVERAGE(J211:N211),"")')
    worksheet.write('T212', '=IFERROR(AVERAGE(J212:N212),"")')
    worksheet.write('T213', '=IFERROR(AVERAGE(J213:N213),"")')
    worksheet.write('T214', '=IFERROR(AVERAGE(J214:N214),"")')
    worksheet.write('T215', '=IFERROR(AVERAGE(J215:N215),"")')
    worksheet.write('T216', '=IFERROR(AVERAGE(J216:N216),"")')
    worksheet.write('T217', '=IFERROR(AVERAGE(J217:N217),"")')
    worksheet.write('T218', '=IFERROR(AVERAGE(J218:N218),"")')
    worksheet.write('T219', '=IFERROR(AVERAGE(J219:N219),"")')
    worksheet.write('T220', '=IFERROR(AVERAGE(J220:N220),"")')
    worksheet.write('T221', '=IFERROR(AVERAGE(J221:N221),"")')
    worksheet.write('T222', '=IFERROR(AVERAGE(J222:N222),"")')
    worksheet.write('T223', '=IFERROR(AVERAGE(J223:N223),"")')
    worksheet.write('T224', '=IFERROR(AVERAGE(J224:N224),"")')
    worksheet.write('T225', '=IFERROR(AVERAGE(J225:N225),"")')
    worksheet.write('T226', '=IFERROR(AVERAGE(J226:N226),"")')
    worksheet.write('T227', '=IFERROR(AVERAGE(J227:N227),"")')
    worksheet.write('T228', '=IFERROR(AVERAGE(J228:N228),"")')
    worksheet.write('T229', '=IFERROR(AVERAGE(J229:N229),"")')
    worksheet.write('T230', '=IFERROR(AVERAGE(J230:N230),"")')
    worksheet.write('T231', '=IFERROR(AVERAGE(J231:N231),"")')
    worksheet.write('T232', '=IFERROR(AVERAGE(J232:N232),"")')
    worksheet.write('T233', '=IFERROR(AVERAGE(J233:N233),"")')
    worksheet.write('T234', '=IFERROR(AVERAGE(J234:N234),"")')
    worksheet.write('T235', '=IFERROR(AVERAGE(J235:N235),"")')
    worksheet.write('T236', '=IFERROR(AVERAGE(J236:N236),"")')
    worksheet.write('T237', '=IFERROR(AVERAGE(J237:N237),"")')
    worksheet.write('T238', '=IFERROR(AVERAGE(J238:N238),"")')
    worksheet.write('T239', '=IFERROR(AVERAGE(J239:N239),"")')
    worksheet.write('T240', '=IFERROR(AVERAGE(J240:N240),"")')
    worksheet.write('T241', '=IFERROR(AVERAGE(J241:N241),"")')
    worksheet.write('T242', '=IFERROR(AVERAGE(J242:N242),"")')
    worksheet.write('T243', '=IFERROR(AVERAGE(J243:N243),"")')
    worksheet.write('T244', '=IFERROR(AVERAGE(J244:N244),"")')
    worksheet.write('T245', '=IFERROR(AVERAGE(J245:N245),"")')
    worksheet.write('T246', '=IFERROR(AVERAGE(J246:N246),"")')
    worksheet.write('T247', '=IFERROR(AVERAGE(J247:N247),"")')
    worksheet.write('T248', '=IFERROR(AVERAGE(J248:N248),"")')
    worksheet.write('T249', '=IFERROR(AVERAGE(J249:N249),"")')
    worksheet.write('T250', '=IFERROR(AVERAGE(J250:N250),"")')
    worksheet.write('T251', '=IFERROR(AVERAGE(J251:N251),"")')
    worksheet.write('T252', '=IFERROR(AVERAGE(J252:N252),"")')
    worksheet.write('T253', '=IFERROR(AVERAGE(J253:N253),"")')
    worksheet.write('T254', '=IFERROR(AVERAGE(J254:N254),"")')
    worksheet.write('T255', '=IFERROR(AVERAGE(J255:N255),"")')
    worksheet.write('T256', '=IFERROR(AVERAGE(J256:N256),"")')
    worksheet.write('T257', '=IFERROR(AVERAGE(J257:N257),"")')
    worksheet.write('T258', '=IFERROR(AVERAGE(J258:N258),"")')
    worksheet.write('T259', '=IFERROR(AVERAGE(J259:N259),"")')
    worksheet.write('T260', '=IFERROR(AVERAGE(J260:N260),"")')
    worksheet.write('T261', '=IFERROR(AVERAGE(J261:N261),"")')
    worksheet.write('T262', '=IFERROR(AVERAGE(J262:N262),"")')
    worksheet.write('T263', '=IFERROR(AVERAGE(J263:N263),"")')
    worksheet.write('T264', '=IFERROR(AVERAGE(J264:N264),"")')
    worksheet.write('T265', '=IFERROR(AVERAGE(J265:N265),"")')
    worksheet.write('T266', '=IFERROR(AVERAGE(J266:N266),"")')
    worksheet.write('T267', '=IFERROR(AVERAGE(J267:N267),"")')
    worksheet.write('T268', '=IFERROR(AVERAGE(J268:N268),"")')
    worksheet.write('T269', '=IFERROR(AVERAGE(J269:N269),"")')
    worksheet.write('T270', '=IFERROR(AVERAGE(J270:N270),"")')
    worksheet.write('T271', '=IFERROR(AVERAGE(J271:N271),"")')
    worksheet.write('T272', '=IFERROR(AVERAGE(J272:N272),"")')
    worksheet.write('T273', '=IFERROR(AVERAGE(J273:N273),"")')
    worksheet.write('T274', '=IFERROR(AVERAGE(J274:N274),"")')
    worksheet.write('T275', '=IFERROR(AVERAGE(J275:N275),"")')
    worksheet.write('T276', '=IFERROR(AVERAGE(J276:N276),"")')
    worksheet.write('T277', '=IFERROR(AVERAGE(J277:N277),"")')
    worksheet.write('T278', '=IFERROR(AVERAGE(J278:N278),"")')
    worksheet.write('T279', '=IFERROR(AVERAGE(J279:N279),"")')
    worksheet.write('T280', '=IFERROR(AVERAGE(J280:N280),"")')
    worksheet.write('T281', '=IFERROR(AVERAGE(J281:N281),"")')
    worksheet.write('T282', '=IFERROR(AVERAGE(J282:N282),"")')
    worksheet.write('T283', '=IFERROR(AVERAGE(J283:N283),"")')
    worksheet.write('T284', '=IFERROR(AVERAGE(J284:N284),"")')
    worksheet.write('T285', '=IFERROR(AVERAGE(J285:N285),"")')
    worksheet.write('T286', '=IFERROR(AVERAGE(J286:N286),"")')
    worksheet.write('T287', '=IFERROR(AVERAGE(J287:N287),"")')
    worksheet.write('T288', '=IFERROR(AVERAGE(J288:N288),"")')
    worksheet.write('T289', '=IFERROR(AVERAGE(J289:N289),"")')
    worksheet.write('T290', '=IFERROR(AVERAGE(J290:N290),"")')
    worksheet.write('T291', '=IFERROR(AVERAGE(J291:N291),"")')
    worksheet.write('T292', '=IFERROR(AVERAGE(J292:N292),"")')
    worksheet.write('T293', '=IFERROR(AVERAGE(J293:N293),"")')
    worksheet.write('T294', '=IFERROR(AVERAGE(J294:N294),"")')
    worksheet.write('T295', '=IFERROR(AVERAGE(J295:N295),"")')
    worksheet.write('T296', '=IFERROR(AVERAGE(J296:N296),"")')
    worksheet.write('T297', '=IFERROR(AVERAGE(J297:N297),"")')
    worksheet.write('T298', '=IFERROR(AVERAGE(J298:N298),"")')
    worksheet.write('T299', '=IFERROR(AVERAGE(J299:N299),"")')
    worksheet.write('T300', '=IFERROR(AVERAGE(J300:N300),"")')
    worksheet.write('T301', '=IFERROR(AVERAGE(J301:N301),"")')
    worksheet.write('T302', '=IFERROR(AVERAGE(J302:N302),"")')
    worksheet.write('T303', '=IFERROR(AVERAGE(J303:N303),"")')
    worksheet.write('T304', '=IFERROR(AVERAGE(J304:N304),"")')
    worksheet.write('T305', '=IFERROR(AVERAGE(J305:N305),"")')
    worksheet.write('T306', '=IFERROR(AVERAGE(J306:N306),"")')
    worksheet.write('T307', '=IFERROR(AVERAGE(J307:N307),"")')
    worksheet.write('T308', '=IFERROR(AVERAGE(J308:N308),"")')
    worksheet.write('T309', '=IFERROR(AVERAGE(J309:N309),"")')
    worksheet.write('T310', '=IFERROR(AVERAGE(J310:N310),"")')
    worksheet.write('T311', '=IFERROR(AVERAGE(J311:N311),"")')
    worksheet.write('T312', '=IFERROR(AVERAGE(J312:N312),"")')
    worksheet.write('T313', '=IFERROR(AVERAGE(J313:N313),"")')
    worksheet.write('T314', '=IFERROR(AVERAGE(J314:N314),"")')
    worksheet.write('T315', '=IFERROR(AVERAGE(J315:N315),"")')
    worksheet.write('T316', '=IFERROR(AVERAGE(J316:N316),"")')
    worksheet.write('T317', '=IFERROR(AVERAGE(J317:N317),"")')
    worksheet.write('T318', '=IFERROR(AVERAGE(J318:N318),"")')
    worksheet.write('T319', '=IFERROR(AVERAGE(J319:N319),"")')
    worksheet.write('T320', '=IFERROR(AVERAGE(J320:N320),"")')
    worksheet.write('T321', '=IFERROR(AVERAGE(J321:N321),"")')
    worksheet.write('T322', '=IFERROR(AVERAGE(J322:N322),"")')
    worksheet.write('T323', '=IFERROR(AVERAGE(J323:N323),"")')
    worksheet.write('T324', '=IFERROR(AVERAGE(J324:N324),"")')
    worksheet.write('T325', '=IFERROR(AVERAGE(J325:N325),"")')
    worksheet.write('T326', '=IFERROR(AVERAGE(J326:N326),"")')
    worksheet.write('T327', '=IFERROR(AVERAGE(J327:N327),"")')
    worksheet.write('T328', '=IFERROR(AVERAGE(J328:N328),"")')
    worksheet.write('T329', '=IFERROR(AVERAGE(J329:N329),"")')
    worksheet.write('T330', '=IFERROR(AVERAGE(J330:N330),"")')
    worksheet.write('T331', '=IFERROR(AVERAGE(J331:N331),"")')
    worksheet.write('T332', '=IFERROR(AVERAGE(J332:N332),"")')
    worksheet.write('T333', '=IFERROR(AVERAGE(J333:N333),"")')
    worksheet.write('T334', '=IFERROR(AVERAGE(J334:N334),"")')
    worksheet.write('T335', '=IFERROR(AVERAGE(J335:N335),"")')
    worksheet.write('T336', '=IFERROR(AVERAGE(J336:N336),"")')
    worksheet.write('T337', '=IFERROR(AVERAGE(J337:N337),"")')
    worksheet.write('T338', '=IFERROR(AVERAGE(J338:N338),"")')
    worksheet.write('T339', '=IFERROR(AVERAGE(J339:N339),"")')
    worksheet.write('T340', '=IFERROR(AVERAGE(J340:N340),"")')
    worksheet.write('T341', '=IFERROR(AVERAGE(J341:N341),"")')
    worksheet.write('T342', '=IFERROR(AVERAGE(J342:N342),"")')
    worksheet.write('T343', '=IFERROR(AVERAGE(J343:N343),"")')
    worksheet.write('T344', '=IFERROR(AVERAGE(J344:N344),"")')
    worksheet.write('T345', '=IFERROR(AVERAGE(J345:N345),"")')
    worksheet.write('T346', '=IFERROR(AVERAGE(J346:N346),"")')
    worksheet.write('T347', '=IFERROR(AVERAGE(J347:N347),"")')
    worksheet.write('T348', '=IFERROR(AVERAGE(J348:N348),"")')
    worksheet.write('T349', '=IFERROR(AVERAGE(J349:N349),"")')
    worksheet.write('T350', '=IFERROR(AVERAGE(J350:N350),"")')
    worksheet.write('T351', '=IFERROR(AVERAGE(J351:N351),"")')
    worksheet.write('T352', '=IFERROR(AVERAGE(J352:N352),"")')
    worksheet.write('T353', '=IFERROR(AVERAGE(J353:N353),"")')
    worksheet.write('U144', '="SD"')
    worksheet.write('U145', '=IFERROR(STDEV(J145:N145),"")')
    worksheet.write('U146', '=IFERROR(STDEV(J146:N146),"")')
    worksheet.write('U147', '=IFERROR(STDEV(J147:N147),"")')
    worksheet.write('U148', '=IFERROR(STDEV(J148:N148),"")')
    worksheet.write('U149', '=IFERROR(STDEV(J149:N149),"")')
    worksheet.write('U150', '=IFERROR(STDEV(J150:N150),"")')
    worksheet.write('U151', '=IFERROR(STDEV(J151:N151),"")')
    worksheet.write('U152', '=IFERROR(STDEV(J152:N152),"")')
    worksheet.write('U153', '=IFERROR(STDEV(J153:N153),"")')
    worksheet.write('U154', '=IFERROR(STDEV(J154:N154),"")')
    worksheet.write('U155', '=IFERROR(STDEV(J155:N155),"")')
    worksheet.write('U156', '=IFERROR(STDEV(J156:N156),"")')
    worksheet.write('U157', '=IFERROR(STDEV(J157:N157),"")')
    worksheet.write('U158', '=IFERROR(STDEV(J158:N158),"")')
    worksheet.write('U159', '=IFERROR(STDEV(J159:N159),"")')
    worksheet.write('U160', '=IFERROR(STDEV(J160:N160),"")')
    worksheet.write('U161', '=IFERROR(STDEV(J161:N161),"")')
    worksheet.write('U162', '=IFERROR(STDEV(J162:N162),"")')
    worksheet.write('U163', '=IFERROR(STDEV(J163:N163),"")')
    worksheet.write('U164', '=IFERROR(STDEV(J164:N164),"")')
    worksheet.write('U165', '=IFERROR(STDEV(J165:N165),"")')
    worksheet.write('U166', '=IFERROR(STDEV(J166:N166),"")')
    worksheet.write('U167', '=IFERROR(STDEV(J167:N167),"")')
    worksheet.write('U168', '=IFERROR(STDEV(J168:N168),"")')
    worksheet.write('U169', '=IFERROR(STDEV(J169:N169),"")')
    worksheet.write('U170', '=IFERROR(STDEV(J170:N170),"")')
    worksheet.write('U171', '=IFERROR(STDEV(J171:N171),"")')
    worksheet.write('U172', '=IFERROR(STDEV(J172:N172),"")')
    worksheet.write('U173', '=IFERROR(STDEV(J173:N173),"")')
    worksheet.write('U174', '=IFERROR(STDEV(J174:N174),"")')
    worksheet.write('U175', '=IFERROR(STDEV(J175:N175),"")')
    worksheet.write('U176', '=IFERROR(STDEV(J176:N176),"")')
    worksheet.write('U177', '=IFERROR(STDEV(J177:N177),"")')
    worksheet.write('U178', '=IFERROR(STDEV(J178:N178),"")')
    worksheet.write('U179', '=IFERROR(STDEV(J179:N179),"")')
    worksheet.write('U180', '=IFERROR(STDEV(J180:N180),"")')
    worksheet.write('U181', '=IFERROR(STDEV(J181:N181),"")')
    worksheet.write('U182', '=IFERROR(STDEV(J182:N182),"")')
    worksheet.write('U183', '=IFERROR(STDEV(J183:N183),"")')
    worksheet.write('U184', '=IFERROR(STDEV(J184:N184),"")')
    worksheet.write('U185', '=IFERROR(STDEV(J185:N185),"")')
    worksheet.write('U186', '=IFERROR(STDEV(J186:N186),"")')
    worksheet.write('U187', '=IFERROR(STDEV(J187:N187),"")')
    worksheet.write('U188', '=IFERROR(STDEV(J188:N188),"")')
    worksheet.write('U189', '=IFERROR(STDEV(J189:N189),"")')
    worksheet.write('U190', '=IFERROR(STDEV(J190:N190),"")')
    worksheet.write('U191', '=IFERROR(STDEV(J191:N191),"")')
    worksheet.write('U192', '=IFERROR(STDEV(J192:N192),"")')
    worksheet.write('U193', '=IFERROR(STDEV(J193:N193),"")')
    worksheet.write('U194', '=IFERROR(STDEV(J194:N194),"")')
    worksheet.write('U195', '=IFERROR(STDEV(J195:N195),"")')
    worksheet.write('U196', '=IFERROR(STDEV(J196:N196),"")')
    worksheet.write('U197', '=IFERROR(STDEV(J197:N197),"")')
    worksheet.write('U198', '=IFERROR(STDEV(J198:N198),"")')
    worksheet.write('U199', '=IFERROR(STDEV(J199:N199),"")')
    worksheet.write('U200', '=IFERROR(STDEV(J200:N200),"")')
    worksheet.write('U201', '=IFERROR(STDEV(J201:N201),"")')
    worksheet.write('U202', '=IFERROR(STDEV(J202:N202),"")')
    worksheet.write('U203', '=IFERROR(STDEV(J203:N203),"")')
    worksheet.write('U204', '=IFERROR(STDEV(J204:N204),"")')
    worksheet.write('U205', '=IFERROR(STDEV(J205:N205),"")')
    worksheet.write('U206', '=IFERROR(STDEV(J206:N206),"")')
    worksheet.write('U207', '=IFERROR(STDEV(J207:N207),"")')
    worksheet.write('U208', '=IFERROR(STDEV(J208:N208),"")')
    worksheet.write('U209', '=IFERROR(STDEV(J209:N209),"")')
    worksheet.write('U210', '=IFERROR(STDEV(J210:N210),"")')
    worksheet.write('U211', '=IFERROR(STDEV(J211:N211),"")')
    worksheet.write('U212', '=IFERROR(STDEV(J212:N212),"")')
    worksheet.write('U213', '=IFERROR(STDEV(J213:N213),"")')
    worksheet.write('U214', '=IFERROR(STDEV(J214:N214),"")')
    worksheet.write('U215', '=IFERROR(STDEV(J215:N215),"")')
    worksheet.write('U216', '=IFERROR(STDEV(J216:N216),"")')
    worksheet.write('U217', '=IFERROR(STDEV(J217:N217),"")')
    worksheet.write('U218', '=IFERROR(STDEV(J218:N218),"")')
    worksheet.write('U219', '=IFERROR(STDEV(J219:N219),"")')
    worksheet.write('U220', '=IFERROR(STDEV(J220:N220),"")')
    worksheet.write('U221', '=IFERROR(STDEV(J221:N221),"")')
    worksheet.write('U222', '=IFERROR(STDEV(J222:N222),"")')
    worksheet.write('U223', '=IFERROR(STDEV(J223:N223),"")')
    worksheet.write('U224', '=IFERROR(STDEV(J224:N224),"")')
    worksheet.write('U225', '=IFERROR(STDEV(J225:N225),"")')
    worksheet.write('U226', '=IFERROR(STDEV(J226:N226),"")')
    worksheet.write('U227', '=IFERROR(STDEV(J227:N227),"")')
    worksheet.write('U228', '=IFERROR(STDEV(J228:N228),"")')
    worksheet.write('U229', '=IFERROR(STDEV(J229:N229),"")')
    worksheet.write('U230', '=IFERROR(STDEV(J230:N230),"")')
    worksheet.write('U231', '=IFERROR(STDEV(J231:N231),"")')
    worksheet.write('U232', '=IFERROR(STDEV(J232:N232),"")')
    worksheet.write('U233', '=IFERROR(STDEV(J233:N233),"")')
    worksheet.write('U234', '=IFERROR(STDEV(J234:N234),"")')
    worksheet.write('U235', '=IFERROR(STDEV(J235:N235),"")')
    worksheet.write('U236', '=IFERROR(STDEV(J236:N236),"")')
    worksheet.write('U237', '=IFERROR(STDEV(J237:N237),"")')
    worksheet.write('U238', '=IFERROR(STDEV(J238:N238),"")')
    worksheet.write('U239', '=IFERROR(STDEV(J239:N239),"")')
    worksheet.write('U240', '=IFERROR(STDEV(J240:N240),"")')
    worksheet.write('U241', '=IFERROR(STDEV(J241:N241),"")')
    worksheet.write('U242', '=IFERROR(STDEV(J242:N242),"")')
    worksheet.write('U243', '=IFERROR(STDEV(J243:N243),"")')
    worksheet.write('U244', '=IFERROR(STDEV(J244:N244),"")')
    worksheet.write('U245', '=IFERROR(STDEV(J245:N245),"")')
    worksheet.write('U246', '=IFERROR(STDEV(J246:N246),"")')
    worksheet.write('U247', '=IFERROR(STDEV(J247:N247),"")')
    worksheet.write('U248', '=IFERROR(STDEV(J248:N248),"")')
    worksheet.write('U249', '=IFERROR(STDEV(J249:N249),"")')
    worksheet.write('U250', '=IFERROR(STDEV(J250:N250),"")')
    worksheet.write('U251', '=IFERROR(STDEV(J251:N251),"")')
    worksheet.write('U252', '=IFERROR(STDEV(J252:N252),"")')
    worksheet.write('U253', '=IFERROR(STDEV(J253:N253),"")')
    worksheet.write('U254', '=IFERROR(STDEV(J254:N254),"")')
    worksheet.write('U255', '=IFERROR(STDEV(J255:N255),"")')
    worksheet.write('U256', '=IFERROR(STDEV(J256:N256),"")')
    worksheet.write('U257', '=IFERROR(STDEV(J257:N257),"")')
    worksheet.write('U258', '=IFERROR(STDEV(J258:N258),"")')
    worksheet.write('U259', '=IFERROR(STDEV(J259:N259),"")')
    worksheet.write('U260', '=IFERROR(STDEV(J260:N260),"")')
    worksheet.write('U261', '=IFERROR(STDEV(J261:N261),"")')
    worksheet.write('U262', '=IFERROR(STDEV(J262:N262),"")')
    worksheet.write('U263', '=IFERROR(STDEV(J263:N263),"")')
    worksheet.write('U264', '=IFERROR(STDEV(J264:N264),"")')
    worksheet.write('U265', '=IFERROR(STDEV(J265:N265),"")')
    worksheet.write('U266', '=IFERROR(STDEV(J266:N266),"")')
    worksheet.write('U267', '=IFERROR(STDEV(J267:N267),"")')
    worksheet.write('U268', '=IFERROR(STDEV(J268:N268),"")')
    worksheet.write('U269', '=IFERROR(STDEV(J269:N269),"")')
    worksheet.write('U270', '=IFERROR(STDEV(J270:N270),"")')
    worksheet.write('U271', '=IFERROR(STDEV(J271:N271),"")')
    worksheet.write('U272', '=IFERROR(STDEV(J272:N272),"")')
    worksheet.write('U273', '=IFERROR(STDEV(J273:N273),"")')
    worksheet.write('U274', '=IFERROR(STDEV(J274:N274),"")')
    worksheet.write('U275', '=IFERROR(STDEV(J275:N275),"")')
    worksheet.write('U276', '=IFERROR(STDEV(J276:N276),"")')
    worksheet.write('U277', '=IFERROR(STDEV(J277:N277),"")')
    worksheet.write('U278', '=IFERROR(STDEV(J278:N278),"")')
    worksheet.write('U279', '=IFERROR(STDEV(J279:N279),"")')
    worksheet.write('U280', '=IFERROR(STDEV(J280:N280),"")')
    worksheet.write('U281', '=IFERROR(STDEV(J281:N281),"")')
    worksheet.write('U282', '=IFERROR(STDEV(J282:N282),"")')
    worksheet.write('U283', '=IFERROR(STDEV(J283:N283),"")')
    worksheet.write('U284', '=IFERROR(STDEV(J284:N284),"")')
    worksheet.write('U285', '=IFERROR(STDEV(J285:N285),"")')
    worksheet.write('U286', '=IFERROR(STDEV(J286:N286),"")')
    worksheet.write('U287', '=IFERROR(STDEV(J287:N287),"")')
    worksheet.write('U288', '=IFERROR(STDEV(J288:N288),"")')
    worksheet.write('U289', '=IFERROR(STDEV(J289:N289),"")')
    worksheet.write('U290', '=IFERROR(STDEV(J290:N290),"")')
    worksheet.write('U291', '=IFERROR(STDEV(J291:N291),"")')
    worksheet.write('U292', '=IFERROR(STDEV(J292:N292),"")')
    worksheet.write('U293', '=IFERROR(STDEV(J293:N293),"")')
    worksheet.write('U294', '=IFERROR(STDEV(J294:N294),"")')
    worksheet.write('U295', '=IFERROR(STDEV(J295:N295),"")')
    worksheet.write('U296', '=IFERROR(STDEV(J296:N296),"")')
    worksheet.write('U297', '=IFERROR(STDEV(J297:N297),"")')
    worksheet.write('U298', '=IFERROR(STDEV(J298:N298),"")')
    worksheet.write('U299', '=IFERROR(STDEV(J299:N299),"")')
    worksheet.write('U300', '=IFERROR(STDEV(J300:N300),"")')
    worksheet.write('U301', '=IFERROR(STDEV(J301:N301),"")')
    worksheet.write('U302', '=IFERROR(STDEV(J302:N302),"")')
    worksheet.write('U303', '=IFERROR(STDEV(J303:N303),"")')
    worksheet.write('U304', '=IFERROR(STDEV(J304:N304),"")')
    worksheet.write('U305', '=IFERROR(STDEV(J305:N305),"")')
    worksheet.write('U306', '=IFERROR(STDEV(J306:N306),"")')
    worksheet.write('U307', '=IFERROR(STDEV(J307:N307),"")')
    worksheet.write('U308', '=IFERROR(STDEV(J308:N308),"")')
    worksheet.write('U309', '=IFERROR(STDEV(J309:N309),"")')
    worksheet.write('U310', '=IFERROR(STDEV(J310:N310),"")')
    worksheet.write('U311', '=IFERROR(STDEV(J311:N311),"")')
    worksheet.write('U312', '=IFERROR(STDEV(J312:N312),"")')
    worksheet.write('U313', '=IFERROR(STDEV(J313:N313),"")')
    worksheet.write('U314', '=IFERROR(STDEV(J314:N314),"")')
    worksheet.write('U315', '=IFERROR(STDEV(J315:N315),"")')
    worksheet.write('U316', '=IFERROR(STDEV(J316:N316),"")')
    worksheet.write('U317', '=IFERROR(STDEV(J317:N317),"")')
    worksheet.write('U318', '=IFERROR(STDEV(J318:N318),"")')
    worksheet.write('U319', '=IFERROR(STDEV(J319:N319),"")')
    worksheet.write('U320', '=IFERROR(STDEV(J320:N320),"")')
    worksheet.write('U321', '=IFERROR(STDEV(J321:N321),"")')
    worksheet.write('U322', '=IFERROR(STDEV(J322:N322),"")')
    worksheet.write('U323', '=IFERROR(STDEV(J323:N323),"")')
    worksheet.write('U324', '=IFERROR(STDEV(J324:N324),"")')
    worksheet.write('U325', '=IFERROR(STDEV(J325:N325),"")')
    worksheet.write('U326', '=IFERROR(STDEV(J326:N326),"")')
    worksheet.write('U327', '=IFERROR(STDEV(J327:N327),"")')
    worksheet.write('U328', '=IFERROR(STDEV(J328:N328),"")')
    worksheet.write('U329', '=IFERROR(STDEV(J329:N329),"")')
    worksheet.write('U330', '=IFERROR(STDEV(J330:N330),"")')
    worksheet.write('U331', '=IFERROR(STDEV(J331:N331),"")')
    worksheet.write('U332', '=IFERROR(STDEV(J332:N332),"")')
    worksheet.write('U333', '=IFERROR(STDEV(J333:N333),"")')
    worksheet.write('U334', '=IFERROR(STDEV(J334:N334),"")')
    worksheet.write('U335', '=IFERROR(STDEV(J335:N335),"")')
    worksheet.write('U336', '=IFERROR(STDEV(J336:N336),"")')
    worksheet.write('U337', '=IFERROR(STDEV(J337:N337),"")')
    worksheet.write('U338', '=IFERROR(STDEV(J338:N338),"")')
    worksheet.write('U339', '=IFERROR(STDEV(J339:N339),"")')
    worksheet.write('U340', '=IFERROR(STDEV(J340:N340),"")')
    worksheet.write('U341', '=IFERROR(STDEV(J341:N341),"")')
    worksheet.write('U342', '=IFERROR(STDEV(J342:N342),"")')
    worksheet.write('U343', '=IFERROR(STDEV(J343:N343),"")')
    worksheet.write('U344', '=IFERROR(STDEV(J344:N344),"")')
    worksheet.write('U345', '=IFERROR(STDEV(J345:N345),"")')
    worksheet.write('U346', '=IFERROR(STDEV(J346:N346),"")')
    worksheet.write('U347', '=IFERROR(STDEV(J347:N347),"")')
    worksheet.write('U348', '=IFERROR(STDEV(J348:N348),"")')
    worksheet.write('U349', '=IFERROR(STDEV(J349:N349),"")')
    worksheet.write('U350', '=IFERROR(STDEV(J350:N350),"")')
    worksheet.write('U351', '=IFERROR(STDEV(J351:N351),"")')
    worksheet.write('U352', '=IFERROR(STDEV(J352:N352),"")')
    worksheet.write('U353', '=IFERROR(STDEV(J353:N353),"")')
    worksheet.write('V144', '=J144')
    worksheet.write('V145',
                    '=IF(C145="-","",IF(ISBLANK(B145),"",IF(OR(ISNUMBER(FIND("Growth",B145)),ISNUMBER(FIND("Margin",B145))),"",(J145-T145)/U145)))')
    worksheet.write('V146',
                    '=IF(C146="-","",IF(ISBLANK(B146),"",IF(OR(ISNUMBER(FIND("Growth",B146)),ISNUMBER(FIND("Margin",B146))),"",(J146-T146)/U146)))')
    worksheet.write('V147',
                    '=IF(C147="-","",IF(ISBLANK(B147),"",IF(OR(ISNUMBER(FIND("Growth",B147)),ISNUMBER(FIND("Margin",B147))),"",(J147-T147)/U147)))')
    worksheet.write('V148',
                    '=IF(C148="-","",IF(ISBLANK(B148),"",IF(OR(ISNUMBER(FIND("Growth",B148)),ISNUMBER(FIND("Margin",B148))),"",(J148-T148)/U148)))')
    worksheet.write('V149',
                    '=IF(C149="-","",IF(ISBLANK(B149),"",IF(OR(ISNUMBER(FIND("Growth",B149)),ISNUMBER(FIND("Margin",B149))),"",(J149-T149)/U149)))')
    worksheet.write('V150',
                    '=IF(C150="-","",IF(ISBLANK(B150),"",IF(OR(ISNUMBER(FIND("Growth",B150)),ISNUMBER(FIND("Margin",B150))),"",(J150-T150)/U150)))')
    worksheet.write('V151',
                    '=IF(C151="-","",IF(ISBLANK(B151),"",IF(OR(ISNUMBER(FIND("Growth",B151)),ISNUMBER(FIND("Margin",B151))),"",(J151-T151)/U151)))')
    worksheet.write('V152',
                    '=IF(C152="-","",IF(ISBLANK(B152),"",IF(OR(ISNUMBER(FIND("Growth",B152)),ISNUMBER(FIND("Margin",B152))),"",(J152-T152)/U152)))')
    worksheet.write('V153',
                    '=IF(C153="-","",IF(ISBLANK(B153),"",IF(OR(ISNUMBER(FIND("Growth",B153)),ISNUMBER(FIND("Margin",B153))),"",(J153-T153)/U153)))')
    worksheet.write('V154',
                    '=IF(C154="-","",IF(ISBLANK(B154),"",IF(OR(ISNUMBER(FIND("Growth",B154)),ISNUMBER(FIND("Margin",B154))),"",(J154-T154)/U154)))')
    worksheet.write('V155',
                    '=IF(C155="-","",IF(ISBLANK(B155),"",IF(OR(ISNUMBER(FIND("Growth",B155)),ISNUMBER(FIND("Margin",B155))),"",(J155-T155)/U155)))')
    worksheet.write('V156',
                    '=IF(C156="-","",IF(ISBLANK(B156),"",IF(OR(ISNUMBER(FIND("Growth",B156)),ISNUMBER(FIND("Margin",B156))),"",(J156-T156)/U156)))')
    worksheet.write('V157',
                    '=IF(C157="-","",IF(ISBLANK(B157),"",IF(OR(ISNUMBER(FIND("Growth",B157)),ISNUMBER(FIND("Margin",B157))),"",(J157-T157)/U157)))')
    worksheet.write('V158',
                    '=IF(C158="-","",IF(ISBLANK(B158),"",IF(OR(ISNUMBER(FIND("Growth",B158)),ISNUMBER(FIND("Margin",B158))),"",(J158-T158)/U158)))')
    worksheet.write('V159',
                    '=IF(C159="-","",IF(ISBLANK(B159),"",IF(OR(ISNUMBER(FIND("Growth",B159)),ISNUMBER(FIND("Margin",B159))),"",(J159-T159)/U159)))')
    worksheet.write('V160',
                    '=IF(C160="-","",IF(ISBLANK(B160),"",IF(OR(ISNUMBER(FIND("Growth",B160)),ISNUMBER(FIND("Margin",B160))),"",(J160-T160)/U160)))')
    worksheet.write('V161',
                    '=IF(C161="-","",IF(ISBLANK(B161),"",IF(OR(ISNUMBER(FIND("Growth",B161)),ISNUMBER(FIND("Margin",B161))),"",(J161-T161)/U161)))')
    worksheet.write('V162',
                    '=IF(C162="-","",IF(ISBLANK(B162),"",IF(OR(ISNUMBER(FIND("Growth",B162)),ISNUMBER(FIND("Margin",B162))),"",(J162-T162)/U162)))')
    worksheet.write('V163',
                    '=IF(C163="-","",IF(ISBLANK(B163),"",IF(OR(ISNUMBER(FIND("Growth",B163)),ISNUMBER(FIND("Margin",B163))),"",(J163-T163)/U163)))')
    worksheet.write('V164',
                    '=IF(C164="-","",IF(ISBLANK(B164),"",IF(OR(ISNUMBER(FIND("Growth",B164)),ISNUMBER(FIND("Margin",B164))),"",(J164-T164)/U164)))')
    worksheet.write('V165',
                    '=IF(C165="-","",IF(ISBLANK(B165),"",IF(OR(ISNUMBER(FIND("Growth",B165)),ISNUMBER(FIND("Margin",B165))),"",(J165-T165)/U165)))')
    worksheet.write('V166',
                    '=IF(C166="-","",IF(ISBLANK(B166),"",IF(OR(ISNUMBER(FIND("Growth",B166)),ISNUMBER(FIND("Margin",B166))),"",(J166-T166)/U166)))')
    worksheet.write('V167',
                    '=IF(C167="-","",IF(ISBLANK(B167),"",IF(OR(ISNUMBER(FIND("Growth",B167)),ISNUMBER(FIND("Margin",B167))),"",(J167-T167)/U167)))')
    worksheet.write('V168',
                    '=IF(C168="-","",IF(ISBLANK(B168),"",IF(OR(ISNUMBER(FIND("Growth",B168)),ISNUMBER(FIND("Margin",B168))),"",(J168-T168)/U168)))')
    worksheet.write('V169',
                    '=IF(C169="-","",IF(ISBLANK(B169),"",IF(OR(ISNUMBER(FIND("Growth",B169)),ISNUMBER(FIND("Margin",B169))),"",(J169-T169)/U169)))')
    worksheet.write('V170',
                    '=IF(C170="-","",IF(ISBLANK(B170),"",IF(OR(ISNUMBER(FIND("Growth",B170)),ISNUMBER(FIND("Margin",B170))),"",(J170-T170)/U170)))')
    worksheet.write('V171',
                    '=IF(C171="-","",IF(ISBLANK(B171),"",IF(OR(ISNUMBER(FIND("Growth",B171)),ISNUMBER(FIND("Margin",B171))),"",(J171-T171)/U171)))')
    worksheet.write('V172',
                    '=IF(C172="-","",IF(ISBLANK(B172),"",IF(OR(ISNUMBER(FIND("Growth",B172)),ISNUMBER(FIND("Margin",B172))),"",(J172-T172)/U172)))')
    worksheet.write('V173',
                    '=IF(C173="-","",IF(ISBLANK(B173),"",IF(OR(ISNUMBER(FIND("Growth",B173)),ISNUMBER(FIND("Margin",B173))),"",(J173-T173)/U173)))')
    worksheet.write('V174',
                    '=IF(C174="-","",IF(ISBLANK(B174),"",IF(OR(ISNUMBER(FIND("Growth",B174)),ISNUMBER(FIND("Margin",B174))),"",(J174-T174)/U174)))')
    worksheet.write('V175',
                    '=IF(C175="-","",IF(ISBLANK(B175),"",IF(OR(ISNUMBER(FIND("Growth",B175)),ISNUMBER(FIND("Margin",B175))),"",(J175-T175)/U175)))')
    worksheet.write('V176',
                    '=IF(C176="-","",IF(ISBLANK(B176),"",IF(OR(ISNUMBER(FIND("Growth",B176)),ISNUMBER(FIND("Margin",B176))),"",(J176-T176)/U176)))')
    worksheet.write('V177',
                    '=IF(C177="-","",IF(ISBLANK(B177),"",IF(OR(ISNUMBER(FIND("Growth",B177)),ISNUMBER(FIND("Margin",B177))),"",(J177-T177)/U177)))')
    worksheet.write('V178',
                    '=IF(C178="-","",IF(ISBLANK(B178),"",IF(OR(ISNUMBER(FIND("Growth",B178)),ISNUMBER(FIND("Margin",B178))),"",(J178-T178)/U178)))')
    worksheet.write('V179',
                    '=IF(C179="-","",IF(ISBLANK(B179),"",IF(OR(ISNUMBER(FIND("Growth",B179)),ISNUMBER(FIND("Margin",B179))),"",(J179-T179)/U179)))')
    worksheet.write('V180',
                    '=IF(C180="-","",IF(ISBLANK(B180),"",IF(OR(ISNUMBER(FIND("Growth",B180)),ISNUMBER(FIND("Margin",B180))),"",(J180-T180)/U180)))')
    worksheet.write('V181',
                    '=IF(C181="-","",IF(ISBLANK(B181),"",IF(OR(ISNUMBER(FIND("Growth",B181)),ISNUMBER(FIND("Margin",B181))),"",(J181-T181)/U181)))')
    worksheet.write('V182',
                    '=IF(C182="-","",IF(ISBLANK(B182),"",IF(OR(ISNUMBER(FIND("Growth",B182)),ISNUMBER(FIND("Margin",B182))),"",(J182-T182)/U182)))')
    worksheet.write('V183',
                    '=IF(C183="-","",IF(ISBLANK(B183),"",IF(OR(ISNUMBER(FIND("Growth",B183)),ISNUMBER(FIND("Margin",B183))),"",(J183-T183)/U183)))')
    worksheet.write('V184',
                    '=IF(C184="-","",IF(ISBLANK(B184),"",IF(OR(ISNUMBER(FIND("Growth",B184)),ISNUMBER(FIND("Margin",B184))),"",(J184-T184)/U184)))')
    worksheet.write('V185',
                    '=IF(C185="-","",IF(ISBLANK(B185),"",IF(OR(ISNUMBER(FIND("Growth",B185)),ISNUMBER(FIND("Margin",B185))),"",(J185-T185)/U185)))')
    worksheet.write('V186',
                    '=IF(C186="-","",IF(ISBLANK(B186),"",IF(OR(ISNUMBER(FIND("Growth",B186)),ISNUMBER(FIND("Margin",B186))),"",(J186-T186)/U186)))')
    worksheet.write('V187',
                    '=IF(C187="-","",IF(ISBLANK(B187),"",IF(OR(ISNUMBER(FIND("Growth",B187)),ISNUMBER(FIND("Margin",B187))),"",(J187-T187)/U187)))')
    worksheet.write('V188',
                    '=IF(C188="-","",IF(ISBLANK(B188),"",IF(OR(ISNUMBER(FIND("Growth",B188)),ISNUMBER(FIND("Margin",B188))),"",(J188-T188)/U188)))')
    worksheet.write('V189',
                    '=IF(C189="-","",IF(ISBLANK(B189),"",IF(OR(ISNUMBER(FIND("Growth",B189)),ISNUMBER(FIND("Margin",B189))),"",(J189-T189)/U189)))')
    worksheet.write('V190',
                    '=IF(C190="-","",IF(ISBLANK(B190),"",IF(OR(ISNUMBER(FIND("Growth",B190)),ISNUMBER(FIND("Margin",B190))),"",(J190-T190)/U190)))')
    worksheet.write('V191',
                    '=IF(C191="-","",IF(ISBLANK(B191),"",IF(OR(ISNUMBER(FIND("Growth",B191)),ISNUMBER(FIND("Margin",B191))),"",(J191-T191)/U191)))')
    worksheet.write('V192',
                    '=IF(C192="-","",IF(ISBLANK(B192),"",IF(OR(ISNUMBER(FIND("Growth",B192)),ISNUMBER(FIND("Margin",B192))),"",(J192-T192)/U192)))')
    worksheet.write('V193',
                    '=IF(C193="-","",IF(ISBLANK(B193),"",IF(OR(ISNUMBER(FIND("Growth",B193)),ISNUMBER(FIND("Margin",B193))),"",(J193-T193)/U193)))')
    worksheet.write('V194',
                    '=IF(C194="-","",IF(ISBLANK(B194),"",IF(OR(ISNUMBER(FIND("Growth",B194)),ISNUMBER(FIND("Margin",B194))),"",(J194-T194)/U194)))')
    worksheet.write('V195',
                    '=IF(C195="-","",IF(ISBLANK(B195),"",IF(OR(ISNUMBER(FIND("Growth",B195)),ISNUMBER(FIND("Margin",B195))),"",(J195-T195)/U195)))')
    worksheet.write('V196',
                    '=IF(C196="-","",IF(ISBLANK(B196),"",IF(OR(ISNUMBER(FIND("Growth",B196)),ISNUMBER(FIND("Margin",B196))),"",(J196-T196)/U196)))')
    worksheet.write('V197',
                    '=IF(C197="-","",IF(ISBLANK(B197),"",IF(OR(ISNUMBER(FIND("Growth",B197)),ISNUMBER(FIND("Margin",B197))),"",(J197-T197)/U197)))')
    worksheet.write('V198',
                    '=IF(C198="-","",IF(ISBLANK(B198),"",IF(OR(ISNUMBER(FIND("Growth",B198)),ISNUMBER(FIND("Margin",B198))),"",(J198-T198)/U198)))')
    worksheet.write('V199',
                    '=IF(C199="-","",IF(ISBLANK(B199),"",IF(OR(ISNUMBER(FIND("Growth",B199)),ISNUMBER(FIND("Margin",B199))),"",(J199-T199)/U199)))')
    worksheet.write('V200',
                    '=IF(C200="-","",IF(ISBLANK(B200),"",IF(OR(ISNUMBER(FIND("Growth",B200)),ISNUMBER(FIND("Margin",B200))),"",(J200-T200)/U200)))')
    worksheet.write('V201',
                    '=IF(C201="-","",IF(ISBLANK(B201),"",IF(OR(ISNUMBER(FIND("Growth",B201)),ISNUMBER(FIND("Margin",B201))),"",(J201-T201)/U201)))')
    worksheet.write('V202',
                    '=IF(C202="-","",IF(ISBLANK(B202),"",IF(OR(ISNUMBER(FIND("Growth",B202)),ISNUMBER(FIND("Margin",B202))),"",(J202-T202)/U202)))')
    worksheet.write('V203',
                    '=IF(C203="-","",IF(ISBLANK(B203),"",IF(OR(ISNUMBER(FIND("Growth",B203)),ISNUMBER(FIND("Margin",B203))),"",(J203-T203)/U203)))')
    worksheet.write('V204',
                    '=IF(C204="-","",IF(ISBLANK(B204),"",IF(OR(ISNUMBER(FIND("Growth",B204)),ISNUMBER(FIND("Margin",B204))),"",(J204-T204)/U204)))')
    worksheet.write('V205',
                    '=IF(C205="-","",IF(ISBLANK(B205),"",IF(OR(ISNUMBER(FIND("Growth",B205)),ISNUMBER(FIND("Margin",B205))),"",(J205-T205)/U205)))')
    worksheet.write('V206',
                    '=IF(C206="-","",IF(ISBLANK(B206),"",IF(OR(ISNUMBER(FIND("Growth",B206)),ISNUMBER(FIND("Margin",B206))),"",(J206-T206)/U206)))')
    worksheet.write('V207',
                    '=IF(C207="-","",IF(ISBLANK(B207),"",IF(OR(ISNUMBER(FIND("Growth",B207)),ISNUMBER(FIND("Margin",B207))),"",(J207-T207)/U207)))')
    worksheet.write('V208',
                    '=IF(C208="-","",IF(ISBLANK(B208),"",IF(OR(ISNUMBER(FIND("Growth",B208)),ISNUMBER(FIND("Margin",B208))),"",(J208-T208)/U208)))')
    worksheet.write('V209',
                    '=IF(C209="-","",IF(ISBLANK(B209),"",IF(OR(ISNUMBER(FIND("Growth",B209)),ISNUMBER(FIND("Margin",B209))),"",(J209-T209)/U209)))')
    worksheet.write('V210',
                    '=IF(C210="-","",IF(ISBLANK(B210),"",IF(OR(ISNUMBER(FIND("Growth",B210)),ISNUMBER(FIND("Margin",B210))),"",(J210-T210)/U210)))')
    worksheet.write('V211',
                    '=IF(C211="-","",IF(ISBLANK(B211),"",IF(OR(ISNUMBER(FIND("Growth",B211)),ISNUMBER(FIND("Margin",B211))),"",(J211-T211)/U211)))')
    worksheet.write('V212',
                    '=IF(C212="-","",IF(ISBLANK(B212),"",IF(OR(ISNUMBER(FIND("Growth",B212)),ISNUMBER(FIND("Margin",B212))),"",(J212-T212)/U212)))')
    worksheet.write('V213',
                    '=IF(C213="-","",IF(ISBLANK(B213),"",IF(OR(ISNUMBER(FIND("Growth",B213)),ISNUMBER(FIND("Margin",B213))),"",(J213-T213)/U213)))')
    worksheet.write('V214',
                    '=IF(C214="-","",IF(ISBLANK(B214),"",IF(OR(ISNUMBER(FIND("Growth",B214)),ISNUMBER(FIND("Margin",B214))),"",(J214-T214)/U214)))')
    worksheet.write('V215',
                    '=IF(C215="-","",IF(ISBLANK(B215),"",IF(OR(ISNUMBER(FIND("Growth",B215)),ISNUMBER(FIND("Margin",B215))),"",(J215-T215)/U215)))')
    worksheet.write('V216',
                    '=IF(C216="-","",IF(ISBLANK(B216),"",IF(OR(ISNUMBER(FIND("Growth",B216)),ISNUMBER(FIND("Margin",B216))),"",(J216-T216)/U216)))')
    worksheet.write('V217',
                    '=IF(C217="-","",IF(ISBLANK(B217),"",IF(OR(ISNUMBER(FIND("Growth",B217)),ISNUMBER(FIND("Margin",B217))),"",(J217-T217)/U217)))')
    worksheet.write('V218',
                    '=IF(C218="-","",IF(ISBLANK(B218),"",IF(OR(ISNUMBER(FIND("Growth",B218)),ISNUMBER(FIND("Margin",B218))),"",(J218-T218)/U218)))')
    worksheet.write('V219',
                    '=IF(C219="-","",IF(ISBLANK(B219),"",IF(OR(ISNUMBER(FIND("Growth",B219)),ISNUMBER(FIND("Margin",B219))),"",(J219-T219)/U219)))')
    worksheet.write('V220',
                    '=IF(C220="-","",IF(ISBLANK(B220),"",IF(OR(ISNUMBER(FIND("Growth",B220)),ISNUMBER(FIND("Margin",B220))),"",(J220-T220)/U220)))')
    worksheet.write('V221',
                    '=IF(C221="-","",IF(ISBLANK(B221),"",IF(OR(ISNUMBER(FIND("Growth",B221)),ISNUMBER(FIND("Margin",B221))),"",(J221-T221)/U221)))')
    worksheet.write('V222',
                    '=IF(C222="-","",IF(ISBLANK(B222),"",IF(OR(ISNUMBER(FIND("Growth",B222)),ISNUMBER(FIND("Margin",B222))),"",(J222-T222)/U222)))')
    worksheet.write('V223',
                    '=IF(C223="-","",IF(ISBLANK(B223),"",IF(OR(ISNUMBER(FIND("Growth",B223)),ISNUMBER(FIND("Margin",B223))),"",(J223-T223)/U223)))')
    worksheet.write('V224',
                    '=IF(C224="-","",IF(ISBLANK(B224),"",IF(OR(ISNUMBER(FIND("Growth",B224)),ISNUMBER(FIND("Margin",B224))),"",(J224-T224)/U224)))')
    worksheet.write('V225',
                    '=IF(C225="-","",IF(ISBLANK(B225),"",IF(OR(ISNUMBER(FIND("Growth",B225)),ISNUMBER(FIND("Margin",B225))),"",(J225-T225)/U225)))')
    worksheet.write('V226',
                    '=IF(C226="-","",IF(ISBLANK(B226),"",IF(OR(ISNUMBER(FIND("Growth",B226)),ISNUMBER(FIND("Margin",B226))),"",(J226-T226)/U226)))')
    worksheet.write('V227',
                    '=IF(C227="-","",IF(ISBLANK(B227),"",IF(OR(ISNUMBER(FIND("Growth",B227)),ISNUMBER(FIND("Margin",B227))),"",(J227-T227)/U227)))')
    worksheet.write('V228',
                    '=IF(C228="-","",IF(ISBLANK(B228),"",IF(OR(ISNUMBER(FIND("Growth",B228)),ISNUMBER(FIND("Margin",B228))),"",(J228-T228)/U228)))')
    worksheet.write('V229',
                    '=IF(C229="-","",IF(ISBLANK(B229),"",IF(OR(ISNUMBER(FIND("Growth",B229)),ISNUMBER(FIND("Margin",B229))),"",(J229-T229)/U229)))')
    worksheet.write('V230',
                    '=IF(C230="-","",IF(ISBLANK(B230),"",IF(OR(ISNUMBER(FIND("Growth",B230)),ISNUMBER(FIND("Margin",B230))),"",(J230-T230)/U230)))')
    worksheet.write('V231',
                    '=IF(C231="-","",IF(ISBLANK(B231),"",IF(OR(ISNUMBER(FIND("Growth",B231)),ISNUMBER(FIND("Margin",B231))),"",(J231-T231)/U231)))')
    worksheet.write('V232',
                    '=IF(C232="-","",IF(ISBLANK(B232),"",IF(OR(ISNUMBER(FIND("Growth",B232)),ISNUMBER(FIND("Margin",B232))),"",(J232-T232)/U232)))')
    worksheet.write('V233',
                    '=IF(C233="-","",IF(ISBLANK(B233),"",IF(OR(ISNUMBER(FIND("Growth",B233)),ISNUMBER(FIND("Margin",B233))),"",(J233-T233)/U233)))')
    worksheet.write('V234',
                    '=IF(C234="-","",IF(ISBLANK(B234),"",IF(OR(ISNUMBER(FIND("Growth",B234)),ISNUMBER(FIND("Margin",B234))),"",(J234-T234)/U234)))')
    worksheet.write('V235',
                    '=IF(C235="-","",IF(ISBLANK(B235),"",IF(OR(ISNUMBER(FIND("Growth",B235)),ISNUMBER(FIND("Margin",B235))),"",(J235-T235)/U235)))')
    worksheet.write('V236',
                    '=IF(C236="-","",IF(ISBLANK(B236),"",IF(OR(ISNUMBER(FIND("Growth",B236)),ISNUMBER(FIND("Margin",B236))),"",(J236-T236)/U236)))')
    worksheet.write('V237',
                    '=IF(C237="-","",IF(ISBLANK(B237),"",IF(OR(ISNUMBER(FIND("Growth",B237)),ISNUMBER(FIND("Margin",B237))),"",(J237-T237)/U237)))')
    worksheet.write('V238',
                    '=IF(C238="-","",IF(ISBLANK(B238),"",IF(OR(ISNUMBER(FIND("Growth",B238)),ISNUMBER(FIND("Margin",B238))),"",(J238-T238)/U238)))')
    worksheet.write('V239',
                    '=IF(C239="-","",IF(ISBLANK(B239),"",IF(OR(ISNUMBER(FIND("Growth",B239)),ISNUMBER(FIND("Margin",B239))),"",(J239-T239)/U239)))')
    worksheet.write('V240',
                    '=IF(C240="-","",IF(ISBLANK(B240),"",IF(OR(ISNUMBER(FIND("Growth",B240)),ISNUMBER(FIND("Margin",B240))),"",(J240-T240)/U240)))')
    worksheet.write('V241',
                    '=IF(C241="-","",IF(ISBLANK(B241),"",IF(OR(ISNUMBER(FIND("Growth",B241)),ISNUMBER(FIND("Margin",B241))),"",(J241-T241)/U241)))')
    worksheet.write('V242',
                    '=IF(C242="-","",IF(ISBLANK(B242),"",IF(OR(ISNUMBER(FIND("Growth",B242)),ISNUMBER(FIND("Margin",B242))),"",(J242-T242)/U242)))')
    worksheet.write('V243',
                    '=IF(C243="-","",IF(ISBLANK(B243),"",IF(OR(ISNUMBER(FIND("Growth",B243)),ISNUMBER(FIND("Margin",B243))),"",(J243-T243)/U243)))')
    worksheet.write('V244',
                    '=IF(C244="-","",IF(ISBLANK(B244),"",IF(OR(ISNUMBER(FIND("Growth",B244)),ISNUMBER(FIND("Margin",B244))),"",(J244-T244)/U244)))')
    worksheet.write('V245',
                    '=IF(C245="-","",IF(ISBLANK(B245),"",IF(OR(ISNUMBER(FIND("Growth",B245)),ISNUMBER(FIND("Margin",B245))),"",(J245-T245)/U245)))')
    worksheet.write('V246',
                    '=IF(C246="-","",IF(ISBLANK(B246),"",IF(OR(ISNUMBER(FIND("Growth",B246)),ISNUMBER(FIND("Margin",B246))),"",(J246-T246)/U246)))')
    worksheet.write('V247',
                    '=IF(C247="-","",IF(ISBLANK(B247),"",IF(OR(ISNUMBER(FIND("Growth",B247)),ISNUMBER(FIND("Margin",B247))),"",(J247-T247)/U247)))')
    worksheet.write('V248',
                    '=IF(C248="-","",IF(ISBLANK(B248),"",IF(OR(ISNUMBER(FIND("Growth",B248)),ISNUMBER(FIND("Margin",B248))),"",(J248-T248)/U248)))')
    worksheet.write('V249',
                    '=IF(C249="-","",IF(ISBLANK(B249),"",IF(OR(ISNUMBER(FIND("Growth",B249)),ISNUMBER(FIND("Margin",B249))),"",(J249-T249)/U249)))')
    worksheet.write('V250',
                    '=IF(C250="-","",IF(ISBLANK(B250),"",IF(OR(ISNUMBER(FIND("Growth",B250)),ISNUMBER(FIND("Margin",B250))),"",(J250-T250)/U250)))')
    worksheet.write('V251',
                    '=IF(C251="-","",IF(ISBLANK(B251),"",IF(OR(ISNUMBER(FIND("Growth",B251)),ISNUMBER(FIND("Margin",B251))),"",(J251-T251)/U251)))')
    worksheet.write('V252',
                    '=IF(C252="-","",IF(ISBLANK(B252),"",IF(OR(ISNUMBER(FIND("Growth",B252)),ISNUMBER(FIND("Margin",B252))),"",(J252-T252)/U252)))')
    worksheet.write('V253',
                    '=IF(C253="-","",IF(ISBLANK(B253),"",IF(OR(ISNUMBER(FIND("Growth",B253)),ISNUMBER(FIND("Margin",B253))),"",(J253-T253)/U253)))')
    worksheet.write('V254',
                    '=IF(C254="-","",IF(ISBLANK(B254),"",IF(OR(ISNUMBER(FIND("Growth",B254)),ISNUMBER(FIND("Margin",B254))),"",(J254-T254)/U254)))')
    worksheet.write('V255',
                    '=IF(C255="-","",IF(ISBLANK(B255),"",IF(OR(ISNUMBER(FIND("Growth",B255)),ISNUMBER(FIND("Margin",B255))),"",(J255-T255)/U255)))')
    worksheet.write('V256',
                    '=IF(C256="-","",IF(ISBLANK(B256),"",IF(OR(ISNUMBER(FIND("Growth",B256)),ISNUMBER(FIND("Margin",B256))),"",(J256-T256)/U256)))')
    worksheet.write('V257',
                    '=IF(C257="-","",IF(ISBLANK(B257),"",IF(OR(ISNUMBER(FIND("Growth",B257)),ISNUMBER(FIND("Margin",B257))),"",(J257-T257)/U257)))')
    worksheet.write('V258',
                    '=IF(C258="-","",IF(ISBLANK(B258),"",IF(OR(ISNUMBER(FIND("Growth",B258)),ISNUMBER(FIND("Margin",B258))),"",(J258-T258)/U258)))')
    worksheet.write('V259',
                    '=IF(C259="-","",IF(ISBLANK(B259),"",IF(OR(ISNUMBER(FIND("Growth",B259)),ISNUMBER(FIND("Margin",B259))),"",(J259-T259)/U259)))')
    worksheet.write('V260',
                    '=IF(C260="-","",IF(ISBLANK(B260),"",IF(OR(ISNUMBER(FIND("Growth",B260)),ISNUMBER(FIND("Margin",B260))),"",(J260-T260)/U260)))')
    worksheet.write('V261',
                    '=IF(C261="-","",IF(ISBLANK(B261),"",IF(OR(ISNUMBER(FIND("Growth",B261)),ISNUMBER(FIND("Margin",B261))),"",(J261-T261)/U261)))')
    worksheet.write('V262',
                    '=IF(C262="-","",IF(ISBLANK(B262),"",IF(OR(ISNUMBER(FIND("Growth",B262)),ISNUMBER(FIND("Margin",B262))),"",(J262-T262)/U262)))')
    worksheet.write('V263',
                    '=IF(C263="-","",IF(ISBLANK(B263),"",IF(OR(ISNUMBER(FIND("Growth",B263)),ISNUMBER(FIND("Margin",B263))),"",(J263-T263)/U263)))')
    worksheet.write('V264',
                    '=IF(C264="-","",IF(ISBLANK(B264),"",IF(OR(ISNUMBER(FIND("Growth",B264)),ISNUMBER(FIND("Margin",B264))),"",(J264-T264)/U264)))')
    worksheet.write('V265',
                    '=IF(C265="-","",IF(ISBLANK(B265),"",IF(OR(ISNUMBER(FIND("Growth",B265)),ISNUMBER(FIND("Margin",B265))),"",(J265-T265)/U265)))')
    worksheet.write('V266',
                    '=IF(C266="-","",IF(ISBLANK(B266),"",IF(OR(ISNUMBER(FIND("Growth",B266)),ISNUMBER(FIND("Margin",B266))),"",(J266-T266)/U266)))')
    worksheet.write('V267',
                    '=IF(C267="-","",IF(ISBLANK(B267),"",IF(OR(ISNUMBER(FIND("Growth",B267)),ISNUMBER(FIND("Margin",B267))),"",(J267-T267)/U267)))')
    worksheet.write('V268',
                    '=IF(C268="-","",IF(ISBLANK(B268),"",IF(OR(ISNUMBER(FIND("Growth",B268)),ISNUMBER(FIND("Margin",B268))),"",(J268-T268)/U268)))')
    worksheet.write('V269',
                    '=IF(C269="-","",IF(ISBLANK(B269),"",IF(OR(ISNUMBER(FIND("Growth",B269)),ISNUMBER(FIND("Margin",B269))),"",(J269-T269)/U269)))')
    worksheet.write('V270',
                    '=IF(C270="-","",IF(ISBLANK(B270),"",IF(OR(ISNUMBER(FIND("Growth",B270)),ISNUMBER(FIND("Margin",B270))),"",(J270-T270)/U270)))')
    worksheet.write('V271',
                    '=IF(C271="-","",IF(ISBLANK(B271),"",IF(OR(ISNUMBER(FIND("Growth",B271)),ISNUMBER(FIND("Margin",B271))),"",(J271-T271)/U271)))')
    worksheet.write('V272',
                    '=IF(C272="-","",IF(ISBLANK(B272),"",IF(OR(ISNUMBER(FIND("Growth",B272)),ISNUMBER(FIND("Margin",B272))),"",(J272-T272)/U272)))')
    worksheet.write('V273',
                    '=IF(C273="-","",IF(ISBLANK(B273),"",IF(OR(ISNUMBER(FIND("Growth",B273)),ISNUMBER(FIND("Margin",B273))),"",(J273-T273)/U273)))')
    worksheet.write('V274',
                    '=IF(C274="-","",IF(ISBLANK(B274),"",IF(OR(ISNUMBER(FIND("Growth",B274)),ISNUMBER(FIND("Margin",B274))),"",(J274-T274)/U274)))')
    worksheet.write('V275',
                    '=IF(C275="-","",IF(ISBLANK(B275),"",IF(OR(ISNUMBER(FIND("Growth",B275)),ISNUMBER(FIND("Margin",B275))),"",(J275-T275)/U275)))')
    worksheet.write('V276',
                    '=IF(C276="-","",IF(ISBLANK(B276),"",IF(OR(ISNUMBER(FIND("Growth",B276)),ISNUMBER(FIND("Margin",B276))),"",(J276-T276)/U276)))')
    worksheet.write('V277',
                    '=IF(C277="-","",IF(ISBLANK(B277),"",IF(OR(ISNUMBER(FIND("Growth",B277)),ISNUMBER(FIND("Margin",B277))),"",(J277-T277)/U277)))')
    worksheet.write('V278',
                    '=IF(C278="-","",IF(ISBLANK(B278),"",IF(OR(ISNUMBER(FIND("Growth",B278)),ISNUMBER(FIND("Margin",B278))),"",(J278-T278)/U278)))')
    worksheet.write('V279',
                    '=IF(C279="-","",IF(ISBLANK(B279),"",IF(OR(ISNUMBER(FIND("Growth",B279)),ISNUMBER(FIND("Margin",B279))),"",(J279-T279)/U279)))')
    worksheet.write('V280',
                    '=IF(C280="-","",IF(ISBLANK(B280),"",IF(OR(ISNUMBER(FIND("Growth",B280)),ISNUMBER(FIND("Margin",B280))),"",(J280-T280)/U280)))')
    worksheet.write('V281',
                    '=IF(C281="-","",IF(ISBLANK(B281),"",IF(OR(ISNUMBER(FIND("Growth",B281)),ISNUMBER(FIND("Margin",B281))),"",(J281-T281)/U281)))')
    worksheet.write('V282',
                    '=IF(C282="-","",IF(ISBLANK(B282),"",IF(OR(ISNUMBER(FIND("Growth",B282)),ISNUMBER(FIND("Margin",B282))),"",(J282-T282)/U282)))')
    worksheet.write('V283',
                    '=IF(C283="-","",IF(ISBLANK(B283),"",IF(OR(ISNUMBER(FIND("Growth",B283)),ISNUMBER(FIND("Margin",B283))),"",(J283-T283)/U283)))')
    worksheet.write('V284',
                    '=IF(C284="-","",IF(ISBLANK(B284),"",IF(OR(ISNUMBER(FIND("Growth",B284)),ISNUMBER(FIND("Margin",B284))),"",(J284-T284)/U284)))')
    worksheet.write('V285',
                    '=IF(C285="-","",IF(ISBLANK(B285),"",IF(OR(ISNUMBER(FIND("Growth",B285)),ISNUMBER(FIND("Margin",B285))),"",(J285-T285)/U285)))')
    worksheet.write('V286',
                    '=IF(C286="-","",IF(ISBLANK(B286),"",IF(OR(ISNUMBER(FIND("Growth",B286)),ISNUMBER(FIND("Margin",B286))),"",(J286-T286)/U286)))')
    worksheet.write('V287',
                    '=IF(C287="-","",IF(ISBLANK(B287),"",IF(OR(ISNUMBER(FIND("Growth",B287)),ISNUMBER(FIND("Margin",B287))),"",(J287-T287)/U287)))')
    worksheet.write('V288',
                    '=IF(C288="-","",IF(ISBLANK(B288),"",IF(OR(ISNUMBER(FIND("Growth",B288)),ISNUMBER(FIND("Margin",B288))),"",(J288-T288)/U288)))')
    worksheet.write('V289',
                    '=IF(C289="-","",IF(ISBLANK(B289),"",IF(OR(ISNUMBER(FIND("Growth",B289)),ISNUMBER(FIND("Margin",B289))),"",(J289-T289)/U289)))')
    worksheet.write('V290',
                    '=IF(C290="-","",IF(ISBLANK(B290),"",IF(OR(ISNUMBER(FIND("Growth",B290)),ISNUMBER(FIND("Margin",B290))),"",(J290-T290)/U290)))')
    worksheet.write('V291',
                    '=IF(C291="-","",IF(ISBLANK(B291),"",IF(OR(ISNUMBER(FIND("Growth",B291)),ISNUMBER(FIND("Margin",B291))),"",(J291-T291)/U291)))')
    worksheet.write('V292',
                    '=IF(C292="-","",IF(ISBLANK(B292),"",IF(OR(ISNUMBER(FIND("Growth",B292)),ISNUMBER(FIND("Margin",B292))),"",(J292-T292)/U292)))')
    worksheet.write('V293',
                    '=IF(C293="-","",IF(ISBLANK(B293),"",IF(OR(ISNUMBER(FIND("Growth",B293)),ISNUMBER(FIND("Margin",B293))),"",(J293-T293)/U293)))')
    worksheet.write('V294',
                    '=IF(C294="-","",IF(ISBLANK(B294),"",IF(OR(ISNUMBER(FIND("Growth",B294)),ISNUMBER(FIND("Margin",B294))),"",(J294-T294)/U294)))')
    worksheet.write('V295',
                    '=IF(C295="-","",IF(ISBLANK(B295),"",IF(OR(ISNUMBER(FIND("Growth",B295)),ISNUMBER(FIND("Margin",B295))),"",(J295-T295)/U295)))')
    worksheet.write('V296',
                    '=IF(C296="-","",IF(ISBLANK(B296),"",IF(OR(ISNUMBER(FIND("Growth",B296)),ISNUMBER(FIND("Margin",B296))),"",(J296-T296)/U296)))')
    worksheet.write('V297',
                    '=IF(C297="-","",IF(ISBLANK(B297),"",IF(OR(ISNUMBER(FIND("Growth",B297)),ISNUMBER(FIND("Margin",B297))),"",(J297-T297)/U297)))')
    worksheet.write('V298',
                    '=IF(C298="-","",IF(ISBLANK(B298),"",IF(OR(ISNUMBER(FIND("Growth",B298)),ISNUMBER(FIND("Margin",B298))),"",(J298-T298)/U298)))')
    worksheet.write('V299',
                    '=IF(C299="-","",IF(ISBLANK(B299),"",IF(OR(ISNUMBER(FIND("Growth",B299)),ISNUMBER(FIND("Margin",B299))),"",(J299-T299)/U299)))')
    worksheet.write('V300',
                    '=IF(C300="-","",IF(ISBLANK(B300),"",IF(OR(ISNUMBER(FIND("Growth",B300)),ISNUMBER(FIND("Margin",B300))),"",(J300-T300)/U300)))')
    worksheet.write('V301',
                    '=IF(C301="-","",IF(ISBLANK(B301),"",IF(OR(ISNUMBER(FIND("Growth",B301)),ISNUMBER(FIND("Margin",B301))),"",(J301-T301)/U301)))')
    worksheet.write('V302',
                    '=IF(C302="-","",IF(ISBLANK(B302),"",IF(OR(ISNUMBER(FIND("Growth",B302)),ISNUMBER(FIND("Margin",B302))),"",(J302-T302)/U302)))')
    worksheet.write('V303',
                    '=IF(C303="-","",IF(ISBLANK(B303),"",IF(OR(ISNUMBER(FIND("Growth",B303)),ISNUMBER(FIND("Margin",B303))),"",(J303-T303)/U303)))')
    worksheet.write('V304',
                    '=IF(C304="-","",IF(ISBLANK(B304),"",IF(OR(ISNUMBER(FIND("Growth",B304)),ISNUMBER(FIND("Margin",B304))),"",(J304-T304)/U304)))')
    worksheet.write('V305',
                    '=IF(C305="-","",IF(ISBLANK(B305),"",IF(OR(ISNUMBER(FIND("Growth",B305)),ISNUMBER(FIND("Margin",B305))),"",(J305-T305)/U305)))')
    worksheet.write('V306',
                    '=IF(C306="-","",IF(ISBLANK(B306),"",IF(OR(ISNUMBER(FIND("Growth",B306)),ISNUMBER(FIND("Margin",B306))),"",(J306-T306)/U306)))')
    worksheet.write('V307',
                    '=IF(C307="-","",IF(ISBLANK(B307),"",IF(OR(ISNUMBER(FIND("Growth",B307)),ISNUMBER(FIND("Margin",B307))),"",(J307-T307)/U307)))')
    worksheet.write('V308',
                    '=IF(C308="-","",IF(ISBLANK(B308),"",IF(OR(ISNUMBER(FIND("Growth",B308)),ISNUMBER(FIND("Margin",B308))),"",(J308-T308)/U308)))')
    worksheet.write('V309',
                    '=IF(C309="-","",IF(ISBLANK(B309),"",IF(OR(ISNUMBER(FIND("Growth",B309)),ISNUMBER(FIND("Margin",B309))),"",(J309-T309)/U309)))')
    worksheet.write('V310',
                    '=IF(C310="-","",IF(ISBLANK(B310),"",IF(OR(ISNUMBER(FIND("Growth",B310)),ISNUMBER(FIND("Margin",B310))),"",(J310-T310)/U310)))')
    worksheet.write('V311',
                    '=IF(C311="-","",IF(ISBLANK(B311),"",IF(OR(ISNUMBER(FIND("Growth",B311)),ISNUMBER(FIND("Margin",B311))),"",(J311-T311)/U311)))')
    worksheet.write('V312',
                    '=IF(C312="-","",IF(ISBLANK(B312),"",IF(OR(ISNUMBER(FIND("Growth",B312)),ISNUMBER(FIND("Margin",B312))),"",(J312-T312)/U312)))')
    worksheet.write('V313',
                    '=IF(C313="-","",IF(ISBLANK(B313),"",IF(OR(ISNUMBER(FIND("Growth",B313)),ISNUMBER(FIND("Margin",B313))),"",(J313-T313)/U313)))')
    worksheet.write('V314',
                    '=IF(C314="-","",IF(ISBLANK(B314),"",IF(OR(ISNUMBER(FIND("Growth",B314)),ISNUMBER(FIND("Margin",B314))),"",(J314-T314)/U314)))')
    worksheet.write('V315',
                    '=IF(C315="-","",IF(ISBLANK(B315),"",IF(OR(ISNUMBER(FIND("Growth",B315)),ISNUMBER(FIND("Margin",B315))),"",(J315-T315)/U315)))')
    worksheet.write('V316',
                    '=IF(C316="-","",IF(ISBLANK(B316),"",IF(OR(ISNUMBER(FIND("Growth",B316)),ISNUMBER(FIND("Margin",B316))),"",(J316-T316)/U316)))')
    worksheet.write('V317',
                    '=IF(C317="-","",IF(ISBLANK(B317),"",IF(OR(ISNUMBER(FIND("Growth",B317)),ISNUMBER(FIND("Margin",B317))),"",(J317-T317)/U317)))')
    worksheet.write('V318',
                    '=IF(C318="-","",IF(ISBLANK(B318),"",IF(OR(ISNUMBER(FIND("Growth",B318)),ISNUMBER(FIND("Margin",B318))),"",(J318-T318)/U318)))')
    worksheet.write('V319',
                    '=IF(C319="-","",IF(ISBLANK(B319),"",IF(OR(ISNUMBER(FIND("Growth",B319)),ISNUMBER(FIND("Margin",B319))),"",(J319-T319)/U319)))')
    worksheet.write('V320',
                    '=IF(C320="-","",IF(ISBLANK(B320),"",IF(OR(ISNUMBER(FIND("Growth",B320)),ISNUMBER(FIND("Margin",B320))),"",(J320-T320)/U320)))')
    worksheet.write('V321',
                    '=IF(C321="-","",IF(ISBLANK(B321),"",IF(OR(ISNUMBER(FIND("Growth",B321)),ISNUMBER(FIND("Margin",B321))),"",(J321-T321)/U321)))')
    worksheet.write('V322',
                    '=IF(C322="-","",IF(ISBLANK(B322),"",IF(OR(ISNUMBER(FIND("Growth",B322)),ISNUMBER(FIND("Margin",B322))),"",(J322-T322)/U322)))')
    worksheet.write('V323',
                    '=IF(C323="-","",IF(ISBLANK(B323),"",IF(OR(ISNUMBER(FIND("Growth",B323)),ISNUMBER(FIND("Margin",B323))),"",(J323-T323)/U323)))')
    worksheet.write('V324',
                    '=IF(C324="-","",IF(ISBLANK(B324),"",IF(OR(ISNUMBER(FIND("Growth",B324)),ISNUMBER(FIND("Margin",B324))),"",(J324-T324)/U324)))')
    worksheet.write('V325',
                    '=IF(C325="-","",IF(ISBLANK(B325),"",IF(OR(ISNUMBER(FIND("Growth",B325)),ISNUMBER(FIND("Margin",B325))),"",(J325-T325)/U325)))')
    worksheet.write('V326',
                    '=IF(C326="-","",IF(ISBLANK(B326),"",IF(OR(ISNUMBER(FIND("Growth",B326)),ISNUMBER(FIND("Margin",B326))),"",(J326-T326)/U326)))')
    worksheet.write('V327',
                    '=IF(C327="-","",IF(ISBLANK(B327),"",IF(OR(ISNUMBER(FIND("Growth",B327)),ISNUMBER(FIND("Margin",B327))),"",(J327-T327)/U327)))')
    worksheet.write('V328',
                    '=IF(C328="-","",IF(ISBLANK(B328),"",IF(OR(ISNUMBER(FIND("Growth",B328)),ISNUMBER(FIND("Margin",B328))),"",(J328-T328)/U328)))')
    worksheet.write('V329',
                    '=IF(C329="-","",IF(ISBLANK(B329),"",IF(OR(ISNUMBER(FIND("Growth",B329)),ISNUMBER(FIND("Margin",B329))),"",(J329-T329)/U329)))')
    worksheet.write('V330',
                    '=IF(C330="-","",IF(ISBLANK(B330),"",IF(OR(ISNUMBER(FIND("Growth",B330)),ISNUMBER(FIND("Margin",B330))),"",(J330-T330)/U330)))')
    worksheet.write('V331',
                    '=IF(C331="-","",IF(ISBLANK(B331),"",IF(OR(ISNUMBER(FIND("Growth",B331)),ISNUMBER(FIND("Margin",B331))),"",(J331-T331)/U331)))')
    worksheet.write('V332',
                    '=IF(C332="-","",IF(ISBLANK(B332),"",IF(OR(ISNUMBER(FIND("Growth",B332)),ISNUMBER(FIND("Margin",B332))),"",(J332-T332)/U332)))')
    worksheet.write('V333',
                    '=IF(C333="-","",IF(ISBLANK(B333),"",IF(OR(ISNUMBER(FIND("Growth",B333)),ISNUMBER(FIND("Margin",B333))),"",(J333-T333)/U333)))')
    worksheet.write('V334',
                    '=IF(C334="-","",IF(ISBLANK(B334),"",IF(OR(ISNUMBER(FIND("Growth",B334)),ISNUMBER(FIND("Margin",B334))),"",(J334-T334)/U334)))')
    worksheet.write('V335',
                    '=IF(C335="-","",IF(ISBLANK(B335),"",IF(OR(ISNUMBER(FIND("Growth",B335)),ISNUMBER(FIND("Margin",B335))),"",(J335-T335)/U335)))')
    worksheet.write('V336',
                    '=IF(C336="-","",IF(ISBLANK(B336),"",IF(OR(ISNUMBER(FIND("Growth",B336)),ISNUMBER(FIND("Margin",B336))),"",(J336-T336)/U336)))')
    worksheet.write('V337',
                    '=IF(C337="-","",IF(ISBLANK(B337),"",IF(OR(ISNUMBER(FIND("Growth",B337)),ISNUMBER(FIND("Margin",B337))),"",(J337-T337)/U337)))')
    worksheet.write('V338',
                    '=IF(C338="-","",IF(ISBLANK(B338),"",IF(OR(ISNUMBER(FIND("Growth",B338)),ISNUMBER(FIND("Margin",B338))),"",(J338-T338)/U338)))')
    worksheet.write('V339',
                    '=IF(C339="-","",IF(ISBLANK(B339),"",IF(OR(ISNUMBER(FIND("Growth",B339)),ISNUMBER(FIND("Margin",B339))),"",(J339-T339)/U339)))')
    worksheet.write('V340',
                    '=IF(C340="-","",IF(ISBLANK(B340),"",IF(OR(ISNUMBER(FIND("Growth",B340)),ISNUMBER(FIND("Margin",B340))),"",(J340-T340)/U340)))')
    worksheet.write('V341',
                    '=IF(C341="-","",IF(ISBLANK(B341),"",IF(OR(ISNUMBER(FIND("Growth",B341)),ISNUMBER(FIND("Margin",B341))),"",(J341-T341)/U341)))')
    worksheet.write('V342',
                    '=IF(C342="-","",IF(ISBLANK(B342),"",IF(OR(ISNUMBER(FIND("Growth",B342)),ISNUMBER(FIND("Margin",B342))),"",(J342-T342)/U342)))')
    worksheet.write('V343',
                    '=IF(C343="-","",IF(ISBLANK(B343),"",IF(OR(ISNUMBER(FIND("Growth",B343)),ISNUMBER(FIND("Margin",B343))),"",(J343-T343)/U343)))')
    worksheet.write('V344',
                    '=IF(C344="-","",IF(ISBLANK(B344),"",IF(OR(ISNUMBER(FIND("Growth",B344)),ISNUMBER(FIND("Margin",B344))),"",(J344-T344)/U344)))')
    worksheet.write('V345',
                    '=IF(C345="-","",IF(ISBLANK(B345),"",IF(OR(ISNUMBER(FIND("Growth",B345)),ISNUMBER(FIND("Margin",B345))),"",(J345-T345)/U345)))')
    worksheet.write('V346',
                    '=IF(C346="-","",IF(ISBLANK(B346),"",IF(OR(ISNUMBER(FIND("Growth",B346)),ISNUMBER(FIND("Margin",B346))),"",(J346-T346)/U346)))')
    worksheet.write('V347',
                    '=IF(C347="-","",IF(ISBLANK(B347),"",IF(OR(ISNUMBER(FIND("Growth",B347)),ISNUMBER(FIND("Margin",B347))),"",(J347-T347)/U347)))')
    worksheet.write('V348',
                    '=IF(C348="-","",IF(ISBLANK(B348),"",IF(OR(ISNUMBER(FIND("Growth",B348)),ISNUMBER(FIND("Margin",B348))),"",(J348-T348)/U348)))')
    worksheet.write('V349',
                    '=IF(C349="-","",IF(ISBLANK(B349),"",IF(OR(ISNUMBER(FIND("Growth",B349)),ISNUMBER(FIND("Margin",B349))),"",(J349-T349)/U349)))')
    worksheet.write('V350',
                    '=IF(C350="-","",IF(ISBLANK(B350),"",IF(OR(ISNUMBER(FIND("Growth",B350)),ISNUMBER(FIND("Margin",B350))),"",(J350-T350)/U350)))')
    worksheet.write('V351',
                    '=IF(C351="-","",IF(ISBLANK(B351),"",IF(OR(ISNUMBER(FIND("Growth",B351)),ISNUMBER(FIND("Margin",B351))),"",(J351-T351)/U351)))')
    worksheet.write('V352',
                    '=IF(C352="-","",IF(ISBLANK(B352),"",IF(OR(ISNUMBER(FIND("Growth",B352)),ISNUMBER(FIND("Margin",B352))),"",(J352-T352)/U352)))')
    worksheet.write('V353',
                    '=IF(C353="-","",IF(ISBLANK(B353),"",IF(OR(ISNUMBER(FIND("Growth",B353)),ISNUMBER(FIND("Margin",B353))),"",(J353-T353)/U353)))')
    worksheet.write('W144', '=K144')
    worksheet.write('W145', '=IF(OR(D145="-",ISBLANK(D145)),"",(K145-T145)/U145)')
    worksheet.write('W146', '=IF(OR(D146="-",ISBLANK(D146)),"",(K146-T146)/U146)')
    worksheet.write('W147', '=IF(OR(D147="-",ISBLANK(D147)),"",(K147-T147)/U147)')
    worksheet.write('W148', '=IF(OR(D148="-",ISBLANK(D148)),"",(K148-T148)/U148)')
    worksheet.write('W149', '=IF(OR(D149="-",ISBLANK(D149)),"",(K149-T149)/U149)')
    worksheet.write('W150', '=IF(OR(D150="-",ISBLANK(D150)),"",(K150-T150)/U150)')
    worksheet.write('W151', '=IF(OR(D151="-",ISBLANK(D151)),"",(K151-T151)/U151)')
    worksheet.write('W152', '=IF(OR(D152="-",ISBLANK(D152)),"",(K152-T152)/U152)')
    worksheet.write('W153', '=IF(OR(D153="-",ISBLANK(D153)),"",(K153-T153)/U153)')
    worksheet.write('W154', '=IF(OR(D154="-",ISBLANK(D154)),"",(K154-T154)/U154)')
    worksheet.write('W155', '=IF(OR(D155="-",ISBLANK(D155)),"",(K155-T155)/U155)')
    worksheet.write('W156', '=IF(OR(D156="-",ISBLANK(D156)),"",(K156-T156)/U156)')
    worksheet.write('W157', '=IF(OR(D157="-",ISBLANK(D157)),"",(K157-T157)/U157)')
    worksheet.write('W158', '=IF(OR(D158="-",ISBLANK(D158)),"",(K158-T158)/U158)')
    worksheet.write('W159', '=IF(OR(D159="-",ISBLANK(D159)),"",(K159-T159)/U159)')
    worksheet.write('W160', '=IF(OR(D160="-",ISBLANK(D160)),"",(K160-T160)/U160)')
    worksheet.write('W161', '=IF(OR(D161="-",ISBLANK(D161)),"",(K161-T161)/U161)')
    worksheet.write('W162', '=IF(OR(D162="-",ISBLANK(D162)),"",(K162-T162)/U162)')
    worksheet.write('W163', '=IF(OR(D163="-",ISBLANK(D163)),"",(K163-T163)/U163)')
    worksheet.write('W164', '=IF(OR(D164="-",ISBLANK(D164)),"",(K164-T164)/U164)')
    worksheet.write('W165', '=IF(OR(D165="-",ISBLANK(D165)),"",(K165-T165)/U165)')
    worksheet.write('W166', '=IF(OR(D166="-",ISBLANK(D166)),"",(K166-T166)/U166)')
    worksheet.write('W167', '=IF(OR(D167="-",ISBLANK(D167)),"",(K167-T167)/U167)')
    worksheet.write('W168', '=IF(OR(D168="-",ISBLANK(D168)),"",(K168-T168)/U168)')
    worksheet.write('W169', '=IF(OR(D169="-",ISBLANK(D169)),"",(K169-T169)/U169)')
    worksheet.write('W170', '=IF(OR(D170="-",ISBLANK(D170)),"",(K170-T170)/U170)')
    worksheet.write('W171', '=IF(OR(D171="-",ISBLANK(D171)),"",(K171-T171)/U171)')
    worksheet.write('W172', '=IF(OR(D172="-",ISBLANK(D172)),"",(K172-T172)/U172)')
    worksheet.write('W173', '=IF(OR(D173="-",ISBLANK(D173)),"",(K173-T173)/U173)')
    worksheet.write('W174', '=IF(OR(D174="-",ISBLANK(D174)),"",(K174-T174)/U174)')
    worksheet.write('W175', '=IF(OR(D175="-",ISBLANK(D175)),"",(K175-T175)/U175)')
    worksheet.write('W176', '=IF(OR(D176="-",ISBLANK(D176)),"",(K176-T176)/U176)')
    worksheet.write('W177', '=IF(OR(D177="-",ISBLANK(D177)),"",(K177-T177)/U177)')
    worksheet.write('W178', '=IF(OR(D178="-",ISBLANK(D178)),"",(K178-T178)/U178)')
    worksheet.write('W179', '=IF(OR(D179="-",ISBLANK(D179)),"",(K179-T179)/U179)')
    worksheet.write('W180', '=IF(OR(D180="-",ISBLANK(D180)),"",(K180-T180)/U180)')
    worksheet.write('W181', '=IF(OR(D181="-",ISBLANK(D181)),"",(K181-T181)/U181)')
    worksheet.write('W182', '=IF(OR(D182="-",ISBLANK(D182)),"",(K182-T182)/U182)')
    worksheet.write('W183', '=IF(OR(D183="-",ISBLANK(D183)),"",(K183-T183)/U183)')
    worksheet.write('W184', '=IF(OR(D184="-",ISBLANK(D184)),"",(K184-T184)/U184)')
    worksheet.write('W185', '=IF(OR(D185="-",ISBLANK(D185)),"",(K185-T185)/U185)')
    worksheet.write('W186', '=IF(OR(D186="-",ISBLANK(D186)),"",(K186-T186)/U186)')
    worksheet.write('W187', '=IF(OR(D187="-",ISBLANK(D187)),"",(K187-T187)/U187)')
    worksheet.write('W188', '=IF(OR(D188="-",ISBLANK(D188)),"",(K188-T188)/U188)')
    worksheet.write('W189', '=IF(OR(D189="-",ISBLANK(D189)),"",(K189-T189)/U189)')
    worksheet.write('W190', '=IF(OR(D190="-",ISBLANK(D190)),"",(K190-T190)/U190)')
    worksheet.write('W191', '=IF(OR(D191="-",ISBLANK(D191)),"",(K191-T191)/U191)')
    worksheet.write('W192', '=IF(OR(D192="-",ISBLANK(D192)),"",(K192-T192)/U192)')
    worksheet.write('W193', '=IF(OR(D193="-",ISBLANK(D193)),"",(K193-T193)/U193)')
    worksheet.write('W194', '=IF(OR(D194="-",ISBLANK(D194)),"",(K194-T194)/U194)')
    worksheet.write('W195', '=IF(OR(D195="-",ISBLANK(D195)),"",(K195-T195)/U195)')
    worksheet.write('W196', '=IF(OR(D196="-",ISBLANK(D196)),"",(K196-T196)/U196)')
    worksheet.write('W197', '=IF(OR(D197="-",ISBLANK(D197)),"",(K197-T197)/U197)')
    worksheet.write('W198', '=IF(OR(D198="-",ISBLANK(D198)),"",(K198-T198)/U198)')
    worksheet.write('W199', '=IF(OR(D199="-",ISBLANK(D199)),"",(K199-T199)/U199)')
    worksheet.write('W200', '=IF(OR(D200="-",ISBLANK(D200)),"",(K200-T200)/U200)')
    worksheet.write('W201', '=IF(OR(D201="-",ISBLANK(D201)),"",(K201-T201)/U201)')
    worksheet.write('W202', '=IF(OR(D202="-",ISBLANK(D202)),"",(K202-T202)/U202)')
    worksheet.write('W203', '=IF(OR(D203="-",ISBLANK(D203)),"",(K203-T203)/U203)')
    worksheet.write('W204', '=IF(OR(D204="-",ISBLANK(D204)),"",(K204-T204)/U204)')
    worksheet.write('W205', '=IF(OR(D205="-",ISBLANK(D205)),"",(K205-T205)/U205)')
    worksheet.write('W206', '=IF(OR(D206="-",ISBLANK(D206)),"",(K206-T206)/U206)')
    worksheet.write('W207', '=IF(OR(D207="-",ISBLANK(D207)),"",(K207-T207)/U207)')
    worksheet.write('W208', '=IF(OR(D208="-",ISBLANK(D208)),"",(K208-T208)/U208)')
    worksheet.write('W209', '=IF(OR(D209="-",ISBLANK(D209)),"",(K209-T209)/U209)')
    worksheet.write('W210', '=IF(OR(D210="-",ISBLANK(D210)),"",(K210-T210)/U210)')
    worksheet.write('W211', '=IF(OR(D211="-",ISBLANK(D211)),"",(K211-T211)/U211)')
    worksheet.write('W212', '=IF(OR(D212="-",ISBLANK(D212)),"",(K212-T212)/U212)')
    worksheet.write('W213', '=IF(OR(D213="-",ISBLANK(D213)),"",(K213-T213)/U213)')
    worksheet.write('W214', '=IF(OR(D214="-",ISBLANK(D214)),"",(K214-T214)/U214)')
    worksheet.write('W215', '=IF(OR(D215="-",ISBLANK(D215)),"",(K215-T215)/U215)')
    worksheet.write('W216', '=IF(OR(D216="-",ISBLANK(D216)),"",(K216-T216)/U216)')
    worksheet.write('W217', '=IF(OR(D217="-",ISBLANK(D217)),"",(K217-T217)/U217)')
    worksheet.write('W218', '=IF(OR(D218="-",ISBLANK(D218)),"",(K218-T218)/U218)')
    worksheet.write('W219', '=IF(OR(D219="-",ISBLANK(D219)),"",(K219-T219)/U219)')
    worksheet.write('W220', '=IF(OR(D220="-",ISBLANK(D220)),"",(K220-T220)/U220)')
    worksheet.write('W221', '=IF(OR(D221="-",ISBLANK(D221)),"",(K221-T221)/U221)')
    worksheet.write('W222', '=IF(OR(D222="-",ISBLANK(D222)),"",(K222-T222)/U222)')
    worksheet.write('W223', '=IF(OR(D223="-",ISBLANK(D223)),"",(K223-T223)/U223)')
    worksheet.write('W224', '=IF(OR(D224="-",ISBLANK(D224)),"",(K224-T224)/U224)')
    worksheet.write('W225', '=IF(OR(D225="-",ISBLANK(D225)),"",(K225-T225)/U225)')
    worksheet.write('W226', '=IF(OR(D226="-",ISBLANK(D226)),"",(K226-T226)/U226)')
    worksheet.write('W227', '=IF(OR(D227="-",ISBLANK(D227)),"",(K227-T227)/U227)')
    worksheet.write('W228', '=IF(OR(D228="-",ISBLANK(D228)),"",(K228-T228)/U228)')
    worksheet.write('W229', '=IF(OR(D229="-",ISBLANK(D229)),"",(K229-T229)/U229)')
    worksheet.write('W230', '=IF(OR(D230="-",ISBLANK(D230)),"",(K230-T230)/U230)')
    worksheet.write('W231', '=IF(OR(D231="-",ISBLANK(D231)),"",(K231-T231)/U231)')
    worksheet.write('W232', '=IF(OR(D232="-",ISBLANK(D232)),"",(K232-T232)/U232)')
    worksheet.write('W233', '=IF(OR(D233="-",ISBLANK(D233)),"",(K233-T233)/U233)')
    worksheet.write('W234', '=IF(OR(D234="-",ISBLANK(D234)),"",(K234-T234)/U234)')
    worksheet.write('W235', '=IF(OR(D235="-",ISBLANK(D235)),"",(K235-T235)/U235)')
    worksheet.write('W236', '=IF(OR(D236="-",ISBLANK(D236)),"",(K236-T236)/U236)')
    worksheet.write('W237', '=IF(OR(D237="-",ISBLANK(D237)),"",(K237-T237)/U237)')
    worksheet.write('W238', '=IF(OR(D238="-",ISBLANK(D238)),"",(K238-T238)/U238)')
    worksheet.write('W239', '=IF(OR(D239="-",ISBLANK(D239)),"",(K239-T239)/U239)')
    worksheet.write('W240', '=IF(OR(D240="-",ISBLANK(D240)),"",(K240-T240)/U240)')
    worksheet.write('W241', '=IF(OR(D241="-",ISBLANK(D241)),"",(K241-T241)/U241)')
    worksheet.write('W242', '=IF(OR(D242="-",ISBLANK(D242)),"",(K242-T242)/U242)')
    worksheet.write('W243', '=IF(OR(D243="-",ISBLANK(D243)),"",(K243-T243)/U243)')
    worksheet.write('W244', '=IF(OR(D244="-",ISBLANK(D244)),"",(K244-T244)/U244)')
    worksheet.write('W245', '=IF(OR(D245="-",ISBLANK(D245)),"",(K245-T245)/U245)')
    worksheet.write('W246', '=IF(OR(D246="-",ISBLANK(D246)),"",(K246-T246)/U246)')
    worksheet.write('W247', '=IF(OR(D247="-",ISBLANK(D247)),"",(K247-T247)/U247)')
    worksheet.write('W248', '=IF(OR(D248="-",ISBLANK(D248)),"",(K248-T248)/U248)')
    worksheet.write('W249', '=IF(OR(D249="-",ISBLANK(D249)),"",(K249-T249)/U249)')
    worksheet.write('W250', '=IF(OR(D250="-",ISBLANK(D250)),"",(K250-T250)/U250)')
    worksheet.write('W251', '=IF(OR(D251="-",ISBLANK(D251)),"",(K251-T251)/U251)')
    worksheet.write('W252', '=IF(OR(D252="-",ISBLANK(D252)),"",(K252-T252)/U252)')
    worksheet.write('W253', '=IF(OR(D253="-",ISBLANK(D253)),"",(K253-T253)/U253)')
    worksheet.write('W254', '=IF(OR(D254="-",ISBLANK(D254)),"",(K254-T254)/U254)')
    worksheet.write('W255', '=IF(OR(D255="-",ISBLANK(D255)),"",(K255-T255)/U255)')
    worksheet.write('W256', '=IF(OR(D256="-",ISBLANK(D256)),"",(K256-T256)/U256)')
    worksheet.write('W257', '=IF(OR(D257="-",ISBLANK(D257)),"",(K257-T257)/U257)')
    worksheet.write('W258', '=IF(OR(D258="-",ISBLANK(D258)),"",(K258-T258)/U258)')
    worksheet.write('W259', '=IF(OR(D259="-",ISBLANK(D259)),"",(K259-T259)/U259)')
    worksheet.write('W260', '=IF(OR(D260="-",ISBLANK(D260)),"",(K260-T260)/U260)')
    worksheet.write('W261', '=IF(OR(D261="-",ISBLANK(D261)),"",(K261-T261)/U261)')
    worksheet.write('W262', '=IF(OR(D262="-",ISBLANK(D262)),"",(K262-T262)/U262)')
    worksheet.write('W263', '=IF(OR(D263="-",ISBLANK(D263)),"",(K263-T263)/U263)')
    worksheet.write('W264', '=IF(OR(D264="-",ISBLANK(D264)),"",(K264-T264)/U264)')
    worksheet.write('W265', '=IF(OR(D265="-",ISBLANK(D265)),"",(K265-T265)/U265)')
    worksheet.write('W266', '=IF(OR(D266="-",ISBLANK(D266)),"",(K266-T266)/U266)')
    worksheet.write('W267', '=IF(OR(D267="-",ISBLANK(D267)),"",(K267-T267)/U267)')
    worksheet.write('W268', '=IF(OR(D268="-",ISBLANK(D268)),"",(K268-T268)/U268)')
    worksheet.write('W269', '=IF(OR(D269="-",ISBLANK(D269)),"",(K269-T269)/U269)')
    worksheet.write('W270', '=IF(OR(D270="-",ISBLANK(D270)),"",(K270-T270)/U270)')
    worksheet.write('W271', '=IF(OR(D271="-",ISBLANK(D271)),"",(K271-T271)/U271)')
    worksheet.write('W272', '=IF(OR(D272="-",ISBLANK(D272)),"",(K272-T272)/U272)')
    worksheet.write('W273', '=IF(OR(D273="-",ISBLANK(D273)),"",(K273-T273)/U273)')
    worksheet.write('W274', '=IF(OR(D274="-",ISBLANK(D274)),"",(K274-T274)/U274)')
    worksheet.write('W275', '=IF(OR(D275="-",ISBLANK(D275)),"",(K275-T275)/U275)')
    worksheet.write('W276', '=IF(OR(D276="-",ISBLANK(D276)),"",(K276-T276)/U276)')
    worksheet.write('W277', '=IF(OR(D277="-",ISBLANK(D277)),"",(K277-T277)/U277)')
    worksheet.write('W278', '=IF(OR(D278="-",ISBLANK(D278)),"",(K278-T278)/U278)')
    worksheet.write('W279', '=IF(OR(D279="-",ISBLANK(D279)),"",(K279-T279)/U279)')
    worksheet.write('W280', '=IF(OR(D280="-",ISBLANK(D280)),"",(K280-T280)/U280)')
    worksheet.write('W281', '=IF(OR(D281="-",ISBLANK(D281)),"",(K281-T281)/U281)')
    worksheet.write('W282', '=IF(OR(D282="-",ISBLANK(D282)),"",(K282-T282)/U282)')
    worksheet.write('W283', '=IF(OR(D283="-",ISBLANK(D283)),"",(K283-T283)/U283)')
    worksheet.write('W284', '=IF(OR(D284="-",ISBLANK(D284)),"",(K284-T284)/U284)')
    worksheet.write('W285', '=IF(OR(D285="-",ISBLANK(D285)),"",(K285-T285)/U285)')
    worksheet.write('W286', '=IF(OR(D286="-",ISBLANK(D286)),"",(K286-T286)/U286)')
    worksheet.write('W287', '=IF(OR(D287="-",ISBLANK(D287)),"",(K287-T287)/U287)')
    worksheet.write('W288', '=IF(OR(D288="-",ISBLANK(D288)),"",(K288-T288)/U288)')
    worksheet.write('W289', '=IF(OR(D289="-",ISBLANK(D289)),"",(K289-T289)/U289)')
    worksheet.write('W290', '=IF(OR(D290="-",ISBLANK(D290)),"",(K290-T290)/U290)')
    worksheet.write('W291', '=IF(OR(D291="-",ISBLANK(D291)),"",(K291-T291)/U291)')
    worksheet.write('W292', '=IF(OR(D292="-",ISBLANK(D292)),"",(K292-T292)/U292)')
    worksheet.write('W293', '=IF(OR(D293="-",ISBLANK(D293)),"",(K293-T293)/U293)')
    worksheet.write('W294', '=IF(OR(D294="-",ISBLANK(D294)),"",(K294-T294)/U294)')
    worksheet.write('W295', '=IF(OR(D295="-",ISBLANK(D295)),"",(K295-T295)/U295)')
    worksheet.write('W296', '=IF(OR(D296="-",ISBLANK(D296)),"",(K296-T296)/U296)')
    worksheet.write('W297', '=IF(OR(D297="-",ISBLANK(D297)),"",(K297-T297)/U297)')
    worksheet.write('W298', '=IF(OR(D298="-",ISBLANK(D298)),"",(K298-T298)/U298)')
    worksheet.write('W299', '=IF(OR(D299="-",ISBLANK(D299)),"",(K299-T299)/U299)')
    worksheet.write('W300', '=IF(OR(D300="-",ISBLANK(D300)),"",(K300-T300)/U300)')
    worksheet.write('W301', '=IF(OR(D301="-",ISBLANK(D301)),"",(K301-T301)/U301)')
    worksheet.write('W302', '=IF(OR(D302="-",ISBLANK(D302)),"",(K302-T302)/U302)')
    worksheet.write('W303', '=IF(OR(D303="-",ISBLANK(D303)),"",(K303-T303)/U303)')
    worksheet.write('W304', '=IF(OR(D304="-",ISBLANK(D304)),"",(K304-T304)/U304)')
    worksheet.write('W305', '=IF(OR(D305="-",ISBLANK(D305)),"",(K305-T305)/U305)')
    worksheet.write('W306', '=IF(OR(D306="-",ISBLANK(D306)),"",(K306-T306)/U306)')
    worksheet.write('W307', '=IF(OR(D307="-",ISBLANK(D307)),"",(K307-T307)/U307)')
    worksheet.write('W308', '=IF(OR(D308="-",ISBLANK(D308)),"",(K308-T308)/U308)')
    worksheet.write('W309', '=IF(OR(D309="-",ISBLANK(D309)),"",(K309-T309)/U309)')
    worksheet.write('W310', '=IF(OR(D310="-",ISBLANK(D310)),"",(K310-T310)/U310)')
    worksheet.write('W311', '=IF(OR(D311="-",ISBLANK(D311)),"",(K311-T311)/U311)')
    worksheet.write('W312', '=IF(OR(D312="-",ISBLANK(D312)),"",(K312-T312)/U312)')
    worksheet.write('W313', '=IF(OR(D313="-",ISBLANK(D313)),"",(K313-T313)/U313)')
    worksheet.write('W314', '=IF(OR(D314="-",ISBLANK(D314)),"",(K314-T314)/U314)')
    worksheet.write('W315', '=IF(OR(D315="-",ISBLANK(D315)),"",(K315-T315)/U315)')
    worksheet.write('W316', '=IF(OR(D316="-",ISBLANK(D316)),"",(K316-T316)/U316)')
    worksheet.write('W317', '=IF(OR(D317="-",ISBLANK(D317)),"",(K317-T317)/U317)')
    worksheet.write('W318', '=IF(OR(D318="-",ISBLANK(D318)),"",(K318-T318)/U318)')
    worksheet.write('W319', '=IF(OR(D319="-",ISBLANK(D319)),"",(K319-T319)/U319)')
    worksheet.write('W320', '=IF(OR(D320="-",ISBLANK(D320)),"",(K320-T320)/U320)')
    worksheet.write('W321', '=IF(OR(D321="-",ISBLANK(D321)),"",(K321-T321)/U321)')
    worksheet.write('W322', '=IF(OR(D322="-",ISBLANK(D322)),"",(K322-T322)/U322)')
    worksheet.write('W323', '=IF(OR(D323="-",ISBLANK(D323)),"",(K323-T323)/U323)')
    worksheet.write('W324', '=IF(OR(D324="-",ISBLANK(D324)),"",(K324-T324)/U324)')
    worksheet.write('W325', '=IF(OR(D325="-",ISBLANK(D325)),"",(K325-T325)/U325)')
    worksheet.write('W326', '=IF(OR(D326="-",ISBLANK(D326)),"",(K326-T326)/U326)')
    worksheet.write('W327', '=IF(OR(D327="-",ISBLANK(D327)),"",(K327-T327)/U327)')
    worksheet.write('W328', '=IF(OR(D328="-",ISBLANK(D328)),"",(K328-T328)/U328)')
    worksheet.write('W329', '=IF(OR(D329="-",ISBLANK(D329)),"",(K329-T329)/U329)')
    worksheet.write('W330', '=IF(OR(D330="-",ISBLANK(D330)),"",(K330-T330)/U330)')
    worksheet.write('W331', '=IF(OR(D331="-",ISBLANK(D331)),"",(K331-T331)/U331)')
    worksheet.write('W332', '=IF(OR(D332="-",ISBLANK(D332)),"",(K332-T332)/U332)')
    worksheet.write('W333', '=IF(OR(D333="-",ISBLANK(D333)),"",(K333-T333)/U333)')
    worksheet.write('W334', '=IF(OR(D334="-",ISBLANK(D334)),"",(K334-T334)/U334)')
    worksheet.write('W335', '=IF(OR(D335="-",ISBLANK(D335)),"",(K335-T335)/U335)')
    worksheet.write('W336', '=IF(OR(D336="-",ISBLANK(D336)),"",(K336-T336)/U336)')
    worksheet.write('W337', '=IF(OR(D337="-",ISBLANK(D337)),"",(K337-T337)/U337)')
    worksheet.write('W338', '=IF(OR(D338="-",ISBLANK(D338)),"",(K338-T338)/U338)')
    worksheet.write('W339', '=IF(OR(D339="-",ISBLANK(D339)),"",(K339-T339)/U339)')
    worksheet.write('W340', '=IF(OR(D340="-",ISBLANK(D340)),"",(K340-T340)/U340)')
    worksheet.write('W341', '=IF(OR(D341="-",ISBLANK(D341)),"",(K341-T341)/U341)')
    worksheet.write('W342', '=IF(OR(D342="-",ISBLANK(D342)),"",(K342-T342)/U342)')
    worksheet.write('W343', '=IF(OR(D343="-",ISBLANK(D343)),"",(K343-T343)/U343)')
    worksheet.write('W344', '=IF(OR(D344="-",ISBLANK(D344)),"",(K344-T344)/U344)')
    worksheet.write('W345', '=IF(OR(D345="-",ISBLANK(D345)),"",(K345-T345)/U345)')
    worksheet.write('W346', '=IF(OR(D346="-",ISBLANK(D346)),"",(K346-T346)/U346)')
    worksheet.write('W347', '=IF(OR(D347="-",ISBLANK(D347)),"",(K347-T347)/U347)')
    worksheet.write('W348', '=IF(OR(D348="-",ISBLANK(D348)),"",(K348-T348)/U348)')
    worksheet.write('W349', '=IF(OR(D349="-",ISBLANK(D349)),"",(K349-T349)/U349)')
    worksheet.write('W350', '=IF(OR(D350="-",ISBLANK(D350)),"",(K350-T350)/U350)')
    worksheet.write('W351', '=IF(OR(D351="-",ISBLANK(D351)),"",(K351-T351)/U351)')
    worksheet.write('W352', '=IF(OR(D352="-",ISBLANK(D352)),"",(K352-T352)/U352)')
    worksheet.write('W353', '=IF(OR(D353="-",ISBLANK(D353)),"",(K353-T353)/U353)')
    worksheet.write('X144', '=L144')
    worksheet.write('X145', '=IF(OR(E145="-",ISBLANK(E145)),"",(L145-T145)/U145)')
    worksheet.write('X146', '=IF(OR(E146="-",ISBLANK(E146)),"",(L146-T146)/U146)')
    worksheet.write('X147', '=IF(OR(E147="-",ISBLANK(E147)),"",(L147-T147)/U147)')
    worksheet.write('X148', '=IF(OR(E148="-",ISBLANK(E148)),"",(L148-T148)/U148)')
    worksheet.write('X149', '=IF(OR(E149="-",ISBLANK(E149)),"",(L149-T149)/U149)')
    worksheet.write('X150', '=IF(OR(E150="-",ISBLANK(E150)),"",(L150-T150)/U150)')
    worksheet.write('X151', '=IF(OR(E151="-",ISBLANK(E151)),"",(L151-T151)/U151)')
    worksheet.write('X152', '=IF(OR(E152="-",ISBLANK(E152)),"",(L152-T152)/U152)')
    worksheet.write('X153', '=IF(OR(E153="-",ISBLANK(E153)),"",(L153-T153)/U153)')
    worksheet.write('X154', '=IF(OR(E154="-",ISBLANK(E154)),"",(L154-T154)/U154)')
    worksheet.write('X155', '=IF(OR(E155="-",ISBLANK(E155)),"",(L155-T155)/U155)')
    worksheet.write('X156', '=IF(OR(E156="-",ISBLANK(E156)),"",(L156-T156)/U156)')
    worksheet.write('X157', '=IF(OR(E157="-",ISBLANK(E157)),"",(L157-T157)/U157)')
    worksheet.write('X158', '=IF(OR(E158="-",ISBLANK(E158)),"",(L158-T158)/U158)')
    worksheet.write('X159', '=IF(OR(E159="-",ISBLANK(E159)),"",(L159-T159)/U159)')
    worksheet.write('X160', '=IF(OR(E160="-",ISBLANK(E160)),"",(L160-T160)/U160)')
    worksheet.write('X161', '=IF(OR(E161="-",ISBLANK(E161)),"",(L161-T161)/U161)')
    worksheet.write('X162', '=IF(OR(E162="-",ISBLANK(E162)),"",(L162-T162)/U162)')
    worksheet.write('X163', '=IF(OR(E163="-",ISBLANK(E163)),"",(L163-T163)/U163)')
    worksheet.write('X164', '=IF(OR(E164="-",ISBLANK(E164)),"",(L164-T164)/U164)')
    worksheet.write('X165', '=IF(OR(E165="-",ISBLANK(E165)),"",(L165-T165)/U165)')
    worksheet.write('X166', '=IF(OR(E166="-",ISBLANK(E166)),"",(L166-T166)/U166)')
    worksheet.write('X167', '=IF(OR(E167="-",ISBLANK(E167)),"",(L167-T167)/U167)')
    worksheet.write('X168', '=IF(OR(E168="-",ISBLANK(E168)),"",(L168-T168)/U168)')
    worksheet.write('X169', '=IF(OR(E169="-",ISBLANK(E169)),"",(L169-T169)/U169)')
    worksheet.write('X170', '=IF(OR(E170="-",ISBLANK(E170)),"",(L170-T170)/U170)')
    worksheet.write('X171', '=IF(OR(E171="-",ISBLANK(E171)),"",(L171-T171)/U171)')
    worksheet.write('X172', '=IF(OR(E172="-",ISBLANK(E172)),"",(L172-T172)/U172)')
    worksheet.write('X173', '=IF(OR(E173="-",ISBLANK(E173)),"",(L173-T173)/U173)')
    worksheet.write('X174', '=IF(OR(E174="-",ISBLANK(E174)),"",(L174-T174)/U174)')
    worksheet.write('X175', '=IF(OR(E175="-",ISBLANK(E175)),"",(L175-T175)/U175)')
    worksheet.write('X176', '=IF(OR(E176="-",ISBLANK(E176)),"",(L176-T176)/U176)')
    worksheet.write('X177', '=IF(OR(E177="-",ISBLANK(E177)),"",(L177-T177)/U177)')
    worksheet.write('X178', '=IF(OR(E178="-",ISBLANK(E178)),"",(L178-T178)/U178)')
    worksheet.write('X179', '=IF(OR(E179="-",ISBLANK(E179)),"",(L179-T179)/U179)')
    worksheet.write('X180', '=IF(OR(E180="-",ISBLANK(E180)),"",(L180-T180)/U180)')
    worksheet.write('X181', '=IF(OR(E181="-",ISBLANK(E181)),"",(L181-T181)/U181)')
    worksheet.write('X182', '=IF(OR(E182="-",ISBLANK(E182)),"",(L182-T182)/U182)')
    worksheet.write('X183', '=IF(OR(E183="-",ISBLANK(E183)),"",(L183-T183)/U183)')
    worksheet.write('X184', '=IF(OR(E184="-",ISBLANK(E184)),"",(L184-T184)/U184)')
    worksheet.write('X185', '=IF(OR(E185="-",ISBLANK(E185)),"",(L185-T185)/U185)')
    worksheet.write('X186', '=IF(OR(E186="-",ISBLANK(E186)),"",(L186-T186)/U186)')
    worksheet.write('X187', '=IF(OR(E187="-",ISBLANK(E187)),"",(L187-T187)/U187)')
    worksheet.write('X188', '=IF(OR(E188="-",ISBLANK(E188)),"",(L188-T188)/U188)')
    worksheet.write('X189', '=IF(OR(E189="-",ISBLANK(E189)),"",(L189-T189)/U189)')
    worksheet.write('X190', '=IF(OR(E190="-",ISBLANK(E190)),"",(L190-T190)/U190)')
    worksheet.write('X191', '=IF(OR(E191="-",ISBLANK(E191)),"",(L191-T191)/U191)')
    worksheet.write('X192', '=IF(OR(E192="-",ISBLANK(E192)),"",(L192-T192)/U192)')
    worksheet.write('X193', '=IF(OR(E193="-",ISBLANK(E193)),"",(L193-T193)/U193)')
    worksheet.write('X194', '=IF(OR(E194="-",ISBLANK(E194)),"",(L194-T194)/U194)')
    worksheet.write('X195', '=IF(OR(E195="-",ISBLANK(E195)),"",(L195-T195)/U195)')
    worksheet.write('X196', '=IF(OR(E196="-",ISBLANK(E196)),"",(L196-T196)/U196)')
    worksheet.write('X197', '=IF(OR(E197="-",ISBLANK(E197)),"",(L197-T197)/U197)')
    worksheet.write('X198', '=IF(OR(E198="-",ISBLANK(E198)),"",(L198-T198)/U198)')
    worksheet.write('X199', '=IF(OR(E199="-",ISBLANK(E199)),"",(L199-T199)/U199)')
    worksheet.write('X200', '=IF(OR(E200="-",ISBLANK(E200)),"",(L200-T200)/U200)')
    worksheet.write('X201', '=IF(OR(E201="-",ISBLANK(E201)),"",(L201-T201)/U201)')
    worksheet.write('X202', '=IF(OR(E202="-",ISBLANK(E202)),"",(L202-T202)/U202)')
    worksheet.write('X203', '=IF(OR(E203="-",ISBLANK(E203)),"",(L203-T203)/U203)')
    worksheet.write('X204', '=IF(OR(E204="-",ISBLANK(E204)),"",(L204-T204)/U204)')
    worksheet.write('X205', '=IF(OR(E205="-",ISBLANK(E205)),"",(L205-T205)/U205)')
    worksheet.write('X206', '=IF(OR(E206="-",ISBLANK(E206)),"",(L206-T206)/U206)')
    worksheet.write('X207', '=IF(OR(E207="-",ISBLANK(E207)),"",(L207-T207)/U207)')
    worksheet.write('X208', '=IF(OR(E208="-",ISBLANK(E208)),"",(L208-T208)/U208)')
    worksheet.write('X209', '=IF(OR(E209="-",ISBLANK(E209)),"",(L209-T209)/U209)')
    worksheet.write('X210', '=IF(OR(E210="-",ISBLANK(E210)),"",(L210-T210)/U210)')
    worksheet.write('X211', '=IF(OR(E211="-",ISBLANK(E211)),"",(L211-T211)/U211)')
    worksheet.write('X212', '=IF(OR(E212="-",ISBLANK(E212)),"",(L212-T212)/U212)')
    worksheet.write('X213', '=IF(OR(E213="-",ISBLANK(E213)),"",(L213-T213)/U213)')
    worksheet.write('X214', '=IF(OR(E214="-",ISBLANK(E214)),"",(L214-T214)/U214)')
    worksheet.write('X215', '=IF(OR(E215="-",ISBLANK(E215)),"",(L215-T215)/U215)')
    worksheet.write('X216', '=IF(OR(E216="-",ISBLANK(E216)),"",(L216-T216)/U216)')
    worksheet.write('X217', '=IF(OR(E217="-",ISBLANK(E217)),"",(L217-T217)/U217)')
    worksheet.write('X218', '=IF(OR(E218="-",ISBLANK(E218)),"",(L218-T218)/U218)')
    worksheet.write('X219', '=IF(OR(E219="-",ISBLANK(E219)),"",(L219-T219)/U219)')
    worksheet.write('X220', '=IF(OR(E220="-",ISBLANK(E220)),"",(L220-T220)/U220)')
    worksheet.write('X221', '=IF(OR(E221="-",ISBLANK(E221)),"",(L221-T221)/U221)')
    worksheet.write('X222', '=IF(OR(E222="-",ISBLANK(E222)),"",(L222-T222)/U222)')
    worksheet.write('X223', '=IF(OR(E223="-",ISBLANK(E223)),"",(L223-T223)/U223)')
    worksheet.write('X224', '=IF(OR(E224="-",ISBLANK(E224)),"",(L224-T224)/U224)')
    worksheet.write('X225', '=IF(OR(E225="-",ISBLANK(E225)),"",(L225-T225)/U225)')
    worksheet.write('X226', '=IF(OR(E226="-",ISBLANK(E226)),"",(L226-T226)/U226)')
    worksheet.write('X227', '=IF(OR(E227="-",ISBLANK(E227)),"",(L227-T227)/U227)')
    worksheet.write('X228', '=IF(OR(E228="-",ISBLANK(E228)),"",(L228-T228)/U228)')
    worksheet.write('X229', '=IF(OR(E229="-",ISBLANK(E229)),"",(L229-T229)/U229)')
    worksheet.write('X230', '=IF(OR(E230="-",ISBLANK(E230)),"",(L230-T230)/U230)')
    worksheet.write('X231', '=IF(OR(E231="-",ISBLANK(E231)),"",(L231-T231)/U231)')
    worksheet.write('X232', '=IF(OR(E232="-",ISBLANK(E232)),"",(L232-T232)/U232)')
    worksheet.write('X233', '=IF(OR(E233="-",ISBLANK(E233)),"",(L233-T233)/U233)')
    worksheet.write('X234', '=IF(OR(E234="-",ISBLANK(E234)),"",(L234-T234)/U234)')
    worksheet.write('X235', '=IF(OR(E235="-",ISBLANK(E235)),"",(L235-T235)/U235)')
    worksheet.write('X236', '=IF(OR(E236="-",ISBLANK(E236)),"",(L236-T236)/U236)')
    worksheet.write('X237', '=IF(OR(E237="-",ISBLANK(E237)),"",(L237-T237)/U237)')
    worksheet.write('X238', '=IF(OR(E238="-",ISBLANK(E238)),"",(L238-T238)/U238)')
    worksheet.write('X239', '=IF(OR(E239="-",ISBLANK(E239)),"",(L239-T239)/U239)')
    worksheet.write('X240', '=IF(OR(E240="-",ISBLANK(E240)),"",(L240-T240)/U240)')
    worksheet.write('X241', '=IF(OR(E241="-",ISBLANK(E241)),"",(L241-T241)/U241)')
    worksheet.write('X242', '=IF(OR(E242="-",ISBLANK(E242)),"",(L242-T242)/U242)')
    worksheet.write('X243', '=IF(OR(E243="-",ISBLANK(E243)),"",(L243-T243)/U243)')
    worksheet.write('X244', '=IF(OR(E244="-",ISBLANK(E244)),"",(L244-T244)/U244)')
    worksheet.write('X245', '=IF(OR(E245="-",ISBLANK(E245)),"",(L245-T245)/U245)')
    worksheet.write('X246', '=IF(OR(E246="-",ISBLANK(E246)),"",(L246-T246)/U246)')
    worksheet.write('X247', '=IF(OR(E247="-",ISBLANK(E247)),"",(L247-T247)/U247)')
    worksheet.write('X248', '=IF(OR(E248="-",ISBLANK(E248)),"",(L248-T248)/U248)')
    worksheet.write('X249', '=IF(OR(E249="-",ISBLANK(E249)),"",(L249-T249)/U249)')
    worksheet.write('X250', '=IF(OR(E250="-",ISBLANK(E250)),"",(L250-T250)/U250)')
    worksheet.write('X251', '=IF(OR(E251="-",ISBLANK(E251)),"",(L251-T251)/U251)')
    worksheet.write('X252', '=IF(OR(E252="-",ISBLANK(E252)),"",(L252-T252)/U252)')
    worksheet.write('X253', '=IF(OR(E253="-",ISBLANK(E253)),"",(L253-T253)/U253)')
    worksheet.write('X254', '=IF(OR(E254="-",ISBLANK(E254)),"",(L254-T254)/U254)')
    worksheet.write('X255', '=IF(OR(E255="-",ISBLANK(E255)),"",(L255-T255)/U255)')
    worksheet.write('X256', '=IF(OR(E256="-",ISBLANK(E256)),"",(L256-T256)/U256)')
    worksheet.write('X257', '=IF(OR(E257="-",ISBLANK(E257)),"",(L257-T257)/U257)')
    worksheet.write('X258', '=IF(OR(E258="-",ISBLANK(E258)),"",(L258-T258)/U258)')
    worksheet.write('X259', '=IF(OR(E259="-",ISBLANK(E259)),"",(L259-T259)/U259)')
    worksheet.write('X260', '=IF(OR(E260="-",ISBLANK(E260)),"",(L260-T260)/U260)')
    worksheet.write('X261', '=IF(OR(E261="-",ISBLANK(E261)),"",(L261-T261)/U261)')
    worksheet.write('X262', '=IF(OR(E262="-",ISBLANK(E262)),"",(L262-T262)/U262)')
    worksheet.write('X263', '=IF(OR(E263="-",ISBLANK(E263)),"",(L263-T263)/U263)')
    worksheet.write('X264', '=IF(OR(E264="-",ISBLANK(E264)),"",(L264-T264)/U264)')
    worksheet.write('X265', '=IF(OR(E265="-",ISBLANK(E265)),"",(L265-T265)/U265)')
    worksheet.write('X266', '=IF(OR(E266="-",ISBLANK(E266)),"",(L266-T266)/U266)')
    worksheet.write('X267', '=IF(OR(E267="-",ISBLANK(E267)),"",(L267-T267)/U267)')
    worksheet.write('X268', '=IF(OR(E268="-",ISBLANK(E268)),"",(L268-T268)/U268)')
    worksheet.write('X269', '=IF(OR(E269="-",ISBLANK(E269)),"",(L269-T269)/U269)')
    worksheet.write('X270', '=IF(OR(E270="-",ISBLANK(E270)),"",(L270-T270)/U270)')
    worksheet.write('X271', '=IF(OR(E271="-",ISBLANK(E271)),"",(L271-T271)/U271)')
    worksheet.write('X272', '=IF(OR(E272="-",ISBLANK(E272)),"",(L272-T272)/U272)')
    worksheet.write('X273', '=IF(OR(E273="-",ISBLANK(E273)),"",(L273-T273)/U273)')
    worksheet.write('X274', '=IF(OR(E274="-",ISBLANK(E274)),"",(L274-T274)/U274)')
    worksheet.write('X275', '=IF(OR(E275="-",ISBLANK(E275)),"",(L275-T275)/U275)')
    worksheet.write('X276', '=IF(OR(E276="-",ISBLANK(E276)),"",(L276-T276)/U276)')
    worksheet.write('X277', '=IF(OR(E277="-",ISBLANK(E277)),"",(L277-T277)/U277)')
    worksheet.write('X278', '=IF(OR(E278="-",ISBLANK(E278)),"",(L278-T278)/U278)')
    worksheet.write('X279', '=IF(OR(E279="-",ISBLANK(E279)),"",(L279-T279)/U279)')
    worksheet.write('X280', '=IF(OR(E280="-",ISBLANK(E280)),"",(L280-T280)/U280)')
    worksheet.write('X281', '=IF(OR(E281="-",ISBLANK(E281)),"",(L281-T281)/U281)')
    worksheet.write('X282', '=IF(OR(E282="-",ISBLANK(E282)),"",(L282-T282)/U282)')
    worksheet.write('X283', '=IF(OR(E283="-",ISBLANK(E283)),"",(L283-T283)/U283)')
    worksheet.write('X284', '=IF(OR(E284="-",ISBLANK(E284)),"",(L284-T284)/U284)')
    worksheet.write('X285', '=IF(OR(E285="-",ISBLANK(E285)),"",(L285-T285)/U285)')
    worksheet.write('X286', '=IF(OR(E286="-",ISBLANK(E286)),"",(L286-T286)/U286)')
    worksheet.write('X287', '=IF(OR(E287="-",ISBLANK(E287)),"",(L287-T287)/U287)')
    worksheet.write('X288', '=IF(OR(E288="-",ISBLANK(E288)),"",(L288-T288)/U288)')
    worksheet.write('X289', '=IF(OR(E289="-",ISBLANK(E289)),"",(L289-T289)/U289)')
    worksheet.write('X290', '=IF(OR(E290="-",ISBLANK(E290)),"",(L290-T290)/U290)')
    worksheet.write('X291', '=IF(OR(E291="-",ISBLANK(E291)),"",(L291-T291)/U291)')
    worksheet.write('X292', '=IF(OR(E292="-",ISBLANK(E292)),"",(L292-T292)/U292)')
    worksheet.write('X293', '=IF(OR(E293="-",ISBLANK(E293)),"",(L293-T293)/U293)')
    worksheet.write('X294', '=IF(OR(E294="-",ISBLANK(E294)),"",(L294-T294)/U294)')
    worksheet.write('X295', '=IF(OR(E295="-",ISBLANK(E295)),"",(L295-T295)/U295)')
    worksheet.write('X296', '=IF(OR(E296="-",ISBLANK(E296)),"",(L296-T296)/U296)')
    worksheet.write('X297', '=IF(OR(E297="-",ISBLANK(E297)),"",(L297-T297)/U297)')
    worksheet.write('X298', '=IF(OR(E298="-",ISBLANK(E298)),"",(L298-T298)/U298)')
    worksheet.write('X299', '=IF(OR(E299="-",ISBLANK(E299)),"",(L299-T299)/U299)')
    worksheet.write('X300', '=IF(OR(E300="-",ISBLANK(E300)),"",(L300-T300)/U300)')
    worksheet.write('X301', '=IF(OR(E301="-",ISBLANK(E301)),"",(L301-T301)/U301)')
    worksheet.write('X302', '=IF(OR(E302="-",ISBLANK(E302)),"",(L302-T302)/U302)')
    worksheet.write('X303', '=IF(OR(E303="-",ISBLANK(E303)),"",(L303-T303)/U303)')
    worksheet.write('X304', '=IF(OR(E304="-",ISBLANK(E304)),"",(L304-T304)/U304)')
    worksheet.write('X305', '=IF(OR(E305="-",ISBLANK(E305)),"",(L305-T305)/U305)')
    worksheet.write('X306', '=IF(OR(E306="-",ISBLANK(E306)),"",(L306-T306)/U306)')
    worksheet.write('X307', '=IF(OR(E307="-",ISBLANK(E307)),"",(L307-T307)/U307)')
    worksheet.write('X308', '=IF(OR(E308="-",ISBLANK(E308)),"",(L308-T308)/U308)')
    worksheet.write('X309', '=IF(OR(E309="-",ISBLANK(E309)),"",(L309-T309)/U309)')
    worksheet.write('X310', '=IF(OR(E310="-",ISBLANK(E310)),"",(L310-T310)/U310)')
    worksheet.write('X311', '=IF(OR(E311="-",ISBLANK(E311)),"",(L311-T311)/U311)')
    worksheet.write('X312', '=IF(OR(E312="-",ISBLANK(E312)),"",(L312-T312)/U312)')
    worksheet.write('X313', '=IF(OR(E313="-",ISBLANK(E313)),"",(L313-T313)/U313)')
    worksheet.write('X314', '=IF(OR(E314="-",ISBLANK(E314)),"",(L314-T314)/U314)')
    worksheet.write('X315', '=IF(OR(E315="-",ISBLANK(E315)),"",(L315-T315)/U315)')
    worksheet.write('X316', '=IF(OR(E316="-",ISBLANK(E316)),"",(L316-T316)/U316)')
    worksheet.write('X317', '=IF(OR(E317="-",ISBLANK(E317)),"",(L317-T317)/U317)')
    worksheet.write('X318', '=IF(OR(E318="-",ISBLANK(E318)),"",(L318-T318)/U318)')
    worksheet.write('X319', '=IF(OR(E319="-",ISBLANK(E319)),"",(L319-T319)/U319)')
    worksheet.write('X320', '=IF(OR(E320="-",ISBLANK(E320)),"",(L320-T320)/U320)')
    worksheet.write('X321', '=IF(OR(E321="-",ISBLANK(E321)),"",(L321-T321)/U321)')
    worksheet.write('X322', '=IF(OR(E322="-",ISBLANK(E322)),"",(L322-T322)/U322)')
    worksheet.write('X323', '=IF(OR(E323="-",ISBLANK(E323)),"",(L323-T323)/U323)')
    worksheet.write('X324', '=IF(OR(E324="-",ISBLANK(E324)),"",(L324-T324)/U324)')
    worksheet.write('X325', '=IF(OR(E325="-",ISBLANK(E325)),"",(L325-T325)/U325)')
    worksheet.write('X326', '=IF(OR(E326="-",ISBLANK(E326)),"",(L326-T326)/U326)')
    worksheet.write('X327', '=IF(OR(E327="-",ISBLANK(E327)),"",(L327-T327)/U327)')
    worksheet.write('X328', '=IF(OR(E328="-",ISBLANK(E328)),"",(L328-T328)/U328)')
    worksheet.write('X329', '=IF(OR(E329="-",ISBLANK(E329)),"",(L329-T329)/U329)')
    worksheet.write('X330', '=IF(OR(E330="-",ISBLANK(E330)),"",(L330-T330)/U330)')
    worksheet.write('X331', '=IF(OR(E331="-",ISBLANK(E331)),"",(L331-T331)/U331)')
    worksheet.write('X332', '=IF(OR(E332="-",ISBLANK(E332)),"",(L332-T332)/U332)')
    worksheet.write('X333', '=IF(OR(E333="-",ISBLANK(E333)),"",(L333-T333)/U333)')
    worksheet.write('X334', '=IF(OR(E334="-",ISBLANK(E334)),"",(L334-T334)/U334)')
    worksheet.write('X335', '=IF(OR(E335="-",ISBLANK(E335)),"",(L335-T335)/U335)')
    worksheet.write('X336', '=IF(OR(E336="-",ISBLANK(E336)),"",(L336-T336)/U336)')
    worksheet.write('X337', '=IF(OR(E337="-",ISBLANK(E337)),"",(L337-T337)/U337)')
    worksheet.write('X338', '=IF(OR(E338="-",ISBLANK(E338)),"",(L338-T338)/U338)')
    worksheet.write('X339', '=IF(OR(E339="-",ISBLANK(E339)),"",(L339-T339)/U339)')
    worksheet.write('X340', '=IF(OR(E340="-",ISBLANK(E340)),"",(L340-T340)/U340)')
    worksheet.write('X341', '=IF(OR(E341="-",ISBLANK(E341)),"",(L341-T341)/U341)')
    worksheet.write('X342', '=IF(OR(E342="-",ISBLANK(E342)),"",(L342-T342)/U342)')
    worksheet.write('X343', '=IF(OR(E343="-",ISBLANK(E343)),"",(L343-T343)/U343)')
    worksheet.write('X344', '=IF(OR(E344="-",ISBLANK(E344)),"",(L344-T344)/U344)')
    worksheet.write('X345', '=IF(OR(E345="-",ISBLANK(E345)),"",(L345-T345)/U345)')
    worksheet.write('X346', '=IF(OR(E346="-",ISBLANK(E346)),"",(L346-T346)/U346)')
    worksheet.write('X347', '=IF(OR(E347="-",ISBLANK(E347)),"",(L347-T347)/U347)')
    worksheet.write('X348', '=IF(OR(E348="-",ISBLANK(E348)),"",(L348-T348)/U348)')
    worksheet.write('X349', '=IF(OR(E349="-",ISBLANK(E349)),"",(L349-T349)/U349)')
    worksheet.write('X350', '=IF(OR(E350="-",ISBLANK(E350)),"",(L350-T350)/U350)')
    worksheet.write('X351', '=IF(OR(E351="-",ISBLANK(E351)),"",(L351-T351)/U351)')
    worksheet.write('X352', '=IF(OR(E352="-",ISBLANK(E352)),"",(L352-T352)/U352)')
    worksheet.write('X353', '=IF(OR(E353="-",ISBLANK(E353)),"",(L353-T353)/U353)')
    worksheet.write('Y144', '=M144')
    worksheet.write('Y145', '=IF(OR(F145="-",ISBLANK(F145)),"",(M145-T145)/U145)')
    worksheet.write('Y146', '=IF(OR(F146="-",ISBLANK(F146)),"",(M146-T146)/U146)')
    worksheet.write('Y147', '=IF(OR(F147="-",ISBLANK(F147)),"",(M147-T147)/U147)')
    worksheet.write('Y148', '=IF(OR(F148="-",ISBLANK(F148)),"",(M148-T148)/U148)')
    worksheet.write('Y149', '=IF(OR(F149="-",ISBLANK(F149)),"",(M149-T149)/U149)')
    worksheet.write('Y150', '=IF(OR(F150="-",ISBLANK(F150)),"",(M150-T150)/U150)')
    worksheet.write('Y151', '=IF(OR(F151="-",ISBLANK(F151)),"",(M151-T151)/U151)')
    worksheet.write('Y152', '=IF(OR(F152="-",ISBLANK(F152)),"",(M152-T152)/U152)')
    worksheet.write('Y153', '=IF(OR(F153="-",ISBLANK(F153)),"",(M153-T153)/U153)')
    worksheet.write('Y154', '=IF(OR(F154="-",ISBLANK(F154)),"",(M154-T154)/U154)')
    worksheet.write('Y155', '=IF(OR(F155="-",ISBLANK(F155)),"",(M155-T155)/U155)')
    worksheet.write('Y156', '=IF(OR(F156="-",ISBLANK(F156)),"",(M156-T156)/U156)')
    worksheet.write('Y157', '=IF(OR(F157="-",ISBLANK(F157)),"",(M157-T157)/U157)')
    worksheet.write('Y158', '=IF(OR(F158="-",ISBLANK(F158)),"",(M158-T158)/U158)')
    worksheet.write('Y159', '=IF(OR(F159="-",ISBLANK(F159)),"",(M159-T159)/U159)')
    worksheet.write('Y160', '=IF(OR(F160="-",ISBLANK(F160)),"",(M160-T160)/U160)')
    worksheet.write('Y161', '=IF(OR(F161="-",ISBLANK(F161)),"",(M161-T161)/U161)')
    worksheet.write('Y162', '=IF(OR(F162="-",ISBLANK(F162)),"",(M162-T162)/U162)')
    worksheet.write('Y163', '=IF(OR(F163="-",ISBLANK(F163)),"",(M163-T163)/U163)')
    worksheet.write('Y164', '=IF(OR(F164="-",ISBLANK(F164)),"",(M164-T164)/U164)')
    worksheet.write('Y165', '=IF(OR(F165="-",ISBLANK(F165)),"",(M165-T165)/U165)')
    worksheet.write('Y166', '=IF(OR(F166="-",ISBLANK(F166)),"",(M166-T166)/U166)')
    worksheet.write('Y167', '=IF(OR(F167="-",ISBLANK(F167)),"",(M167-T167)/U167)')
    worksheet.write('Y168', '=IF(OR(F168="-",ISBLANK(F168)),"",(M168-T168)/U168)')
    worksheet.write('Y169', '=IF(OR(F169="-",ISBLANK(F169)),"",(M169-T169)/U169)')
    worksheet.write('Y170', '=IF(OR(F170="-",ISBLANK(F170)),"",(M170-T170)/U170)')
    worksheet.write('Y171', '=IF(OR(F171="-",ISBLANK(F171)),"",(M171-T171)/U171)')
    worksheet.write('Y172', '=IF(OR(F172="-",ISBLANK(F172)),"",(M172-T172)/U172)')
    worksheet.write('Y173', '=IF(OR(F173="-",ISBLANK(F173)),"",(M173-T173)/U173)')
    worksheet.write('Y174', '=IF(OR(F174="-",ISBLANK(F174)),"",(M174-T174)/U174)')
    worksheet.write('Y175', '=IF(OR(F175="-",ISBLANK(F175)),"",(M175-T175)/U175)')
    worksheet.write('Y176', '=IF(OR(F176="-",ISBLANK(F176)),"",(M176-T176)/U176)')
    worksheet.write('Y177', '=IF(OR(F177="-",ISBLANK(F177)),"",(M177-T177)/U177)')
    worksheet.write('Y178', '=IF(OR(F178="-",ISBLANK(F178)),"",(M178-T178)/U178)')
    worksheet.write('Y179', '=IF(OR(F179="-",ISBLANK(F179)),"",(M179-T179)/U179)')
    worksheet.write('Y180', '=IF(OR(F180="-",ISBLANK(F180)),"",(M180-T180)/U180)')
    worksheet.write('Y181', '=IF(OR(F181="-",ISBLANK(F181)),"",(M181-T181)/U181)')
    worksheet.write('Y182', '=IF(OR(F182="-",ISBLANK(F182)),"",(M182-T182)/U182)')
    worksheet.write('Y183', '=IF(OR(F183="-",ISBLANK(F183)),"",(M183-T183)/U183)')
    worksheet.write('Y184', '=IF(OR(F184="-",ISBLANK(F184)),"",(M184-T184)/U184)')
    worksheet.write('Y185', '=IF(OR(F185="-",ISBLANK(F185)),"",(M185-T185)/U185)')
    worksheet.write('Y186', '=IF(OR(F186="-",ISBLANK(F186)),"",(M186-T186)/U186)')
    worksheet.write('Y187', '=IF(OR(F187="-",ISBLANK(F187)),"",(M187-T187)/U187)')
    worksheet.write('Y188', '=IF(OR(F188="-",ISBLANK(F188)),"",(M188-T188)/U188)')
    worksheet.write('Y189', '=IF(OR(F189="-",ISBLANK(F189)),"",(M189-T189)/U189)')
    worksheet.write('Y190', '=IF(OR(F190="-",ISBLANK(F190)),"",(M190-T190)/U190)')
    worksheet.write('Y191', '=IF(OR(F191="-",ISBLANK(F191)),"",(M191-T191)/U191)')
    worksheet.write('Y192', '=IF(OR(F192="-",ISBLANK(F192)),"",(M192-T192)/U192)')
    worksheet.write('Y193', '=IF(OR(F193="-",ISBLANK(F193)),"",(M193-T193)/U193)')
    worksheet.write('Y194', '=IF(OR(F194="-",ISBLANK(F194)),"",(M194-T194)/U194)')
    worksheet.write('Y195', '=IF(OR(F195="-",ISBLANK(F195)),"",(M195-T195)/U195)')
    worksheet.write('Y196', '=IF(OR(F196="-",ISBLANK(F196)),"",(M196-T196)/U196)')
    worksheet.write('Y197', '=IF(OR(F197="-",ISBLANK(F197)),"",(M197-T197)/U197)')
    worksheet.write('Y198', '=IF(OR(F198="-",ISBLANK(F198)),"",(M198-T198)/U198)')
    worksheet.write('Y199', '=IF(OR(F199="-",ISBLANK(F199)),"",(M199-T199)/U199)')
    worksheet.write('Y200', '=IF(OR(F200="-",ISBLANK(F200)),"",(M200-T200)/U200)')
    worksheet.write('Y201', '=IF(OR(F201="-",ISBLANK(F201)),"",(M201-T201)/U201)')
    worksheet.write('Y202', '=IF(OR(F202="-",ISBLANK(F202)),"",(M202-T202)/U202)')
    worksheet.write('Y203', '=IF(OR(F203="-",ISBLANK(F203)),"",(M203-T203)/U203)')
    worksheet.write('Y204', '=IF(OR(F204="-",ISBLANK(F204)),"",(M204-T204)/U204)')
    worksheet.write('Y205', '=IF(OR(F205="-",ISBLANK(F205)),"",(M205-T205)/U205)')
    worksheet.write('Y206', '=IF(OR(F206="-",ISBLANK(F206)),"",(M206-T206)/U206)')
    worksheet.write('Y207', '=IF(OR(F207="-",ISBLANK(F207)),"",(M207-T207)/U207)')
    worksheet.write('Y208', '=IF(OR(F208="-",ISBLANK(F208)),"",(M208-T208)/U208)')
    worksheet.write('Y209', '=IF(OR(F209="-",ISBLANK(F209)),"",(M209-T209)/U209)')
    worksheet.write('Y210', '=IF(OR(F210="-",ISBLANK(F210)),"",(M210-T210)/U210)')
    worksheet.write('Y211', '=IF(OR(F211="-",ISBLANK(F211)),"",(M211-T211)/U211)')
    worksheet.write('Y212', '=IF(OR(F212="-",ISBLANK(F212)),"",(M212-T212)/U212)')
    worksheet.write('Y213', '=IF(OR(F213="-",ISBLANK(F213)),"",(M213-T213)/U213)')
    worksheet.write('Y214', '=IF(OR(F214="-",ISBLANK(F214)),"",(M214-T214)/U214)')
    worksheet.write('Y215', '=IF(OR(F215="-",ISBLANK(F215)),"",(M215-T215)/U215)')
    worksheet.write('Y216', '=IF(OR(F216="-",ISBLANK(F216)),"",(M216-T216)/U216)')
    worksheet.write('Y217', '=IF(OR(F217="-",ISBLANK(F217)),"",(M217-T217)/U217)')
    worksheet.write('Y218', '=IF(OR(F218="-",ISBLANK(F218)),"",(M218-T218)/U218)')
    worksheet.write('Y219', '=IF(OR(F219="-",ISBLANK(F219)),"",(M219-T219)/U219)')
    worksheet.write('Y220', '=IF(OR(F220="-",ISBLANK(F220)),"",(M220-T220)/U220)')
    worksheet.write('Y221', '=IF(OR(F221="-",ISBLANK(F221)),"",(M221-T221)/U221)')
    worksheet.write('Y222', '=IF(OR(F222="-",ISBLANK(F222)),"",(M222-T222)/U222)')
    worksheet.write('Y223', '=IF(OR(F223="-",ISBLANK(F223)),"",(M223-T223)/U223)')
    worksheet.write('Y224', '=IF(OR(F224="-",ISBLANK(F224)),"",(M224-T224)/U224)')
    worksheet.write('Y225', '=IF(OR(F225="-",ISBLANK(F225)),"",(M225-T225)/U225)')
    worksheet.write('Y226', '=IF(OR(F226="-",ISBLANK(F226)),"",(M226-T226)/U226)')
    worksheet.write('Y227', '=IF(OR(F227="-",ISBLANK(F227)),"",(M227-T227)/U227)')
    worksheet.write('Y228', '=IF(OR(F228="-",ISBLANK(F228)),"",(M228-T228)/U228)')
    worksheet.write('Y229', '=IF(OR(F229="-",ISBLANK(F229)),"",(M229-T229)/U229)')
    worksheet.write('Y230', '=IF(OR(F230="-",ISBLANK(F230)),"",(M230-T230)/U230)')
    worksheet.write('Y231', '=IF(OR(F231="-",ISBLANK(F231)),"",(M231-T231)/U231)')
    worksheet.write('Y232', '=IF(OR(F232="-",ISBLANK(F232)),"",(M232-T232)/U232)')
    worksheet.write('Y233', '=IF(OR(F233="-",ISBLANK(F233)),"",(M233-T233)/U233)')
    worksheet.write('Y234', '=IF(OR(F234="-",ISBLANK(F234)),"",(M234-T234)/U234)')
    worksheet.write('Y235', '=IF(OR(F235="-",ISBLANK(F235)),"",(M235-T235)/U235)')
    worksheet.write('Y236', '=IF(OR(F236="-",ISBLANK(F236)),"",(M236-T236)/U236)')
    worksheet.write('Y237', '=IF(OR(F237="-",ISBLANK(F237)),"",(M237-T237)/U237)')
    worksheet.write('Y238', '=IF(OR(F238="-",ISBLANK(F238)),"",(M238-T238)/U238)')
    worksheet.write('Y239', '=IF(OR(F239="-",ISBLANK(F239)),"",(M239-T239)/U239)')
    worksheet.write('Y240', '=IF(OR(F240="-",ISBLANK(F240)),"",(M240-T240)/U240)')
    worksheet.write('Y241', '=IF(OR(F241="-",ISBLANK(F241)),"",(M241-T241)/U241)')
    worksheet.write('Y242', '=IF(OR(F242="-",ISBLANK(F242)),"",(M242-T242)/U242)')
    worksheet.write('Y243', '=IF(OR(F243="-",ISBLANK(F243)),"",(M243-T243)/U243)')
    worksheet.write('Y244', '=IF(OR(F244="-",ISBLANK(F244)),"",(M244-T244)/U244)')
    worksheet.write('Y245', '=IF(OR(F245="-",ISBLANK(F245)),"",(M245-T245)/U245)')
    worksheet.write('Y246', '=IF(OR(F246="-",ISBLANK(F246)),"",(M246-T246)/U246)')
    worksheet.write('Y247', '=IF(OR(F247="-",ISBLANK(F247)),"",(M247-T247)/U247)')
    worksheet.write('Y248', '=IF(OR(F248="-",ISBLANK(F248)),"",(M248-T248)/U248)')
    worksheet.write('Y249', '=IF(OR(F249="-",ISBLANK(F249)),"",(M249-T249)/U249)')
    worksheet.write('Y250', '=IF(OR(F250="-",ISBLANK(F250)),"",(M250-T250)/U250)')
    worksheet.write('Y251', '=IF(OR(F251="-",ISBLANK(F251)),"",(M251-T251)/U251)')
    worksheet.write('Y252', '=IF(OR(F252="-",ISBLANK(F252)),"",(M252-T252)/U252)')
    worksheet.write('Y253', '=IF(OR(F253="-",ISBLANK(F253)),"",(M253-T253)/U253)')
    worksheet.write('Y254', '=IF(OR(F254="-",ISBLANK(F254)),"",(M254-T254)/U254)')
    worksheet.write('Y255', '=IF(OR(F255="-",ISBLANK(F255)),"",(M255-T255)/U255)')
    worksheet.write('Y256', '=IF(OR(F256="-",ISBLANK(F256)),"",(M256-T256)/U256)')
    worksheet.write('Y257', '=IF(OR(F257="-",ISBLANK(F257)),"",(M257-T257)/U257)')
    worksheet.write('Y258', '=IF(OR(F258="-",ISBLANK(F258)),"",(M258-T258)/U258)')
    worksheet.write('Y259', '=IF(OR(F259="-",ISBLANK(F259)),"",(M259-T259)/U259)')
    worksheet.write('Y260', '=IF(OR(F260="-",ISBLANK(F260)),"",(M260-T260)/U260)')
    worksheet.write('Y261', '=IF(OR(F261="-",ISBLANK(F261)),"",(M261-T261)/U261)')
    worksheet.write('Y262', '=IF(OR(F262="-",ISBLANK(F262)),"",(M262-T262)/U262)')
    worksheet.write('Y263', '=IF(OR(F263="-",ISBLANK(F263)),"",(M263-T263)/U263)')
    worksheet.write('Y264', '=IF(OR(F264="-",ISBLANK(F264)),"",(M264-T264)/U264)')
    worksheet.write('Y265', '=IF(OR(F265="-",ISBLANK(F265)),"",(M265-T265)/U265)')
    worksheet.write('Y266', '=IF(OR(F266="-",ISBLANK(F266)),"",(M266-T266)/U266)')
    worksheet.write('Y267', '=IF(OR(F267="-",ISBLANK(F267)),"",(M267-T267)/U267)')
    worksheet.write('Y268', '=IF(OR(F268="-",ISBLANK(F268)),"",(M268-T268)/U268)')
    worksheet.write('Y269', '=IF(OR(F269="-",ISBLANK(F269)),"",(M269-T269)/U269)')
    worksheet.write('Y270', '=IF(OR(F270="-",ISBLANK(F270)),"",(M270-T270)/U270)')
    worksheet.write('Y271', '=IF(OR(F271="-",ISBLANK(F271)),"",(M271-T271)/U271)')
    worksheet.write('Y272', '=IF(OR(F272="-",ISBLANK(F272)),"",(M272-T272)/U272)')
    worksheet.write('Y273', '=IF(OR(F273="-",ISBLANK(F273)),"",(M273-T273)/U273)')
    worksheet.write('Y274', '=IF(OR(F274="-",ISBLANK(F274)),"",(M274-T274)/U274)')
    worksheet.write('Y275', '=IF(OR(F275="-",ISBLANK(F275)),"",(M275-T275)/U275)')
    worksheet.write('Y276', '=IF(OR(F276="-",ISBLANK(F276)),"",(M276-T276)/U276)')
    worksheet.write('Y277', '=IF(OR(F277="-",ISBLANK(F277)),"",(M277-T277)/U277)')
    worksheet.write('Y278', '=IF(OR(F278="-",ISBLANK(F278)),"",(M278-T278)/U278)')
    worksheet.write('Y279', '=IF(OR(F279="-",ISBLANK(F279)),"",(M279-T279)/U279)')
    worksheet.write('Y280', '=IF(OR(F280="-",ISBLANK(F280)),"",(M280-T280)/U280)')
    worksheet.write('Y281', '=IF(OR(F281="-",ISBLANK(F281)),"",(M281-T281)/U281)')
    worksheet.write('Y282', '=IF(OR(F282="-",ISBLANK(F282)),"",(M282-T282)/U282)')
    worksheet.write('Y283', '=IF(OR(F283="-",ISBLANK(F283)),"",(M283-T283)/U283)')
    worksheet.write('Y284', '=IF(OR(F284="-",ISBLANK(F284)),"",(M284-T284)/U284)')
    worksheet.write('Y285', '=IF(OR(F285="-",ISBLANK(F285)),"",(M285-T285)/U285)')
    worksheet.write('Y286', '=IF(OR(F286="-",ISBLANK(F286)),"",(M286-T286)/U286)')
    worksheet.write('Y287', '=IF(OR(F287="-",ISBLANK(F287)),"",(M287-T287)/U287)')
    worksheet.write('Y288', '=IF(OR(F288="-",ISBLANK(F288)),"",(M288-T288)/U288)')
    worksheet.write('Y289', '=IF(OR(F289="-",ISBLANK(F289)),"",(M289-T289)/U289)')
    worksheet.write('Y290', '=IF(OR(F290="-",ISBLANK(F290)),"",(M290-T290)/U290)')
    worksheet.write('Y291', '=IF(OR(F291="-",ISBLANK(F291)),"",(M291-T291)/U291)')
    worksheet.write('Y292', '=IF(OR(F292="-",ISBLANK(F292)),"",(M292-T292)/U292)')
    worksheet.write('Y293', '=IF(OR(F293="-",ISBLANK(F293)),"",(M293-T293)/U293)')
    worksheet.write('Y294', '=IF(OR(F294="-",ISBLANK(F294)),"",(M294-T294)/U294)')
    worksheet.write('Y295', '=IF(OR(F295="-",ISBLANK(F295)),"",(M295-T295)/U295)')
    worksheet.write('Y296', '=IF(OR(F296="-",ISBLANK(F296)),"",(M296-T296)/U296)')
    worksheet.write('Y297', '=IF(OR(F297="-",ISBLANK(F297)),"",(M297-T297)/U297)')
    worksheet.write('Y298', '=IF(OR(F298="-",ISBLANK(F298)),"",(M298-T298)/U298)')
    worksheet.write('Y299', '=IF(OR(F299="-",ISBLANK(F299)),"",(M299-T299)/U299)')
    worksheet.write('Y300', '=IF(OR(F300="-",ISBLANK(F300)),"",(M300-T300)/U300)')
    worksheet.write('Y301', '=IF(OR(F301="-",ISBLANK(F301)),"",(M301-T301)/U301)')
    worksheet.write('Y302', '=IF(OR(F302="-",ISBLANK(F302)),"",(M302-T302)/U302)')
    worksheet.write('Y303', '=IF(OR(F303="-",ISBLANK(F303)),"",(M303-T303)/U303)')
    worksheet.write('Y304', '=IF(OR(F304="-",ISBLANK(F304)),"",(M304-T304)/U304)')
    worksheet.write('Y305', '=IF(OR(F305="-",ISBLANK(F305)),"",(M305-T305)/U305)')
    worksheet.write('Y306', '=IF(OR(F306="-",ISBLANK(F306)),"",(M306-T306)/U306)')
    worksheet.write('Y307', '=IF(OR(F307="-",ISBLANK(F307)),"",(M307-T307)/U307)')
    worksheet.write('Y308', '=IF(OR(F308="-",ISBLANK(F308)),"",(M308-T308)/U308)')
    worksheet.write('Y309', '=IF(OR(F309="-",ISBLANK(F309)),"",(M309-T309)/U309)')
    worksheet.write('Y310', '=IF(OR(F310="-",ISBLANK(F310)),"",(M310-T310)/U310)')
    worksheet.write('Y311', '=IF(OR(F311="-",ISBLANK(F311)),"",(M311-T311)/U311)')
    worksheet.write('Y312', '=IF(OR(F312="-",ISBLANK(F312)),"",(M312-T312)/U312)')
    worksheet.write('Y313', '=IF(OR(F313="-",ISBLANK(F313)),"",(M313-T313)/U313)')
    worksheet.write('Y314', '=IF(OR(F314="-",ISBLANK(F314)),"",(M314-T314)/U314)')
    worksheet.write('Y315', '=IF(OR(F315="-",ISBLANK(F315)),"",(M315-T315)/U315)')
    worksheet.write('Y316', '=IF(OR(F316="-",ISBLANK(F316)),"",(M316-T316)/U316)')
    worksheet.write('Y317', '=IF(OR(F317="-",ISBLANK(F317)),"",(M317-T317)/U317)')
    worksheet.write('Y318', '=IF(OR(F318="-",ISBLANK(F318)),"",(M318-T318)/U318)')
    worksheet.write('Y319', '=IF(OR(F319="-",ISBLANK(F319)),"",(M319-T319)/U319)')
    worksheet.write('Y320', '=IF(OR(F320="-",ISBLANK(F320)),"",(M320-T320)/U320)')
    worksheet.write('Y321', '=IF(OR(F321="-",ISBLANK(F321)),"",(M321-T321)/U321)')
    worksheet.write('Y322', '=IF(OR(F322="-",ISBLANK(F322)),"",(M322-T322)/U322)')
    worksheet.write('Y323', '=IF(OR(F323="-",ISBLANK(F323)),"",(M323-T323)/U323)')
    worksheet.write('Y324', '=IF(OR(F324="-",ISBLANK(F324)),"",(M324-T324)/U324)')
    worksheet.write('Y325', '=IF(OR(F325="-",ISBLANK(F325)),"",(M325-T325)/U325)')
    worksheet.write('Y326', '=IF(OR(F326="-",ISBLANK(F326)),"",(M326-T326)/U326)')
    worksheet.write('Y327', '=IF(OR(F327="-",ISBLANK(F327)),"",(M327-T327)/U327)')
    worksheet.write('Y328', '=IF(OR(F328="-",ISBLANK(F328)),"",(M328-T328)/U328)')
    worksheet.write('Y329', '=IF(OR(F329="-",ISBLANK(F329)),"",(M329-T329)/U329)')
    worksheet.write('Y330', '=IF(OR(F330="-",ISBLANK(F330)),"",(M330-T330)/U330)')
    worksheet.write('Y331', '=IF(OR(F331="-",ISBLANK(F331)),"",(M331-T331)/U331)')
    worksheet.write('Y332', '=IF(OR(F332="-",ISBLANK(F332)),"",(M332-T332)/U332)')
    worksheet.write('Y333', '=IF(OR(F333="-",ISBLANK(F333)),"",(M333-T333)/U333)')
    worksheet.write('Y334', '=IF(OR(F334="-",ISBLANK(F334)),"",(M334-T334)/U334)')
    worksheet.write('Y335', '=IF(OR(F335="-",ISBLANK(F335)),"",(M335-T335)/U335)')
    worksheet.write('Y336', '=IF(OR(F336="-",ISBLANK(F336)),"",(M336-T336)/U336)')
    worksheet.write('Y337', '=IF(OR(F337="-",ISBLANK(F337)),"",(M337-T337)/U337)')
    worksheet.write('Y338', '=IF(OR(F338="-",ISBLANK(F338)),"",(M338-T338)/U338)')
    worksheet.write('Y339', '=IF(OR(F339="-",ISBLANK(F339)),"",(M339-T339)/U339)')
    worksheet.write('Y340', '=IF(OR(F340="-",ISBLANK(F340)),"",(M340-T340)/U340)')
    worksheet.write('Y341', '=IF(OR(F341="-",ISBLANK(F341)),"",(M341-T341)/U341)')
    worksheet.write('Y342', '=IF(OR(F342="-",ISBLANK(F342)),"",(M342-T342)/U342)')
    worksheet.write('Y343', '=IF(OR(F343="-",ISBLANK(F343)),"",(M343-T343)/U343)')
    worksheet.write('Y344', '=IF(OR(F344="-",ISBLANK(F344)),"",(M344-T344)/U344)')
    worksheet.write('Y345', '=IF(OR(F345="-",ISBLANK(F345)),"",(M345-T345)/U345)')
    worksheet.write('Y346', '=IF(OR(F346="-",ISBLANK(F346)),"",(M346-T346)/U346)')
    worksheet.write('Y347', '=IF(OR(F347="-",ISBLANK(F347)),"",(M347-T347)/U347)')
    worksheet.write('Y348', '=IF(OR(F348="-",ISBLANK(F348)),"",(M348-T348)/U348)')
    worksheet.write('Y349', '=IF(OR(F349="-",ISBLANK(F349)),"",(M349-T349)/U349)')
    worksheet.write('Y350', '=IF(OR(F350="-",ISBLANK(F350)),"",(M350-T350)/U350)')
    worksheet.write('Y351', '=IF(OR(F351="-",ISBLANK(F351)),"",(M351-T351)/U351)')
    worksheet.write('Y352', '=IF(OR(F352="-",ISBLANK(F352)),"",(M352-T352)/U352)')
    worksheet.write('Y353', '=IF(OR(F353="-",ISBLANK(F353)),"",(M353-T353)/U353)')
    worksheet.write('Z144', '=N144')
    worksheet.write('Z145', '=IFERROR(IF(OR(G145="-",ISBLANK(G145)),"",(N145-T145)/U145),"")')
    worksheet.write('Z146', '=IFERROR(IF(OR(G146="-",ISBLANK(G146)),"",(N146-T146)/U146),"")')
    worksheet.write('Z147', '=IFERROR(IF(OR(G147="-",ISBLANK(G147)),"",(N147-T147)/U147),"")')
    worksheet.write('Z148', '=IFERROR(IF(OR(G148="-",ISBLANK(G148)),"",(N148-T148)/U148),"")')
    worksheet.write('Z149', '=IFERROR(IF(OR(G149="-",ISBLANK(G149)),"",(N149-T149)/U149),"")')
    worksheet.write('Z150', '=IFERROR(IF(OR(G150="-",ISBLANK(G150)),"",(N150-T150)/U150),"")')
    worksheet.write('Z151', '=IFERROR(IF(OR(G151="-",ISBLANK(G151)),"",(N151-T151)/U151),"")')
    worksheet.write('Z152', '=IFERROR(IF(OR(G152="-",ISBLANK(G152)),"",(N152-T152)/U152),"")')
    worksheet.write('Z153', '=IFERROR(IF(OR(G153="-",ISBLANK(G153)),"",(N153-T153)/U153),"")')
    worksheet.write('Z154', '=IFERROR(IF(OR(G154="-",ISBLANK(G154)),"",(N154-T154)/U154),"")')
    worksheet.write('Z155', '=IFERROR(IF(OR(G155="-",ISBLANK(G155)),"",(N155-T155)/U155),"")')
    worksheet.write('Z156', '=IFERROR(IF(OR(G156="-",ISBLANK(G156)),"",(N156-T156)/U156),"")')
    worksheet.write('Z157', '=IFERROR(IF(OR(G157="-",ISBLANK(G157)),"",(N157-T157)/U157),"")')
    worksheet.write('Z158', '=IFERROR(IF(OR(G158="-",ISBLANK(G158)),"",(N158-T158)/U158),"")')
    worksheet.write('Z159', '=IFERROR(IF(OR(G159="-",ISBLANK(G159)),"",(N159-T159)/U159),"")')
    worksheet.write('Z160', '=IFERROR(IF(OR(G160="-",ISBLANK(G160)),"",(N160-T160)/U160),"")')
    worksheet.write('Z161', '=IFERROR(IF(OR(G161="-",ISBLANK(G161)),"",(N161-T161)/U161),"")')
    worksheet.write('Z162', '=IFERROR(IF(OR(G162="-",ISBLANK(G162)),"",(N162-T162)/U162),"")')
    worksheet.write('Z163', '=IFERROR(IF(OR(G163="-",ISBLANK(G163)),"",(N163-T163)/U163),"")')
    worksheet.write('Z164', '=IFERROR(IF(OR(G164="-",ISBLANK(G164)),"",(N164-T164)/U164),"")')
    worksheet.write('Z165', '=IFERROR(IF(OR(G165="-",ISBLANK(G165)),"",(N165-T165)/U165),"")')
    worksheet.write('Z166', '=IFERROR(IF(OR(G166="-",ISBLANK(G166)),"",(N166-T166)/U166),"")')
    worksheet.write('Z167', '=IFERROR(IF(OR(G167="-",ISBLANK(G167)),"",(N167-T167)/U167),"")')
    worksheet.write('Z168', '=IFERROR(IF(OR(G168="-",ISBLANK(G168)),"",(N168-T168)/U168),"")')
    worksheet.write('Z169', '=IFERROR(IF(OR(G169="-",ISBLANK(G169)),"",(N169-T169)/U169),"")')
    worksheet.write('Z170', '=IFERROR(IF(OR(G170="-",ISBLANK(G170)),"",(N170-T170)/U170),"")')
    worksheet.write('Z171', '=IFERROR(IF(OR(G171="-",ISBLANK(G171)),"",(N171-T171)/U171),"")')
    worksheet.write('Z172', '=IFERROR(IF(OR(G172="-",ISBLANK(G172)),"",(N172-T172)/U172),"")')
    worksheet.write('Z173', '=IFERROR(IF(OR(G173="-",ISBLANK(G173)),"",(N173-T173)/U173),"")')
    worksheet.write('Z174', '=IFERROR(IF(OR(G174="-",ISBLANK(G174)),"",(N174-T174)/U174),"")')
    worksheet.write('Z175', '=IFERROR(IF(OR(G175="-",ISBLANK(G175)),"",(N175-T175)/U175),"")')
    worksheet.write('Z176', '=IFERROR(IF(OR(G176="-",ISBLANK(G176)),"",(N176-T176)/U176),"")')
    worksheet.write('Z177', '=IFERROR(IF(OR(G177="-",ISBLANK(G177)),"",(N177-T177)/U177),"")')
    worksheet.write('Z178', '=IFERROR(IF(OR(G178="-",ISBLANK(G178)),"",(N178-T178)/U178),"")')
    worksheet.write('Z179', '=IFERROR(IF(OR(G179="-",ISBLANK(G179)),"",(N179-T179)/U179),"")')
    worksheet.write('Z180', '=IFERROR(IF(OR(G180="-",ISBLANK(G180)),"",(N180-T180)/U180),"")')
    worksheet.write('Z181', '=IFERROR(IF(OR(G181="-",ISBLANK(G181)),"",(N181-T181)/U181),"")')
    worksheet.write('Z182', '=IFERROR(IF(OR(G182="-",ISBLANK(G182)),"",(N182-T182)/U182),"")')
    worksheet.write('Z183', '=IFERROR(IF(OR(G183="-",ISBLANK(G183)),"",(N183-T183)/U183),"")')
    worksheet.write('Z184', '=IFERROR(IF(OR(G184="-",ISBLANK(G184)),"",(N184-T184)/U184),"")')
    worksheet.write('Z185', '=IFERROR(IF(OR(G185="-",ISBLANK(G185)),"",(N185-T185)/U185),"")')
    worksheet.write('Z186', '=IFERROR(IF(OR(G186="-",ISBLANK(G186)),"",(N186-T186)/U186),"")')
    worksheet.write('Z187', '=IFERROR(IF(OR(G187="-",ISBLANK(G187)),"",(N187-T187)/U187),"")')
    worksheet.write('Z188', '=IFERROR(IF(OR(G188="-",ISBLANK(G188)),"",(N188-T188)/U188),"")')
    worksheet.write('Z189', '=IFERROR(IF(OR(G189="-",ISBLANK(G189)),"",(N189-T189)/U189),"")')
    worksheet.write('Z190', '=IFERROR(IF(OR(G190="-",ISBLANK(G190)),"",(N190-T190)/U190),"")')
    worksheet.write('Z191', '=IFERROR(IF(OR(G191="-",ISBLANK(G191)),"",(N191-T191)/U191),"")')
    worksheet.write('Z192', '=IFERROR(IF(OR(G192="-",ISBLANK(G192)),"",(N192-T192)/U192),"")')
    worksheet.write('Z193', '=IFERROR(IF(OR(G193="-",ISBLANK(G193)),"",(N193-T193)/U193),"")')
    worksheet.write('Z194', '=IFERROR(IF(OR(G194="-",ISBLANK(G194)),"",(N194-T194)/U194),"")')
    worksheet.write('Z195', '=IFERROR(IF(OR(G195="-",ISBLANK(G195)),"",(N195-T195)/U195),"")')
    worksheet.write('Z196', '=IFERROR(IF(OR(G196="-",ISBLANK(G196)),"",(N196-T196)/U196),"")')
    worksheet.write('Z197', '=IFERROR(IF(OR(G197="-",ISBLANK(G197)),"",(N197-T197)/U197),"")')
    worksheet.write('Z198', '=IFERROR(IF(OR(G198="-",ISBLANK(G198)),"",(N198-T198)/U198),"")')
    worksheet.write('Z199', '=IFERROR(IF(OR(G199="-",ISBLANK(G199)),"",(N199-T199)/U199),"")')
    worksheet.write('Z200', '=IFERROR(IF(OR(G200="-",ISBLANK(G200)),"",(N200-T200)/U200),"")')
    worksheet.write('Z201', '=IFERROR(IF(OR(G201="-",ISBLANK(G201)),"",(N201-T201)/U201),"")')
    worksheet.write('Z202', '=IFERROR(IF(OR(G202="-",ISBLANK(G202)),"",(N202-T202)/U202),"")')
    worksheet.write('Z203', '=IFERROR(IF(OR(G203="-",ISBLANK(G203)),"",(N203-T203)/U203),"")')
    worksheet.write('Z204', '=IFERROR(IF(OR(G204="-",ISBLANK(G204)),"",(N204-T204)/U204),"")')
    worksheet.write('Z205', '=IFERROR(IF(OR(G205="-",ISBLANK(G205)),"",(N205-T205)/U205),"")')
    worksheet.write('Z206', '=IFERROR(IF(OR(G206="-",ISBLANK(G206)),"",(N206-T206)/U206),"")')
    worksheet.write('Z207', '=IFERROR(IF(OR(G207="-",ISBLANK(G207)),"",(N207-T207)/U207),"")')
    worksheet.write('Z208', '=IFERROR(IF(OR(G208="-",ISBLANK(G208)),"",(N208-T208)/U208),"")')
    worksheet.write('Z209', '=IFERROR(IF(OR(G209="-",ISBLANK(G209)),"",(N209-T209)/U209),"")')
    worksheet.write('Z210', '=IFERROR(IF(OR(G210="-",ISBLANK(G210)),"",(N210-T210)/U210),"")')
    worksheet.write('Z211', '=IFERROR(IF(OR(G211="-",ISBLANK(G211)),"",(N211-T211)/U211),"")')
    worksheet.write('Z212', '=IFERROR(IF(OR(G212="-",ISBLANK(G212)),"",(N212-T212)/U212),"")')
    worksheet.write('Z213', '=IFERROR(IF(OR(G213="-",ISBLANK(G213)),"",(N213-T213)/U213),"")')
    worksheet.write('Z214', '=IFERROR(IF(OR(G214="-",ISBLANK(G214)),"",(N214-T214)/U214),"")')
    worksheet.write('Z215', '=IFERROR(IF(OR(G215="-",ISBLANK(G215)),"",(N215-T215)/U215),"")')
    worksheet.write('Z216', '=IFERROR(IF(OR(G216="-",ISBLANK(G216)),"",(N216-T216)/U216),"")')
    worksheet.write('Z217', '=IFERROR(IF(OR(G217="-",ISBLANK(G217)),"",(N217-T217)/U217),"")')
    worksheet.write('Z218', '=IFERROR(IF(OR(G218="-",ISBLANK(G218)),"",(N218-T218)/U218),"")')
    worksheet.write('Z219', '=IFERROR(IF(OR(G219="-",ISBLANK(G219)),"",(N219-T219)/U219),"")')
    worksheet.write('Z220', '=IFERROR(IF(OR(G220="-",ISBLANK(G220)),"",(N220-T220)/U220),"")')
    worksheet.write('Z221', '=IFERROR(IF(OR(G221="-",ISBLANK(G221)),"",(N221-T221)/U221),"")')
    worksheet.write('Z222', '=IFERROR(IF(OR(G222="-",ISBLANK(G222)),"",(N222-T222)/U222),"")')
    worksheet.write('Z223', '=IFERROR(IF(OR(G223="-",ISBLANK(G223)),"",(N223-T223)/U223),"")')
    worksheet.write('Z224', '=IFERROR(IF(OR(G224="-",ISBLANK(G224)),"",(N224-T224)/U224),"")')
    worksheet.write('Z225', '=IFERROR(IF(OR(G225="-",ISBLANK(G225)),"",(N225-T225)/U225),"")')
    worksheet.write('Z226', '=IFERROR(IF(OR(G226="-",ISBLANK(G226)),"",(N226-T226)/U226),"")')
    worksheet.write('Z227', '=IFERROR(IF(OR(G227="-",ISBLANK(G227)),"",(N227-T227)/U227),"")')
    worksheet.write('Z228', '=IFERROR(IF(OR(G228="-",ISBLANK(G228)),"",(N228-T228)/U228),"")')
    worksheet.write('Z229', '=IFERROR(IF(OR(G229="-",ISBLANK(G229)),"",(N229-T229)/U229),"")')
    worksheet.write('Z230', '=IFERROR(IF(OR(G230="-",ISBLANK(G230)),"",(N230-T230)/U230),"")')
    worksheet.write('Z231', '=IFERROR(IF(OR(G231="-",ISBLANK(G231)),"",(N231-T231)/U231),"")')
    worksheet.write('Z232', '=IFERROR(IF(OR(G232="-",ISBLANK(G232)),"",(N232-T232)/U232),"")')
    worksheet.write('Z233', '=IFERROR(IF(OR(G233="-",ISBLANK(G233)),"",(N233-T233)/U233),"")')
    worksheet.write('Z234', '=IFERROR(IF(OR(G234="-",ISBLANK(G234)),"",(N234-T234)/U234),"")')
    worksheet.write('Z235', '=IFERROR(IF(OR(G235="-",ISBLANK(G235)),"",(N235-T235)/U235),"")')
    worksheet.write('Z236', '=IFERROR(IF(OR(G236="-",ISBLANK(G236)),"",(N236-T236)/U236),"")')
    worksheet.write('Z237', '=IFERROR(IF(OR(G237="-",ISBLANK(G237)),"",(N237-T237)/U237),"")')
    worksheet.write('Z238', '=IFERROR(IF(OR(G238="-",ISBLANK(G238)),"",(N238-T238)/U238),"")')
    worksheet.write('Z239', '=IFERROR(IF(OR(G239="-",ISBLANK(G239)),"",(N239-T239)/U239),"")')
    worksheet.write('Z240', '=IFERROR(IF(OR(G240="-",ISBLANK(G240)),"",(N240-T240)/U240),"")')
    worksheet.write('Z241', '=IFERROR(IF(OR(G241="-",ISBLANK(G241)),"",(N241-T241)/U241),"")')
    worksheet.write('Z242', '=IFERROR(IF(OR(G242="-",ISBLANK(G242)),"",(N242-T242)/U242),"")')
    worksheet.write('Z243', '=IFERROR(IF(OR(G243="-",ISBLANK(G243)),"",(N243-T243)/U243),"")')
    worksheet.write('Z244', '=IFERROR(IF(OR(G244="-",ISBLANK(G244)),"",(N244-T244)/U244),"")')
    worksheet.write('Z245', '=IFERROR(IF(OR(G245="-",ISBLANK(G245)),"",(N245-T245)/U245),"")')
    worksheet.write('Z246', '=IFERROR(IF(OR(G246="-",ISBLANK(G246)),"",(N246-T246)/U246),"")')
    worksheet.write('Z247', '=IFERROR(IF(OR(G247="-",ISBLANK(G247)),"",(N247-T247)/U247),"")')
    worksheet.write('Z248', '=IFERROR(IF(OR(G248="-",ISBLANK(G248)),"",(N248-T248)/U248),"")')
    worksheet.write('Z249', '=IFERROR(IF(OR(G249="-",ISBLANK(G249)),"",(N249-T249)/U249),"")')
    worksheet.write('Z250', '=IFERROR(IF(OR(G250="-",ISBLANK(G250)),"",(N250-T250)/U250),"")')
    worksheet.write('Z251', '=IFERROR(IF(OR(G251="-",ISBLANK(G251)),"",(N251-T251)/U251),"")')
    worksheet.write('Z252', '=IFERROR(IF(OR(G252="-",ISBLANK(G252)),"",(N252-T252)/U252),"")')
    worksheet.write('Z253', '=IFERROR(IF(OR(G253="-",ISBLANK(G253)),"",(N253-T253)/U253),"")')
    worksheet.write('Z254', '=IFERROR(IF(OR(G254="-",ISBLANK(G254)),"",(N254-T254)/U254),"")')
    worksheet.write('Z255', '=IFERROR(IF(OR(G255="-",ISBLANK(G255)),"",(N255-T255)/U255),"")')
    worksheet.write('Z256', '=IFERROR(IF(OR(G256="-",ISBLANK(G256)),"",(N256-T256)/U256),"")')
    worksheet.write('Z257', '=IFERROR(IF(OR(G257="-",ISBLANK(G257)),"",(N257-T257)/U257),"")')
    worksheet.write('Z258', '=IFERROR(IF(OR(G258="-",ISBLANK(G258)),"",(N258-T258)/U258),"")')
    worksheet.write('Z259', '=IFERROR(IF(OR(G259="-",ISBLANK(G259)),"",(N259-T259)/U259),"")')
    worksheet.write('Z260', '=IFERROR(IF(OR(G260="-",ISBLANK(G260)),"",(N260-T260)/U260),"")')
    worksheet.write('Z261', '=IFERROR(IF(OR(G261="-",ISBLANK(G261)),"",(N261-T261)/U261),"")')
    worksheet.write('Z262', '=IFERROR(IF(OR(G262="-",ISBLANK(G262)),"",(N262-T262)/U262),"")')
    worksheet.write('Z263', '=IFERROR(IF(OR(G263="-",ISBLANK(G263)),"",(N263-T263)/U263),"")')
    worksheet.write('Z264', '=IFERROR(IF(OR(G264="-",ISBLANK(G264)),"",(N264-T264)/U264),"")')
    worksheet.write('Z265', '=IFERROR(IF(OR(G265="-",ISBLANK(G265)),"",(N265-T265)/U265),"")')
    worksheet.write('Z266', '=IFERROR(IF(OR(G266="-",ISBLANK(G266)),"",(N266-T266)/U266),"")')
    worksheet.write('Z267', '=IFERROR(IF(OR(G267="-",ISBLANK(G267)),"",(N267-T267)/U267),"")')
    worksheet.write('Z268', '=IFERROR(IF(OR(G268="-",ISBLANK(G268)),"",(N268-T268)/U268),"")')
    worksheet.write('Z269', '=IFERROR(IF(OR(G269="-",ISBLANK(G269)),"",(N269-T269)/U269),"")')
    worksheet.write('Z270', '=IFERROR(IF(OR(G270="-",ISBLANK(G270)),"",(N270-T270)/U270),"")')
    worksheet.write('Z271', '=IFERROR(IF(OR(G271="-",ISBLANK(G271)),"",(N271-T271)/U271),"")')
    worksheet.write('Z272', '=IFERROR(IF(OR(G272="-",ISBLANK(G272)),"",(N272-T272)/U272),"")')
    worksheet.write('Z273', '=IFERROR(IF(OR(G273="-",ISBLANK(G273)),"",(N273-T273)/U273),"")')
    worksheet.write('Z274', '=IFERROR(IF(OR(G274="-",ISBLANK(G274)),"",(N274-T274)/U274),"")')
    worksheet.write('Z275', '=IFERROR(IF(OR(G275="-",ISBLANK(G275)),"",(N275-T275)/U275),"")')
    worksheet.write('Z276', '=IFERROR(IF(OR(G276="-",ISBLANK(G276)),"",(N276-T276)/U276),"")')
    worksheet.write('Z277', '=IFERROR(IF(OR(G277="-",ISBLANK(G277)),"",(N277-T277)/U277),"")')
    worksheet.write('Z278', '=IFERROR(IF(OR(G278="-",ISBLANK(G278)),"",(N278-T278)/U278),"")')
    worksheet.write('Z279', '=IFERROR(IF(OR(G279="-",ISBLANK(G279)),"",(N279-T279)/U279),"")')
    worksheet.write('Z280', '=IFERROR(IF(OR(G280="-",ISBLANK(G280)),"",(N280-T280)/U280),"")')
    worksheet.write('Z281', '=IFERROR(IF(OR(G281="-",ISBLANK(G281)),"",(N281-T281)/U281),"")')
    worksheet.write('Z282', '=IFERROR(IF(OR(G282="-",ISBLANK(G282)),"",(N282-T282)/U282),"")')
    worksheet.write('Z283', '=IFERROR(IF(OR(G283="-",ISBLANK(G283)),"",(N283-T283)/U283),"")')
    worksheet.write('Z284', '=IFERROR(IF(OR(G284="-",ISBLANK(G284)),"",(N284-T284)/U284),"")')
    worksheet.write('Z285', '=IFERROR(IF(OR(G285="-",ISBLANK(G285)),"",(N285-T285)/U285),"")')
    worksheet.write('Z286', '=IFERROR(IF(OR(G286="-",ISBLANK(G286)),"",(N286-T286)/U286),"")')
    worksheet.write('Z287', '=IFERROR(IF(OR(G287="-",ISBLANK(G287)),"",(N287-T287)/U287),"")')
    worksheet.write('Z288', '=IFERROR(IF(OR(G288="-",ISBLANK(G288)),"",(N288-T288)/U288),"")')
    worksheet.write('Z289', '=IFERROR(IF(OR(G289="-",ISBLANK(G289)),"",(N289-T289)/U289),"")')
    worksheet.write('Z290', '=IFERROR(IF(OR(G290="-",ISBLANK(G290)),"",(N290-T290)/U290),"")')
    worksheet.write('Z291', '=IFERROR(IF(OR(G291="-",ISBLANK(G291)),"",(N291-T291)/U291),"")')
    worksheet.write('Z292', '=IFERROR(IF(OR(G292="-",ISBLANK(G292)),"",(N292-T292)/U292),"")')
    worksheet.write('Z293', '=IFERROR(IF(OR(G293="-",ISBLANK(G293)),"",(N293-T293)/U293),"")')
    worksheet.write('Z294', '=IFERROR(IF(OR(G294="-",ISBLANK(G294)),"",(N294-T294)/U294),"")')
    worksheet.write('Z295', '=IFERROR(IF(OR(G295="-",ISBLANK(G295)),"",(N295-T295)/U295),"")')
    worksheet.write('Z296', '=IFERROR(IF(OR(G296="-",ISBLANK(G296)),"",(N296-T296)/U296),"")')
    worksheet.write('Z297', '=IFERROR(IF(OR(G297="-",ISBLANK(G297)),"",(N297-T297)/U297),"")')
    worksheet.write('Z298', '=IFERROR(IF(OR(G298="-",ISBLANK(G298)),"",(N298-T298)/U298),"")')
    worksheet.write('Z299', '=IFERROR(IF(OR(G299="-",ISBLANK(G299)),"",(N299-T299)/U299),"")')
    worksheet.write('Z300', '=IFERROR(IF(OR(G300="-",ISBLANK(G300)),"",(N300-T300)/U300),"")')
    worksheet.write('Z301', '=IFERROR(IF(OR(G301="-",ISBLANK(G301)),"",(N301-T301)/U301),"")')
    worksheet.write('Z302', '=IFERROR(IF(OR(G302="-",ISBLANK(G302)),"",(N302-T302)/U302),"")')
    worksheet.write('Z303', '=IFERROR(IF(OR(G303="-",ISBLANK(G303)),"",(N303-T303)/U303),"")')
    worksheet.write('Z304', '=IFERROR(IF(OR(G304="-",ISBLANK(G304)),"",(N304-T304)/U304),"")')
    worksheet.write('Z305', '=IFERROR(IF(OR(G305="-",ISBLANK(G305)),"",(N305-T305)/U305),"")')
    worksheet.write('Z306', '=IFERROR(IF(OR(G306="-",ISBLANK(G306)),"",(N306-T306)/U306),"")')
    worksheet.write('Z307', '=IFERROR(IF(OR(G307="-",ISBLANK(G307)),"",(N307-T307)/U307),"")')
    worksheet.write('Z308', '=IFERROR(IF(OR(G308="-",ISBLANK(G308)),"",(N308-T308)/U308),"")')
    worksheet.write('Z309', '=IFERROR(IF(OR(G309="-",ISBLANK(G309)),"",(N309-T309)/U309),"")')
    worksheet.write('Z310', '=IFERROR(IF(OR(G310="-",ISBLANK(G310)),"",(N310-T310)/U310),"")')
    worksheet.write('Z311', '=IFERROR(IF(OR(G311="-",ISBLANK(G311)),"",(N311-T311)/U311),"")')
    worksheet.write('Z312', '=IFERROR(IF(OR(G312="-",ISBLANK(G312)),"",(N312-T312)/U312),"")')
    worksheet.write('Z313', '=IFERROR(IF(OR(G313="-",ISBLANK(G313)),"",(N313-T313)/U313),"")')
    worksheet.write('Z314', '=IFERROR(IF(OR(G314="-",ISBLANK(G314)),"",(N314-T314)/U314),"")')
    worksheet.write('Z315', '=IFERROR(IF(OR(G315="-",ISBLANK(G315)),"",(N315-T315)/U315),"")')
    worksheet.write('Z316', '=IFERROR(IF(OR(G316="-",ISBLANK(G316)),"",(N316-T316)/U316),"")')
    worksheet.write('Z317', '=IFERROR(IF(OR(G317="-",ISBLANK(G317)),"",(N317-T317)/U317),"")')
    worksheet.write('Z318', '=IFERROR(IF(OR(G318="-",ISBLANK(G318)),"",(N318-T318)/U318),"")')
    worksheet.write('Z319', '=IFERROR(IF(OR(G319="-",ISBLANK(G319)),"",(N319-T319)/U319),"")')
    worksheet.write('Z320', '=IFERROR(IF(OR(G320="-",ISBLANK(G320)),"",(N320-T320)/U320),"")')
    worksheet.write('Z321', '=IFERROR(IF(OR(G321="-",ISBLANK(G321)),"",(N321-T321)/U321),"")')
    worksheet.write('Z322', '=IFERROR(IF(OR(G322="-",ISBLANK(G322)),"",(N322-T322)/U322),"")')
    worksheet.write('Z323', '=IFERROR(IF(OR(G323="-",ISBLANK(G323)),"",(N323-T323)/U323),"")')
    worksheet.write('Z324', '=IFERROR(IF(OR(G324="-",ISBLANK(G324)),"",(N324-T324)/U324),"")')
    worksheet.write('Z325', '=IFERROR(IF(OR(G325="-",ISBLANK(G325)),"",(N325-T325)/U325),"")')
    worksheet.write('Z326', '=IFERROR(IF(OR(G326="-",ISBLANK(G326)),"",(N326-T326)/U326),"")')
    worksheet.write('Z327', '=IFERROR(IF(OR(G327="-",ISBLANK(G327)),"",(N327-T327)/U327),"")')
    worksheet.write('Z328', '=IFERROR(IF(OR(G328="-",ISBLANK(G328)),"",(N328-T328)/U328),"")')
    worksheet.write('Z329', '=IFERROR(IF(OR(G329="-",ISBLANK(G329)),"",(N329-T329)/U329),"")')
    worksheet.write('Z330', '=IFERROR(IF(OR(G330="-",ISBLANK(G330)),"",(N330-T330)/U330),"")')
    worksheet.write('Z331', '=IFERROR(IF(OR(G331="-",ISBLANK(G331)),"",(N331-T331)/U331),"")')
    worksheet.write('Z332', '=IFERROR(IF(OR(G332="-",ISBLANK(G332)),"",(N332-T332)/U332),"")')
    worksheet.write('Z333', '=IFERROR(IF(OR(G333="-",ISBLANK(G333)),"",(N333-T333)/U333),"")')
    worksheet.write('Z334', '=IFERROR(IF(OR(G334="-",ISBLANK(G334)),"",(N334-T334)/U334),"")')
    worksheet.write('Z335', '=IFERROR(IF(OR(G335="-",ISBLANK(G335)),"",(N335-T335)/U335),"")')
    worksheet.write('Z336', '=IFERROR(IF(OR(G336="-",ISBLANK(G336)),"",(N336-T336)/U336),"")')
    worksheet.write('Z337', '=IFERROR(IF(OR(G337="-",ISBLANK(G337)),"",(N337-T337)/U337),"")')
    worksheet.write('Z338', '=IFERROR(IF(OR(G338="-",ISBLANK(G338)),"",(N338-T338)/U338),"")')
    worksheet.write('Z339', '=IFERROR(IF(OR(G339="-",ISBLANK(G339)),"",(N339-T339)/U339),"")')
    worksheet.write('Z340', '=IFERROR(IF(OR(G340="-",ISBLANK(G340)),"",(N340-T340)/U340),"")')
    worksheet.write('Z341', '=IFERROR(IF(OR(G341="-",ISBLANK(G341)),"",(N341-T341)/U341),"")')
    worksheet.write('Z342', '=IFERROR(IF(OR(G342="-",ISBLANK(G342)),"",(N342-T342)/U342),"")')
    worksheet.write('Z343', '=IFERROR(IF(OR(G343="-",ISBLANK(G343)),"",(N343-T343)/U343),"")')
    worksheet.write('Z344', '=IFERROR(IF(OR(G344="-",ISBLANK(G344)),"",(N344-T344)/U344),"")')
    worksheet.write('Z345', '=IFERROR(IF(OR(G345="-",ISBLANK(G345)),"",(N345-T345)/U345),"")')
    worksheet.write('Z346', '=IFERROR(IF(OR(G346="-",ISBLANK(G346)),"",(N346-T346)/U346),"")')
    worksheet.write('Z347', '=IFERROR(IF(OR(G347="-",ISBLANK(G347)),"",(N347-T347)/U347),"")')
    worksheet.write('Z348', '=IFERROR(IF(OR(G348="-",ISBLANK(G348)),"",(N348-T348)/U348),"")')
    worksheet.write('Z349', '=IFERROR(IF(OR(G349="-",ISBLANK(G349)),"",(N349-T349)/U349),"")')
    worksheet.write('Z350', '=IFERROR(IF(OR(G350="-",ISBLANK(G350)),"",(N350-T350)/U350),"")')
    worksheet.write('Z351', '=IFERROR(IF(OR(G351="-",ISBLANK(G351)),"",(N351-T351)/U351),"")')
    worksheet.write('Z352', '=IFERROR(IF(OR(G352="-",ISBLANK(G352)),"",(N352-T352)/U352),"")')
    worksheet.write('Z353', '=IFERROR(IF(OR(G353="-",ISBLANK(G353)),"",(N353-T353)/U353),"")')
    worksheet.write('AA144', '="Max z"')
    worksheet.write('AA145',
                    '=IF(MAX(MAX(V145:Z145),ABS(MIN(V145:Z145)))=ABS(MIN(V145:Z145)),MIN(V145:Z145),MAX(V145:Z145))')
    worksheet.write('AA146',
                    '=IF(MAX(MAX(V146:Z146),ABS(MIN(V146:Z146)))=ABS(MIN(V146:Z146)),MIN(V146:Z146),MAX(V146:Z146))')
    worksheet.write('AA147',
                    '=IF(MAX(MAX(V147:Z147),ABS(MIN(V147:Z147)))=ABS(MIN(V147:Z147)),MIN(V147:Z147),MAX(V147:Z147))')
    worksheet.write('AA148',
                    '=IF(MAX(MAX(V148:Z148),ABS(MIN(V148:Z148)))=ABS(MIN(V148:Z148)),MIN(V148:Z148),MAX(V148:Z148))')
    worksheet.write('AA149',
                    '=IF(MAX(MAX(V149:Z149),ABS(MIN(V149:Z149)))=ABS(MIN(V149:Z149)),MIN(V149:Z149),MAX(V149:Z149))')
    worksheet.write('AA150',
                    '=IF(MAX(MAX(V150:Z150),ABS(MIN(V150:Z150)))=ABS(MIN(V150:Z150)),MIN(V150:Z150),MAX(V150:Z150))')
    worksheet.write('AA151',
                    '=IF(MAX(MAX(V151:Z151),ABS(MIN(V151:Z151)))=ABS(MIN(V151:Z151)),MIN(V151:Z151),MAX(V151:Z151))')
    worksheet.write('AA152',
                    '=IF(MAX(MAX(V152:Z152),ABS(MIN(V152:Z152)))=ABS(MIN(V152:Z152)),MIN(V152:Z152),MAX(V152:Z152))')
    worksheet.write('AA153',
                    '=IF(MAX(MAX(V153:Z153),ABS(MIN(V153:Z153)))=ABS(MIN(V153:Z153)),MIN(V153:Z153),MAX(V153:Z153))')
    worksheet.write('AA154',
                    '=IF(MAX(MAX(V154:Z154),ABS(MIN(V154:Z154)))=ABS(MIN(V154:Z154)),MIN(V154:Z154),MAX(V154:Z154))')
    worksheet.write('AA155',
                    '=IF(MAX(MAX(V155:Z155),ABS(MIN(V155:Z155)))=ABS(MIN(V155:Z155)),MIN(V155:Z155),MAX(V155:Z155))')
    worksheet.write('AA156',
                    '=IF(MAX(MAX(V156:Z156),ABS(MIN(V156:Z156)))=ABS(MIN(V156:Z156)),MIN(V156:Z156),MAX(V156:Z156))')
    worksheet.write('AA157',
                    '=IF(MAX(MAX(V157:Z157),ABS(MIN(V157:Z157)))=ABS(MIN(V157:Z157)),MIN(V157:Z157),MAX(V157:Z157))')
    worksheet.write('AA158',
                    '=IF(MAX(MAX(V158:Z158),ABS(MIN(V158:Z158)))=ABS(MIN(V158:Z158)),MIN(V158:Z158),MAX(V158:Z158))')
    worksheet.write('AA159',
                    '=IF(MAX(MAX(V159:Z159),ABS(MIN(V159:Z159)))=ABS(MIN(V159:Z159)),MIN(V159:Z159),MAX(V159:Z159))')
    worksheet.write('AA160',
                    '=IF(MAX(MAX(V160:Z160),ABS(MIN(V160:Z160)))=ABS(MIN(V160:Z160)),MIN(V160:Z160),MAX(V160:Z160))')
    worksheet.write('AA161',
                    '=IF(MAX(MAX(V161:Z161),ABS(MIN(V161:Z161)))=ABS(MIN(V161:Z161)),MIN(V161:Z161),MAX(V161:Z161))')
    worksheet.write('AA162',
                    '=IF(MAX(MAX(V162:Z162),ABS(MIN(V162:Z162)))=ABS(MIN(V162:Z162)),MIN(V162:Z162),MAX(V162:Z162))')
    worksheet.write('AA163',
                    '=IF(MAX(MAX(V163:Z163),ABS(MIN(V163:Z163)))=ABS(MIN(V163:Z163)),MIN(V163:Z163),MAX(V163:Z163))')
    worksheet.write('AA164',
                    '=IF(MAX(MAX(V164:Z164),ABS(MIN(V164:Z164)))=ABS(MIN(V164:Z164)),MIN(V164:Z164),MAX(V164:Z164))')
    worksheet.write('AA165',
                    '=IF(MAX(MAX(V165:Z165),ABS(MIN(V165:Z165)))=ABS(MIN(V165:Z165)),MIN(V165:Z165),MAX(V165:Z165))')
    worksheet.write('AA166',
                    '=IF(MAX(MAX(V166:Z166),ABS(MIN(V166:Z166)))=ABS(MIN(V166:Z166)),MIN(V166:Z166),MAX(V166:Z166))')
    worksheet.write('AA167',
                    '=IF(MAX(MAX(V167:Z167),ABS(MIN(V167:Z167)))=ABS(MIN(V167:Z167)),MIN(V167:Z167),MAX(V167:Z167))')
    worksheet.write('AA168',
                    '=IF(MAX(MAX(V168:Z168),ABS(MIN(V168:Z168)))=ABS(MIN(V168:Z168)),MIN(V168:Z168),MAX(V168:Z168))')
    worksheet.write('AA169',
                    '=IF(MAX(MAX(V169:Z169),ABS(MIN(V169:Z169)))=ABS(MIN(V169:Z169)),MIN(V169:Z169),MAX(V169:Z169))')
    worksheet.write('AA170',
                    '=IF(MAX(MAX(V170:Z170),ABS(MIN(V170:Z170)))=ABS(MIN(V170:Z170)),MIN(V170:Z170),MAX(V170:Z170))')
    worksheet.write('AA171',
                    '=IF(MAX(MAX(V171:Z171),ABS(MIN(V171:Z171)))=ABS(MIN(V171:Z171)),MIN(V171:Z171),MAX(V171:Z171))')
    worksheet.write('AA172',
                    '=IF(MAX(MAX(V172:Z172),ABS(MIN(V172:Z172)))=ABS(MIN(V172:Z172)),MIN(V172:Z172),MAX(V172:Z172))')
    worksheet.write('AA173',
                    '=IF(MAX(MAX(V173:Z173),ABS(MIN(V173:Z173)))=ABS(MIN(V173:Z173)),MIN(V173:Z173),MAX(V173:Z173))')
    worksheet.write('AA174',
                    '=IF(MAX(MAX(V174:Z174),ABS(MIN(V174:Z174)))=ABS(MIN(V174:Z174)),MIN(V174:Z174),MAX(V174:Z174))')
    worksheet.write('AA175',
                    '=IF(MAX(MAX(V175:Z175),ABS(MIN(V175:Z175)))=ABS(MIN(V175:Z175)),MIN(V175:Z175),MAX(V175:Z175))')
    worksheet.write('AA176',
                    '=IF(MAX(MAX(V176:Z176),ABS(MIN(V176:Z176)))=ABS(MIN(V176:Z176)),MIN(V176:Z176),MAX(V176:Z176))')
    worksheet.write('AA177',
                    '=IF(MAX(MAX(V177:Z177),ABS(MIN(V177:Z177)))=ABS(MIN(V177:Z177)),MIN(V177:Z177),MAX(V177:Z177))')
    worksheet.write('AA178',
                    '=IF(MAX(MAX(V178:Z178),ABS(MIN(V178:Z178)))=ABS(MIN(V178:Z178)),MIN(V178:Z178),MAX(V178:Z178))')
    worksheet.write('AA179',
                    '=IF(MAX(MAX(V179:Z179),ABS(MIN(V179:Z179)))=ABS(MIN(V179:Z179)),MIN(V179:Z179),MAX(V179:Z179))')
    worksheet.write('AA180',
                    '=IF(MAX(MAX(V180:Z180),ABS(MIN(V180:Z180)))=ABS(MIN(V180:Z180)),MIN(V180:Z180),MAX(V180:Z180))')
    worksheet.write('AA181',
                    '=IF(MAX(MAX(V181:Z181),ABS(MIN(V181:Z181)))=ABS(MIN(V181:Z181)),MIN(V181:Z181),MAX(V181:Z181))')
    worksheet.write('AA182',
                    '=IF(MAX(MAX(V182:Z182),ABS(MIN(V182:Z182)))=ABS(MIN(V182:Z182)),MIN(V182:Z182),MAX(V182:Z182))')
    worksheet.write('AA183',
                    '=IF(MAX(MAX(V183:Z183),ABS(MIN(V183:Z183)))=ABS(MIN(V183:Z183)),MIN(V183:Z183),MAX(V183:Z183))')
    worksheet.write('AA184',
                    '=IF(MAX(MAX(V184:Z184),ABS(MIN(V184:Z184)))=ABS(MIN(V184:Z184)),MIN(V184:Z184),MAX(V184:Z184))')
    worksheet.write('AA185',
                    '=IF(MAX(MAX(V185:Z185),ABS(MIN(V185:Z185)))=ABS(MIN(V185:Z185)),MIN(V185:Z185),MAX(V185:Z185))')
    worksheet.write('AA186',
                    '=IF(MAX(MAX(V186:Z186),ABS(MIN(V186:Z186)))=ABS(MIN(V186:Z186)),MIN(V186:Z186),MAX(V186:Z186))')
    worksheet.write('AA187',
                    '=IF(MAX(MAX(V187:Z187),ABS(MIN(V187:Z187)))=ABS(MIN(V187:Z187)),MIN(V187:Z187),MAX(V187:Z187))')
    worksheet.write('AA188',
                    '=IF(MAX(MAX(V188:Z188),ABS(MIN(V188:Z188)))=ABS(MIN(V188:Z188)),MIN(V188:Z188),MAX(V188:Z188))')
    worksheet.write('AA189',
                    '=IF(MAX(MAX(V189:Z189),ABS(MIN(V189:Z189)))=ABS(MIN(V189:Z189)),MIN(V189:Z189),MAX(V189:Z189))')
    worksheet.write('AA190',
                    '=IF(MAX(MAX(V190:Z190),ABS(MIN(V190:Z190)))=ABS(MIN(V190:Z190)),MIN(V190:Z190),MAX(V190:Z190))')
    worksheet.write('AA191',
                    '=IF(MAX(MAX(V191:Z191),ABS(MIN(V191:Z191)))=ABS(MIN(V191:Z191)),MIN(V191:Z191),MAX(V191:Z191))')
    worksheet.write('AA192',
                    '=IF(MAX(MAX(V192:Z192),ABS(MIN(V192:Z192)))=ABS(MIN(V192:Z192)),MIN(V192:Z192),MAX(V192:Z192))')
    worksheet.write('AA193',
                    '=IF(MAX(MAX(V193:Z193),ABS(MIN(V193:Z193)))=ABS(MIN(V193:Z193)),MIN(V193:Z193),MAX(V193:Z193))')
    worksheet.write('AA194',
                    '=IF(MAX(MAX(V194:Z194),ABS(MIN(V194:Z194)))=ABS(MIN(V194:Z194)),MIN(V194:Z194),MAX(V194:Z194))')
    worksheet.write('AA195',
                    '=IF(MAX(MAX(V195:Z195),ABS(MIN(V195:Z195)))=ABS(MIN(V195:Z195)),MIN(V195:Z195),MAX(V195:Z195))')
    worksheet.write('AA196',
                    '=IF(MAX(MAX(V196:Z196),ABS(MIN(V196:Z196)))=ABS(MIN(V196:Z196)),MIN(V196:Z196),MAX(V196:Z196))')
    worksheet.write('AA197',
                    '=IF(MAX(MAX(V197:Z197),ABS(MIN(V197:Z197)))=ABS(MIN(V197:Z197)),MIN(V197:Z197),MAX(V197:Z197))')
    worksheet.write('AA198',
                    '=IF(MAX(MAX(V198:Z198),ABS(MIN(V198:Z198)))=ABS(MIN(V198:Z198)),MIN(V198:Z198),MAX(V198:Z198))')
    worksheet.write('AA199',
                    '=IF(MAX(MAX(V199:Z199),ABS(MIN(V199:Z199)))=ABS(MIN(V199:Z199)),MIN(V199:Z199),MAX(V199:Z199))')
    worksheet.write('AA200',
                    '=IF(MAX(MAX(V200:Z200),ABS(MIN(V200:Z200)))=ABS(MIN(V200:Z200)),MIN(V200:Z200),MAX(V200:Z200))')
    worksheet.write('AA201',
                    '=IF(MAX(MAX(V201:Z201),ABS(MIN(V201:Z201)))=ABS(MIN(V201:Z201)),MIN(V201:Z201),MAX(V201:Z201))')
    worksheet.write('AA202',
                    '=IF(MAX(MAX(V202:Z202),ABS(MIN(V202:Z202)))=ABS(MIN(V202:Z202)),MIN(V202:Z202),MAX(V202:Z202))')
    worksheet.write('AA203',
                    '=IF(MAX(MAX(V203:Z203),ABS(MIN(V203:Z203)))=ABS(MIN(V203:Z203)),MIN(V203:Z203),MAX(V203:Z203))')
    worksheet.write('AA204',
                    '=IF(MAX(MAX(V204:Z204),ABS(MIN(V204:Z204)))=ABS(MIN(V204:Z204)),MIN(V204:Z204),MAX(V204:Z204))')
    worksheet.write('AA205',
                    '=IF(MAX(MAX(V205:Z205),ABS(MIN(V205:Z205)))=ABS(MIN(V205:Z205)),MIN(V205:Z205),MAX(V205:Z205))')
    worksheet.write('AA206',
                    '=IF(MAX(MAX(V206:Z206),ABS(MIN(V206:Z206)))=ABS(MIN(V206:Z206)),MIN(V206:Z206),MAX(V206:Z206))')
    worksheet.write('AA207',
                    '=IF(MAX(MAX(V207:Z207),ABS(MIN(V207:Z207)))=ABS(MIN(V207:Z207)),MIN(V207:Z207),MAX(V207:Z207))')
    worksheet.write('AA208',
                    '=IF(MAX(MAX(V208:Z208),ABS(MIN(V208:Z208)))=ABS(MIN(V208:Z208)),MIN(V208:Z208),MAX(V208:Z208))')
    worksheet.write('AA209',
                    '=IF(MAX(MAX(V209:Z209),ABS(MIN(V209:Z209)))=ABS(MIN(V209:Z209)),MIN(V209:Z209),MAX(V209:Z209))')
    worksheet.write('AA210',
                    '=IF(MAX(MAX(V210:Z210),ABS(MIN(V210:Z210)))=ABS(MIN(V210:Z210)),MIN(V210:Z210),MAX(V210:Z210))')
    worksheet.write('AA211',
                    '=IF(MAX(MAX(V211:Z211),ABS(MIN(V211:Z211)))=ABS(MIN(V211:Z211)),MIN(V211:Z211),MAX(V211:Z211))')
    worksheet.write('AA212',
                    '=IF(MAX(MAX(V212:Z212),ABS(MIN(V212:Z212)))=ABS(MIN(V212:Z212)),MIN(V212:Z212),MAX(V212:Z212))')
    worksheet.write('AA213',
                    '=IF(MAX(MAX(V213:Z213),ABS(MIN(V213:Z213)))=ABS(MIN(V213:Z213)),MIN(V213:Z213),MAX(V213:Z213))')
    worksheet.write('AA214',
                    '=IF(MAX(MAX(V214:Z214),ABS(MIN(V214:Z214)))=ABS(MIN(V214:Z214)),MIN(V214:Z214),MAX(V214:Z214))')
    worksheet.write('AA215',
                    '=IF(MAX(MAX(V215:Z215),ABS(MIN(V215:Z215)))=ABS(MIN(V215:Z215)),MIN(V215:Z215),MAX(V215:Z215))')
    worksheet.write('AA216',
                    '=IF(MAX(MAX(V216:Z216),ABS(MIN(V216:Z216)))=ABS(MIN(V216:Z216)),MIN(V216:Z216),MAX(V216:Z216))')
    worksheet.write('AA217',
                    '=IF(MAX(MAX(V217:Z217),ABS(MIN(V217:Z217)))=ABS(MIN(V217:Z217)),MIN(V217:Z217),MAX(V217:Z217))')
    worksheet.write('AA218',
                    '=IF(MAX(MAX(V218:Z218),ABS(MIN(V218:Z218)))=ABS(MIN(V218:Z218)),MIN(V218:Z218),MAX(V218:Z218))')
    worksheet.write('AA219',
                    '=IF(MAX(MAX(V219:Z219),ABS(MIN(V219:Z219)))=ABS(MIN(V219:Z219)),MIN(V219:Z219),MAX(V219:Z219))')
    worksheet.write('AA220',
                    '=IF(MAX(MAX(V220:Z220),ABS(MIN(V220:Z220)))=ABS(MIN(V220:Z220)),MIN(V220:Z220),MAX(V220:Z220))')
    worksheet.write('AA221',
                    '=IF(MAX(MAX(V221:Z221),ABS(MIN(V221:Z221)))=ABS(MIN(V221:Z221)),MIN(V221:Z221),MAX(V221:Z221))')
    worksheet.write('AA222',
                    '=IF(MAX(MAX(V222:Z222),ABS(MIN(V222:Z222)))=ABS(MIN(V222:Z222)),MIN(V222:Z222),MAX(V222:Z222))')
    worksheet.write('AA223',
                    '=IF(MAX(MAX(V223:Z223),ABS(MIN(V223:Z223)))=ABS(MIN(V223:Z223)),MIN(V223:Z223),MAX(V223:Z223))')
    worksheet.write('AA224',
                    '=IF(MAX(MAX(V224:Z224),ABS(MIN(V224:Z224)))=ABS(MIN(V224:Z224)),MIN(V224:Z224),MAX(V224:Z224))')
    worksheet.write('AA225',
                    '=IF(MAX(MAX(V225:Z225),ABS(MIN(V225:Z225)))=ABS(MIN(V225:Z225)),MIN(V225:Z225),MAX(V225:Z225))')
    worksheet.write('AA226',
                    '=IF(MAX(MAX(V226:Z226),ABS(MIN(V226:Z226)))=ABS(MIN(V226:Z226)),MIN(V226:Z226),MAX(V226:Z226))')
    worksheet.write('AA227',
                    '=IF(MAX(MAX(V227:Z227),ABS(MIN(V227:Z227)))=ABS(MIN(V227:Z227)),MIN(V227:Z227),MAX(V227:Z227))')
    worksheet.write('AA228',
                    '=IF(MAX(MAX(V228:Z228),ABS(MIN(V228:Z228)))=ABS(MIN(V228:Z228)),MIN(V228:Z228),MAX(V228:Z228))')
    worksheet.write('AA229',
                    '=IF(MAX(MAX(V229:Z229),ABS(MIN(V229:Z229)))=ABS(MIN(V229:Z229)),MIN(V229:Z229),MAX(V229:Z229))')
    worksheet.write('AA230',
                    '=IF(MAX(MAX(V230:Z230),ABS(MIN(V230:Z230)))=ABS(MIN(V230:Z230)),MIN(V230:Z230),MAX(V230:Z230))')
    worksheet.write('AA231',
                    '=IF(MAX(MAX(V231:Z231),ABS(MIN(V231:Z231)))=ABS(MIN(V231:Z231)),MIN(V231:Z231),MAX(V231:Z231))')
    worksheet.write('AA232',
                    '=IF(MAX(MAX(V232:Z232),ABS(MIN(V232:Z232)))=ABS(MIN(V232:Z232)),MIN(V232:Z232),MAX(V232:Z232))')
    worksheet.write('AA233',
                    '=IF(MAX(MAX(V233:Z233),ABS(MIN(V233:Z233)))=ABS(MIN(V233:Z233)),MIN(V233:Z233),MAX(V233:Z233))')
    worksheet.write('AA234',
                    '=IF(MAX(MAX(V234:Z234),ABS(MIN(V234:Z234)))=ABS(MIN(V234:Z234)),MIN(V234:Z234),MAX(V234:Z234))')
    worksheet.write('AA235',
                    '=IF(MAX(MAX(V235:Z235),ABS(MIN(V235:Z235)))=ABS(MIN(V235:Z235)),MIN(V235:Z235),MAX(V235:Z235))')
    worksheet.write('AA236',
                    '=IF(MAX(MAX(V236:Z236),ABS(MIN(V236:Z236)))=ABS(MIN(V236:Z236)),MIN(V236:Z236),MAX(V236:Z236))')
    worksheet.write('AA237',
                    '=IF(MAX(MAX(V237:Z237),ABS(MIN(V237:Z237)))=ABS(MIN(V237:Z237)),MIN(V237:Z237),MAX(V237:Z237))')
    worksheet.write('AA238',
                    '=IF(MAX(MAX(V238:Z238),ABS(MIN(V238:Z238)))=ABS(MIN(V238:Z238)),MIN(V238:Z238),MAX(V238:Z238))')
    worksheet.write('AA239',
                    '=IF(MAX(MAX(V239:Z239),ABS(MIN(V239:Z239)))=ABS(MIN(V239:Z239)),MIN(V239:Z239),MAX(V239:Z239))')
    worksheet.write('AA240',
                    '=IF(MAX(MAX(V240:Z240),ABS(MIN(V240:Z240)))=ABS(MIN(V240:Z240)),MIN(V240:Z240),MAX(V240:Z240))')
    worksheet.write('AA241',
                    '=IF(MAX(MAX(V241:Z241),ABS(MIN(V241:Z241)))=ABS(MIN(V241:Z241)),MIN(V241:Z241),MAX(V241:Z241))')
    worksheet.write('AA242',
                    '=IF(MAX(MAX(V242:Z242),ABS(MIN(V242:Z242)))=ABS(MIN(V242:Z242)),MIN(V242:Z242),MAX(V242:Z242))')
    worksheet.write('AA243',
                    '=IF(MAX(MAX(V243:Z243),ABS(MIN(V243:Z243)))=ABS(MIN(V243:Z243)),MIN(V243:Z243),MAX(V243:Z243))')
    worksheet.write('AA244',
                    '=IF(MAX(MAX(V244:Z244),ABS(MIN(V244:Z244)))=ABS(MIN(V244:Z244)),MIN(V244:Z244),MAX(V244:Z244))')
    worksheet.write('AA245',
                    '=IF(MAX(MAX(V245:Z245),ABS(MIN(V245:Z245)))=ABS(MIN(V245:Z245)),MIN(V245:Z245),MAX(V245:Z245))')
    worksheet.write('AA246',
                    '=IF(MAX(MAX(V246:Z246),ABS(MIN(V246:Z246)))=ABS(MIN(V246:Z246)),MIN(V246:Z246),MAX(V246:Z246))')
    worksheet.write('AA247',
                    '=IF(MAX(MAX(V247:Z247),ABS(MIN(V247:Z247)))=ABS(MIN(V247:Z247)),MIN(V247:Z247),MAX(V247:Z247))')
    worksheet.write('AA248',
                    '=IF(MAX(MAX(V248:Z248),ABS(MIN(V248:Z248)))=ABS(MIN(V248:Z248)),MIN(V248:Z248),MAX(V248:Z248))')
    worksheet.write('AA249',
                    '=IF(MAX(MAX(V249:Z249),ABS(MIN(V249:Z249)))=ABS(MIN(V249:Z249)),MIN(V249:Z249),MAX(V249:Z249))')
    worksheet.write('AA250',
                    '=IF(MAX(MAX(V250:Z250),ABS(MIN(V250:Z250)))=ABS(MIN(V250:Z250)),MIN(V250:Z250),MAX(V250:Z250))')
    worksheet.write('AA251',
                    '=IF(MAX(MAX(V251:Z251),ABS(MIN(V251:Z251)))=ABS(MIN(V251:Z251)),MIN(V251:Z251),MAX(V251:Z251))')
    worksheet.write('AA252',
                    '=IF(MAX(MAX(V252:Z252),ABS(MIN(V252:Z252)))=ABS(MIN(V252:Z252)),MIN(V252:Z252),MAX(V252:Z252))')
    worksheet.write('AA253',
                    '=IF(MAX(MAX(V253:Z253),ABS(MIN(V253:Z253)))=ABS(MIN(V253:Z253)),MIN(V253:Z253),MAX(V253:Z253))')
    worksheet.write('AA254',
                    '=IF(MAX(MAX(V254:Z254),ABS(MIN(V254:Z254)))=ABS(MIN(V254:Z254)),MIN(V254:Z254),MAX(V254:Z254))')
    worksheet.write('AA255',
                    '=IF(MAX(MAX(V255:Z255),ABS(MIN(V255:Z255)))=ABS(MIN(V255:Z255)),MIN(V255:Z255),MAX(V255:Z255))')
    worksheet.write('AA256',
                    '=IF(MAX(MAX(V256:Z256),ABS(MIN(V256:Z256)))=ABS(MIN(V256:Z256)),MIN(V256:Z256),MAX(V256:Z256))')
    worksheet.write('AA257',
                    '=IF(MAX(MAX(V257:Z257),ABS(MIN(V257:Z257)))=ABS(MIN(V257:Z257)),MIN(V257:Z257),MAX(V257:Z257))')
    worksheet.write('AA258',
                    '=IF(MAX(MAX(V258:Z258),ABS(MIN(V258:Z258)))=ABS(MIN(V258:Z258)),MIN(V258:Z258),MAX(V258:Z258))')
    worksheet.write('AA259',
                    '=IF(MAX(MAX(V259:Z259),ABS(MIN(V259:Z259)))=ABS(MIN(V259:Z259)),MIN(V259:Z259),MAX(V259:Z259))')
    worksheet.write('AA260',
                    '=IF(MAX(MAX(V260:Z260),ABS(MIN(V260:Z260)))=ABS(MIN(V260:Z260)),MIN(V260:Z260),MAX(V260:Z260))')
    worksheet.write('AA261',
                    '=IF(MAX(MAX(V261:Z261),ABS(MIN(V261:Z261)))=ABS(MIN(V261:Z261)),MIN(V261:Z261),MAX(V261:Z261))')
    worksheet.write('AA262',
                    '=IF(MAX(MAX(V262:Z262),ABS(MIN(V262:Z262)))=ABS(MIN(V262:Z262)),MIN(V262:Z262),MAX(V262:Z262))')
    worksheet.write('AA263',
                    '=IF(MAX(MAX(V263:Z263),ABS(MIN(V263:Z263)))=ABS(MIN(V263:Z263)),MIN(V263:Z263),MAX(V263:Z263))')
    worksheet.write('AA264',
                    '=IF(MAX(MAX(V264:Z264),ABS(MIN(V264:Z264)))=ABS(MIN(V264:Z264)),MIN(V264:Z264),MAX(V264:Z264))')
    worksheet.write('AA265',
                    '=IF(MAX(MAX(V265:Z265),ABS(MIN(V265:Z265)))=ABS(MIN(V265:Z265)),MIN(V265:Z265),MAX(V265:Z265))')
    worksheet.write('AA266',
                    '=IF(MAX(MAX(V266:Z266),ABS(MIN(V266:Z266)))=ABS(MIN(V266:Z266)),MIN(V266:Z266),MAX(V266:Z266))')
    worksheet.write('AA267',
                    '=IF(MAX(MAX(V267:Z267),ABS(MIN(V267:Z267)))=ABS(MIN(V267:Z267)),MIN(V267:Z267),MAX(V267:Z267))')
    worksheet.write('AA268',
                    '=IF(MAX(MAX(V268:Z268),ABS(MIN(V268:Z268)))=ABS(MIN(V268:Z268)),MIN(V268:Z268),MAX(V268:Z268))')
    worksheet.write('AA269',
                    '=IF(MAX(MAX(V269:Z269),ABS(MIN(V269:Z269)))=ABS(MIN(V269:Z269)),MIN(V269:Z269),MAX(V269:Z269))')
    worksheet.write('AA270',
                    '=IF(MAX(MAX(V270:Z270),ABS(MIN(V270:Z270)))=ABS(MIN(V270:Z270)),MIN(V270:Z270),MAX(V270:Z270))')
    worksheet.write('AA271',
                    '=IF(MAX(MAX(V271:Z271),ABS(MIN(V271:Z271)))=ABS(MIN(V271:Z271)),MIN(V271:Z271),MAX(V271:Z271))')
    worksheet.write('AA272',
                    '=IF(MAX(MAX(V272:Z272),ABS(MIN(V272:Z272)))=ABS(MIN(V272:Z272)),MIN(V272:Z272),MAX(V272:Z272))')
    worksheet.write('AA273',
                    '=IF(MAX(MAX(V273:Z273),ABS(MIN(V273:Z273)))=ABS(MIN(V273:Z273)),MIN(V273:Z273),MAX(V273:Z273))')
    worksheet.write('AA274',
                    '=IF(MAX(MAX(V274:Z274),ABS(MIN(V274:Z274)))=ABS(MIN(V274:Z274)),MIN(V274:Z274),MAX(V274:Z274))')
    worksheet.write('AA275',
                    '=IF(MAX(MAX(V275:Z275),ABS(MIN(V275:Z275)))=ABS(MIN(V275:Z275)),MIN(V275:Z275),MAX(V275:Z275))')
    worksheet.write('AA276',
                    '=IF(MAX(MAX(V276:Z276),ABS(MIN(V276:Z276)))=ABS(MIN(V276:Z276)),MIN(V276:Z276),MAX(V276:Z276))')
    worksheet.write('AA277',
                    '=IF(MAX(MAX(V277:Z277),ABS(MIN(V277:Z277)))=ABS(MIN(V277:Z277)),MIN(V277:Z277),MAX(V277:Z277))')
    worksheet.write('AA278',
                    '=IF(MAX(MAX(V278:Z278),ABS(MIN(V278:Z278)))=ABS(MIN(V278:Z278)),MIN(V278:Z278),MAX(V278:Z278))')
    worksheet.write('AA279',
                    '=IF(MAX(MAX(V279:Z279),ABS(MIN(V279:Z279)))=ABS(MIN(V279:Z279)),MIN(V279:Z279),MAX(V279:Z279))')
    worksheet.write('AA280',
                    '=IF(MAX(MAX(V280:Z280),ABS(MIN(V280:Z280)))=ABS(MIN(V280:Z280)),MIN(V280:Z280),MAX(V280:Z280))')
    worksheet.write('AA281',
                    '=IF(MAX(MAX(V281:Z281),ABS(MIN(V281:Z281)))=ABS(MIN(V281:Z281)),MIN(V281:Z281),MAX(V281:Z281))')
    worksheet.write('AA282',
                    '=IF(MAX(MAX(V282:Z282),ABS(MIN(V282:Z282)))=ABS(MIN(V282:Z282)),MIN(V282:Z282),MAX(V282:Z282))')
    worksheet.write('AA283',
                    '=IF(MAX(MAX(V283:Z283),ABS(MIN(V283:Z283)))=ABS(MIN(V283:Z283)),MIN(V283:Z283),MAX(V283:Z283))')
    worksheet.write('AA284',
                    '=IF(MAX(MAX(V284:Z284),ABS(MIN(V284:Z284)))=ABS(MIN(V284:Z284)),MIN(V284:Z284),MAX(V284:Z284))')
    worksheet.write('AA285',
                    '=IF(MAX(MAX(V285:Z285),ABS(MIN(V285:Z285)))=ABS(MIN(V285:Z285)),MIN(V285:Z285),MAX(V285:Z285))')
    worksheet.write('AA286',
                    '=IF(MAX(MAX(V286:Z286),ABS(MIN(V286:Z286)))=ABS(MIN(V286:Z286)),MIN(V286:Z286),MAX(V286:Z286))')
    worksheet.write('AA287',
                    '=IF(MAX(MAX(V287:Z287),ABS(MIN(V287:Z287)))=ABS(MIN(V287:Z287)),MIN(V287:Z287),MAX(V287:Z287))')
    worksheet.write('AA288',
                    '=IF(MAX(MAX(V288:Z288),ABS(MIN(V288:Z288)))=ABS(MIN(V288:Z288)),MIN(V288:Z288),MAX(V288:Z288))')
    worksheet.write('AA289',
                    '=IF(MAX(MAX(V289:Z289),ABS(MIN(V289:Z289)))=ABS(MIN(V289:Z289)),MIN(V289:Z289),MAX(V289:Z289))')
    worksheet.write('AA290',
                    '=IF(MAX(MAX(V290:Z290),ABS(MIN(V290:Z290)))=ABS(MIN(V290:Z290)),MIN(V290:Z290),MAX(V290:Z290))')
    worksheet.write('AA291',
                    '=IF(MAX(MAX(V291:Z291),ABS(MIN(V291:Z291)))=ABS(MIN(V291:Z291)),MIN(V291:Z291),MAX(V291:Z291))')
    worksheet.write('AA292',
                    '=IF(MAX(MAX(V292:Z292),ABS(MIN(V292:Z292)))=ABS(MIN(V292:Z292)),MIN(V292:Z292),MAX(V292:Z292))')
    worksheet.write('AA293',
                    '=IF(MAX(MAX(V293:Z293),ABS(MIN(V293:Z293)))=ABS(MIN(V293:Z293)),MIN(V293:Z293),MAX(V293:Z293))')
    worksheet.write('AA294',
                    '=IF(MAX(MAX(V294:Z294),ABS(MIN(V294:Z294)))=ABS(MIN(V294:Z294)),MIN(V294:Z294),MAX(V294:Z294))')
    worksheet.write('AA295',
                    '=IF(MAX(MAX(V295:Z295),ABS(MIN(V295:Z295)))=ABS(MIN(V295:Z295)),MIN(V295:Z295),MAX(V295:Z295))')
    worksheet.write('AA296',
                    '=IF(MAX(MAX(V296:Z296),ABS(MIN(V296:Z296)))=ABS(MIN(V296:Z296)),MIN(V296:Z296),MAX(V296:Z296))')
    worksheet.write('AA297',
                    '=IF(MAX(MAX(V297:Z297),ABS(MIN(V297:Z297)))=ABS(MIN(V297:Z297)),MIN(V297:Z297),MAX(V297:Z297))')
    worksheet.write('AA298',
                    '=IF(MAX(MAX(V298:Z298),ABS(MIN(V298:Z298)))=ABS(MIN(V298:Z298)),MIN(V298:Z298),MAX(V298:Z298))')
    worksheet.write('AA299',
                    '=IF(MAX(MAX(V299:Z299),ABS(MIN(V299:Z299)))=ABS(MIN(V299:Z299)),MIN(V299:Z299),MAX(V299:Z299))')
    worksheet.write('AA300',
                    '=IF(MAX(MAX(V300:Z300),ABS(MIN(V300:Z300)))=ABS(MIN(V300:Z300)),MIN(V300:Z300),MAX(V300:Z300))')
    worksheet.write('AA301',
                    '=IF(MAX(MAX(V301:Z301),ABS(MIN(V301:Z301)))=ABS(MIN(V301:Z301)),MIN(V301:Z301),MAX(V301:Z301))')
    worksheet.write('AA302',
                    '=IF(MAX(MAX(V302:Z302),ABS(MIN(V302:Z302)))=ABS(MIN(V302:Z302)),MIN(V302:Z302),MAX(V302:Z302))')
    worksheet.write('AA303',
                    '=IF(MAX(MAX(V303:Z303),ABS(MIN(V303:Z303)))=ABS(MIN(V303:Z303)),MIN(V303:Z303),MAX(V303:Z303))')
    worksheet.write('AA304',
                    '=IF(MAX(MAX(V304:Z304),ABS(MIN(V304:Z304)))=ABS(MIN(V304:Z304)),MIN(V304:Z304),MAX(V304:Z304))')
    worksheet.write('AA305',
                    '=IF(MAX(MAX(V305:Z305),ABS(MIN(V305:Z305)))=ABS(MIN(V305:Z305)),MIN(V305:Z305),MAX(V305:Z305))')
    worksheet.write('AA306',
                    '=IF(MAX(MAX(V306:Z306),ABS(MIN(V306:Z306)))=ABS(MIN(V306:Z306)),MIN(V306:Z306),MAX(V306:Z306))')
    worksheet.write('AA307',
                    '=IF(MAX(MAX(V307:Z307),ABS(MIN(V307:Z307)))=ABS(MIN(V307:Z307)),MIN(V307:Z307),MAX(V307:Z307))')
    worksheet.write('AA308',
                    '=IF(MAX(MAX(V308:Z308),ABS(MIN(V308:Z308)))=ABS(MIN(V308:Z308)),MIN(V308:Z308),MAX(V308:Z308))')
    worksheet.write('AA309',
                    '=IF(MAX(MAX(V309:Z309),ABS(MIN(V309:Z309)))=ABS(MIN(V309:Z309)),MIN(V309:Z309),MAX(V309:Z309))')
    worksheet.write('AA310',
                    '=IF(MAX(MAX(V310:Z310),ABS(MIN(V310:Z310)))=ABS(MIN(V310:Z310)),MIN(V310:Z310),MAX(V310:Z310))')
    worksheet.write('AA311',
                    '=IF(MAX(MAX(V311:Z311),ABS(MIN(V311:Z311)))=ABS(MIN(V311:Z311)),MIN(V311:Z311),MAX(V311:Z311))')
    worksheet.write('AA312',
                    '=IF(MAX(MAX(V312:Z312),ABS(MIN(V312:Z312)))=ABS(MIN(V312:Z312)),MIN(V312:Z312),MAX(V312:Z312))')
    worksheet.write('AA313',
                    '=IF(MAX(MAX(V313:Z313),ABS(MIN(V313:Z313)))=ABS(MIN(V313:Z313)),MIN(V313:Z313),MAX(V313:Z313))')
    worksheet.write('AA314',
                    '=IF(MAX(MAX(V314:Z314),ABS(MIN(V314:Z314)))=ABS(MIN(V314:Z314)),MIN(V314:Z314),MAX(V314:Z314))')
    worksheet.write('AA315',
                    '=IF(MAX(MAX(V315:Z315),ABS(MIN(V315:Z315)))=ABS(MIN(V315:Z315)),MIN(V315:Z315),MAX(V315:Z315))')
    worksheet.write('AA316',
                    '=IF(MAX(MAX(V316:Z316),ABS(MIN(V316:Z316)))=ABS(MIN(V316:Z316)),MIN(V316:Z316),MAX(V316:Z316))')
    worksheet.write('AA317',
                    '=IF(MAX(MAX(V317:Z317),ABS(MIN(V317:Z317)))=ABS(MIN(V317:Z317)),MIN(V317:Z317),MAX(V317:Z317))')
    worksheet.write('AA318',
                    '=IF(MAX(MAX(V318:Z318),ABS(MIN(V318:Z318)))=ABS(MIN(V318:Z318)),MIN(V318:Z318),MAX(V318:Z318))')
    worksheet.write('AA319',
                    '=IF(MAX(MAX(V319:Z319),ABS(MIN(V319:Z319)))=ABS(MIN(V319:Z319)),MIN(V319:Z319),MAX(V319:Z319))')
    worksheet.write('AA320',
                    '=IF(MAX(MAX(V320:Z320),ABS(MIN(V320:Z320)))=ABS(MIN(V320:Z320)),MIN(V320:Z320),MAX(V320:Z320))')
    worksheet.write('AA321',
                    '=IF(MAX(MAX(V321:Z321),ABS(MIN(V321:Z321)))=ABS(MIN(V321:Z321)),MIN(V321:Z321),MAX(V321:Z321))')
    worksheet.write('AA322',
                    '=IF(MAX(MAX(V322:Z322),ABS(MIN(V322:Z322)))=ABS(MIN(V322:Z322)),MIN(V322:Z322),MAX(V322:Z322))')
    worksheet.write('AA323',
                    '=IF(MAX(MAX(V323:Z323),ABS(MIN(V323:Z323)))=ABS(MIN(V323:Z323)),MIN(V323:Z323),MAX(V323:Z323))')
    worksheet.write('AA324',
                    '=IF(MAX(MAX(V324:Z324),ABS(MIN(V324:Z324)))=ABS(MIN(V324:Z324)),MIN(V324:Z324),MAX(V324:Z324))')
    worksheet.write('AA325',
                    '=IF(MAX(MAX(V325:Z325),ABS(MIN(V325:Z325)))=ABS(MIN(V325:Z325)),MIN(V325:Z325),MAX(V325:Z325))')
    worksheet.write('AA326',
                    '=IF(MAX(MAX(V326:Z326),ABS(MIN(V326:Z326)))=ABS(MIN(V326:Z326)),MIN(V326:Z326),MAX(V326:Z326))')
    worksheet.write('AA327',
                    '=IF(MAX(MAX(V327:Z327),ABS(MIN(V327:Z327)))=ABS(MIN(V327:Z327)),MIN(V327:Z327),MAX(V327:Z327))')
    worksheet.write('AA328',
                    '=IF(MAX(MAX(V328:Z328),ABS(MIN(V328:Z328)))=ABS(MIN(V328:Z328)),MIN(V328:Z328),MAX(V328:Z328))')
    worksheet.write('AA329',
                    '=IF(MAX(MAX(V329:Z329),ABS(MIN(V329:Z329)))=ABS(MIN(V329:Z329)),MIN(V329:Z329),MAX(V329:Z329))')
    worksheet.write('AA330',
                    '=IF(MAX(MAX(V330:Z330),ABS(MIN(V330:Z330)))=ABS(MIN(V330:Z330)),MIN(V330:Z330),MAX(V330:Z330))')
    worksheet.write('AA331',
                    '=IF(MAX(MAX(V331:Z331),ABS(MIN(V331:Z331)))=ABS(MIN(V331:Z331)),MIN(V331:Z331),MAX(V331:Z331))')
    worksheet.write('AA332',
                    '=IF(MAX(MAX(V332:Z332),ABS(MIN(V332:Z332)))=ABS(MIN(V332:Z332)),MIN(V332:Z332),MAX(V332:Z332))')
    worksheet.write('AA333',
                    '=IF(MAX(MAX(V333:Z333),ABS(MIN(V333:Z333)))=ABS(MIN(V333:Z333)),MIN(V333:Z333),MAX(V333:Z333))')
    worksheet.write('AA334',
                    '=IF(MAX(MAX(V334:Z334),ABS(MIN(V334:Z334)))=ABS(MIN(V334:Z334)),MIN(V334:Z334),MAX(V334:Z334))')
    worksheet.write('AA335',
                    '=IF(MAX(MAX(V335:Z335),ABS(MIN(V335:Z335)))=ABS(MIN(V335:Z335)),MIN(V335:Z335),MAX(V335:Z335))')
    worksheet.write('AA336',
                    '=IF(MAX(MAX(V336:Z336),ABS(MIN(V336:Z336)))=ABS(MIN(V336:Z336)),MIN(V336:Z336),MAX(V336:Z336))')
    worksheet.write('AA337',
                    '=IF(MAX(MAX(V337:Z337),ABS(MIN(V337:Z337)))=ABS(MIN(V337:Z337)),MIN(V337:Z337),MAX(V337:Z337))')
    worksheet.write('AA338',
                    '=IF(MAX(MAX(V338:Z338),ABS(MIN(V338:Z338)))=ABS(MIN(V338:Z338)),MIN(V338:Z338),MAX(V338:Z338))')
    worksheet.write('AA339',
                    '=IF(MAX(MAX(V339:Z339),ABS(MIN(V339:Z339)))=ABS(MIN(V339:Z339)),MIN(V339:Z339),MAX(V339:Z339))')
    worksheet.write('AA340',
                    '=IF(MAX(MAX(V340:Z340),ABS(MIN(V340:Z340)))=ABS(MIN(V340:Z340)),MIN(V340:Z340),MAX(V340:Z340))')
    worksheet.write('AA341',
                    '=IF(MAX(MAX(V341:Z341),ABS(MIN(V341:Z341)))=ABS(MIN(V341:Z341)),MIN(V341:Z341),MAX(V341:Z341))')
    worksheet.write('AA342',
                    '=IF(MAX(MAX(V342:Z342),ABS(MIN(V342:Z342)))=ABS(MIN(V342:Z342)),MIN(V342:Z342),MAX(V342:Z342))')
    worksheet.write('AA343',
                    '=IF(MAX(MAX(V343:Z343),ABS(MIN(V343:Z343)))=ABS(MIN(V343:Z343)),MIN(V343:Z343),MAX(V343:Z343))')
    worksheet.write('AA344',
                    '=IF(MAX(MAX(V344:Z344),ABS(MIN(V344:Z344)))=ABS(MIN(V344:Z344)),MIN(V344:Z344),MAX(V344:Z344))')
    worksheet.write('AA345',
                    '=IF(MAX(MAX(V345:Z345),ABS(MIN(V345:Z345)))=ABS(MIN(V345:Z345)),MIN(V345:Z345),MAX(V345:Z345))')
    worksheet.write('AA346',
                    '=IF(MAX(MAX(V346:Z346),ABS(MIN(V346:Z346)))=ABS(MIN(V346:Z346)),MIN(V346:Z346),MAX(V346:Z346))')
    worksheet.write('AA347',
                    '=IF(MAX(MAX(V347:Z347),ABS(MIN(V347:Z347)))=ABS(MIN(V347:Z347)),MIN(V347:Z347),MAX(V347:Z347))')
    worksheet.write('AA348',
                    '=IF(MAX(MAX(V348:Z348),ABS(MIN(V348:Z348)))=ABS(MIN(V348:Z348)),MIN(V348:Z348),MAX(V348:Z348))')
    worksheet.write('AA349',
                    '=IF(MAX(MAX(V349:Z349),ABS(MIN(V349:Z349)))=ABS(MIN(V349:Z349)),MIN(V349:Z349),MAX(V349:Z349))')
    worksheet.write('AA350',
                    '=IF(MAX(MAX(V350:Z350),ABS(MIN(V350:Z350)))=ABS(MIN(V350:Z350)),MIN(V350:Z350),MAX(V350:Z350))')
    worksheet.write('AA351',
                    '=IF(MAX(MAX(V351:Z351),ABS(MIN(V351:Z351)))=ABS(MIN(V351:Z351)),MIN(V351:Z351),MAX(V351:Z351))')
    worksheet.write('AA352',
                    '=IF(MAX(MAX(V352:Z352),ABS(MIN(V352:Z352)))=ABS(MIN(V352:Z352)),MIN(V352:Z352),MAX(V352:Z352))')
    worksheet.write('AA353',
                    '=IF(MAX(MAX(V353:Z353),ABS(MIN(V353:Z353)))=ABS(MIN(V353:Z353)),MIN(V353:Z353),MAX(V353:Z353))')
    worksheet.write('AB144', '="Max z Year"')
    worksheet.write('AB145', '=IFERROR(V144+MATCH(AA145,V145:Z145,0)-1,"")')
    worksheet.write('AB146', '=IFERROR(V144+MATCH(AA146,V146:Z146,0)-1,"")')
    worksheet.write('AB147', '=IFERROR(V144+MATCH(AA147,V147:Z147,0)-1,"")')
    worksheet.write('AB148', '=IFERROR(V144+MATCH(AA148,V148:Z148,0)-1,"")')
    worksheet.write('AB149', '=IFERROR(V144+MATCH(AA149,V149:Z149,0)-1,"")')
    worksheet.write('AB150', '=IFERROR(V144+MATCH(AA150,V150:Z150,0)-1,"")')
    worksheet.write('AB151', '=IFERROR(V144+MATCH(AA151,V151:Z151,0)-1,"")')
    worksheet.write('AB152', '=IFERROR(V144+MATCH(AA152,V152:Z152,0)-1,"")')
    worksheet.write('AB153', '=IFERROR(V144+MATCH(AA153,V153:Z153,0)-1,"")')
    worksheet.write('AB154', '=IFERROR(V144+MATCH(AA154,V154:Z154,0)-1,"")')
    worksheet.write('AB155', '=IFERROR(V144+MATCH(AA155,V155:Z155,0)-1,"")')
    worksheet.write('AB156', '=IFERROR(V144+MATCH(AA156,V156:Z156,0)-1,"")')
    worksheet.write('AB157', '=IFERROR(V144+MATCH(AA157,V157:Z157,0)-1,"")')
    worksheet.write('AB158', '=IFERROR(V144+MATCH(AA158,V158:Z158,0)-1,"")')
    worksheet.write('AB159', '=IFERROR(V144+MATCH(AA159,V159:Z159,0)-1,"")')
    worksheet.write('AB160', '=IFERROR(V144+MATCH(AA160,V160:Z160,0)-1,"")')
    worksheet.write('AB161', '=IFERROR(V144+MATCH(AA161,V161:Z161,0)-1,"")')
    worksheet.write('AB162', '=IFERROR(V144+MATCH(AA162,V162:Z162,0)-1,"")')
    worksheet.write('AB163', '=IFERROR(V144+MATCH(AA163,V163:Z163,0)-1,"")')
    worksheet.write('AB164', '=IFERROR(V144+MATCH(AA164,V164:Z164,0)-1,"")')
    worksheet.write('AB165', '=IFERROR(V144+MATCH(AA165,V165:Z165,0)-1,"")')
    worksheet.write('AB166', '=IFERROR(V144+MATCH(AA166,V166:Z166,0)-1,"")')
    worksheet.write('AB167', '=IFERROR(V144+MATCH(AA167,V167:Z167,0)-1,"")')
    worksheet.write('AB168', '=IFERROR(V144+MATCH(AA168,V168:Z168,0)-1,"")')
    worksheet.write('AB169', '=IFERROR(V144+MATCH(AA169,V169:Z169,0)-1,"")')
    worksheet.write('AB170', '=IFERROR(V144+MATCH(AA170,V170:Z170,0)-1,"")')
    worksheet.write('AB171', '=IFERROR(V144+MATCH(AA171,V171:Z171,0)-1,"")')
    worksheet.write('AB172', '=IFERROR(V144+MATCH(AA172,V172:Z172,0)-1,"")')
    worksheet.write('AB173', '=IFERROR(V144+MATCH(AA173,V173:Z173,0)-1,"")')
    worksheet.write('AB174', '=IFERROR(V144+MATCH(AA174,V174:Z174,0)-1,"")')
    worksheet.write('AB175', '=IFERROR(V144+MATCH(AA175,V175:Z175,0)-1,"")')
    worksheet.write('AB176', '=IFERROR(V144+MATCH(AA176,V176:Z176,0)-1,"")')
    worksheet.write('AB177', '=IFERROR(V144+MATCH(AA177,V177:Z177,0)-1,"")')
    worksheet.write('AB178', '=IFERROR(V144+MATCH(AA178,V178:Z178,0)-1,"")')
    worksheet.write('AB179', '=IFERROR(V144+MATCH(AA179,V179:Z179,0)-1,"")')
    worksheet.write('AB180', '=IFERROR(V144+MATCH(AA180,V180:Z180,0)-1,"")')
    worksheet.write('AB181', '=IFERROR(V144+MATCH(AA181,V181:Z181,0)-1,"")')
    worksheet.write('AB182', '=IFERROR(V144+MATCH(AA182,V182:Z182,0)-1,"")')
    worksheet.write('AB183', '=IFERROR(V144+MATCH(AA183,V183:Z183,0)-1,"")')
    worksheet.write('AB184', '=IFERROR(V144+MATCH(AA184,V184:Z184,0)-1,"")')
    worksheet.write('AB185', '=IFERROR(V144+MATCH(AA185,V185:Z185,0)-1,"")')
    worksheet.write('AB186', '=IFERROR(V144+MATCH(AA186,V186:Z186,0)-1,"")')
    worksheet.write('AB187', '=IFERROR(V144+MATCH(AA187,V187:Z187,0)-1,"")')
    worksheet.write('AB188', '=IFERROR(V144+MATCH(AA188,V188:Z188,0)-1,"")')
    worksheet.write('AB189', '=IFERROR(V144+MATCH(AA189,V189:Z189,0)-1,"")')
    worksheet.write('AB190', '=IFERROR(V144+MATCH(AA190,V190:Z190,0)-1,"")')
    worksheet.write('AB191', '=IFERROR(V144+MATCH(AA191,V191:Z191,0)-1,"")')
    worksheet.write('AB192', '=IFERROR(V144+MATCH(AA192,V192:Z192,0)-1,"")')
    worksheet.write('AB193', '=IFERROR(V144+MATCH(AA193,V193:Z193,0)-1,"")')
    worksheet.write('AB194', '=IFERROR(V144+MATCH(AA194,V194:Z194,0)-1,"")')
    worksheet.write('AB195', '=IFERROR(V144+MATCH(AA195,V195:Z195,0)-1,"")')
    worksheet.write('AB196', '=IFERROR(V144+MATCH(AA196,V196:Z196,0)-1,"")')
    worksheet.write('AB197', '=IFERROR(V144+MATCH(AA197,V197:Z197,0)-1,"")')
    worksheet.write('AB198', '=IFERROR(V144+MATCH(AA198,V198:Z198,0)-1,"")')
    worksheet.write('AB199', '=IFERROR(V144+MATCH(AA199,V199:Z199,0)-1,"")')
    worksheet.write('AB200', '=IFERROR(V144+MATCH(AA200,V200:Z200,0)-1,"")')
    worksheet.write('AB201', '=IFERROR(V144+MATCH(AA201,V201:Z201,0)-1,"")')
    worksheet.write('AB202', '=IFERROR(V144+MATCH(AA202,V202:Z202,0)-1,"")')
    worksheet.write('AB203', '=IFERROR(V144+MATCH(AA203,V203:Z203,0)-1,"")')
    worksheet.write('AB204', '=IFERROR(V144+MATCH(AA204,V204:Z204,0)-1,"")')
    worksheet.write('AB205', '=IFERROR(V144+MATCH(AA205,V205:Z205,0)-1,"")')
    worksheet.write('AB206', '=IFERROR(V144+MATCH(AA206,V206:Z206,0)-1,"")')
    worksheet.write('AB207', '=IFERROR(V144+MATCH(AA207,V207:Z207,0)-1,"")')
    worksheet.write('AB208', '=IFERROR(V144+MATCH(AA208,V208:Z208,0)-1,"")')
    worksheet.write('AB209', '=IFERROR(V144+MATCH(AA209,V209:Z209,0)-1,"")')
    worksheet.write('AB210', '=IFERROR(V144+MATCH(AA210,V210:Z210,0)-1,"")')
    worksheet.write('AB211', '=IFERROR(V144+MATCH(AA211,V211:Z211,0)-1,"")')
    worksheet.write('AB212', '=IFERROR(V144+MATCH(AA212,V212:Z212,0)-1,"")')
    worksheet.write('AB213', '=IFERROR(V144+MATCH(AA213,V213:Z213,0)-1,"")')
    worksheet.write('AB214', '=IFERROR(V144+MATCH(AA214,V214:Z214,0)-1,"")')
    worksheet.write('AB215', '=IFERROR(V144+MATCH(AA215,V215:Z215,0)-1,"")')
    worksheet.write('AB216', '=IFERROR(V144+MATCH(AA216,V216:Z216,0)-1,"")')
    worksheet.write('AB217', '=IFERROR(V144+MATCH(AA217,V217:Z217,0)-1,"")')
    worksheet.write('AB218', '=IFERROR(V144+MATCH(AA218,V218:Z218,0)-1,"")')
    worksheet.write('AB219', '=IFERROR(V144+MATCH(AA219,V219:Z219,0)-1,"")')
    worksheet.write('AB220', '=IFERROR(V144+MATCH(AA220,V220:Z220,0)-1,"")')
    worksheet.write('AB221', '=IFERROR(V144+MATCH(AA221,V221:Z221,0)-1,"")')
    worksheet.write('AB222', '=IFERROR(V144+MATCH(AA222,V222:Z222,0)-1,"")')
    worksheet.write('AB223', '=IFERROR(V144+MATCH(AA223,V223:Z223,0)-1,"")')
    worksheet.write('AB224', '=IFERROR(V144+MATCH(AA224,V224:Z224,0)-1,"")')
    worksheet.write('AB225', '=IFERROR(V144+MATCH(AA225,V225:Z225,0)-1,"")')
    worksheet.write('AB226', '=IFERROR(V144+MATCH(AA226,V226:Z226,0)-1,"")')
    worksheet.write('AB227', '=IFERROR(V144+MATCH(AA227,V227:Z227,0)-1,"")')
    worksheet.write('AB228', '=IFERROR(V144+MATCH(AA228,V228:Z228,0)-1,"")')
    worksheet.write('AB229', '=IFERROR(V144+MATCH(AA229,V229:Z229,0)-1,"")')
    worksheet.write('AB230', '=IFERROR(V144+MATCH(AA230,V230:Z230,0)-1,"")')
    worksheet.write('AB231', '=IFERROR(V144+MATCH(AA231,V231:Z231,0)-1,"")')
    worksheet.write('AB232', '=IFERROR(V144+MATCH(AA232,V232:Z232,0)-1,"")')
    worksheet.write('AB233', '=IFERROR(V144+MATCH(AA233,V233:Z233,0)-1,"")')
    worksheet.write('AB234', '=IFERROR(V144+MATCH(AA234,V234:Z234,0)-1,"")')
    worksheet.write('AB235', '=IFERROR(V144+MATCH(AA235,V235:Z235,0)-1,"")')
    worksheet.write('AB236', '=IFERROR(V144+MATCH(AA236,V236:Z236,0)-1,"")')
    worksheet.write('AB237', '=IFERROR(V144+MATCH(AA237,V237:Z237,0)-1,"")')
    worksheet.write('AB238', '=IFERROR(V144+MATCH(AA238,V238:Z238,0)-1,"")')
    worksheet.write('AB239', '=IFERROR(V144+MATCH(AA239,V239:Z239,0)-1,"")')
    worksheet.write('AB240', '=IFERROR(V144+MATCH(AA240,V240:Z240,0)-1,"")')
    worksheet.write('AB241', '=IFERROR(V144+MATCH(AA241,V241:Z241,0)-1,"")')
    worksheet.write('AB242', '=IFERROR(V144+MATCH(AA242,V242:Z242,0)-1,"")')
    worksheet.write('AB243', '=IFERROR(V144+MATCH(AA243,V243:Z243,0)-1,"")')
    worksheet.write('AB244', '=IFERROR(V144+MATCH(AA244,V244:Z244,0)-1,"")')
    worksheet.write('AB245', '=IFERROR(V144+MATCH(AA245,V245:Z245,0)-1,"")')
    worksheet.write('AB246', '=IFERROR(V144+MATCH(AA246,V246:Z246,0)-1,"")')
    worksheet.write('AB247', '=IFERROR(V144+MATCH(AA247,V247:Z247,0)-1,"")')
    worksheet.write('AB248', '=IFERROR(V144+MATCH(AA248,V248:Z248,0)-1,"")')
    worksheet.write('AB249', '=IFERROR(V144+MATCH(AA249,V249:Z249,0)-1,"")')
    worksheet.write('AB250', '=IFERROR(V144+MATCH(AA250,V250:Z250,0)-1,"")')
    worksheet.write('AB251', '=IFERROR(V144+MATCH(AA251,V251:Z251,0)-1,"")')
    worksheet.write('AB252', '=IFERROR(V144+MATCH(AA252,V252:Z252,0)-1,"")')
    worksheet.write('AB253', '=IFERROR(V144+MATCH(AA253,V253:Z253,0)-1,"")')
    worksheet.write('AB254', '=IFERROR(V144+MATCH(AA254,V254:Z254,0)-1,"")')
    worksheet.write('AB255', '=IFERROR(V144+MATCH(AA255,V255:Z255,0)-1,"")')
    worksheet.write('AB256', '=IFERROR(V144+MATCH(AA256,V256:Z256,0)-1,"")')
    worksheet.write('AB257', '=IFERROR(V144+MATCH(AA257,V257:Z257,0)-1,"")')
    worksheet.write('AB258', '=IFERROR(V144+MATCH(AA258,V258:Z258,0)-1,"")')
    worksheet.write('AB259', '=IFERROR(V144+MATCH(AA259,V259:Z259,0)-1,"")')
    worksheet.write('AB260', '=IFERROR(V144+MATCH(AA260,V260:Z260,0)-1,"")')
    worksheet.write('AB261', '=IFERROR(V144+MATCH(AA261,V261:Z261,0)-1,"")')
    worksheet.write('AB262', '=IFERROR(V144+MATCH(AA262,V262:Z262,0)-1,"")')
    worksheet.write('AB263', '=IFERROR(V144+MATCH(AA263,V263:Z263,0)-1,"")')
    worksheet.write('AB264', '=IFERROR(V144+MATCH(AA264,V264:Z264,0)-1,"")')
    worksheet.write('AB265', '=IFERROR(V144+MATCH(AA265,V265:Z265,0)-1,"")')
    worksheet.write('AB266', '=IFERROR(V144+MATCH(AA266,V266:Z266,0)-1,"")')
    worksheet.write('AB267', '=IFERROR(V144+MATCH(AA267,V267:Z267,0)-1,"")')
    worksheet.write('AB268', '=IFERROR(V144+MATCH(AA268,V268:Z268,0)-1,"")')
    worksheet.write('AB269', '=IFERROR(V144+MATCH(AA269,V269:Z269,0)-1,"")')
    worksheet.write('AB270', '=IFERROR(V144+MATCH(AA270,V270:Z270,0)-1,"")')
    worksheet.write('AB271', '=IFERROR(V144+MATCH(AA271,V271:Z271,0)-1,"")')
    worksheet.write('AB272', '=IFERROR(V144+MATCH(AA272,V272:Z272,0)-1,"")')
    worksheet.write('AB273', '=IFERROR(V144+MATCH(AA273,V273:Z273,0)-1,"")')
    worksheet.write('AB274', '=IFERROR(V144+MATCH(AA274,V274:Z274,0)-1,"")')
    worksheet.write('AB275', '=IFERROR(V144+MATCH(AA275,V275:Z275,0)-1,"")')
    worksheet.write('AB276', '=IFERROR(V144+MATCH(AA276,V276:Z276,0)-1,"")')
    worksheet.write('AB277', '=IFERROR(V144+MATCH(AA277,V277:Z277,0)-1,"")')
    worksheet.write('AB278', '=IFERROR(V144+MATCH(AA278,V278:Z278,0)-1,"")')
    worksheet.write('AB279', '=IFERROR(V144+MATCH(AA279,V279:Z279,0)-1,"")')
    worksheet.write('AB280', '=IFERROR(V144+MATCH(AA280,V280:Z280,0)-1,"")')
    worksheet.write('AB281', '=IFERROR(V144+MATCH(AA281,V281:Z281,0)-1,"")')
    worksheet.write('AB282', '=IFERROR(V144+MATCH(AA282,V282:Z282,0)-1,"")')
    worksheet.write('AB283', '=IFERROR(V144+MATCH(AA283,V283:Z283,0)-1,"")')
    worksheet.write('AB284', '=IFERROR(V144+MATCH(AA284,V284:Z284,0)-1,"")')
    worksheet.write('AB285', '=IFERROR(V144+MATCH(AA285,V285:Z285,0)-1,"")')
    worksheet.write('AB286', '=IFERROR(V144+MATCH(AA286,V286:Z286,0)-1,"")')
    worksheet.write('AB287', '=IFERROR(V144+MATCH(AA287,V287:Z287,0)-1,"")')
    worksheet.write('AB288', '=IFERROR(V144+MATCH(AA288,V288:Z288,0)-1,"")')
    worksheet.write('AB289', '=IFERROR(V144+MATCH(AA289,V289:Z289,0)-1,"")')
    worksheet.write('AB290', '=IFERROR(V144+MATCH(AA290,V290:Z290,0)-1,"")')
    worksheet.write('AB291', '=IFERROR(V144+MATCH(AA291,V291:Z291,0)-1,"")')
    worksheet.write('AB292', '=IFERROR(V144+MATCH(AA292,V292:Z292,0)-1,"")')
    worksheet.write('AB293', '=IFERROR(V144+MATCH(AA293,V293:Z293,0)-1,"")')
    worksheet.write('AB294', '=IFERROR(V144+MATCH(AA294,V294:Z294,0)-1,"")')
    worksheet.write('AB295', '=IFERROR(V144+MATCH(AA295,V295:Z295,0)-1,"")')
    worksheet.write('AB296', '=IFERROR(V144+MATCH(AA296,V296:Z296,0)-1,"")')
    worksheet.write('AB297', '=IFERROR(V144+MATCH(AA297,V297:Z297,0)-1,"")')
    worksheet.write('AB298', '=IFERROR(V144+MATCH(AA298,V298:Z298,0)-1,"")')
    worksheet.write('AB299', '=IFERROR(V144+MATCH(AA299,V299:Z299,0)-1,"")')
    worksheet.write('AB300', '=IFERROR(V144+MATCH(AA300,V300:Z300,0)-1,"")')
    worksheet.write('AB301', '=IFERROR(V144+MATCH(AA301,V301:Z301,0)-1,"")')
    worksheet.write('AB302', '=IFERROR(V144+MATCH(AA302,V302:Z302,0)-1,"")')
    worksheet.write('AB303', '=IFERROR(V144+MATCH(AA303,V303:Z303,0)-1,"")')
    worksheet.write('AB304', '=IFERROR(V144+MATCH(AA304,V304:Z304,0)-1,"")')
    worksheet.write('AB305', '=IFERROR(V144+MATCH(AA305,V305:Z305,0)-1,"")')
    worksheet.write('AB306', '=IFERROR(V144+MATCH(AA306,V306:Z306,0)-1,"")')
    worksheet.write('AB307', '=IFERROR(V144+MATCH(AA307,V307:Z307,0)-1,"")')
    worksheet.write('AB308', '=IFERROR(V144+MATCH(AA308,V308:Z308,0)-1,"")')
    worksheet.write('AB309', '=IFERROR(V144+MATCH(AA309,V309:Z309,0)-1,"")')
    worksheet.write('AB310', '=IFERROR(V144+MATCH(AA310,V310:Z310,0)-1,"")')
    worksheet.write('AB311', '=IFERROR(V144+MATCH(AA311,V311:Z311,0)-1,"")')
    worksheet.write('AB312', '=IFERROR(V144+MATCH(AA312,V312:Z312,0)-1,"")')
    worksheet.write('AB313', '=IFERROR(V144+MATCH(AA313,V313:Z313,0)-1,"")')
    worksheet.write('AB314', '=IFERROR(V144+MATCH(AA314,V314:Z314,0)-1,"")')
    worksheet.write('AB315', '=IFERROR(V144+MATCH(AA315,V315:Z315,0)-1,"")')
    worksheet.write('AB316', '=IFERROR(V144+MATCH(AA316,V316:Z316,0)-1,"")')
    worksheet.write('AB317', '=IFERROR(V144+MATCH(AA317,V317:Z317,0)-1,"")')
    worksheet.write('AB318', '=IFERROR(V144+MATCH(AA318,V318:Z318,0)-1,"")')
    worksheet.write('AB319', '=IFERROR(V144+MATCH(AA319,V319:Z319,0)-1,"")')
    worksheet.write('AB320', '=IFERROR(V144+MATCH(AA320,V320:Z320,0)-1,"")')
    worksheet.write('AB321', '=IFERROR(V144+MATCH(AA321,V321:Z321,0)-1,"")')
    worksheet.write('AB322', '=IFERROR(V144+MATCH(AA322,V322:Z322,0)-1,"")')
    worksheet.write('AB323', '=IFERROR(V144+MATCH(AA323,V323:Z323,0)-1,"")')
    worksheet.write('AB324', '=IFERROR(V144+MATCH(AA324,V324:Z324,0)-1,"")')
    worksheet.write('AB325', '=IFERROR(V144+MATCH(AA325,V325:Z325,0)-1,"")')
    worksheet.write('AB326', '=IFERROR(V144+MATCH(AA326,V326:Z326,0)-1,"")')
    worksheet.write('AB327', '=IFERROR(V144+MATCH(AA327,V327:Z327,0)-1,"")')
    worksheet.write('AB328', '=IFERROR(V144+MATCH(AA328,V328:Z328,0)-1,"")')
    worksheet.write('AB329', '=IFERROR(V144+MATCH(AA329,V329:Z329,0)-1,"")')
    worksheet.write('AB330', '=IFERROR(V144+MATCH(AA330,V330:Z330,0)-1,"")')
    worksheet.write('AB331', '=IFERROR(V144+MATCH(AA331,V331:Z331,0)-1,"")')
    worksheet.write('AB332', '=IFERROR(V144+MATCH(AA332,V332:Z332,0)-1,"")')
    worksheet.write('AB333', '=IFERROR(V144+MATCH(AA333,V333:Z333,0)-1,"")')
    worksheet.write('AB334', '=IFERROR(V144+MATCH(AA334,V334:Z334,0)-1,"")')
    worksheet.write('AB335', '=IFERROR(V144+MATCH(AA335,V335:Z335,0)-1,"")')
    worksheet.write('AB336', '=IFERROR(V144+MATCH(AA336,V336:Z336,0)-1,"")')
    worksheet.write('AB337', '=IFERROR(V144+MATCH(AA337,V337:Z337,0)-1,"")')
    worksheet.write('AB338', '=IFERROR(V144+MATCH(AA338,V338:Z338,0)-1,"")')
    worksheet.write('AB339', '=IFERROR(V144+MATCH(AA339,V339:Z339,0)-1,"")')
    worksheet.write('AB340', '=IFERROR(V144+MATCH(AA340,V340:Z340,0)-1,"")')
    worksheet.write('AB341', '=IFERROR(V144+MATCH(AA341,V341:Z341,0)-1,"")')
    worksheet.write('AB342', '=IFERROR(V144+MATCH(AA342,V342:Z342,0)-1,"")')
    worksheet.write('AB343', '=IFERROR(V144+MATCH(AA343,V343:Z343,0)-1,"")')
    worksheet.write('AB344', '=IFERROR(V144+MATCH(AA344,V344:Z344,0)-1,"")')
    worksheet.write('AB345', '=IFERROR(V144+MATCH(AA345,V345:Z345,0)-1,"")')
    worksheet.write('AB346', '=IFERROR(V144+MATCH(AA346,V346:Z346,0)-1,"")')
    worksheet.write('AB347', '=IFERROR(V144+MATCH(AA347,V347:Z347,0)-1,"")')
    worksheet.write('AB348', '=IFERROR(V144+MATCH(AA348,V348:Z348,0)-1,"")')
    worksheet.write('AB349', '=IFERROR(V144+MATCH(AA349,V349:Z349,0)-1,"")')
    worksheet.write('AB350', '=IFERROR(V144+MATCH(AA350,V350:Z350,0)-1,"")')
    worksheet.write('AB351', '=IFERROR(V144+MATCH(AA351,V351:Z351,0)-1,"")')
    worksheet.write('AB352', '=IFERROR(V144+MATCH(AA352,V352:Z352,0)-1,"")')
    worksheet.write('AB353', '=IFERROR(V144+MATCH(AA353,V353:Z353,0)-1,"")')

    worksheet.write('AC144', '="Direction"')
    worksheet.write('AC145', '=IF(AB145<>"",IF(S145=AB145,"Low",IF(AB145=Q145,"High","")),"")')
    worksheet.write('AC146', '=IF(AB146<>"",IF(S146=AB146,"Low",IF(AB146=Q146,"High","")),"")')
    worksheet.write('AC147', '=IF(AB147<>"",IF(S147=AB147,"Low",IF(AB147=Q147,"High","")),"")')
    worksheet.write('AC148', '=IF(AB148<>"",IF(S148=AB148,"Low",IF(AB148=Q148,"High","")),"")')
    worksheet.write('AC149', '=IF(AB149<>"",IF(S149=AB149,"Low",IF(AB149=Q149,"High","")),"")')
    worksheet.write('AC150', '=IF(AB150<>"",IF(S150=AB150,"Low",IF(AB150=Q150,"High","")),"")')
    worksheet.write('AC151', '=IF(AB151<>"",IF(S151=AB151,"Low",IF(AB151=Q151,"High","")),"")')
    worksheet.write('AC152', '=IF(AB152<>"",IF(S152=AB152,"Low",IF(AB152=Q152,"High","")),"")')
    worksheet.write('AC153', '=IF(AB153<>"",IF(S153=AB153,"Low",IF(AB153=Q153,"High","")),"")')
    worksheet.write('AC154', '=IF(AB154<>"",IF(S154=AB154,"Low",IF(AB154=Q154,"High","")),"")')
    worksheet.write('AC155', '=IF(AB155<>"",IF(S155=AB155,"Low",IF(AB155=Q155,"High","")),"")')
    worksheet.write('AC156', '=IF(AB156<>"",IF(S156=AB156,"Low",IF(AB156=Q156,"High","")),"")')
    worksheet.write('AC157', '=IF(AB157<>"",IF(S157=AB157,"Low",IF(AB157=Q157,"High","")),"")')
    worksheet.write('AC158', '=IF(AB158<>"",IF(S158=AB158,"Low",IF(AB158=Q158,"High","")),"")')
    worksheet.write('AC159', '=IF(AB159<>"",IF(S159=AB159,"Low",IF(AB159=Q159,"High","")),"")')
    worksheet.write('AC160', '=IF(AB160<>"",IF(S160=AB160,"Low",IF(AB160=Q160,"High","")),"")')
    worksheet.write('AC161', '=IF(AB161<>"",IF(S161=AB161,"Low",IF(AB161=Q161,"High","")),"")')
    worksheet.write('AC162', '=IF(AB162<>"",IF(S162=AB162,"Low",IF(AB162=Q162,"High","")),"")')
    worksheet.write('AC163', '=IF(AB163<>"",IF(S163=AB163,"Low",IF(AB163=Q163,"High","")),"")')
    worksheet.write('AC164', '=IF(AB164<>"",IF(S164=AB164,"Low",IF(AB164=Q164,"High","")),"")')
    worksheet.write('AC165', '=IF(AB165<>"",IF(S165=AB165,"Low",IF(AB165=Q165,"High","")),"")')
    worksheet.write('AC166', '=IF(AB166<>"",IF(S166=AB166,"Low",IF(AB166=Q166,"High","")),"")')
    worksheet.write('AC167', '=IF(AB167<>"",IF(S167=AB167,"Low",IF(AB167=Q167,"High","")),"")')
    worksheet.write('AC168', '=IF(AB168<>"",IF(S168=AB168,"Low",IF(AB168=Q168,"High","")),"")')
    worksheet.write('AC169', '=IF(AB169<>"",IF(S169=AB169,"Low",IF(AB169=Q169,"High","")),"")')
    worksheet.write('AC170', '=IF(AB170<>"",IF(S170=AB170,"Low",IF(AB170=Q170,"High","")),"")')
    worksheet.write('AC171', '=IF(AB171<>"",IF(S171=AB171,"Low",IF(AB171=Q171,"High","")),"")')
    worksheet.write('AC172', '=IF(AB172<>"",IF(S172=AB172,"Low",IF(AB172=Q172,"High","")),"")')
    worksheet.write('AC173', '=IF(AB173<>"",IF(S173=AB173,"Low",IF(AB173=Q173,"High","")),"")')
    worksheet.write('AC174', '=IF(AB174<>"",IF(S174=AB174,"Low",IF(AB174=Q174,"High","")),"")')
    worksheet.write('AC175', '=IF(AB175<>"",IF(S175=AB175,"Low",IF(AB175=Q175,"High","")),"")')
    worksheet.write('AC176', '=IF(AB176<>"",IF(S176=AB176,"Low",IF(AB176=Q176,"High","")),"")')
    worksheet.write('AC177', '=IF(AB177<>"",IF(S177=AB177,"Low",IF(AB177=Q177,"High","")),"")')
    worksheet.write('AC178', '=IF(AB178<>"",IF(S178=AB178,"Low",IF(AB178=Q178,"High","")),"")')
    worksheet.write('AC179', '=IF(AB179<>"",IF(S179=AB179,"Low",IF(AB179=Q179,"High","")),"")')
    worksheet.write('AC180', '=IF(AB180<>"",IF(S180=AB180,"Low",IF(AB180=Q180,"High","")),"")')
    worksheet.write('AC181', '=IF(AB181<>"",IF(S181=AB181,"Low",IF(AB181=Q181,"High","")),"")')
    worksheet.write('AC182', '=IF(AB182<>"",IF(S182=AB182,"Low",IF(AB182=Q182,"High","")),"")')
    worksheet.write('AC183', '=IF(AB183<>"",IF(S183=AB183,"Low",IF(AB183=Q183,"High","")),"")')
    worksheet.write('AC184', '=IF(AB184<>"",IF(S184=AB184,"Low",IF(AB184=Q184,"High","")),"")')
    worksheet.write('AC185', '=IF(AB185<>"",IF(S185=AB185,"Low",IF(AB185=Q185,"High","")),"")')
    worksheet.write('AC186', '=IF(AB186<>"",IF(S186=AB186,"Low",IF(AB186=Q186,"High","")),"")')
    worksheet.write('AC187', '=IF(AB187<>"",IF(S187=AB187,"Low",IF(AB187=Q187,"High","")),"")')
    worksheet.write('AC188', '=IF(AB188<>"",IF(S188=AB188,"Low",IF(AB188=Q188,"High","")),"")')
    worksheet.write('AC189', '=IF(AB189<>"",IF(S189=AB189,"Low",IF(AB189=Q189,"High","")),"")')
    worksheet.write('AC190', '=IF(AB190<>"",IF(S190=AB190,"Low",IF(AB190=Q190,"High","")),"")')
    worksheet.write('AC191', '=IF(AB191<>"",IF(S191=AB191,"Low",IF(AB191=Q191,"High","")),"")')
    worksheet.write('AC192', '=IF(AB192<>"",IF(S192=AB192,"Low",IF(AB192=Q192,"High","")),"")')
    worksheet.write('AC193', '=IF(AB193<>"",IF(S193=AB193,"Low",IF(AB193=Q193,"High","")),"")')
    worksheet.write('AC194', '=IF(AB194<>"",IF(S194=AB194,"Low",IF(AB194=Q194,"High","")),"")')
    worksheet.write('AC195', '=IF(AB195<>"",IF(S195=AB195,"Low",IF(AB195=Q195,"High","")),"")')
    worksheet.write('AC196', '=IF(AB196<>"",IF(S196=AB196,"Low",IF(AB196=Q196,"High","")),"")')
    worksheet.write('AC197', '=IF(AB197<>"",IF(S197=AB197,"Low",IF(AB197=Q197,"High","")),"")')
    worksheet.write('AC198', '=IF(AB198<>"",IF(S198=AB198,"Low",IF(AB198=Q198,"High","")),"")')
    worksheet.write('AC199', '=IF(AB199<>"",IF(S199=AB199,"Low",IF(AB199=Q199,"High","")),"")')
    worksheet.write('AC200', '=IF(AB200<>"",IF(S200=AB200,"Low",IF(AB200=Q200,"High","")),"")')
    worksheet.write('AC201', '=IF(AB201<>"",IF(S201=AB201,"Low",IF(AB201=Q201,"High","")),"")')
    worksheet.write('AC202', '=IF(AB202<>"",IF(S202=AB202,"Low",IF(AB202=Q202,"High","")),"")')
    worksheet.write('AC203', '=IF(AB203<>"",IF(S203=AB203,"Low",IF(AB203=Q203,"High","")),"")')
    worksheet.write('AC204', '=IF(AB204<>"",IF(S204=AB204,"Low",IF(AB204=Q204,"High","")),"")')
    worksheet.write('AC205', '=IF(AB205<>"",IF(S205=AB205,"Low",IF(AB205=Q205,"High","")),"")')
    worksheet.write('AC206', '=IF(AB206<>"",IF(S206=AB206,"Low",IF(AB206=Q206,"High","")),"")')
    worksheet.write('AC207', '=IF(AB207<>"",IF(S207=AB207,"Low",IF(AB207=Q207,"High","")),"")')
    worksheet.write('AC208', '=IF(AB208<>"",IF(S208=AB208,"Low",IF(AB208=Q208,"High","")),"")')
    worksheet.write('AC209', '=IF(AB209<>"",IF(S209=AB209,"Low",IF(AB209=Q209,"High","")),"")')
    worksheet.write('AC210', '=IF(AB210<>"",IF(S210=AB210,"Low",IF(AB210=Q210,"High","")),"")')
    worksheet.write('AC211', '=IF(AB211<>"",IF(S211=AB211,"Low",IF(AB211=Q211,"High","")),"")')
    worksheet.write('AC212', '=IF(AB212<>"",IF(S212=AB212,"Low",IF(AB212=Q212,"High","")),"")')
    worksheet.write('AC213', '=IF(AB213<>"",IF(S213=AB213,"Low",IF(AB213=Q213,"High","")),"")')
    worksheet.write('AC214', '=IF(AB214<>"",IF(S214=AB214,"Low",IF(AB214=Q214,"High","")),"")')
    worksheet.write('AC215', '=IF(AB215<>"",IF(S215=AB215,"Low",IF(AB215=Q215,"High","")),"")')
    worksheet.write('AC216', '=IF(AB216<>"",IF(S216=AB216,"Low",IF(AB216=Q216,"High","")),"")')
    worksheet.write('AC217', '=IF(AB217<>"",IF(S217=AB217,"Low",IF(AB217=Q217,"High","")),"")')
    worksheet.write('AC218', '=IF(AB218<>"",IF(S218=AB218,"Low",IF(AB218=Q218,"High","")),"")')
    worksheet.write('AC219', '=IF(AB219<>"",IF(S219=AB219,"Low",IF(AB219=Q219,"High","")),"")')
    worksheet.write('AC220', '=IF(AB220<>"",IF(S220=AB220,"Low",IF(AB220=Q220,"High","")),"")')
    worksheet.write('AC221', '=IF(AB221<>"",IF(S221=AB221,"Low",IF(AB221=Q221,"High","")),"")')
    worksheet.write('AC222', '=IF(AB222<>"",IF(S222=AB222,"Low",IF(AB222=Q222,"High","")),"")')
    worksheet.write('AC223', '=IF(AB223<>"",IF(S223=AB223,"Low",IF(AB223=Q223,"High","")),"")')
    worksheet.write('AC224', '=IF(AB224<>"",IF(S224=AB224,"Low",IF(AB224=Q224,"High","")),"")')
    worksheet.write('AC225', '=IF(AB225<>"",IF(S225=AB225,"Low",IF(AB225=Q225,"High","")),"")')
    worksheet.write('AC226', '=IF(AB226<>"",IF(S226=AB226,"Low",IF(AB226=Q226,"High","")),"")')
    worksheet.write('AC227', '=IF(AB227<>"",IF(S227=AB227,"Low",IF(AB227=Q227,"High","")),"")')
    worksheet.write('AC228', '=IF(AB228<>"",IF(S228=AB228,"Low",IF(AB228=Q228,"High","")),"")')
    worksheet.write('AC229', '=IF(AB229<>"",IF(S229=AB229,"Low",IF(AB229=Q229,"High","")),"")')
    worksheet.write('AC230', '=IF(AB230<>"",IF(S230=AB230,"Low",IF(AB230=Q230,"High","")),"")')
    worksheet.write('AC231', '=IF(AB231<>"",IF(S231=AB231,"Low",IF(AB231=Q231,"High","")),"")')
    worksheet.write('AC232', '=IF(AB232<>"",IF(S232=AB232,"Low",IF(AB232=Q232,"High","")),"")')
    worksheet.write('AC233', '=IF(AB233<>"",IF(S233=AB233,"Low",IF(AB233=Q233,"High","")),"")')
    worksheet.write('AC234', '=IF(AB234<>"",IF(S234=AB234,"Low",IF(AB234=Q234,"High","")),"")')
    worksheet.write('AC235', '=IF(AB235<>"",IF(S235=AB235,"Low",IF(AB235=Q235,"High","")),"")')
    worksheet.write('AC236', '=IF(AB236<>"",IF(S236=AB236,"Low",IF(AB236=Q236,"High","")),"")')
    worksheet.write('AC237', '=IF(AB237<>"",IF(S237=AB237,"Low",IF(AB237=Q237,"High","")),"")')
    worksheet.write('AC238', '=IF(AB238<>"",IF(S238=AB238,"Low",IF(AB238=Q238,"High","")),"")')
    worksheet.write('AC239', '=IF(AB239<>"",IF(S239=AB239,"Low",IF(AB239=Q239,"High","")),"")')
    worksheet.write('AC240', '=IF(AB240<>"",IF(S240=AB240,"Low",IF(AB240=Q240,"High","")),"")')
    worksheet.write('AC241', '=IF(AB241<>"",IF(S241=AB241,"Low",IF(AB241=Q241,"High","")),"")')
    worksheet.write('AC242', '=IF(AB242<>"",IF(S242=AB242,"Low",IF(AB242=Q242,"High","")),"")')
    worksheet.write('AC243', '=IF(AB243<>"",IF(S243=AB243,"Low",IF(AB243=Q243,"High","")),"")')
    worksheet.write('AC244', '=IF(AB244<>"",IF(S244=AB244,"Low",IF(AB244=Q244,"High","")),"")')
    worksheet.write('AC245', '=IF(AB245<>"",IF(S245=AB245,"Low",IF(AB245=Q245,"High","")),"")')
    worksheet.write('AC246', '=IF(AB246<>"",IF(S246=AB246,"Low",IF(AB246=Q246,"High","")),"")')
    worksheet.write('AC247', '=IF(AB247<>"",IF(S247=AB247,"Low",IF(AB247=Q247,"High","")),"")')
    worksheet.write('AC248', '=IF(AB248<>"",IF(S248=AB248,"Low",IF(AB248=Q248,"High","")),"")')
    worksheet.write('AC249', '=IF(AB249<>"",IF(S249=AB249,"Low",IF(AB249=Q249,"High","")),"")')
    worksheet.write('AC250', '=IF(AB250<>"",IF(S250=AB250,"Low",IF(AB250=Q250,"High","")),"")')
    worksheet.write('AC251', '=IF(AB251<>"",IF(S251=AB251,"Low",IF(AB251=Q251,"High","")),"")')
    worksheet.write('AC252', '=IF(AB252<>"",IF(S252=AB252,"Low",IF(AB252=Q252,"High","")),"")')
    worksheet.write('AC253', '=IF(AB253<>"",IF(S253=AB253,"Low",IF(AB253=Q253,"High","")),"")')
    worksheet.write('AC254', '=IF(AB254<>"",IF(S254=AB254,"Low",IF(AB254=Q254,"High","")),"")')
    worksheet.write('AC255', '=IF(AB255<>"",IF(S255=AB255,"Low",IF(AB255=Q255,"High","")),"")')
    worksheet.write('AC256', '=IF(AB256<>"",IF(S256=AB256,"Low",IF(AB256=Q256,"High","")),"")')
    worksheet.write('AC257', '=IF(AB257<>"",IF(S257=AB257,"Low",IF(AB257=Q257,"High","")),"")')
    worksheet.write('AC258', '=IF(AB258<>"",IF(S258=AB258,"Low",IF(AB258=Q258,"High","")),"")')
    worksheet.write('AC259', '=IF(AB259<>"",IF(S259=AB259,"Low",IF(AB259=Q259,"High","")),"")')
    worksheet.write('AC260', '=IF(AB260<>"",IF(S260=AB260,"Low",IF(AB260=Q260,"High","")),"")')
    worksheet.write('AC261', '=IF(AB261<>"",IF(S261=AB261,"Low",IF(AB261=Q261,"High","")),"")')
    worksheet.write('AC262', '=IF(AB262<>"",IF(S262=AB262,"Low",IF(AB262=Q262,"High","")),"")')
    worksheet.write('AC263', '=IF(AB263<>"",IF(S263=AB263,"Low",IF(AB263=Q263,"High","")),"")')
    worksheet.write('AC264', '=IF(AB264<>"",IF(S264=AB264,"Low",IF(AB264=Q264,"High","")),"")')
    worksheet.write('AC265', '=IF(AB265<>"",IF(S265=AB265,"Low",IF(AB265=Q265,"High","")),"")')
    worksheet.write('AC266', '=IF(AB266<>"",IF(S266=AB266,"Low",IF(AB266=Q266,"High","")),"")')
    worksheet.write('AC267', '=IF(AB267<>"",IF(S267=AB267,"Low",IF(AB267=Q267,"High","")),"")')
    worksheet.write('AC268', '=IF(AB268<>"",IF(S268=AB268,"Low",IF(AB268=Q268,"High","")),"")')
    worksheet.write('AC269', '=IF(AB269<>"",IF(S269=AB269,"Low",IF(AB269=Q269,"High","")),"")')
    worksheet.write('AC270', '=IF(AB270<>"",IF(S270=AB270,"Low",IF(AB270=Q270,"High","")),"")')
    worksheet.write('AC271', '=IF(AB271<>"",IF(S271=AB271,"Low",IF(AB271=Q271,"High","")),"")')
    worksheet.write('AC272', '=IF(AB272<>"",IF(S272=AB272,"Low",IF(AB272=Q272,"High","")),"")')
    worksheet.write('AC273', '=IF(AB273<>"",IF(S273=AB273,"Low",IF(AB273=Q273,"High","")),"")')
    worksheet.write('AC274', '=IF(AB274<>"",IF(S274=AB274,"Low",IF(AB274=Q274,"High","")),"")')
    worksheet.write('AC275', '=IF(AB275<>"",IF(S275=AB275,"Low",IF(AB275=Q275,"High","")),"")')
    worksheet.write('AC276', '=IF(AB276<>"",IF(S276=AB276,"Low",IF(AB276=Q276,"High","")),"")')
    worksheet.write('AC277', '=IF(AB277<>"",IF(S277=AB277,"Low",IF(AB277=Q277,"High","")),"")')
    worksheet.write('AC278', '=IF(AB278<>"",IF(S278=AB278,"Low",IF(AB278=Q278,"High","")),"")')
    worksheet.write('AC279', '=IF(AB279<>"",IF(S279=AB279,"Low",IF(AB279=Q279,"High","")),"")')
    worksheet.write('AC280', '=IF(AB280<>"",IF(S280=AB280,"Low",IF(AB280=Q280,"High","")),"")')
    worksheet.write('AC281', '=IF(AB281<>"",IF(S281=AB281,"Low",IF(AB281=Q281,"High","")),"")')
    worksheet.write('AC282', '=IF(AB282<>"",IF(S282=AB282,"Low",IF(AB282=Q282,"High","")),"")')
    worksheet.write('AC283', '=IF(AB283<>"",IF(S283=AB283,"Low",IF(AB283=Q283,"High","")),"")')
    worksheet.write('AC284', '=IF(AB284<>"",IF(S284=AB284,"Low",IF(AB284=Q284,"High","")),"")')
    worksheet.write('AC285', '=IF(AB285<>"",IF(S285=AB285,"Low",IF(AB285=Q285,"High","")),"")')
    worksheet.write('AC286', '=IF(AB286<>"",IF(S286=AB286,"Low",IF(AB286=Q286,"High","")),"")')
    worksheet.write('AC287', '=IF(AB287<>"",IF(S287=AB287,"Low",IF(AB287=Q287,"High","")),"")')
    worksheet.write('AC288', '=IF(AB288<>"",IF(S288=AB288,"Low",IF(AB288=Q288,"High","")),"")')
    worksheet.write('AC289', '=IF(AB289<>"",IF(S289=AB289,"Low",IF(AB289=Q289,"High","")),"")')
    worksheet.write('AC290', '=IF(AB290<>"",IF(S290=AB290,"Low",IF(AB290=Q290,"High","")),"")')
    worksheet.write('AC291', '=IF(AB291<>"",IF(S291=AB291,"Low",IF(AB291=Q291,"High","")),"")')
    worksheet.write('AC292', '=IF(AB292<>"",IF(S292=AB292,"Low",IF(AB292=Q292,"High","")),"")')
    worksheet.write('AC293', '=IF(AB293<>"",IF(S293=AB293,"Low",IF(AB293=Q293,"High","")),"")')
    worksheet.write('AC294', '=IF(AB294<>"",IF(S294=AB294,"Low",IF(AB294=Q294,"High","")),"")')
    worksheet.write('AC295', '=IF(AB295<>"",IF(S295=AB295,"Low",IF(AB295=Q295,"High","")),"")')
    worksheet.write('AC296', '=IF(AB296<>"",IF(S296=AB296,"Low",IF(AB296=Q296,"High","")),"")')
    worksheet.write('AC297', '=IF(AB297<>"",IF(S297=AB297,"Low",IF(AB297=Q297,"High","")),"")')
    worksheet.write('AC298', '=IF(AB298<>"",IF(S298=AB298,"Low",IF(AB298=Q298,"High","")),"")')
    worksheet.write('AC299', '=IF(AB299<>"",IF(S299=AB299,"Low",IF(AB299=Q299,"High","")),"")')
    worksheet.write('AC300', '=IF(AB300<>"",IF(S300=AB300,"Low",IF(AB300=Q300,"High","")),"")')
    worksheet.write('AC301', '=IF(AB301<>"",IF(S301=AB301,"Low",IF(AB301=Q301,"High","")),"")')
    worksheet.write('AC302', '=IF(AB302<>"",IF(S302=AB302,"Low",IF(AB302=Q302,"High","")),"")')
    worksheet.write('AC303', '=IF(AB303<>"",IF(S303=AB303,"Low",IF(AB303=Q303,"High","")),"")')
    worksheet.write('AC304', '=IF(AB304<>"",IF(S304=AB304,"Low",IF(AB304=Q304,"High","")),"")')
    worksheet.write('AC305', '=IF(AB305<>"",IF(S305=AB305,"Low",IF(AB305=Q305,"High","")),"")')
    worksheet.write('AC306', '=IF(AB306<>"",IF(S306=AB306,"Low",IF(AB306=Q306,"High","")),"")')
    worksheet.write('AC307', '=IF(AB307<>"",IF(S307=AB307,"Low",IF(AB307=Q307,"High","")),"")')
    worksheet.write('AC308', '=IF(AB308<>"",IF(S308=AB308,"Low",IF(AB308=Q308,"High","")),"")')
    worksheet.write('AC309', '=IF(AB309<>"",IF(S309=AB309,"Low",IF(AB309=Q309,"High","")),"")')
    worksheet.write('AC310', '=IF(AB310<>"",IF(S310=AB310,"Low",IF(AB310=Q310,"High","")),"")')
    worksheet.write('AC311', '=IF(AB311<>"",IF(S311=AB311,"Low",IF(AB311=Q311,"High","")),"")')
    worksheet.write('AC312', '=IF(AB312<>"",IF(S312=AB312,"Low",IF(AB312=Q312,"High","")),"")')
    worksheet.write('AC313', '=IF(AB313<>"",IF(S313=AB313,"Low",IF(AB313=Q313,"High","")),"")')
    worksheet.write('AC314', '=IF(AB314<>"",IF(S314=AB314,"Low",IF(AB314=Q314,"High","")),"")')
    worksheet.write('AC315', '=IF(AB315<>"",IF(S315=AB315,"Low",IF(AB315=Q315,"High","")),"")')
    worksheet.write('AC316', '=IF(AB316<>"",IF(S316=AB316,"Low",IF(AB316=Q316,"High","")),"")')
    worksheet.write('AC317', '=IF(AB317<>"",IF(S317=AB317,"Low",IF(AB317=Q317,"High","")),"")')
    worksheet.write('AC318', '=IF(AB318<>"",IF(S318=AB318,"Low",IF(AB318=Q318,"High","")),"")')
    worksheet.write('AC319', '=IF(AB319<>"",IF(S319=AB319,"Low",IF(AB319=Q319,"High","")),"")')
    worksheet.write('AC320', '=IF(AB320<>"",IF(S320=AB320,"Low",IF(AB320=Q320,"High","")),"")')
    worksheet.write('AC321', '=IF(AB321<>"",IF(S321=AB321,"Low",IF(AB321=Q321,"High","")),"")')
    worksheet.write('AC322', '=IF(AB322<>"",IF(S322=AB322,"Low",IF(AB322=Q322,"High","")),"")')
    worksheet.write('AC323', '=IF(AB323<>"",IF(S323=AB323,"Low",IF(AB323=Q323,"High","")),"")')
    worksheet.write('AC324', '=IF(AB324<>"",IF(S324=AB324,"Low",IF(AB324=Q324,"High","")),"")')
    worksheet.write('AC325', '=IF(AB325<>"",IF(S325=AB325,"Low",IF(AB325=Q325,"High","")),"")')
    worksheet.write('AC326', '=IF(AB326<>"",IF(S326=AB326,"Low",IF(AB326=Q326,"High","")),"")')
    worksheet.write('AC327', '=IF(AB327<>"",IF(S327=AB327,"Low",IF(AB327=Q327,"High","")),"")')
    worksheet.write('AC328', '=IF(AB328<>"",IF(S328=AB328,"Low",IF(AB328=Q328,"High","")),"")')
    worksheet.write('AC329', '=IF(AB329<>"",IF(S329=AB329,"Low",IF(AB329=Q329,"High","")),"")')
    worksheet.write('AC330', '=IF(AB330<>"",IF(S330=AB330,"Low",IF(AB330=Q330,"High","")),"")')
    worksheet.write('AC331', '=IF(AB331<>"",IF(S331=AB331,"Low",IF(AB331=Q331,"High","")),"")')
    worksheet.write('AC332', '=IF(AB332<>"",IF(S332=AB332,"Low",IF(AB332=Q332,"High","")),"")')
    worksheet.write('AC333', '=IF(AB333<>"",IF(S333=AB333,"Low",IF(AB333=Q333,"High","")),"")')
    worksheet.write('AC334', '=IF(AB334<>"",IF(S334=AB334,"Low",IF(AB334=Q334,"High","")),"")')
    worksheet.write('AC335', '=IF(AB335<>"",IF(S335=AB335,"Low",IF(AB335=Q335,"High","")),"")')
    worksheet.write('AC336', '=IF(AB336<>"",IF(S336=AB336,"Low",IF(AB336=Q336,"High","")),"")')
    worksheet.write('AC337', '=IF(AB337<>"",IF(S337=AB337,"Low",IF(AB337=Q337,"High","")),"")')
    worksheet.write('AC338', '=IF(AB338<>"",IF(S338=AB338,"Low",IF(AB338=Q338,"High","")),"")')
    worksheet.write('AC339', '=IF(AB339<>"",IF(S339=AB339,"Low",IF(AB339=Q339,"High","")),"")')
    worksheet.write('AC340', '=IF(AB340<>"",IF(S340=AB340,"Low",IF(AB340=Q340,"High","")),"")')
    worksheet.write('AC341', '=IF(AB341<>"",IF(S341=AB341,"Low",IF(AB341=Q341,"High","")),"")')
    worksheet.write('AC342', '=IF(AB342<>"",IF(S342=AB342,"Low",IF(AB342=Q342,"High","")),"")')
    worksheet.write('AC343', '=IF(AB343<>"",IF(S343=AB343,"Low",IF(AB343=Q343,"High","")),"")')
    worksheet.write('AC344', '=IF(AB344<>"",IF(S344=AB344,"Low",IF(AB344=Q344,"High","")),"")')
    worksheet.write('AC345', '=IF(AB345<>"",IF(S345=AB345,"Low",IF(AB345=Q345,"High","")),"")')
    worksheet.write('AC346', '=IF(AB346<>"",IF(S346=AB346,"Low",IF(AB346=Q346,"High","")),"")')
    worksheet.write('AC347', '=IF(AB347<>"",IF(S347=AB347,"Low",IF(AB347=Q347,"High","")),"")')
    worksheet.write('AC348', '=IF(AB348<>"",IF(S348=AB348,"Low",IF(AB348=Q348,"High","")),"")')
    worksheet.write('AC349', '=IF(AB349<>"",IF(S349=AB349,"Low",IF(AB349=Q349,"High","")),"")')
    worksheet.write('AC350', '=IF(AB350<>"",IF(S350=AB350,"Low",IF(AB350=Q350,"High","")),"")')
    worksheet.write('AC351', '=IF(AB351<>"",IF(S351=AB351,"Low",IF(AB351=Q351,"High","")),"")')
    worksheet.write('AC352', '=IF(AB352<>"",IF(S352=AB352,"Low",IF(AB352=Q352,"High","")),"")')
    worksheet.write('AC353', '=IF(S353=AB353,"Low",IF(AB353=Q353,"High",""))')

    #END Financial Statement Statistical Analysis -----------------------------------------------------------------------------------------------------------------------

    #SUPERTREE ------------------------------------------------------------------------------------------------------------------------------------------------------

    worksheet.write('B450', '="ROIC Super Tree"')
    worksheet.write('D476', '=C144')
    worksheet.write('E476', '=D144')
    worksheet.write('F476', '=E144')
    worksheet.write('G476', '=F144')
    worksheet.write('H476', '=G144')
    worksheet.write('D477', '=J467*(1-J487)')
    worksheet.write('E477', '=K467*(1-K487)')
    worksheet.write('F477', '=L467*(1-L487)')
    worksheet.write('G477', '=M467*(1-M487)')
    worksheet.write('H477', '=N467*(1-N487)')
    worksheet.write('J466', '=D476')
    worksheet.write('K466', '=E476')
    worksheet.write('L466', '=F476')
    worksheet.write('M466', '=G476')
    worksheet.write('N466', '=H476')
    worksheet.write('J467', '=Q462*(1/Q490)')
    worksheet.write('K467', '=R462*(1/R490)')
    worksheet.write('L467', '=S462*(1/S490)')
    worksheet.write('M467', '=T462*(1/T490)')
    worksheet.write('N467', '=U462*(1/U490)')
    worksheet.write('K476', '=RIGHT(D476,2) & "-" & RIGHT(E476,2)')
    worksheet.write('L476', '=RIGHT(E476,2) & "-" & RIGHT(F476,2)')
    worksheet.write('M476', '=RIGHT(F476,2) & "-" & RIGHT(G476,2)')
    worksheet.write('N476', '=RIGHT(G476,2) & "-" & RIGHT(H476,2)')
    worksheet.write('K477', '=E477-D477')
    worksheet.write('L477', '=F477-E477')
    worksheet.write('M477', '=G477-F477')
    worksheet.write('N477', '=H477-G477')
    worksheet.write('J486', '=D476')
    worksheet.write('K486', '=E476')
    worksheet.write('L486', '=F476')
    worksheet.write('M486', '=G476')
    worksheet.write('N486', '=H476')
    worksheet.write('J487',
                    '=(INDIRECT("J" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('K487',
                    '=(INDIRECT("K" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('L487',
                    '=(INDIRECT("L" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('M487',
                    '=(INDIRECT("M" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('N487',
                    '=(INDIRECT("N" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('Q461', '=D476')
    worksheet.write('R461', '=E476')
    worksheet.write('S461', '=F476')
    worksheet.write('T461', '=G476')
    worksheet.write('U461', '=H476')
    worksheet.write('Q462', '=X455-X463-X471')
    worksheet.write('R462', '=Y455-Y463-Y471')
    worksheet.write('S462', '=Z455-Z463-Z471')
    worksheet.write('T462', '=AA455-AA463-AA471')
    worksheet.write('U462', '=AB455-AB463-AB471')
    worksheet.write('Q471', '=K476')
    worksheet.write('R471', '=L476')
    worksheet.write('S471', '=M476')
    worksheet.write('T471', '=N476')
    worksheet.write('Q472', '=K467-J467')
    worksheet.write('R472', '=L467-K467')
    worksheet.write('S472', '=M467-L467')
    worksheet.write('T472', '=N467-M467')
    worksheet.write('Q480', '=K476')
    worksheet.write('R480', '=L476')
    worksheet.write('S480', '=M476')
    worksheet.write('T480', '=N476')
    worksheet.write('Q481', '=K487-J487')
    worksheet.write('R481', '=L487-K487')
    worksheet.write('S481', '=M487-L487')
    worksheet.write('T481', '=N487-M487')
    worksheet.write('Q489', '=D476')
    worksheet.write('R489', '=E476')
    worksheet.write('S489', '=F476')
    worksheet.write('T489', '=G476')
    worksheet.write('U489', '=H476')
    worksheet.write('Q490', '=SUM(X483,X491,X499)')
    worksheet.write('R490', '=SUM(Y483,Y491,Y499)')
    worksheet.write('S490', '=SUM(Z483,Z491,Z499)')
    worksheet.write('T490', '=SUM(AA483,AA491,AA499)')
    worksheet.write('U490', '=SUM(AB483,AB491,AB499)')
    worksheet.write('AE462', '=K476')
    worksheet.write('AF462', '=L476')
    worksheet.write('AG462', '=M476')
    worksheet.write('AH462', '=N476')
    worksheet.write('AE463', '=R462-Q462')
    worksheet.write('AF463', '=S462-R462')
    worksheet.write('AG463', '=T462-S462')
    worksheet.write('AH463', '=U462-T462')
    worksheet.write('AE491', '=K476')
    worksheet.write('AF491', '=L476')
    worksheet.write('AG491', '=M476')
    worksheet.write('AH491', '=N476')
    worksheet.write('AE492', '=R490-Q490')
    worksheet.write('AF492', '=S490-R490')
    worksheet.write('AG492', '=T490-S490')
    worksheet.write('AH492', '=U490-T490')
    worksheet.write('X454', '=D476')
    worksheet.write('Y454', '=E476')
    worksheet.write('Z454', '=F476')
    worksheet.write('AA454', '=G476')


    # BUG FIX
    #=IFERROR((INDIRECT("N" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144)), (1 - (INDIRECT("N" & MATCH("Cost of Goods Sold (COGS) incl. D&A",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))))
    # Uses 1 - Cost of Goods Sold / Sales if Gross Income is unavailable, and Operating Income / Sales otherwise

    worksheet.write('AB454', '=H476')
    worksheet.write('X455',
                    '=IFERROR((INDIRECT("J" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("J" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("J" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('Y455',
                    '=IFERROR((INDIRECT("K" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("K" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("K" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('Z455',
                    '=IFERROR((INDIRECT("L" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("L" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("L" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('AA455',
                    '=IFERROR((INDIRECT("M" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("M" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("M" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('AB455',
                    '=IFERROR((INDIRECT("N" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("N" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("N" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('X462', '=D476')
    worksheet.write('Y462', '=E476')
    worksheet.write('Z462', '=F476')
    worksheet.write('AA462', '=G476')
    worksheet.write('AB462', '=H476')
    worksheet.write('X463',
                    '=(INDIRECT("J" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y463',
                    '=(INDIRECT("K" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z463',
                    '=(INDIRECT("L" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA463',
                    '=(INDIRECT("M" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB463',
                    '=(INDIRECT("N" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('X470', '=D476')
    worksheet.write('Y470', '=E476')
    worksheet.write('Z470', '=F476')
    worksheet.write('AA470', '=G476')
    worksheet.write('AB470', '=H476')
    worksheet.write('X471',
                    '=(INDIRECT("J" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y471',
                    '=(INDIRECT("K" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z471',
                    '=(INDIRECT("L" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA471',
                    '=(INDIRECT("M" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB471',
                    '=(INDIRECT("N" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('X482', '=D476')
    worksheet.write('Y482', '=E476')
    worksheet.write('Z482', '=F476')
    worksheet.write('AA482', '=G476')
    worksheet.write('AB482', '=H476')
    worksheet.write('X483',
                    '=(INDIRECT("J" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("J" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y483',
                    '=(INDIRECT("K" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("K" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z483',
                    '=(INDIRECT("L" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("L" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA483',
                    '=(INDIRECT("M" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("M" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB483',
                    '=(INDIRECT("N" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("N" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('X490', '=D476')
    worksheet.write('Y490', '=E476')
    worksheet.write('Z490', '=F476')
    worksheet.write('AA490', '=G476')
    worksheet.write('AB490', '=H476')
    worksheet.write('X491',
                    '=(INDIRECT("J" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y491',
                    '=(INDIRECT("K" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z491',
                    '=(INDIRECT("L" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA491',
                    '=(INDIRECT("M" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB491',
                    '=(INDIRECT("N" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('X498', '=D476')
    worksheet.write('Y498', '=E476')
    worksheet.write('Z498', '=F476')
    worksheet.write('AA498', '=G476')
    worksheet.write('AB498', '=H476')
    worksheet.write('X499',
                    '=(INDIRECT("J" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y499',
                    '=(INDIRECT("K" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z499',
                    '=(INDIRECT("L" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA499',
                    '=(INDIRECT("M" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB499',
                    '=(INDIRECT("N" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AK453', '=K476')
    worksheet.write('AL453', '=L476')
    worksheet.write('AM453', '=M476')
    worksheet.write('AN453', '=N476')
    worksheet.write('AK454', '=Y455-X455')
    worksheet.write('AL454', '=Z455-Y455')
    worksheet.write('AM454', '=AA455-Z455')
    worksheet.write('AN454', '=AB455-AA455')
    worksheet.write('AK462', '=K476')
    worksheet.write('AL462', '=L476')
    worksheet.write('AM462', '=M476')
    worksheet.write('AN462', '=N476')
    worksheet.write('AK463', '=Y463-X463')
    worksheet.write('AL463', '=Z463-Y463')
    worksheet.write('AM463', '=AA463-Z463')
    worksheet.write('AN463', '=AB463-AA463')
    worksheet.write('AK473', '=K476')
    worksheet.write('AL473', '=L476')
    worksheet.write('AM473', '=M476')
    worksheet.write('AN473', '=N476')
    worksheet.write('AK474', '=Y471-X471')
    worksheet.write('AL474', '=Z471-Y471')
    worksheet.write('AM474', '=AA471-Z471')
    worksheet.write('AN474', '=AB471-AA471')
    worksheet.write('AK483', '=K476')
    worksheet.write('AL483', '=L476')
    worksheet.write('AM483', '=M476')
    worksheet.write('AN483', '=N476')
    worksheet.write('AK484', '=Y483-X483')
    worksheet.write('AL484', '=Z483-Y483')
    worksheet.write('AM484', '=AA483-Z483')
    worksheet.write('AN484', '=AB483-AA483')
    worksheet.write('AK491', '=K476')
    worksheet.write('AL491', '=L476')
    worksheet.write('AM491', '=M476')
    worksheet.write('AN491', '=N476')
    worksheet.write('AK492', '=Y491-X491')
    worksheet.write('AL492', '=Z491-Y491')
    worksheet.write('AM492', '=AA491-Z491')
    worksheet.write('AN492', '=AB491-AA491')
    worksheet.write('AK499', '=K476')
    worksheet.write('AL499', '=L476')
    worksheet.write('AM499', '=M476')
    worksheet.write('AN499', '=N476')
    worksheet.write('AK500', '=Y499-X499')
    worksheet.write('AL500', '=Z499-Y499')
    worksheet.write('AM500', '=AA499-Z499')
    worksheet.write('AN500', '=AB499-AA499')
    worksheet.write('D475', '="EOY ROIC"')
    worksheet.write('J465', '="EOY Pretax ROIC"')
    worksheet.write('K475', '="Change in EOY ROIC"')
    worksheet.write('J485', '="Cash Tax Rate"')
    worksheet.write('Q460', '="Operating Margin"')
    worksheet.write('Q470', '="Change in EOY Pretax ROIC"')
    worksheet.write('Q479', '="Change in Cash Tax Rate"')
    worksheet.write('Q488', '="Invested Capital / Sales"')
    worksheet.write('X453', '="Gross Margin"')
    worksheet.write('X461', '="SGA / Sales"')
    worksheet.write('X469', '="Depreciation / Sales"')
    worksheet.write('X481', '="Op WC / Sales"')
    worksheet.write('X489', '="PPE / Sales"')
    worksheet.write('X497', '="Intangibles / Sales"')
    worksheet.write('AE461', '="Change in Operating Margin"')
    worksheet.write('AE490', '="Change in Invested Capital / Sales"')
    worksheet.write('AK452', '="Change in Gross Margin / Sales"')
    worksheet.write('AK461', '="Change in SGA / Sales"')
    worksheet.write('AK472', '="Change in Depreciation / Sales"')
    worksheet.write('AK482', '="Change in Op WC / Sales"')
    worksheet.write('AK490', '="Change in PPE / Sales"')
    worksheet.write('AK498', '="Change in Intagibles / Sales"')

    #Min/Max

    worksheet.write('D478', '="Max " & D475')
    worksheet.write('D479', '=D478 & " Year"')
    worksheet.write('D480', '="Min " & D475')
    worksheet.write('D481', '=D480 & " Year"')
    worksheet.write('E478', '=MAX(D477:H477)')
    worksheet.write('E479', '=VALUE(D476)+MATCH(E478,D477:H477,0)-1')
    worksheet.write('E480', '=MIN(D477:H477)')
    worksheet.write('E481', '=VALUE(D476)+MATCH(E480,D477:H477,0)-1')
    worksheet.write('J468', '="Max " & J465')
    worksheet.write('J469', '=J468 & " Year"')
    worksheet.write('J470', '="Min " & J465')
    worksheet.write('J471', '=J470 & " Year"')
    worksheet.write('K468', '=MAX(J467:N467)')
    worksheet.write('K469', '=VALUE(J466)+MATCH(K468,J467:N467,0)-1')
    worksheet.write('K470', '=MIN(J467:N467)')
    worksheet.write('K471', '=VALUE(J466)+MATCH(K470,J467:N467,0)-1')
    worksheet.write('J488', '="Max " & J485')
    worksheet.write('J489', '=J488 & " Year"')
    worksheet.write('J490', '="Min " & J485')
    worksheet.write('J491', '=J490 & " Year"')
    worksheet.write('K488', '=MAX(J487:N487)')
    worksheet.write('K489', '=VALUE(J486)+MATCH(K488,J487:N487,0)-1')
    worksheet.write('K490', '=MIN(J487:N487)')
    worksheet.write('K491', '=VALUE(J486)+MATCH(K490,J487:N487,0)-1')
    worksheet.write('K478', '="Max " & K475')
    worksheet.write('K479', '=K478 & " Year"')
    worksheet.write('K480', '="Min " & K475')
    worksheet.write('K481', '=K480 & " Year"')
    worksheet.write('L478', '=MAX(K477:N477)')
    worksheet.write('L479',
                    '=IF(MATCH(L478,K477:N477,0)=1,K476,IF(MATCH(L478,K477:N477,0)=2,L476,IF(MATCH(L478,K477:N477,0)=3,M476,N476)))')
    worksheet.write('L480', '=MIN(K477:N477)')
    worksheet.write('L481',
                    '=IF(MATCH(L480,K477:N477,0)=1,K476,IF(MATCH(L480,K477:N477,0)=2,L476,IF(MATCH(L480,K477:N477,0)=3,M476,N476)))')
    worksheet.write('Q463', '="Max " & Q460')
    worksheet.write('Q464', '=Q463 & " Year"')
    worksheet.write('Q465', '="Min " & Q460')
    worksheet.write('Q466', '=Q465 & " Year"')
    worksheet.write('R463', '=MAX(Q462:U462)')
    worksheet.write('R464', '=VALUE(Q461)+MATCH(R463,Q462:U462,0)-1')
    worksheet.write('R465', '=MIN(Q462:U462)')
    worksheet.write('R466', '=VALUE(Q461)+MATCH(R465,Q462:U462,0)-1')
    worksheet.write('Q473', '="Max " & Q470')
    worksheet.write('Q474', '=Q473 & " Year"')
    worksheet.write('Q475', '="Min " & Q470')
    worksheet.write('Q476', '=Q475 & " Year"')
    worksheet.write('R473', '=MAX(Q472:T472)')
    worksheet.write('R474',
                    '=IF(MATCH(R473,Q472:T472,0)=1,Q471,IF(MATCH(R473,Q472:T472,0)=2,R471,IF(MATCH(R473,Q472:T472,0)=3,S471,T471)))')
    worksheet.write('R475', '=MIN(Q472:T472)')
    worksheet.write('R476',
                    '=IF(MATCH(R475,Q472:T472,0)=1,Q471,IF(MATCH(R475,Q472:T472,0)=2,R471,IF(MATCH(R475,Q472:T472,0)=3,S471,T471)))')
    worksheet.write('Q482', '="Max " & Q479')
    worksheet.write('Q483', '=Q482 & " Year"')
    worksheet.write('Q484', '="Min " & Q479')
    worksheet.write('Q485', '=Q484 & " Year"')
    worksheet.write('R482', '=MAX(Q481:T481)')
    worksheet.write('R483',
                    '=IF(MATCH(R482,Q481:T481,0)=1,Q480,IF(MATCH(R482,Q481:T481,0)=2,R480,IF(MATCH(R482,Q481:T481,0)=3,S480,T480)))')
    worksheet.write('R484', '=MIN(Q481:T481)')
    worksheet.write('R485',
                    '=IF(MATCH(R484,Q481:T481,0)=1,Q480,IF(MATCH(R484,Q481:T481,0)=2,R480,IF(MATCH(R484,Q481:T481,0)=3,S480,T480)))')
    worksheet.write('Q491', '="Max " & Q488')
    worksheet.write('Q492', '=Q491 & " Year"')
    worksheet.write('Q493', '="Min " & Q488')
    worksheet.write('Q494', '=Q493 & " Year"')
    worksheet.write('R491', '=MAX(Q490:U490)')
    worksheet.write('R492', '=VALUE(Q489)+MATCH(R491,Q490:U490,0)-1')
    worksheet.write('R493', '=MIN(Q490:U490)')
    worksheet.write('R494', '=VALUE(Q489)+MATCH(R493,Q490:U490,0)-1')
    worksheet.write('X456', '="Max " & X453')
    worksheet.write('X457', '=X456 & " Year"')
    worksheet.write('X458', '="Min " & X453')
    worksheet.write('X459', '=X458 & " Year"')
    worksheet.write('Y456', '=MAX(X455:AB455)')
    worksheet.write('Y457', '=VALUE(X454)+MATCH(Y456,X455:AB455,0)-1')
    worksheet.write('Y458', '=MIN(X455:AB455)')
    worksheet.write('Y459', '=VALUE(X454)+MATCH(Y458,X455:AB455,0)-1')
    worksheet.write('X464', '="Max " & X461')
    worksheet.write('X465', '=X464 & " Year"')
    worksheet.write('X466', '="Min " & X461')
    worksheet.write('X467', '=X466 & " Year"')
    worksheet.write('Y464', '=MAX(X463:AB463)')
    worksheet.write('Y465', '=VALUE(X462)+MATCH(Y464,X463:AB463,0)-1')
    worksheet.write('Y466', '=MIN(X463:AB463)')
    worksheet.write('Y467', '=VALUE(X462)+MATCH(Y466,X463:AB463,0)-1')
    worksheet.write('X472', '="Max " & X469')
    worksheet.write('X473', '=X472 & " Year"')
    worksheet.write('X474', '="Min " & X469')
    worksheet.write('X475', '=X474 & " Year"')
    worksheet.write('Y472', '=MAX(X471:AB471)')
    worksheet.write('Y473', '=VALUE(X470)+MATCH(Y472,X471:AB471,0)-1')
    worksheet.write('Y474', '=MIN(X471:AB471)')
    worksheet.write('Y475', '=VALUE(X470)+MATCH(Y474,X471:AB471,0)-1')
    worksheet.write('X484', '="Max " & X481')
    worksheet.write('X485', '=X484 & " Year"')
    worksheet.write('X486', '="Min " & X481')
    worksheet.write('X487', '=X486 & " Year"')
    worksheet.write('Y484', '=MAX(X483:AB483)')
    worksheet.write('Y485', '=VALUE(X482)+MATCH(Y484,X483:AB483,0)-1')
    worksheet.write('Y486', '=MIN(X483:AB483)')
    worksheet.write('Y487', '=VALUE(X482)+MATCH(Y486,X483:AB483,0)-1')
    worksheet.write('X492', '="Max " & X489')
    worksheet.write('X493', '=X492 & " Year"')
    worksheet.write('X494', '="Min " & X489')
    worksheet.write('X495', '=X494 & " Year"')
    worksheet.write('Y492', '=MAX(X491:AB491)')
    worksheet.write('Y493', '=VALUE(X490)+MATCH(Y492,X491:AB491,0)-1')
    worksheet.write('Y494', '=MIN(X491:AB491)')
    worksheet.write('Y495', '=VALUE(X490)+MATCH(Y494,X491:AB491,0)-1')
    worksheet.write('X500', '="Max " & X497')
    worksheet.write('X501', '=X500 & " Year"')
    worksheet.write('X502', '="Min " & X497')
    worksheet.write('X503', '=X502 & " Year"')
    worksheet.write('Y500', '=MAX(X499:AB499)')
    worksheet.write('Y501', '=VALUE(X498)+MATCH(Y500,X499:AB499,0)-1')
    worksheet.write('Y502', '=MIN(X499:AB499)')
    worksheet.write('Y503', '=VALUE(X498)+MATCH(Y502,X499:AB499,0)-1')
    worksheet.write('AE464', '="Max " & AE461')
    worksheet.write('AE465', '=AE464 & " Year"')
    worksheet.write('AE466', '="Min " & AE461')
    worksheet.write('AE467', '=AE466 & " Year"')
    worksheet.write('AF464', '=MAX(AE463:AH463)')
    worksheet.write('AF465',
                    '=IF(MATCH(AF464,AE463:AH463,0)=1,AE462,IF(MATCH(AF464,AE463:AH463,0)=2,AF462,IF(MATCH(AF464,AE463:AH463,0)=3,AG462,AH462)))')
    worksheet.write('AF466', '=MIN(AE463:AH463)')
    worksheet.write('AF467',
                    '=IF(MATCH(AF466,AE463:AH463,0)=1,AE462,IF(MATCH(AF466,AE463:AH463,0)=2,AF462,IF(MATCH(AF466,AE463:AH463,0)=3,AG462,AH462)))')
    worksheet.write('AE493', '="Max " & AE490')
    worksheet.write('AE494', '=AE493 & " Year"')
    worksheet.write('AE495', '="Min " & AE490')
    worksheet.write('AE496', '=AE495 & " Year"')
    worksheet.write('AF493', '=MAX(AE492:AH492)')
    worksheet.write('AF494',
                    '=IF(MATCH(AF493,AE492:AH492,0)=1,AE491,IF(MATCH(AF493,AE492:AH492,0)=2,AF491,IF(MATCH(AF493,AE492:AH492,0)=3,AG491,AH491)))')
    worksheet.write('AF495', '=MIN(AE492:AH492)')
    worksheet.write('AF496',
                    '=IF(MATCH(AF495,AE492:AH492,0)=1,AE491,IF(MATCH(AF495,AE492:AH492,0)=2,AF491,IF(MATCH(AF495,AE492:AH492,0)=3,AG491,AH491)))')
    worksheet.write('AK455', '="Max " & AK452')
    worksheet.write('AK456', '=AK455 & " Year"')
    worksheet.write('AK457', '="Min " & AK452')
    worksheet.write('AK458', '=AK457 & " Year"')
    worksheet.write('AL455', '=MAX(AK454:AN454)')
    worksheet.write('AL456',
                    '=IF(MATCH(AL455,AK454:AN454,0)=1,AK453,IF(MATCH(AL455,AK454:AN454,0)=2,AL453,IF(MATCH(AL455,AK454:AN454,0)=3,AM453,AN453)))')
    worksheet.write('AL457', '=MIN(AK454:AN454)')
    worksheet.write('AL458',
                    '=IF(MATCH(AL457,AK454:AN454,0)=1,AK453,IF(MATCH(AL457,AK454:AN454,0)=2,AL453,IF(MATCH(AL457,AK454:AN454,0)=3,AM453,AN453)))')
    worksheet.write('AK464', '="Max " & AK461')
    worksheet.write('AK465', '=AK464 & " Year"')
    worksheet.write('AK466', '="Min " & AK461')
    worksheet.write('AK467', '=AK466 & " Year"')
    worksheet.write('AL464', '=MAX(AK463:AN463)')
    worksheet.write('AL465',
                    '=IF(MATCH(AL464,AK463:AN463,0)=1,AK462,IF(MATCH(AL464,AK463:AN463,0)=2,AL462,IF(MATCH(AL464,AK463:AN463,0)=3,AM462,AN462)))')
    worksheet.write('AL466', '=MIN(AK463:AN463)')
    worksheet.write('AL467',
                    '=IF(MATCH(AL466,AK463:AN463,0)=1,AK462,IF(MATCH(AL466,AK463:AN463,0)=2,AL462,IF(MATCH(AL466,AK463:AN463,0)=3,AM462,AN462)))')
    worksheet.write('AK475', '="Max " & AK472')
    worksheet.write('AK476', '=AK475 & " Year"')
    worksheet.write('AK477', '="Min " & AK472')
    worksheet.write('AK478', '=AK477 & " Year"')
    worksheet.write('AL475', '=MAX(AK474:AN474)')
    worksheet.write('AL476',
                    '=IF(MATCH(AL475,AK474:AN474,0)=1,AK473,IF(MATCH(AL475,AK474:AN474,0)=2,AL473,IF(MATCH(AL475,AK474:AN474,0)=3,AM473,AN473)))')
    worksheet.write('AL477', '=MIN(AK474:AN474)')
    worksheet.write('AL478',
                    '=IF(MATCH(AL477,AK474:AN474,0)=1,AK473,IF(MATCH(AL477,AK474:AN474,0)=2,AL473,IF(MATCH(AL477,AK474:AN474,0)=3,AM473,AN473)))')
    worksheet.write('AK485', '="Max " & AK482')
    worksheet.write('AK486', '=AK485 & " Year"')
    worksheet.write('AK487', '="Min " & AK482')
    worksheet.write('AK488', '=AK487 & " Year"')
    worksheet.write('AL485', '=MAX(AK484:AN484)')
    worksheet.write('AL486',
                    '=IF(MATCH(AL485,AK484:AN484,0)=1,AK483,IF(MATCH(AL485,AK484:AN484,0)=2,AL483,IF(MATCH(AL485,AK484:AN484,0)=3,AM483,AN483)))')
    worksheet.write('AL487', '=MIN(AK484:AN484)')
    worksheet.write('AL488',
                    '=IF(MATCH(AL487,AK484:AN484,0)=1,AK483,IF(MATCH(AL487,AK484:AN484,0)=2,AL483,IF(MATCH(AL487,AK484:AN484,0)=3,AM483,AN483)))')
    worksheet.write('AK493', '="Max " & AK490')
    worksheet.write('AK494', '=AK493 & " Year"')
    worksheet.write('AK495', '="Min " & AK490')
    worksheet.write('AK496', '=AK495 & " Year"')
    worksheet.write('AL493', '=MAX(AK492:AN492)')
    worksheet.write('AL494',
                    '=IF(MATCH(AL493,AK492:AN492,0)=1,AK491,IF(MATCH(AL493,AK492:AN492,0)=2,AL491,IF(MATCH(AL493,AK492:AN492,0)=3,AM491,AN491)))')
    worksheet.write('AL495', '=MIN(AK492:AN492)')
    worksheet.write('AL496',
                    '=IF(MATCH(AL495,AK492:AN492,0)=1,AK491,IF(MATCH(AL495,AK492:AN492,0)=2,AL491,IF(MATCH(AL495,AK492:AN492,0)=3,AM491,AN491)))')
    worksheet.write('AK501', '="Max " & AK498')
    worksheet.write('AK502', '=AK501 & " Year"')
    worksheet.write('AK503', '="Min " & AK498')
    worksheet.write('AK504', '=AK503 & " Year"')
    worksheet.write('AL501', '=MAX(AK500:AN500)')
    worksheet.write('AL502',
                    '=IF(MATCH(AL501,AK500:AN500,0)=1,AK499,IF(MATCH(AL501,AK500:AN500,0)=2,AL499,IF(MATCH(AL501,AK500:AN500,0)=3,AM499,AN499)))')
    worksheet.write('AL503', '=MIN(AK500:AN500)')
    worksheet.write('AL504',
                    '=IF(MATCH(AL503,AK500:AN500,0)=1,AK499,IF(MATCH(AL503,AK500:AN500,0)=2,AL499,IF(MATCH(AL503,AK500:AN500,0)=3,AM499,AN499)))')

    #Effect on Min/Max

    worksheet.write('F478', '="Cash Tax  Effect on Max"')
    worksheet.write('F480', '="Cash Tax  Effect on Min"')
    worksheet.write('G478', '=IF(E479=K491,"Max ROIC in same year as Min Cash Tax","Inconclusive Effect")')
    worksheet.write('G480', '=IF(E481=K489,"Min ROIC in same year as Max Cash Tax","Inconclusive Effect")')
    worksheet.write('L468', '="OM Effect on Max"')
    worksheet.write('L469', '="IC Effect on Max"')
    worksheet.write('L470', '="OM Effect on Min"')
    worksheet.write('L471', '="IC Effect on Min"')
    worksheet.write('M468', '=IF(K469=R464,"Max ROIC in same year as Max OM","Inconclusive Effect")')
    worksheet.write('M469', '=IF(K469=R494,"Max ROIC in same year as Min IC","Inconclusive Effect")')
    worksheet.write('M470', '=IF(K471=R466,"Min ROIC in same year as Min OM","Inconclusive Effect")')
    worksheet.write('M471', '=IF(K471=R492,"Min ROIC in same year as Max IC","Inconclusive Effect")')
    worksheet.write('S463', '="GM Effect on Max"')
    worksheet.write('S464', '="SGA Effect on Max"')
    worksheet.write('S465', '="Dep Effect on Max"')
    worksheet.write('S466', '="GM Effect on Min"')
    worksheet.write('S467', '="SGA Effect on Min"')
    worksheet.write('S468', '="Dep Effect on Min"')
    worksheet.write('T463', '=IF(R464=Y457,"Max OM in same year as Max GM","Inconclusive Effect")')
    worksheet.write('T464', '=IF(R464=Y467,"Max OM in same year as Min SGA","Inconclusive Effect")')
    worksheet.write('T465', '=IF(R464=Y475,"Max OM in same year as Min Depr","Inconclusive Effect")')
    worksheet.write('T466', '=IF(R466=Y459,"Min OM in same year as Min GM","Inconclusive Effect")')
    worksheet.write('T467', '=IF(R466=Y465,"Min OM in same year as Max SGA","Inconclusive Effect")')
    worksheet.write('T468', '=IF(R466=Y473,"Min OM in same year as Max Dep","Inconclusive Effect")')
    worksheet.write('S489', '=F476')
    worksheet.write('S490', '=SUM(Z483,Z491,Z499)')
    worksheet.write('S491', '="Op WC Effect on Max"')
    worksheet.write('S492', '="PPE Effect on Max"')
    worksheet.write('S493', '="Intangibles Effect on Max"')
    worksheet.write('S494', '="Op WC Effect on Min"')
    worksheet.write('S495', '="PPE Effect on Min"')
    worksheet.write('S496', '="Intangibles Effect on Min"')
    worksheet.write('T489', '=G476')
    worksheet.write('T490', '=SUM(AA483,AA491,AA499)')
    worksheet.write('T491', '=IF(R492=Y485,"Max IC in same year as Max Op WC","Inconclusive Effect")')
    worksheet.write('T492', '=IF(R492=Y493,"Max IC in same year as Max PPE","Inconclusive Effect")')
    worksheet.write('T493', '=IF(R492=Y501,"Max IC in same year as Max Intangibles","Inconclusive Effect")')
    worksheet.write('T494', '=IF(R494=Y487,"Min IC in same year as Min Op WC","Inconclusive Effect")')
    worksheet.write('T495', '=IF(R494=Y495,"Min IC in same year as Min PPE","Inconclusive Effect")')
    worksheet.write('T496', '=IF(R494=Y503,"Min IC in same year as Min Intangibles","Inconclusive Effect")')

    #Correlation

    worksheet.write('D482', '="Correlation with OM"')
    worksheet.write('D483', '="Correlation with IC"')
    worksheet.write('D484', '="Correlation with GM"')
    worksheet.write('D485', '="Correlation with SGA"')
    worksheet.write('D486', '="Correlation with Dep"')
    worksheet.write('D487', '="Correlation with Op WC"')
    worksheet.write('D488', '="Correlation with PPE"')
    worksheet.write('D489', '="Correlation with Intangibles"')
    worksheet.write('E482', '=CORREL(D477:H477,Q462:U462)')
    worksheet.write('E483', '=CORREL(D477:H477,Q490:U490)')
    worksheet.write('E484', '=CORREL(D477:H477,X455:AB455)')
    worksheet.write('E485', '=CORREL(D477:H477,X463:AB463)')
    worksheet.write('E486', '=CORREL(D477:H477,X471:AB471)')
    worksheet.write('E487', '=CORREL(D477:H477,X483:AB483)')
    worksheet.write('E488', '=CORREL(D477:H477,X491:AB491)')
    worksheet.write('E489', '=CORREL(D477:H477,X499:AB499)')
    worksheet.write('N468', '="Correlation with OM"')
    worksheet.write('N469', '="Correlation with IC"')
    worksheet.write('O468', '=CORREL(J467:N467,Q462:U462)')
    worksheet.write('O469', '=CORREL(J467:N467,Q490:U490)')
    worksheet.write('U463', '="Correlation with GM"')
    worksheet.write('U464', '="Correlation with SGA"')
    worksheet.write('U465', '="Correlation with Dep"')
    worksheet.write('V463', '=CORREL(Q462:U462,X455:AB455)')
    worksheet.write('V464', '=CORREL(Q462:U462,X463:AB463)')
    worksheet.write('V465', '=CORREL(Q462:U462,X471:AB471)')
    worksheet.write('U491', '="Correlation with Op WC"')
    worksheet.write('U492', '="Correlation with PPE"')
    worksheet.write('U493', '="Correlation with Intangibles"')
    worksheet.write('V491', '=CORREL(Q490:U490,X483:AB483)')
    worksheet.write('V492', '=CORREL(Q490:U490,X491:AB491)')
    worksheet.write('V493', '=CORREL(Q490:U490,X499:AB499)')
    worksheet.write('C508', '=D475')
    worksheet.write('C509', '=J465')
    worksheet.write('C510', '=J485')
    worksheet.write('C511', '=Q460')
    worksheet.write('C512', '=Q488')
    worksheet.write('C513', '=X453')
    worksheet.write('C514', '=X461')
    worksheet.write('C515', '=X469')
    worksheet.write('C516', '=X481')
    worksheet.write('C517', '=X489')
    worksheet.write('C518', '=X497')
    worksheet.write('D507', '=D476')
    worksheet.write('D508', '=D477')
    worksheet.write('D509', '=J467')
    worksheet.write('D510', '=J487')
    worksheet.write('D511', '=Q462')
    worksheet.write('D512', '=Q490')
    worksheet.write('D513', '=X455')
    worksheet.write('D514', '=X463')
    worksheet.write('D515', '=X471')
    worksheet.write('D516', '=X483')
    worksheet.write('D517', '=X491')
    worksheet.write('D518', '=X499')
    worksheet.write('E507', '=E476')
    worksheet.write('E508', '=E477')
    worksheet.write('E509', '=K467')
    worksheet.write('E510', '=K487')
    worksheet.write('E511', '=R462')
    worksheet.write('E512', '=R490')
    worksheet.write('E513', '=Y455')
    worksheet.write('E514', '=Y463')
    worksheet.write('E515', '=Y471')
    worksheet.write('E516', '=Y483')
    worksheet.write('E517', '=Y491')
    worksheet.write('E518', '=Y499')
    worksheet.write('F507', '=F476')
    worksheet.write('F508', '=F477')
    worksheet.write('F509', '=L467')
    worksheet.write('F510', '=L487')
    worksheet.write('F511', '=S462')
    worksheet.write('F512', '=S490')
    worksheet.write('F513', '=Z455')
    worksheet.write('F514', '=Z463')
    worksheet.write('F515', '=Z471')
    worksheet.write('F516', '=Z483')
    worksheet.write('F517', '=Z491')
    worksheet.write('F518', '=Z499')
    worksheet.write('G507', '=G476')
    worksheet.write('G508', '=G477')
    worksheet.write('G509', '=M467')
    worksheet.write('G510', '=M487')
    worksheet.write('G511', '=T462')
    worksheet.write('G512', '=T490')
    worksheet.write('G513', '=AA455')
    worksheet.write('G514', '=AA463')
    worksheet.write('G515', '=AA471')
    worksheet.write('G516', '=AA483')
    worksheet.write('G517', '=AA491')
    worksheet.write('G518', '=AA499')
    worksheet.write('H507', '=H476')
    worksheet.write('H508', '=H477')
    worksheet.write('H509', '=N467')
    worksheet.write('H510', '=N487')
    worksheet.write('H511', '=U462')
    worksheet.write('H512', '=U490')
    worksheet.write('H513', '=AB455')
    worksheet.write('H514', '=AB463')
    worksheet.write('H515', '=AB471')
    worksheet.write('H516', '=AB483')
    worksheet.write('H517', '=AB491')
    worksheet.write('H518', '=AB499')
    worksheet.write('I507', '="Average"')
    worksheet.write('I508', '=AVERAGE(D508:H508)')
    worksheet.write('I509', '=AVERAGE(D509:H509)')
    worksheet.write('I510', '=AVERAGE(D510:H510)')
    worksheet.write('I511', '=AVERAGE(D511:H511)')
    worksheet.write('I512', '=AVERAGE(D512:H512)')
    worksheet.write('I513', '=AVERAGE(D513:H513)')
    worksheet.write('I514', '=AVERAGE(D514:H514)')
    worksheet.write('I515', '=AVERAGE(D515:H515)')
    worksheet.write('I516', '=AVERAGE(D516:H516)')
    worksheet.write('I517', '=AVERAGE(D517:H517)')
    worksheet.write('I518', '=AVERAGE(D518:H518)')
    worksheet.write('I519', '="Max SD"')
    worksheet.write('I520', '="Max SD Item"')
    worksheet.write('I521', '="Min SD"')
    worksheet.write('I522', '="Min SD Item"')
    worksheet.write('J507', '="SD"')
    worksheet.write('J508', '=STDEV(D508:H508)')
    worksheet.write('J509', '=STDEV(D509:H509)')
    worksheet.write('J510', '=STDEV(D510:H510)')
    worksheet.write('J511', '=STDEV(D511:H511)')
    worksheet.write('J512', '=STDEV(D512:H512)')
    worksheet.write('J513', '=STDEV(D513:H513)')
    worksheet.write('J514', '=STDEV(D514:H514)')
    worksheet.write('J515', '=STDEV(D515:H515)')
    worksheet.write('J516', '=STDEV(D516:H516)')
    worksheet.write('J517', '=STDEV(D517:H517)')
    worksheet.write('J518', '=STDEV(D518:H518)')
    worksheet.write('J519', '=MAX(J508:J518)')
    worksheet.write('J520', '=INDIRECT("C" & 507 + MATCH(J519,J508:J518,0))')
    worksheet.write('J521', '=MIN(J508:J518)')
    worksheet.write('J522', '=INDIRECT("C" & 507 + MATCH(J521,J508:J518,0))')
    worksheet.write('K507', '=D507')
    worksheet.write('K508', '=(D508-I508)/J508')
    worksheet.write('K509', '=(D509-I509)/J509')
    worksheet.write('K510', '=(D510-I510)/J510')
    worksheet.write('K511', '=(D511-I511)/J511')
    worksheet.write('K512', '=(D512-I512)/J512')
    worksheet.write('K513', '=(D513-I513)/J513')
    worksheet.write('K514', '=(D514-I514)/J514')
    worksheet.write('K515', '=(D515-I515)/J515')
    worksheet.write('K516', '=(D516-I516)/J516')
    worksheet.write('K517', '=(D517-I517)/J517')
    worksheet.write('K518', '=(D518-I518)/J518')
    worksheet.write('L507', '=E507')
    worksheet.write('L508', '=(E508-I508)/J508')
    worksheet.write('L509', '=(E509-I509)/J509')
    worksheet.write('L510', '=(E510-I510)/J510')
    worksheet.write('L511', '=(E511-I511)/J511')
    worksheet.write('L512', '=(E512-I512)/J512')
    worksheet.write('L513', '=(E513-I513)/J513')
    worksheet.write('L514', '=(E514-I514)/J514')
    worksheet.write('L515', '=(E515-I515)/J515')
    worksheet.write('L516', '=(E516-I516)/J516')
    worksheet.write('L517', '=(E517-I517)/J517')
    worksheet.write('L518', '=(E518-I518)/J518')
    worksheet.write('M507', '=F507')
    worksheet.write('M508', '=(F508-I508)/J508')
    worksheet.write('M509', '=(F509-I509)/J509')
    worksheet.write('M510', '=(F510-I510)/J510')
    worksheet.write('M511', '=(F511-I511)/J511')
    worksheet.write('M512', '=(F512-I512)/J512')
    worksheet.write('M513', '=(F513-I513)/J513')
    worksheet.write('M514', '=(F514-I514)/J514')
    worksheet.write('M515', '=(F515-I515)/J515')
    worksheet.write('M516', '=(F516-I516)/J516')
    worksheet.write('M517', '=(F517-I517)/J517')
    worksheet.write('M518', '=(F518-I518)/J518')
    worksheet.write('N507', '=G507')
    worksheet.write('N508', '=(G508-I508)/J508')
    worksheet.write('N509', '=(G509-I509)/J509')
    worksheet.write('N510', '=(G510-I510)/J510')
    worksheet.write('N511', '=(G511-I511)/J511')
    worksheet.write('N512', '=(G512-I512)/J512')
    worksheet.write('N513', '=(G513-I513)/J513')
    worksheet.write('N514', '=(G514-I514)/J514')
    worksheet.write('N515', '=(G515-I515)/J515')
    worksheet.write('N516', '=(G516-I516)/J516')
    worksheet.write('N517', '=(G517-I517)/J517')
    worksheet.write('N518', '=(G518-I518)/J518')
    worksheet.write('O507', '=H507')
    worksheet.write('O508', '=(H508-I508)/J508')
    worksheet.write('O509', '=(H509-I509)/J509')
    worksheet.write('O510', '=(H510-I510)/J510')
    worksheet.write('O511', '=(H511-I511)/J511')
    worksheet.write('O512', '=(H512-I512)/J512')
    worksheet.write('O513', '=(H513-I513)/J513')
    worksheet.write('O514', '=(H514-I514)/J514')
    worksheet.write('O515', '=(H515-I515)/J515')
    worksheet.write('O516', '=(H516-I516)/J516')
    worksheet.write('O517', '=(H517-I517)/J517')
    worksheet.write('O518', '=(H518-I518)/J518')
    worksheet.write('P508',
                    '=K507 + MATCH(IF(MAX(MAX(K508:O508),ABS(MIN(K508:O508)))=ABS(MIN(K508:O508)), MIN(K508:O508),MAX(K508:O508)),K508:O508,0) - 1')
    worksheet.write('P509',
                    '=K507 + MATCH(IF(MAX(MAX(K509:O509),ABS(MIN(K509:O509)))=ABS(MIN(K509:O509)), MIN(K509:O509),MAX(K509:O509)),K509:O509,0) - 1')
    worksheet.write('P510',
                    '=K507 + MATCH(IF(MAX(MAX(K510:O510),ABS(MIN(K510:O510)))=ABS(MIN(K510:O510)), MIN(K510:O510),MAX(K510:O510)),K510:O510,0) - 1')
    worksheet.write('P511',
                    '=K507 + MATCH(IF(MAX(MAX(K511:O511),ABS(MIN(K511:O511)))=ABS(MIN(K511:O511)), MIN(K511:O511),MAX(K511:O511)),K511:O511,0) - 1')
    worksheet.write('P512',
                    '=K507 + MATCH(IF(MAX(MAX(K512:O512),ABS(MIN(K512:O512)))=ABS(MIN(K512:O512)), MIN(K512:O512),MAX(K512:O512)),K512:O512,0) - 1')
    worksheet.write('P513',
                    '=K507 + MATCH(IF(MAX(MAX(K513:O513),ABS(MIN(K513:O513)))=ABS(MIN(K513:O513)), MIN(K513:O513),MAX(K513:O513)),K513:O513,0) - 1')
    worksheet.write('P514',
                    '=K507 + MATCH(IF(MAX(MAX(K514:O514),ABS(MIN(K514:O514)))=ABS(MIN(K514:O514)), MIN(K514:O514),MAX(K514:O514)),K514:O514,0) - 1')
    worksheet.write('P515',
                    '=K507 + MATCH(IF(MAX(MAX(K515:O515),ABS(MIN(K515:O515)))=ABS(MIN(K515:O515)), MIN(K515:O515),MAX(K515:O515)),K515:O515,0) - 1')
    worksheet.write('P516',
                    '=K507 + MATCH(IF(MAX(MAX(K516:O516),ABS(MIN(K516:O516)))=ABS(MIN(K516:O516)), MIN(K516:O516),MAX(K516:O516)),K516:O516,0) - 1')
    worksheet.write('P517',
                    '=K507 + MATCH(IF(MAX(MAX(K517:O517),ABS(MIN(K517:O517)))=ABS(MIN(K517:O517)), MIN(K517:O517),MAX(K517:O517)),K517:O517,0) - 1')
    worksheet.write('P518',
                    '=K507 + MATCH(IF(MAX(MAX(K518:O518),ABS(MIN(K518:O518)))=ABS(MIN(K518:O518)), MIN(K518:O518),MAX(K518:O518)),K518:O518,0) - 1')

    #END SUPERTREE --------------------------------------------------------------------------------------------------------------------------------------------------


    #streamline this
    worksheet.write('K12','=D78')
    worksheet.write('K13','=D89')


    #additons
    #=IF(STDEV(J146:N146)<0.1,IF(COUNTIF(J146:N146,">0")=5,"pos_trend"),"") where J146:N146 is a range of annual data
        # checks if the data isn't deviating too much, and if it isn't, whether there is a positive trend
worksheet = workbook.get_worksheet_by_name("Output Text")
n_row = 1
for ticker in tickers:
    worksheet.write('B' + str(n_row), '=' + ticker + '!K17')
    n_row += 1

workbook.close()

writer = pd.ExcelWriter('Output.xlsx', engine='openpyxl')
wb = load_workbook('Output.xlsx')

# create pandas excel writer for dfs
writer.book = wb

# have to tell pandas that we alerady have sheets, and what they are
writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

# insert dataframes

x = -1
offset = 0 #make global
for ticker in tickers:
    offset = 6
    x += 1

    try:
        df = pd.DataFrame(data_dict[tickers[x]][2][1][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1

        df = pd.DataFrame(data_dict[tickers[x]][2][1][1])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 2
    except:
        pass

    try:
        df = pd.DataFrame(data_dict[tickers[x]][4])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][4])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][5])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][5])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][6])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][6])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][7])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][7])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][8])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][8])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][9])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][9])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][12][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][12][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][14][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][14][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][16][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][16][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][18][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][18][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][20][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][20][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][22][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][22][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][24][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][24][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][26][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][26][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][28][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][28][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][30][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][30][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][32][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][32][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][38][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][38][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][39][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][39][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][43][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][43][0])"

    #for the 2 asset tables issue
    try:
        if is_df(data_dict[tickers[x]][44][0]):
            try:
                df = pd.DataFrame(data_dict[tickers[x]][44][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][44][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][46][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][46][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][50][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][50][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][52][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][52][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][54][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][54][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][55][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][56][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][57][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][57][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][59][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][59][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][61][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][61][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][63][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][63][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][65][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][65][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][67][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][67][0])"

        else:
            try:
                df = pd.DataFrame(data_dict[tickers[x]][45][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][45][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][47][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][47][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][51][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][51][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][53][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][53][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][54][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][54][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][56][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][56][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][58][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][58][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][60][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][60][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][62][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][62][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][64][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][64][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][66][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][66][0])"
    except:
        print "index out of range"



writer.save()

print "Program took", (time.time() - start_time)/60 , "minutes to run"





