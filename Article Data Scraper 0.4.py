#02035
from bs4 import BeautifulSoup, Comment
import requests
import pandas as pd
import numpy as np
import itertools
from tabulate import tabulate
import types
import datetime
import time
import xlsxwriter
import csv
from openpyxl import load_workbook
import requests.packages.urllib3
requests.packages.urllib3.disable_warnings()


#dictionary to store all data
data_dict = {}

d_valid = True
#---------------------------- Get earnings announcements for a given date ----------------------------------------------
# date = ""
# while d_valid == True:
#     date = str(raw_input("Enter a date or type 'exit' (Format: YYYY-MM-DD): "))
#     try:
#         datetime.datetime.strptime(date, '%Y-%m-%d')
#         d_valid = True
#     except:
#         if date =='exit':
#             print "bye"
#             exit()
#         else:
#             print "Nah" + "\n"

start_time = time.time()
time_one = time.time() #----------------------------------------------------------------------------------------------------------------------

def pretty_print(table_title, table_to_print):
    df = pd.read_html(str(table_to_print))
    print table_title, "\n", (tabulate(df[0], headers='keys', tablefmt='psql'))

def replace_with_newlines(element):
    text = ''
    for elem in element.recursiveChildGenerator():
        if isinstance(elem, types.StringTypes):
            text += elem.strip()
        elif elem.name == 'br':
            text += '\n'
    return text

def removekey(d, key):
    r = dict(d)
    del r[key]
    return r

date = "2017-07-24"

url = "finance.yahoo.com/calendar/earnings?day=" + date
r  = requests.get("https://" + url)
data = r.text
soup = BeautifulSoup(data, "html5lib" )
data_table = soup.find('table', {'class':'data-table'})

headers = []
cells = []
tickers = []

#pretty_print('Earnings announcements on ' + date, data_table)
data_dict['Tickers'] = pd.read_html(str(data_table))

cols = data_table.find_all('td')
#ths = data_table.find_all('th')
#for th in ths[1::]:
#    headers.append(th.get_text())
for col in cols[1::]:
    if col != "-":
        cells.append(col.get_text())
    else:
        cells.append("NULL")

w = ''
spl = [list(y) for x, y in itertools.groupby(cells, lambda z: z == w) if not x]

cap = 0
for line in spl:
    tickers.append(line[0])
    cap += 1
    if cap > 100:
        break

print "tickers found"
#print len(tickers)
time_two = time.time() #----------------------------------------------------------------------------------------------------------------------

#-----------------------add more info for known tickers-----------------------------------------------------------------
#https://finance.yahoo.com/quote/TICKER?p=TICKER - stock page

#print "-----------------------------SUMMARY-----------------------------"

ticker_urls = []

#Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "?p=" + ticker)
    #initialize dictionary structure
    data_dict[ticker] = []

x = -1
#Run for each URL
for ticker_url in ticker_urls:
    x += 1
    r  = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    summary_table = soup.find_all('table',['class','W(100%)'])



    #Summary------------------------------------------------------------------------------------------------------------

    #Error if the url is bad
    try:

        cur_price = soup.find('span',['class','Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)'])
        # Filter out stocks under $20
        if float(cur_price.get_text()) < 20:
            #print tickers[x], "was removed from search (price below $20)"
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        # remove companies with a market cap over 10B
        market_cap = 0
        market_caps = summary_table[1].find_all('tr')
        market_cap_string = market_caps[0].get_text()[10:]

        if market_cap_string[-1:] == 'M':
            market_cap = float(market_cap_string[:-1]) * 1000000
        elif market_cap_string[-1:] == 'B':
            market_cap = float(market_cap_string[:-1]) * 1000000000

        if market_cap > 10000000000:
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        #print "Company: " + c_name.get_text()
        #print "Current Price: " + cur_price.get_text()
        #pretty_print("Summary", summary_table)
        data_dict[tickers[x]].extend((c_name.get_text(),cur_price.get_text(),("Summary",pd.read_html(str(summary_table)))))

    except:
        #print "Stock Quote Unavailable"
        #print tickers[x], "removed from search"
        data_dict = removekey(data_dict, tickers[x])
        tickers.remove(tickers[x])
        x -= 1


print "summary complete"
#print len(tickers)

time_three = time.time() #----------------------------------------------------------------------------------------------------------------------

#print "-----------------------------ANALYSTS-----------------------------"
#Analysts------------------------------------------------------------------------------------------------------------

ticker_urls = []

# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/analysts?p=" + ticker)
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        #print c_name.get_text()

        data_dict[tickers[x]].append(c_name.get_text())
        analyst_est_sect = soup.find('section', id='quote-leaf-comp')
        analyst_tables = analyst_est_sect.find_all('table', ['class', 'W(100%)'])
        if len(analyst_tables) == 0:
            #print tickers[x], "was removed (no analyst estimates available)"
            data_dict = removekey(data_dict, tickers[x])
            tickers.remove(tickers[x])
            x -= 1
            continue
        for a_table in analyst_tables:
            #pretty_print("",a_table)
            data_dict[tickers[x]].extend(pd.read_html(str(a_table)))

            # here is an alternative if the  4 lines above break
            # all_tables = soup.find_all('table', class_='W(100%)')
            # for table in all_tables:
            #     pretty_print("This Table", table

    except:
        data_dict = removekey(data_dict, tickers[x])
        #print tickers[x], "removed (analyst estimates unavailable)"
        tickers.remove(tickers[x])
        x -= 1

print "analysts complete"
#print len(tickers)

#print "-----------------------------STATISTICS-----------------------------"
#Statistics---------------------------------------------------------------------------------------------------------
ticker_urls = []
stat_heads = []
# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/key-statistics?p=" + ticker)

z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    z += 1 #serves as x
    x = 0
    try:
        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        stat_tables = soup.find_all('table', ['class', 'table-qsp-stats Mt(10px)'])

        #print c_name.get_text()
        data_dict[tickers[z]].append(c_name.get_text())
        stat_headers = soup.find_all(True, {'class':['Pt(20px)','Pt(6px) Pstart(20px)','Fz(s) Mt(20px)']})
        for stat_header in stat_headers:
            stat_heads.append(stat_header.get_text())


        for stat_table in stat_tables:
            if stat_heads[x] == "Financial Highlights" or stat_heads[x] == "Trading Information":
                #pretty_print(stat_heads[x] + "\n" + "\n" + stat_heads[x+1],stat_table)
                data_dict[tickers[z]].extend((stat_heads[x] + "\n" + "\n" + stat_heads[x+1], pd.read_html(str(stat_table))))
                x += 1
            else:
                #pretty_print(stat_heads[x], stat_table)
                data_dict[tickers[z]].extend((stat_heads[x], pd.read_html(str(stat_table))))
            x += 1
    except:
        #print "Stock Stats Unavailable"
        pass

print "statistics complete"
time_four = time.time() #----------------------------------------------------------------------------------------------------------------------
#print "-----------------------------PROFILE-----------------------------"
#Profile------------------------------------------------------------------------------------------------------------
ticker_urls = []

# Stock Quote Page
for ticker in tickers:
    ticker_urls.append("https://finance.yahoo.com/quote/" + ticker + "/profile?p=" + ticker)

profile_info = []
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # strip comments
        comments = soup.find_all(text=lambda text: isinstance(text, Comment))
        [comment.extract() for comment in comments]

        # Company Name
        c_name = soup.find('h1', ['class', 'D(ib)'])
        geo_con_web = replace_with_newlines(soup.find('p', ['class', 'D(ib) W(47.727%) Pend(40px)']))
        sec_ind_emp = replace_with_newlines(soup.find('p', ['class', 'D(ib) Va(t)']))
        execs_table = soup.find_all('table', ['class', 'W(100%)'])
        #execs = execs_table[0].findChildren()
        descript = replace_with_newlines(soup.find('p',['class','Mt(15px) Lh(1.6)']))

        #print c_name.get_text()
        #pretty_print("Executives", execs_table)
        #print geo_con_web.splitlines()
        #print sec_ind_emp.splitlines()
        #print "Description: " + descript
        data_dict[tickers[x]].extend((c_name.get_text(), pd.read_html(str(execs_table)), geo_con_web.splitlines(),sec_ind_emp.splitlines(),("Description", descript)))
    except:
        #print "Stock Profile Unavailable"
        pass

print "profile complete"

time_five = time.time() #----------------------------------------------------------------------------------------------------------------------
#print "-----------------------------FINANCIALS-----------------------------"
# Financials (marketwatch.com)----------------------------------------------------------------------------------------
ticker_urls = []

# INCOME STATEMENT-----------------------------------------------------
for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials")
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    x += 1
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Income Statement"
        data_dict[tickers[x]].extend((c_name.get_text(),"Income Statement"))
        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            #pretty_print("", fin_table)
            data_dict[tickers[x]].append(pd.read_html(str(fin_table)))
    except:
        #print "Stock Financials Unavailable"
        pass


print "income statements complete"

time_six = time.time()  # ----------------------------------------------------------------------------------------------------------------------
# BALANCE SHEET ----------------------------------------------------
ticker_urls = []

for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials/balance-sheet")
z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    fin_heads = []
    x = 0
    y = 0
    z += 1 #serving as x
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Balance Sheet"
        data_dict[tickers[z]].extend((c_name.get_text(), "Balance Sheet"))
        fin_headers = soup.find_all('h2')
        for fin_header in fin_headers:
            fin_heads.append(fin_header.get_text())
        # only want the 2nd and 3rd
        fin_heads.pop(0)

        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            if len(fin_tables) == 3:
                if x == 0 or x == 2:
                    #pretty_print(fin_heads[y], fin_table)
                    data_dict[tickers[z]].extend((fin_heads[y], pd.read_html(str(fin_table))))
                    y += 1
                else:
                    #pretty_print("", fin_table)
                    data_dict[tickers[z]].append(pd.read_html(str(fin_table)))
            elif len(fin_tables) == 2:
                #pretty_print(fin_heads[x], fin_table)
                data_dict[tickers[z]].extend((fin_heads[x], pd.read_html(str(fin_table))))
            else:
                #pretty_print("", fin_table)
                data_dict[tickers[z]].append(pd.read_html(str(fin_table)))
            x += 1
    except:
        pass
        #print "Stock Financials Unavailable"

print "balance sheets complete"

time_seven = time.time()  # ----------------------------------------------------------------------------------------------------------------------
# CASH FLOW STATEMENT -----------------------------------------------

ticker_urls = []

for ticker in tickers:
    ticker_urls.append("http://www.marketwatch.com/investing/stock/" + ticker + "/financials/cash-flow")
z = -1
# Run for each URL
for ticker_url in ticker_urls:
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    fin_heads = []
    x = 0
    y = 0
    z += 1 #serving as x
    try:
        # Company Name
        c_name = soup.find('h1', id='instrumentname')
        #print c_name.get_text()
        #print "Cash Flow Statement"
        data_dict[tickers[z]].extend((c_name.get_text(), "Cash Flow Statement"))
        fin_headers = soup.find_all('h2')
        for fin_header in fin_headers:
            fin_heads.append(fin_header.get_text())
        # only want the 2nd and 3rd
        fin_heads.pop(0)

        fin_tables = soup.find_all('table', ['class', 'crDataTable'])
        for fin_table in fin_tables:
            #pretty_print(fin_heads[x], fin_table)
            data_dict[tickers[z]].extend((fin_heads[x], pd.read_html(str(fin_table))))
            x += 1

    except:
        #print "Stock Financials Unavailable"
        pass

print "cash flow statements complete"

time_eight = time.time()  # ----------------------------------------------------------------------------------------------------------------------

#Historical Data ----- (potentially download) ----------------------------------------------------------------------
#print  "-----------------------------HISTORICAL DATA (download file)-----------------------------"
#print "To be completed"

time_nine = time.time() #----------------------------------------------------------------------------------------------------------------------

print time_two - time_one, time_three - time_two, time_four - time_three, time_five - time_four, time_six - time_five, time_seven - time_six, time_eight - time_seven, time_nine - time_eight
#print data_dict



# Find comps (ycharts.com)------------------------------------------------------------------------------------------------------------------------------------
ticker_urls = []
for ticker in tickers:
    ticker_urls.append("http://www.ycharts.com/companies/" + ticker)

comp_dict = {}
comps = []
x = -1
# Run for each URL
for ticker_url in ticker_urls:
    x +=1
    r = requests.get(ticker_url)
    data = r.text
    soup = BeautifulSoup(data, "html5lib")
    comp_urls = []
    comps = []
    try:
        comp_table = soup.find('table', ['class', 'relCompSect'])
        table = pd.read_html(str(comp_table))
        data_dict[tickers[x]].append(table)
        #print table[0][0] prints the first column (comp tickers)
        #print table[0][1] prints the second column (comp names)

        #this next part doesn't work
        for tick_num in range(0,4):
            comp_urls.append("https://finance.yahoo.com/quote/" + table[0][0][tick_num] + "/key-statistics?p=" + table[0][0][tick_num])
        w = -1
        for comp_url in comp_urls:
            w += 1
            r = requests.get(comp_url)
            data = r.text
            soup = BeautifulSoup(data, "html5lib")
            try:
                # Company Name
                c_name = soup.find('h1', ['class', 'D(ib)'])
                stat_tables = soup.find_all('table', ['class', 'table-qsp-stats Mt(10px)'])
                data_dict[tickers[x]].extend((c_name.get_text(), pd.read_html(str(stat_tables[0]))))
                comps.append(c_name.get_text())
                #left off
            except:
                print "comp info not found"


        comp_dict[tickers[x]] = comps
    except:
        print "no comps found"
    #data_dict[tickers[x]].append(comps)
    #print comps

print "comps complete"

print data_dict
def is_df(var):
    if isinstance(var, pd.DataFrame):
        return True
    else:
        return False

#Write to Excel ------------------------------------------------------------



#create excel workbook
workbook = xlsxwriter.Workbook('Output.xlsx')

#Create output sheet
workbook.add_worksheet("Output Text")

#add formats
format_dict = {}

x = -1
# create one sheet for each ticker
for ticker in tickers:
    x += 1
    workbook.add_worksheet(ticker)
    worksheet = workbook.get_worksheet_by_name(ticker)
    # write to each sheet
    worksheet.write('B1', "Company Name")
    worksheet.write('C1', data_dict[tickers[x]][0])
    worksheet.write('B2', "Current Price")
    worksheet.write('C2', data_dict[tickers[x]][1])
    worksheet.write('B4',data_dict[tickers[x]][2][0])
    #worksheet.write('E1', pd.DataFrame(data_dict[tickers[x]][2][1]))

    #worksheet.write('F1', '="Comps"')
    #try:
    #    for comp in range(0,len(comp_dict[ticker[x]])):
    #        worksheet.write('F' + str(comp+2),comp_dict[ticker[x]][comp])
    #except:
    #    pass


    #adjust column width
    worksheet.set_column('A:A', 0)
    worksheet.set_column('B:G', 20)
    worksheet.write('D12','=LEFT(C12,FIND("-",C12)-2)')
    worksheet.write('E12','=TRIM(RIGHT(C12,FIND("-",C12)-1))')
    worksheet.write('K2', '=LEFT(C1,FIND("(",C1) - 2)')
    worksheet.write('D1','=TRIM(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(IFERROR(REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("PLC",UPPER(LEFT(C1,FIND("(",C1) - 2))),3,"" ), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND(", INC.",UPPER(LEFT(C1,FIND("(",C1) - 2))),6,"" )), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORPORATION",UPPER(LEFT(C1,FIND("(",C1) - 2))),11,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORP.",UPPER(LEFT(C1,FIND("(",C1) - 2))),5,"" )), REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("COMPANY",UPPER(LEFT(C1,FIND("(",C1) - 2))),7,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("INC.",UPPER(LEFT(C1,FIND("(",C1) - 2))),4,"" )),REPLACE(LEFT(C1,FIND("(",C1) - 2), FIND("CORP",UPPER(LEFT(C1,FIND("(",C1) - 2))),4,"" )),LEFT(C1,FIND("(",C1) - 2)))')

    worksheet.write('K3', '=" is scheduled to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('L3', '=" is slated to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('M3', '=" will report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('N3', '=" reports earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('O3', '=" plans to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')
    worksheet.write('P3', '=" is going to report earnings "&IFERROR("between "&LEFT(C20,FIND("-",C20)-2)&" and "&RIGHT(C20,FIND("-",C20)-2),"on "&C20)')

    worksheet.write('K4', '="The stock is currently trading at " & TEXT(C2,"$####.00") & ", " & IF(C2-C7=0, "at the same price" & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s close")), IF(C2-C7>0, "up " & TEXT((C7-C2)/C7*-1,"##.##%") & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s close")), "down " & TEXT((C7-C2)/C7*1,"##.##%") & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s close", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "slightly", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s close", IF((C7-C8)/C7 <0.01, "slightly below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s close")) ))')
    worksheet.write('L4', '="The current stock price is " & TEXT(C2,"$####.00") & ", " & IF(C2-C7=0, "at the same price" & " after opening " & IF(C8-C7=0, "at the same price as yesterday\'s closing price", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "a fraction", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s closing price", IF((C7-C8)/C7 <0.01, "a bit below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s closing price")), IF(C2-C7>0, "up " & TEXT((C7-C2)/C7*-1,"##.##%") & " after opening " & IF(C8-C7=0, "at yesterday\'s closing price", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "a fraction", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s closing price", IF((C7-C8)/C7 <0.01, "a bit below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s closing price")), "down " & TEXT((C7-C2)/C7*1,"##.##%") & " after opening " & IF(C8-C7=0, "at yesterday\'s closing price", IF(C8-C7>0, "up " & IF((C7-C8)/C7*-1 <0.01, "a fraction", TEXT((C7-C8)/C7*-1,"##.##%")) & " over yesterday\'s closing price", IF((C7-C8)/C7 <0.01, "a bit below", "down " & TEXT((C7-C8)/C7*1,"##.##%") & " from") & " yesterday\'s closing price")) ))')

    worksheet.write('K5', '="The one year target estimate for " & D1 & " is " & TEXT(C23,"$####.00")')
    worksheet.write('L5', '=D1 &" has a one year target estimate of " & TEXT(C23,"$####.00")')
    worksheet.write('L5', '=D1 & " is expected to be trading at " & TEXT(C23, "$####.00") & ", based on target estimates"')

    worksheet.write('K6', '=" which would be " & IF(OR(LEFT(ABS((C23-C2)/C2*100),1)="8",LEFT(ABS((C23-C2)/C2*100),2)="18"), "an ", "a ")  &TEXT(ABS((C23-C2)/C2),"####.00%")&IF((C23-C2)>0," increase over"," decrease from")&" the current price"')
    worksheet.write('K7', '="Earnings are expected to " & IF(C28=D28, "remain constant over the next quarter", IF( D28>C28,  "increase by " & TEXT((D28-C28)/C28*100,"##.##") & "% over last quarter", "decrease by " & TEXT((D28-C28)/C28*-100,"##.##") & "% from last quarter")) & " based on the average of " & $C$27 & " analyst estimates (Yahoo Finance)"')
    worksheet.write('K8', '=IF(VALUE(C2)=D12, "The stock is trading at an all-time low",IF(VALUE(C2) =E12,"The stock is trading at an all-time high",IF(VALUE(C2)<D12+(E12-D12)/3, "The stock is trading in the low end of its 52-week range",IF(VALUE(C2)<D12+2*(E12-D12)/3, "The stock is trading near the middle of its 52 week range", "The stock is trading in the high end of its 52-week range"))) )')

    worksheet.write('K9', '="Over the last 4 quarters, we\'ve seen a positive earnings surprise " & 4 -COUNTIF(C45:F45,"-*") & IF(4 - COUNTIF(C45:F45,"-*")=1, " time,"," times,") & " and a negative earnings surprise " & COUNTIF(C45:F45,"-*") & IF(COUNTIF(C45:F45,"-*")=1, " time", " times")')
    worksheet.write('K9', '="Over the last 4 quarters, there" & IF(4 - COUNTIF(C45:F45,"-*")=1, " has"," have") & " been" & IF(4 - COUNTIF(C45:F45,"-*")=1, " a,","") & " positive earnings surprise" & IF(4 - COUNTIF(C45:F45,"-*")=1, " ","s ") & 4 -COUNTIF(C45:F45,"-*") & IF(4 - COUNTIF(C45:F45,"-*")=1, " time,"," times,") & " and a negative earnings surprise " & COUNTIF(C45:F45,"-*") & IF(COUNTIF(C45:F45,"-*")=1, " time", " times")')

    worksheet.write('K10','=IF(F48=F52,"",IF(F48>F52, "EPS estimates have increased by " & TEXT(F48-F52,"$0.00") & " in the 2 months leading up to the earnings report", "EPS estimates have decreased by " & TEXT(ABS(F48-F52),"$0.00") & " in the 2 months leading up to the earnings report"))')
    worksheet.write('K11','=IF(B145="Interest Income",U42, K42)')

    worksheet.write('J3','=RANDBETWEEN(1,6)')
    worksheet.write('J4','=RANDBETWEEN(1,2)')
    worksheet.write('J5','=RANDBETWEEN(1,2)')

    #Paragraph 1
    worksheet.write('K17','=K2 & IF(J3=1, K3,IF(J3=2,L3,IF(J3=3,M3,IF(J3=4,N3,IF(J3=5,O3,IF(J3=6,P3)))))) & ". " & IF(J4=1,K4,IF(J4=2,L4)) & ". " & IF(J5=1,K5,IF(J5=2,L5)) & K6 & ". " & K7 & ". " & K8 & ". " & K9 & "."')

    #converts financial statement text into numbers
    for y in range(60,400):
        #worksheet.write('J' + str(y), '=IF(TRIM(C' + str(y) + ')="-", "N/A", IF(RIGHT(C' + str(y) + ',1)="M",1000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="B",1000000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="%",0.01*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),C' + str(y) + '))))')
        #worksheet.write('K' + str(y), '=IF(TRIM(D' + str(y) + ')="-", "N/A", IF(RIGHT(D' + str(y) + ',1)="M",1000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="B",1000000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="%",0.01*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),D' + str(y) + '))))')
        #worksheet.write('L' + str(y), '=IF(TRIM(E' + str(y) + ')="-", "N/A", IF(RIGHT(E' + str(y) + ',1)="M",1000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="B",1000000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="%",0.01*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),E' + str(y) + '))))')
        #worksheet.write('M' + str(y), '=IF(TRIM(F' + str(y) + ')="-", "N/A", IF(RIGHT(F' + str(y) + ',1)="M",1000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="B",1000000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="%",0.01*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),F' + str(y) + '))))')
        #worksheet.write('N' + str(y), '=IF(TRIM(G' + str(y) + ')="-", "N/A", IF(RIGHT(G' + str(y) + ',1)="M",1000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="B",1000000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="%",0.01*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),G' + str(y) + '))))')

        #bug fix
        worksheet.write('J' + str(y), '=IFERROR(IF(TRIM(C' + str(y) + ')="-", "N/A", IF(RIGHT(C' + str(y) + ',1)=")",IF(RIGHT(C' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(C' + str(y) + ',2,LEN(C' + str(y) + ')-3)),IF(RIGHT(C' + str(y) + ',2)="M)",-1000000*VALUE(MID(C' + str(y) + ',2,LEN(C' + str(y) + ')-3)),IF(RIGHT(C' + str(y) + ',2)="B)",-1000000000*VALUE(MID(C' + str(y) + ',2,LEN(C' + str(y) + ')-3)),IF(RIGHT(C' + str(y) + ',2)="k)",-1000*VALUE(MID(C' + str(y) + ',2,LEN(C' + str(y) + ')-3)),VALUE(SUBSTITUTE(C' + str(y) + ',",","")))))),IF(RIGHT(C' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="M",1000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="B",1000000000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="%",0.01*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),IF(RIGHT(C' + str(y) + ',1)="k",1000*VALUE(LEFT(C' + str(y) + ',LEN(C' + str(y) + ')-1)),VALUE(SUBSTITUTE(C' + str(y) + ',",",""))))))))),"N/A")')
        worksheet.write('K' + str(y), '=IFERROR(IF(TRIM(D' + str(y) + ')="-", "N/A", IF(RIGHT(D' + str(y) + ',1)=")",IF(RIGHT(D' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(D' + str(y) + ',2,LEN(D' + str(y) + ')-3)),IF(RIGHT(D' + str(y) + ',2)="M)",-1000000*VALUE(MID(D' + str(y) + ',2,LEN(D' + str(y) + ')-3)),IF(RIGHT(D' + str(y) + ',2)="B)",-1000000000*VALUE(MID(D' + str(y) + ',2,LEN(D' + str(y) + ')-3)),IF(RIGHT(D' + str(y) + ',2)="k)",-1000*VALUE(MID(D' + str(y) + ',2,LEN(D' + str(y) + ')-3)),VALUE(SUBSTITUTE(D' + str(y) + ',",","")))))),IF(RIGHT(D' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="M",1000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="B",1000000000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="%",0.01*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),IF(RIGHT(D' + str(y) + ',1)="k",1000*VALUE(LEFT(D' + str(y) + ',LEN(D' + str(y) + ')-1)),VALUE(SUBSTITUTE(D' + str(y) + ',",",""))))))))),"N/A")')
        worksheet.write('L' + str(y), '=IFERROR(IF(TRIM(E' + str(y) + ')="-", "N/A", IF(RIGHT(E' + str(y) + ',1)=")",IF(RIGHT(E' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(E' + str(y) + ',2,LEN(E' + str(y) + ')-3)),IF(RIGHT(E' + str(y) + ',2)="M)",-1000000*VALUE(MID(E' + str(y) + ',2,LEN(E' + str(y) + ')-3)),IF(RIGHT(E' + str(y) + ',2)="B)",-1000000000*VALUE(MID(E' + str(y) + ',2,LEN(E' + str(y) + ')-3)),IF(RIGHT(E' + str(y) + ',2)="k)",-1000*VALUE(MID(E' + str(y) + ',2,LEN(E' + str(y) + ')-3)),VALUE(SUBSTITUTE(E' + str(y) + ',",","")))))),IF(RIGHT(E' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="M",1000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="B",1000000000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="%",0.01*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),IF(RIGHT(E' + str(y) + ',1)="k",1000*VALUE(LEFT(E' + str(y) + ',LEN(E' + str(y) + ')-1)),VALUE(SUBSTITUTE(E' + str(y) + ',",",""))))))))),"N/A")')
        worksheet.write('M' + str(y), '=IFERROR(IF(TRIM(F' + str(y) + ')="-", "N/A", IF(RIGHT(F' + str(y) + ',1)=")",IF(RIGHT(F' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(F' + str(y) + ',2,LEN(F' + str(y) + ')-3)),IF(RIGHT(F' + str(y) + ',2)="M)",-1000000*VALUE(MID(F' + str(y) + ',2,LEN(F' + str(y) + ')-3)),IF(RIGHT(F' + str(y) + ',2)="B)",-1000000000*VALUE(MID(F' + str(y) + ',2,LEN(F' + str(y) + ')-3)),IF(RIGHT(F' + str(y) + ',2)="k)",-1000*VALUE(MID(F' + str(y) + ',2,LEN(F' + str(y) + ')-3)),VALUE(SUBSTITUTE(F' + str(y) + ',",","")))))),IF(RIGHT(F' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="M",1000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="B",1000000000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="%",0.01*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),IF(RIGHT(F' + str(y) + ',1)="k",1000*VALUE(LEFT(F' + str(y) + ',LEN(F' + str(y) + ')-1)),VALUE(SUBSTITUTE(F' + str(y) + ',",",""))))))))),"N/A")')
        worksheet.write('N' + str(y), '=IFERROR(IF(TRIM(G' + str(y) + ')="-", "N/A", IF(RIGHT(G' + str(y) + ',1)=")",IF(RIGHT(G' + str(y) + ',2)="T)",-1000000000000*VALUE(MID(G' + str(y) + ',2,LEN(G' + str(y) + ')-3)),IF(RIGHT(G' + str(y) + ',2)="M)",-1000000*VALUE(MID(G' + str(y) + ',2,LEN(G' + str(y) + ')-3)),IF(RIGHT(G' + str(y) + ',2)="B)",-1000000000*VALUE(MID(G' + str(y) + ',2,LEN(G' + str(y) + ')-3)),IF(RIGHT(G' + str(y) + ',2)="k)",-1000*VALUE(MID(G' + str(y) + ',2,LEN(G' + str(y) + ')-3)),VALUE(SUBSTITUTE(G' + str(y) + ',",","")))))),IF(RIGHT(G' + str(y) + ',1)="T",1000000000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="M",1000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="B",1000000000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="%",0.01*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),IF(RIGHT(G' + str(y) + ',1)="k",1000*VALUE(LEFT(G' + str(y) + ',LEN(G' + str(y) + ')-1)),VALUE(SUBSTITUTE(G' + str(y) + ',",",""))))))))),"N/A")')

        worksheet.write('I' + str(y), '=IF(AND(K' + str(y) + '> J' + str(y)+ ', L' + str(y) + '> K' +str(y)+ ', M' + str(y) + '> L' +str(y)+ ', N' + str(y) +'> M' + str(y) + '), "pos_trend", IF(AND(K' + str(y) +'< J' +str(y)+ ', L' + str(y) +'< K' +str(y)+ ', M' + str(y) +'< L' +str(y)+ ', N' + str(y) + '< M' + str(y) + '), "neg_trend", "N/A"))')

    #trend analysis
    for row in range(22, 39):
        worksheet.write('J' + str(row), '=IF(K' + str(row) + ' <> "",' + str(row - 21) + ', 0)')
        worksheet.write('L' + str(row), '=IF(EXACT(K' + str(row) + ',UPPER(K' + str(row) + ')),K' + str(row) + ',LOWER(K' + str(row) + '))')
        worksheet.write('M22', '=L22')
        worksheet.write('M' + str(row + 1), '=IF(L' + str(row + 1) + '<>"", M' + str(row) + ' & ", " & L' + str(row + 1) + ',M' + str(row) + ')')
        #bank version
        worksheet.write('T' + str(row), '=IF(U' + str(row) + ' <> "",' + str(row - 21) + ', 0)')
        worksheet.write('V' + str(row), '=IF(EXACT(U' + str(row) + ',UPPER(U' + str(row) + ')),U' + str(row) + ',LOWER(U' + str(row) + '))')
        worksheet.write('W22', '=V22')
        worksheet.write('W' + str(row + 1), '=IF(V' + str(row + 1) + '<>"", W' + str(row) + ' & ", " & V' + str(row + 1) + ',W' + str(row) + ')')

    worksheet.write('K22', '=IF(I145="pos_trend","Revenue","")')
    worksheet.write('K23', '=IF(I146="pos_trend",B146,"")')
    worksheet.write('K24', '=IF(I153="pos_trend",B153,"")')
    worksheet.write('K25', '=IF(I154="pos_trend",B154,"")')
    worksheet.write('K26', '=IF(I155="pos_trend",B155,"")')
    worksheet.write('K27', '=IF(I172="pos_trend",B172,"")')
    worksheet.write('K28', '=IF(I173="pos_trend",B173,"")')
    worksheet.write('K29', '=IF(I174="pos_trend",B174,"")')
    worksheet.write('K30', '=IF(I185="pos_trend",B185,"")')
    worksheet.write('K31', '=IF(I186="pos_trend",B186,"")')
    worksheet.write('K32', '=IF(I187="pos_trend",B187,"")')
    worksheet.write('K33', '=IF(I195="pos_trend",B195,"")')
    worksheet.write('K34', '=IF(I196="pos_trend",B196,"")')
    worksheet.write('K35', '=IF(I201="pos_trend",B201,"")')
    worksheet.write('K36', '=IF(I202="pos_trend",B202,"")')
    worksheet.write('K37', '=IF(I203="pos_trend",B203,"")')
    worksheet.write('K38', '=IF(I351="pos_trend",B351,"")')
    worksheet.write('K39', '=IF(I352="pos_trend",B352,"")')
    worksheet.write('K40', '=VLOOKUP(J40,J22:K39,2,FALSE)')
    worksheet.write('K42', '=SUBSTITUTE(IF(M40<>"", D1 & " has managed to increase " & M40 & " each year since " & C144, "No positive trends")," , "," ")')

    worksheet.write('J40', '=MAX(J22:J39)')
    worksheet.write('M40', '=IF(IFERROR(FIND(",",M39),TRUE)=TRUE,M39,IF(NOT(EXACT(K40,UPPER(K40))),SUBSTITUTE(M39,LOWER(K40),"and "&LOWER(K40)),SUBSTITUTE(M39,K40,"and "&K40)))')

    #bank version
    worksheet.write('U22', '=IF(AND(B145 = "Interest Income",I145="pos_trend"), "Interest Income","")')
    worksheet.write('U23', '=IF(I151="pos_trend",B151,"")')
    worksheet.write('U24', '=IF(I161="pos_trend",B161,"")')
    worksheet.write('U25', '=IF(I162="pos_trend",B162,"")')
    worksheet.write('U26', '=IF(I167="pos_trend",B167,"")')
    worksheet.write('U27', '=IF(I170="pos_trend",B170,"")')
    worksheet.write('U28', '=IF(I171="pos_trend",B171,"")')
    worksheet.write('U29', '=IF(I172="pos_trend",B172,"")')
    worksheet.write('U30', '=IF(I178="pos_trend",B178,"")')
    worksheet.write('U31', '=IF(I199="pos_trend",B199,"")')
    worksheet.write('U32', '=IF(I209="pos_trend",B209,"")')
    worksheet.write('U33', '=IF(I231="pos_trend",B231,"")')
    worksheet.write('U34', '=IF(I251="pos_trend",B251,"")')
    worksheet.write('U35', '=IF(I279="pos_trend",B279,"")')
    worksheet.write('U36', '=IF(I336="pos_trend",B336,"")')
    worksheet.write('U37', '=IF(I235="pos_trend",B235,"")')
    worksheet.write('U38', '=IF(I236="pos_trend",B236,"")')
    worksheet.write('U40', '=VLOOKUP(T40,T22:U39,2,FALSE)')
    worksheet.write('U42', '=SUBSTITUTE(IF(W40<>"", D1 & " has managed to increase " & W40 & " each year since " & C144, "No positive trends")," , "," ")')

    worksheet.write('T40', '=MAX(T22:T39)')
    worksheet.write('W40', '=IF(IFERROR(FIND(",",W39),TRUE)=TRUE,W39,IF(NOT(EXACT(U40,UPPER(U40))),SUBSTITUTE(W39,LOWER(U40),"and "&LOWER(U40)),SUBSTITUTE(W39,U40,"and "&U40)))')

    #for row in range(348,356):
    #    worksheet.write('E' + str(row),'=TRIM(IF(ISNUMBER(VALUE(RIGHT(B'+str(row)+',1))),REPLACE(B'+str(row)+',LEN(B'+str(row)+'),1,""),B'+str(row)+'))')
    #    worksheet.write('E' + str(row),'=AVERAGE(VALUE(J'+str(row)+'),VALUE(J'+str(row+11)+'),VALUE(J'+str(row+22)+'),VALUE(J'+str(row+33)+'))')


    #comps
    for z in range(500,510):
        try:
            worksheet.write('C' + str(z), data_dict[tickers[x]][56][z-500])
        except:
            pass

        try:
            worksheet.write('C' + str(z), data_dict[tickers[x]][55][z-500])
        except:
            pass

    worksheet.write('D67','Comp Average')
    worksheet.write('D68', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B68,B69:B500,0)+68))),VALUE(INDIRECT("J"&(MATCH(B68,B69:B500,0)+79))),VALUE(INDIRECT("J"&(MATCH(B68,B69:B500,0)+90))),VALUE(INDIRECT("J"&(MATCH(B68,B69:B500,0)+101))))')
    worksheet.write('D69', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B69,B70:B501,0)+69))),VALUE(INDIRECT("J"&(MATCH(B69,B70:B501,0)+80))),VALUE(INDIRECT("J"&(MATCH(B69,B70:B501,0)+91))),VALUE(INDIRECT("J"&(MATCH(B69,B70:B501,0)+102))))')
    worksheet.write('D70', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B70,B71:B502,0)+70))),VALUE(INDIRECT("J"&(MATCH(B70,B71:B502,0)+81))),VALUE(INDIRECT("J"&(MATCH(B70,B71:B502,0)+92))),VALUE(INDIRECT("J"&(MATCH(B70,B71:B502,0)+103))))')
    worksheet.write('D71', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B71,B72:B503,0)+71))),VALUE(INDIRECT("J"&(MATCH(B71,B72:B503,0)+82))),VALUE(INDIRECT("J"&(MATCH(B71,B72:B503,0)+93))),VALUE(INDIRECT("J"&(MATCH(B71,B72:B503,0)+104))))')
    worksheet.write('D72', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B72,B73:B504,0)+72))),VALUE(INDIRECT("J"&(MATCH(B72,B73:B504,0)+83))),VALUE(INDIRECT("J"&(MATCH(B72,B73:B504,0)+94))),VALUE(INDIRECT("J"&(MATCH(B72,B73:B504,0)+105))))')
    worksheet.write('D73', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B73,B74:B505,0)+73))),VALUE(INDIRECT("J"&(MATCH(B73,B74:B505,0)+84))),VALUE(INDIRECT("J"&(MATCH(B73,B74:B505,0)+95))),VALUE(INDIRECT("J"&(MATCH(B73,B74:B505,0)+106))))')
    worksheet.write('D74', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B74,B75:B506,0)+74))),VALUE(INDIRECT("J"&(MATCH(B74,B75:B506,0)+85))),VALUE(INDIRECT("J"&(MATCH(B74,B75:B506,0)+96))),VALUE(INDIRECT("J"&(MATCH(B74,B75:B506,0)+107))))')
    worksheet.write('D75', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B75,B76:B507,0)+75))),VALUE(INDIRECT("J"&(MATCH(B75,B76:B507,0)+86))),VALUE(INDIRECT("J"&(MATCH(B75,B76:B507,0)+97))),VALUE(INDIRECT("J"&(MATCH(B75,B76:B507,0)+108))))')
    worksheet.write('D76', '=AVERAGE(VALUE(INDIRECT("J"&(MATCH(B76,B77:B508,0)+76))),VALUE(INDIRECT("J"&(MATCH(B76,B77:B508,0)+87))),VALUE(INDIRECT("J"&(MATCH(B76,B77:B508,0)+98))),VALUE(INDIRECT("J"&(MATCH(B76,B77:B508,0)+109))))')

    worksheet.write('E67', '=C1')
    worksheet.write('E68', '=IF(AND(C68<>"",D68<>0),IF(VALUE(J68)>VALUE(K68),"above average","below average"),"no data")')
    worksheet.write('E69', '=IF(AND(C69<>"",D69<>0),IF(VALUE(J69)>VALUE(K69),"above average","below average"),"no data")')
    worksheet.write('E70', '=IF(AND(C70<>"",D70<>0),IF(VALUE(J70)>VALUE(K70),"above average","below average"),"no data")')
    worksheet.write('E71', '=IF(AND(C71<>"",D71<>0),IF(VALUE(J71)>VALUE(K71),"above average","below average"),"no data")')
    worksheet.write('E72', '=IF(AND(C72<>"",D72<>0),IF(VALUE(J72)>VALUE(K72),"above average","below average"),"no data")')
    worksheet.write('E73', '=IF(AND(C73<>"",D73<>0),IF(VALUE(J73)>VALUE(K73),"above average","below average"),"no data")')
    worksheet.write('E74', '=IF(AND(C74<>"",D74<>0),IF(VALUE(J74)>VALUE(K74),"above average","below average"),"no data")')
    worksheet.write('E75', '=IF(AND(C75<>"",D75<>0),IF(VALUE(J75)>VALUE(K75),"above average","below average"),"no data")')
    worksheet.write('E76', '=IF(AND(C76<>"",D76<>0),IF(VALUE(J76)>VALUE(K76),"above average","below average"),"no data")')

    worksheet.write('F70', '=IF(E70="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B70,1))),REPLACE(B70,LEN(B70),1,""),B70))),"")')
    worksheet.write('F71', '=IF(E71="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B71,1))),REPLACE(B71,LEN(B71),1,""),B71))),"")')
    worksheet.write('F72', '=IF(E72="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B72,1))),REPLACE(B72,LEN(B72),1,""),B72))),"")')
    worksheet.write('F73', '=IF(E73="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B73,1))),REPLACE(B73,LEN(B73),1,""),B73))),"")')
    worksheet.write('F74', '=IF(E74="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B74,1))),REPLACE(B74,LEN(B74),1,""),B74))),"")')
    worksheet.write('F75', '=IF(E75="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B75,1))),REPLACE(B75,LEN(B75),1,""),B75))),"")')
    worksheet.write('F76', '=IF(E76="above average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B76,1))),REPLACE(B76,LEN(B76),1,""),B76))),"")')
    worksheet.write('F77', '=IF(F76="",IF(F75="",IF(F74="",IF(F73="",IF(F72="",IF(F71="",IFERROR(LEFT(F70,FIND("(",F70) - 2),F70),IFERROR(LEFT(F71,FIND("(",F71) - 2),F71)),IFERROR(LEFT(F72,FIND("(",F72) - 2),F72)),IFERROR(LEFT(F73,FIND("(",F73) - 2),F73)),IFERROR(LEFT(F74,FIND("(",F74) - 2),F74)),IFERROR(LEFT(F75,FIND("(",F75) - 2),F75)),IFERROR(LEFT(F76,FIND("(",F76) - 2),F76))')

    worksheet.write('G70', '=IFERROR(LEFT(F70,FIND("(",F70) - 2),F70)')
    worksheet.write('G71', '=IF(F71<>"", G70 & ", " & IFERROR(LEFT(F71,FIND("(",F71) - 2),F71),G70)')
    worksheet.write('G72', '=IF(F72<>"", G71 & ", " & IFERROR(LEFT(F72,FIND("(",F72) - 2),F72),G71)')
    worksheet.write('G73', '=IF(F73<>"", G72 & ", " & IFERROR(LEFT(F73,FIND("(",F73) - 2),F73),G72)')
    worksheet.write('G74', '=IF(F74<>"", G73 & ", " & IFERROR(LEFT(F74,FIND("(",F74) - 2),F74),G73)')
    worksheet.write('G75', '=IF(F75<>"", G74 & ", " & IFERROR(LEFT(F75,FIND("(",F75) - 2),F75),G74)')
    worksheet.write('G76', '=IF(F76<>"", G75 & ", " & IFERROR(LEFT(F76,FIND("(",F76) - 2),F76),G75)')
    worksheet.write('G77', '=TRIM(IF(LEFT(G76,1)=",",REPLACE(G76,1,1,""),SUBSTITUTE(G76,F77, "and " & F77)))')

    worksheet.write('D78', '=IF(COUNTIF(E70:E76,"=above average")>0,"There are some indications that "&D1&" may be overvalued. The company has a higher " & G77 & " than the comparable average", "Inconclusive")')

    worksheet.write('F81','=IF(E70="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B70,1))),REPLACE(B70,LEN(B70),1,""),B70))),"")')
    worksheet.write('F82','=IF(E71="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B71,1))),REPLACE(B71,LEN(B71),1,""),B71))),"")')
    worksheet.write('F83','=IF(E72="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B72,1))),REPLACE(B72,LEN(B72),1,""),B72))),"")')
    worksheet.write('F84','=IF(E73="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B73,1))),REPLACE(B73,LEN(B73),1,""),B73))),"")')
    worksheet.write('F85','=IF(E74="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B74,1))),REPLACE(B74,LEN(B74),1,""),B74))),"")')
    worksheet.write('F86','=IF(E75="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B75,1))),REPLACE(B75,LEN(B75),1,""),B75))),"")')
    worksheet.write('F87','=IF(E76="below average",LOWER(TRIM(IF(ISNUMBER(VALUE(RIGHT(B76,1))),REPLACE(B76,LEN(B76),1,""),B76))),"")')
    worksheet.write('F88','=IF(F87="",IF(F86="",IF(F85="",IF(F84="",IF(F83="",IF(F82="",IFERROR(LEFT(F81,FIND("(",F81) - 2),F81),IFERROR(LEFT(F82,FIND("(",F82) - 2),F82)),IFERROR(LEFT(F83,FIND("(",F83) - 2),F83)),IFERROR(LEFT(F84,FIND("(",F84) - 2),F84)),IFERROR(LEFT(F85,FIND("(",F85) - 2),F85)),IFERROR(LEFT(F86,FIND("(",F86) - 2),F86)),IFERROR(LEFT(F87,FIND("(",F87) - 2),F87))')

    worksheet.write('G81','=IFERROR(LEFT(F81,FIND("(",F81) - 2),F81)')
    worksheet.write('G82','=IF(F82<>"", G81 & ", " & IFERROR(LEFT(F82,FIND("(",F82) - 2),F82),G81)')
    worksheet.write('G83','=IF(F83<>"", G82 & ", " & IFERROR(LEFT(F83,FIND("(",F83) - 2),F83),G82)')
    worksheet.write('G84','=IF(F84<>"", G83 & ", " & IFERROR(LEFT(F84,FIND("(",F84) - 2),F84),G83)')
    worksheet.write('G85','=IF(F85<>"", G84 & ", " & IFERROR(LEFT(F85,FIND("(",F85) - 2),F85),G84)')
    worksheet.write('G86','=IF(F86<>"", G85 & ", " & IFERROR(LEFT(F86,FIND("(",F86) - 2),F86),G85)')
    worksheet.write('G87','=IF(F87<>"", G86 & ", " & IFERROR(LEFT(F87,FIND("(",F87) - 2),F87),G86)')
    worksheet.write('G88','=TRIM(IF(LEFT(G87,1)=",",REPLACE(G87,1,1,""),SUBSTITUTE(G87,F88, "and " & F88)))')

    worksheet.write('D89','=IF(COUNTIF(E70:E76,"=below average")>0,"There are some indications that "&D1&" may be undervalued. The company has a lower " & G88 & " than the comparable average", "Inconclusive")')


    #SUPERTREE ------------------------------------------------------------------------------------------------------------------------------------------------------

    worksheet.write('B450', '="ROIC Super Tree"')
    worksheet.write('D476', '=C144')
    worksheet.write('E476', '=D144')
    worksheet.write('F476', '=E144')
    worksheet.write('G476', '=F144')
    worksheet.write('H476', '=G144')
    worksheet.write('D477', '=J467*(1-J487)')
    worksheet.write('E477', '=K467*(1-K487)')
    worksheet.write('F477', '=L467*(1-L487)')
    worksheet.write('G477', '=M467*(1-M487)')
    worksheet.write('H477', '=N467*(1-N487)')
    worksheet.write('J466', '=D476')
    worksheet.write('K466', '=E476')
    worksheet.write('L466', '=F476')
    worksheet.write('M466', '=G476')
    worksheet.write('N466', '=H476')
    worksheet.write('J467', '=Q462*(1/Q490)')
    worksheet.write('K467', '=R462*(1/R490)')
    worksheet.write('L467', '=S462*(1/S490)')
    worksheet.write('M467', '=T462*(1/T490)')
    worksheet.write('N467', '=U462*(1/U490)')
    worksheet.write('K476', '=RIGHT(D476,2) & "-" & RIGHT(E476,2)')
    worksheet.write('L476', '=RIGHT(E476,2) & "-" & RIGHT(F476,2)')
    worksheet.write('M476', '=RIGHT(F476,2) & "-" & RIGHT(G476,2)')
    worksheet.write('N476', '=RIGHT(G476,2) & "-" & RIGHT(H476,2)')
    worksheet.write('K477', '=E477-D477')
    worksheet.write('L477', '=F477-E477')
    worksheet.write('M477', '=G477-F477')
    worksheet.write('N477', '=H477-G477')
    worksheet.write('J486', '=D476')
    worksheet.write('K486', '=E476')
    worksheet.write('L486', '=F476')
    worksheet.write('M486', '=G476')
    worksheet.write('N486', '=H476')
    worksheet.write('J487',
                    '=(INDIRECT("J" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('K487',
                    '=(INDIRECT("K" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('L487',
                    '=(INDIRECT("L" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('M487',
                    '=(INDIRECT("M" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('N487',
                    '=(INDIRECT("N" & MATCH("Income Tax",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Pretax Income",B145:B403,0) +144))')
    worksheet.write('Q461', '=D476')
    worksheet.write('R461', '=E476')
    worksheet.write('S461', '=F476')
    worksheet.write('T461', '=G476')
    worksheet.write('U461', '=H476')
    worksheet.write('Q462', '=X455-X463-X471')
    worksheet.write('R462', '=Y455-Y463-Y471')
    worksheet.write('S462', '=Z455-Z463-Z471')
    worksheet.write('T462', '=AA455-AA463-AA471')
    worksheet.write('U462', '=AB455-AB463-AB471')
    worksheet.write('Q471', '=K476')
    worksheet.write('R471', '=L476')
    worksheet.write('S471', '=M476')
    worksheet.write('T471', '=N476')
    worksheet.write('Q472', '=K467-J467')
    worksheet.write('R472', '=L467-K467')
    worksheet.write('S472', '=M467-L467')
    worksheet.write('T472', '=N467-M467')
    worksheet.write('Q480', '=K476')
    worksheet.write('R480', '=L476')
    worksheet.write('S480', '=M476')
    worksheet.write('T480', '=N476')
    worksheet.write('Q481', '=K487-J487')
    worksheet.write('R481', '=L487-K487')
    worksheet.write('S481', '=M487-L487')
    worksheet.write('T481', '=N487-M487')
    worksheet.write('Q489', '=D476')
    worksheet.write('R489', '=E476')
    worksheet.write('S489', '=F476')
    worksheet.write('T489', '=G476')
    worksheet.write('U489', '=H476')
    worksheet.write('Q490', '=SUM(X483,X491,X499)')
    worksheet.write('R490', '=SUM(Y483,Y491,Y499)')
    worksheet.write('S490', '=SUM(Z483,Z491,Z499)')
    worksheet.write('T490', '=SUM(AA483,AA491,AA499)')
    worksheet.write('U490', '=SUM(AB483,AB491,AB499)')
    worksheet.write('AE462', '=K476')
    worksheet.write('AF462', '=L476')
    worksheet.write('AG462', '=M476')
    worksheet.write('AH462', '=N476')
    worksheet.write('AE463', '=R462-Q462')
    worksheet.write('AF463', '=S462-R462')
    worksheet.write('AG463', '=T462-S462')
    worksheet.write('AH463', '=U462-T462')
    worksheet.write('AE491', '=K476')
    worksheet.write('AF491', '=L476')
    worksheet.write('AG491', '=M476')
    worksheet.write('AH491', '=N476')
    worksheet.write('AE492', '=R490-Q490')
    worksheet.write('AF492', '=S490-R490')
    worksheet.write('AG492', '=T490-S490')
    worksheet.write('AH492', '=U490-T490')
    worksheet.write('X454', '=D476')
    worksheet.write('Y454', '=E476')
    worksheet.write('Z454', '=F476')
    worksheet.write('AA454', '=G476')


    # BUG FIX
    #=IFERROR((INDIRECT("N" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144)), (1 - (INDIRECT("N" & MATCH("Cost of Goods Sold (COGS) incl. D&A",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))))
    # Uses 1 - Cost of Goods Sold / Sales if Gross Income is unavailable, and Operating Income / Sales otherwise

    worksheet.write('AB454', '=H476')
    worksheet.write('X455',
                    '=IFERROR((INDIRECT("J" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("J" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("J" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('Y455',
                    '=IFERROR((INDIRECT("K" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("K" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("K" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('Z455',
                    '=IFERROR((INDIRECT("L" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("L" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("L" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('AA455',
                    '=IFERROR((INDIRECT("M" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("M" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("M" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('AB455',
                    '=IFERROR((INDIRECT("N" & MATCH("Gross Income",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144)), IFERROR((1 - (INDIRECT("N" & MATCH("Cost of Goods Sold*",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))),(INDIRECT("N" & MATCH("Operating Income",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))))')
    worksheet.write('X462', '=D476')
    worksheet.write('Y462', '=E476')
    worksheet.write('Z462', '=F476')
    worksheet.write('AA462', '=G476')
    worksheet.write('AB462', '=H476')
    worksheet.write('X463',
                    '=(INDIRECT("J" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y463',
                    '=(INDIRECT("K" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z463',
                    '=(INDIRECT("L" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA463',
                    '=(INDIRECT("M" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB463',
                    '=(INDIRECT("N" & MATCH("SG&A Expense",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('X470', '=D476')
    worksheet.write('Y470', '=E476')
    worksheet.write('Z470', '=F476')
    worksheet.write('AA470', '=G476')
    worksheet.write('AB470', '=H476')
    worksheet.write('X471',
                    '=(INDIRECT("J" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y471',
                    '=(INDIRECT("K" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z471',
                    '=(INDIRECT("L" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA471',
                    '=(INDIRECT("M" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB471',
                    '=(INDIRECT("N" & MATCH("Depreciation & Amortization Expense",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('X482', '=D476')
    worksheet.write('Y482', '=E476')
    worksheet.write('Z482', '=F476')
    worksheet.write('AA482', '=G476')
    worksheet.write('AB482', '=H476')
    worksheet.write('X483',
                    '=(INDIRECT("J" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("J" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y483',
                    '=(INDIRECT("K" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("K" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z483',
                    '=(INDIRECT("L" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("L" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA483',
                    '=(INDIRECT("M" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("M" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB483',
                    '=(INDIRECT("N" & MATCH("Total Current Assets",B145:B403,0) +144) - INDIRECT("N" & MATCH("Total Current Liabilities",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('X490', '=D476')
    worksheet.write('Y490', '=E476')
    worksheet.write('Z490', '=F476')
    worksheet.write('AA490', '=G476')
    worksheet.write('AB490', '=H476')
    worksheet.write('X491',
                    '=(INDIRECT("J" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y491',
                    '=(INDIRECT("K" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z491',
                    '=(INDIRECT("L" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA491',
                    '=(INDIRECT("M" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB491',
                    '=(INDIRECT("N" & MATCH("Net Property, Plant & Equipment",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('X498', '=D476')
    worksheet.write('Y498', '=E476')
    worksheet.write('Z498', '=F476')
    worksheet.write('AA498', '=G476')
    worksheet.write('AB498', '=H476')
    worksheet.write('X499',
                    '=(INDIRECT("J" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("J" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Y499',
                    '=(INDIRECT("K" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("K" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('Z499',
                    '=(INDIRECT("L" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("L" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AA499',
                    '=(INDIRECT("M" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("M" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AB499',
                    '=(INDIRECT("N" & MATCH("Intangible Assets",B145:B403,0) +144))/(INDIRECT("N" & MATCH("Sales/Revenue",B145:B403,0) +144))')
    worksheet.write('AK453', '=K476')
    worksheet.write('AL453', '=L476')
    worksheet.write('AM453', '=M476')
    worksheet.write('AN453', '=N476')
    worksheet.write('AK454', '=Y455-X455')
    worksheet.write('AL454', '=Z455-Y455')
    worksheet.write('AM454', '=AA455-Z455')
    worksheet.write('AN454', '=AB455-AA455')
    worksheet.write('AK462', '=K476')
    worksheet.write('AL462', '=L476')
    worksheet.write('AM462', '=M476')
    worksheet.write('AN462', '=N476')
    worksheet.write('AK463', '=Y463-X463')
    worksheet.write('AL463', '=Z463-Y463')
    worksheet.write('AM463', '=AA463-Z463')
    worksheet.write('AN463', '=AB463-AA463')
    worksheet.write('AK473', '=K476')
    worksheet.write('AL473', '=L476')
    worksheet.write('AM473', '=M476')
    worksheet.write('AN473', '=N476')
    worksheet.write('AK474', '=Y471-X471')
    worksheet.write('AL474', '=Z471-Y471')
    worksheet.write('AM474', '=AA471-Z471')
    worksheet.write('AN474', '=AB471-AA471')
    worksheet.write('AK483', '=K476')
    worksheet.write('AL483', '=L476')
    worksheet.write('AM483', '=M476')
    worksheet.write('AN483', '=N476')
    worksheet.write('AK484', '=Y483-X483')
    worksheet.write('AL484', '=Z483-Y483')
    worksheet.write('AM484', '=AA483-Z483')
    worksheet.write('AN484', '=AB483-AA483')
    worksheet.write('AK491', '=K476')
    worksheet.write('AL491', '=L476')
    worksheet.write('AM491', '=M476')
    worksheet.write('AN491', '=N476')
    worksheet.write('AK492', '=Y491-X491')
    worksheet.write('AL492', '=Z491-Y491')
    worksheet.write('AM492', '=AA491-Z491')
    worksheet.write('AN492', '=AB491-AA491')
    worksheet.write('AK499', '=K476')
    worksheet.write('AL499', '=L476')
    worksheet.write('AM499', '=M476')
    worksheet.write('AN499', '=N476')
    worksheet.write('AK500', '=Y499-X499')
    worksheet.write('AL500', '=Z499-Y499')
    worksheet.write('AM500', '=AA499-Z499')
    worksheet.write('AN500', '=AB499-AA499')
    worksheet.write('D475', '="EOY ROIC"')
    worksheet.write('J465', '="EOY Pretax ROIC"')
    worksheet.write('K475', '="Change in EOY ROIC"')
    worksheet.write('J485', '="Cash Tax Rate"')
    worksheet.write('Q460', '="Operating Margin"')
    worksheet.write('Q470', '="Change in EOY Pretax ROIC"')
    worksheet.write('Q479', '="Change in Cash Tax Rate"')
    worksheet.write('Q488', '="Invested Capital / Sales"')
    worksheet.write('X453', '="Gross Margin"')
    worksheet.write('X461', '="SGA / Sales"')
    worksheet.write('X469', '="Depreciation / Sales"')
    worksheet.write('X481', '="Op WC / Sales"')
    worksheet.write('X489', '="PPE / Sales"')
    worksheet.write('X497', '="Intangibles / Sales"')
    worksheet.write('AE461', '="Change in Operating Margin"')
    worksheet.write('AE490', '="Change in Invested Capital / Sales"')
    worksheet.write('AK452', '="Change in Gross Margin / Sales"')
    worksheet.write('AK461', '="Change in SGA / Sales"')
    worksheet.write('AK472', '="Change in Depreciation / Sales"')
    worksheet.write('AK482', '="Change in Op WC / Sales"')
    worksheet.write('AK490', '="Change in PPE / Sales"')
    worksheet.write('AK498', '="Change in Intagibles / Sales"')

    #END SUPERTREE --------------------------------------------------------------------------------------------------------------------------------------------------


    #streamline this
    worksheet.write('K12','=D78')
    worksheet.write('K13','=D89')


    #additons
    #=IF(STDEV.P(J146:N146)<0.1,IF(COUNTIF(J146:N146,">0")=5,"pos_trend"),"") where J146:N146 is a range of annual data
        # checks if the data isn't deviating too much, and if it isn't, whether there is a positive trend
worksheet = workbook.get_worksheet_by_name("Output Text")
n_row = 1
for ticker in tickers:
    worksheet.write('B' + str(n_row), '=' + ticker + '!K17')
    n_row += 1

workbook.close()

writer = pd.ExcelWriter('Output.xlsx', engine='openpyxl')
wb = load_workbook('Output.xlsx')

# create pandas excel writer for dfs
writer.book = wb

# have to tell pandas that we alerady have sheets, and what they are
writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

# insert dataframes

x = -1
offset = 0 #make global
for ticker in tickers:
    offset = 6
    x += 1

    try:
        df = pd.DataFrame(data_dict[tickers[x]][2][1][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1

        df = pd.DataFrame(data_dict[tickers[x]][2][1][1])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 2
    except:
        pass

    try:
        df = pd.DataFrame(data_dict[tickers[x]][4])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][4])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][5])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][5])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][6])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][6])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][7])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][7])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][8])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][8])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][9])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][9])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][12][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][12][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][14][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][14][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][16][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][16][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][18][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][18][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][20][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][20][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][22][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][22][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][24][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][24][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][26][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][26][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][28][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 1
    except:
        print "pd.DataFrame(data_dict[tickers[x]][28][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][30][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
        offset += len(df.index) + 2
    except:
        print "pd.DataFrame(data_dict[tickers[x]][30][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][32][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][32][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][38][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][38][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][39][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][39][0])"

    try:
        df = pd.DataFrame(data_dict[tickers[x]][43][0])
        df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
        offset += len(df.index) + 2
    except:
        print ticker, "pd.DataFrame(data_dict[tickers[x]][43][0])"

    #for the 2 asset tables issue
    try:
        if is_df(data_dict[tickers[x]][44][0]):
            try:
                df = pd.DataFrame(data_dict[tickers[x]][44][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][44][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][46][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][46][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][50][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][50][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][52][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][52][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][54][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][54][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][55][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][56][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][57][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][57][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][59][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][59][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][61][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][61][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][63][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][63][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][65][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][65][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][67][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][67][0])"

        else:
            try:
                df = pd.DataFrame(data_dict[tickers[x]][45][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][45][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][47][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][47][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][51][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print ticker, "pd.DataFrame(data_dict[tickers[x]][51][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][53][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=True)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][53][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][54][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][54][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][56][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][56][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][58][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][58][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][60][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][60][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][62][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][62][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][64][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][64][0])"

            try:
                df = pd.DataFrame(data_dict[tickers[x]][66][0])
                df.to_excel(writer, sheet_name=ticker, startrow=offset, header=False)
                offset += len(df.index) + 2
            except:
                print "pd.DataFrame(data_dict[tickers[x]][66][0])"
    except:
        print "index out of range"



writer.save()

print "Program took", (time.time() - start_time)/60 , "minutes to run"




#Add T to number converter





